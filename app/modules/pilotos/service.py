import math
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models import Pilotos, Usuario
from app.shared.formatters import format_phone_br, only_digits


REGIOES = {"NORTE", "SUL", "LESTE", "OESTE", "CENTRO", "SULDESTE", "CENTRO-OESTE"}


def validate_piloto_data(nome_piloto: str, regiao: str, telefone: str):
    errors = {}
    telefone_digits = only_digits(telefone)

    if not nome_piloto:
        errors["nome_piloto"] = "Informe o nome do piloto."

    if regiao and regiao not in REGIOES:
        errors["regiao"] = "Selecione uma regiao valida."

    if telefone and len(telefone_digits) not in (10, 11):
        errors["telefone"] = "Telefone deve ter 10 ou 11 digitos (com DDD)."

    return errors, telefone_digits


def piloto_duplicado(nome_piloto: str, telefone_digits: str, exclude_id=None):
    if not nome_piloto:
        return None

    query = Pilotos.query.filter(db.func.lower(Pilotos.nome_piloto) == nome_piloto.lower())
    if telefone_digits:
        query = query.filter(Pilotos.telefone == telefone_digits)
    if exclude_id is not None:
        query = query.filter(Pilotos.id != exclude_id)
    return query.first()


def login_em_uso(login: str, exclude_user_id=None):
    if not login:
        return None

    query = Usuario.query.filter(db.func.lower(Usuario.login) == login.lower())
    if exclude_user_id is not None:
        query = query.filter(Usuario.id != exclude_user_id)
    return query.first()


def build_pilotos_query(user_tipo: str, regiao: str, telefone: str, q: str, sort: str):
    query = Pilotos.query

    if regiao:
        query = query.filter(Pilotos.regiao == regiao)

    if telefone:
        query = query.filter(Pilotos.telefone.ilike(f"%{only_digits(telefone)}%"))

    if q:
        like = f"%{q}%"
        q_digits = only_digits(q)
        query = query.filter(
            db.or_(
                Pilotos.nome_piloto.ilike(like),
                Pilotos.regiao.ilike(like),
                Pilotos.telefone.ilike(f"%{q_digits}%") if q_digits else db.false(),
            )
        )

    if sort == "nome_desc":
        query = query.order_by(Pilotos.nome_piloto.desc())
    elif sort == "id_desc":
        query = query.order_by(Pilotos.id.desc())
    elif sort == "id_asc":
        query = query.order_by(Pilotos.id.asc())
    else:
        query = query.order_by(Pilotos.nome_piloto.asc())

    return query


def build_pilotos_filters(q, regiao, telefone, sort, page, per_page, total, total_pages):
    return {
        "q": q,
        "regiao": regiao,
        "telefone": telefone,
        "sort": sort,
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
    }


def serialize_pilotos(rows):
    return [
        {
            "id": piloto.id,
            "nome_piloto": piloto.nome_piloto,
            "regiao": piloto.regiao or "-",
            "telefone_fmt": format_phone_br(piloto.telefone or "") or "-",
            "telefone_digits": only_digits(piloto.telefone or ""),
        }
        for piloto in rows
    ]


def build_pilotos_export(rows, user_tipo: str, uvis_regiao: str):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pilotos"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet["A1"] = "Relatorio de Pilotos"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sheet["A2"].font = Font(color="6B7280")

    if user_tipo in {"uvis", "regional"}:
        sheet["A3"] = f"Regiao do usuario: {uvis_regiao}"
        sheet["A3"].font = Font(color="6B7280")

    start_row = 5 if user_tipo in {"uvis", "regional"} else 4
    headers = ["ID", "Nome", "Regiao", "Telefone"]

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for row_idx, piloto in enumerate(rows, start=start_row + 1):
        values = [
            piloto.id,
            piloto.nome_piloto,
            piloto.regiao or "",
            format_phone_br(piloto.telefone or "") or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = center_align if col_idx == 1 else text_align
            if col_idx == 4:
                cell.number_format = "@"

    last_row = start_row + len(rows)
    last_col = len(headers)
    sheet.freeze_panes = sheet[f"A{start_row + 1}"]
    sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(last_col)}{max(last_row, start_row)}"

    max_widths = {1: 8, 2: 30, 3: 14, 4: 20}
    for col_idx in range(1, last_col + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(start_row + 1, last_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_widths.get(col_idx, 35))

    zebra_fill = PatternFill("solid", fgColor="F9FAFB")
    for row_idx in range(start_row + 1, last_row + 1):
        if (row_idx - (start_row + 1)) % 2 == 1:
            for col_idx in range(1, last_col + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = zebra_fill

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"pilotos_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    return output, filename


def normalize_page(value, default=1):
    try:
        return max(1, int(value or default))
    except ValueError:
        return default


def normalize_per_page(value, default=20):
    try:
        per_page = int(value or default)
    except ValueError:
        per_page = default
    return 10 if per_page < 10 else 50 if per_page > 50 else per_page


def normalize_pagination(total, page, per_page):
    total_pages = max(1, math.ceil(total / per_page))
    if page > total_pages:
        page = total_pages
    return page, total_pages
