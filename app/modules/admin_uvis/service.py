from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import or_

from app.extensions import db
from app.models import EquipeUvis, Notificacao, PilotoUvis, Usuario
from app.shared.access import ADMIN_PANEL_VIEW_TYPES, apply_prefeitura_scope, apply_regiao_scope
from app.shared.query_filters import id_search_clause
from app.shared.access import (
    ADMIN_PANEL_VIEW_TYPES,
    GLOBAL_ADMIN_USER_TYPES,
    apply_prefeitura_scope,
    apply_regiao_scope,
)


def can_access_admin_uvis(user) -> bool:
    return getattr(user, "tipo_usuario", None) in ADMIN_PANEL_VIEW_TYPES


def is_admin_user(user) -> bool:
    return getattr(user, "tipo_usuario", None) in GLOBAL_ADMIN_USER_TYPES


def is_admin_or_prefeitura_admin(user) -> bool:
    return getattr(user, "tipo_usuario", None) in GLOBAL_ADMIN_USER_TYPES | {"prefeitura_admin"}


def is_uvis_user(user) -> bool:
    return getattr(user, "tipo_usuario", None) == "uvis"


def login_em_uso(login: str, exclude_user_id=None):
    if not login:
        return None

    query = Usuario.query.filter(Usuario.login == login)
    if exclude_user_id is not None:
        query = query.filter(Usuario.id != exclude_user_id)
    return query.first()


def build_uvis_query(user, q: str, regiao: str, codigo_setor: str, prefeitura_id=None):
    query = Usuario.query.filter(Usuario.tipo_usuario == "uvis")
    query = apply_prefeitura_scope(query, user, Usuario.prefeitura_id)
    query = apply_regiao_scope(query, user, Usuario.regiao)

    if prefeitura_id:
        query = query.filter(Usuario.prefeitura_id == prefeitura_id)

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                id_search_clause(Usuario.id, q),
                Usuario.nome_uvis.ilike(like),
                Usuario.login.ilike(like),
            )
        )

    if regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{regiao}%"))

    if codigo_setor:
        query = query.filter(Usuario.codigo_setor.ilike(f"%{codigo_setor}%"))

    return query.order_by(Usuario.nome_uvis.asc())


def validate_new_uvis(nome_uvis: str, login: str, senha: str, confirmar: str):
    if not nome_uvis or not login or not senha:
        return "warning", "Preencha: Nome da UVIS, Login e Senha."

    if senha != confirmar:
        return "warning", "As senhas nao conferem."

    if login_em_uso(login):
        return "danger", "Esse login ja esta em uso. Escolha outro."

    return None, None


def validate_edit_uvis(nome_uvis: str, login: str, senha: str, confirmar: str, uvis_id: int):
    if not nome_uvis or not login:
        return "warning", "Preencha: Nome da UVIS e Login."

    if senha and senha != confirmar:
        return "warning", "As senhas nao conferem."

    if login_em_uso(login, exclude_user_id=uvis_id):
        return "danger", "Esse login ja esta em uso. Escolha outro."

    return None, None


def delete_uvis_user(uvis):
    team_account_ids = [
        user_id
        for (user_id,) in db.session.query(Usuario.id)
        .filter(Usuario.equipe_uvis_uvis_usuario_id == uvis.id)
        .all()
    ]

    if team_account_ids:
        Notificacao.query.filter(Notificacao.usuario_id.in_(team_account_ids)).delete(
            synchronize_session=False
        )
        Usuario.query.filter(Usuario.id.in_(team_account_ids)).delete(synchronize_session=False)

    Notificacao.query.filter(Notificacao.usuario_id == uvis.id).delete(synchronize_session=False)
    EquipeUvis.query.filter(EquipeUvis.uvis_usuario_id == uvis.id).delete(synchronize_session=False)
    PilotoUvis.query.filter(PilotoUvis.uvis_usuario_id == uvis.id).delete(synchronize_session=False)
    db.session.delete(uvis)


def build_uvis_export(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "UVIS"

    title_font = Font(bold=True, size=14)
    meta_font = Font(size=10, color="666666")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    zebra_fill = PatternFill("solid", fgColor="F3F6FA")

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    ws["A1"] = "UVIS Cadastradas"
    ws["A1"].font = title_font

    ws["A3"] = f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = meta_font

    start_header_row = 5
    headers = ["ID", "Nome", "Regiao", "Login"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    start_data_row = start_header_row + 1
    for index, uvis in enumerate(rows):
        row_number = start_data_row + index
        values = [uvis.id, uvis.nome_uvis, uvis.regiao, uvis.login]

        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_number, column=column, value=value)
            cell.border = border
            cell.alignment = center if column == 1 else left

            if index % 2 == 1:
                cell.fill = zebra_fill

    end_data_row = start_data_row + len(rows) - 1
    if rows:
        ws.auto_filter.ref = f"A{start_header_row}:D{end_data_row}"
        ws.freeze_panes = f"A{start_data_row}"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16

    total_row = end_data_row + 2
    ws.cell(row=total_row, column=1, value="Total de UVIS:").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(rows)).font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"uvis_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    return output, filename
