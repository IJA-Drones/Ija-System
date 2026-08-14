import hashlib
import json
import math
import os
import re
import unicodedata
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

from app.clients.google_maps_client import reverse_geocode_lat_lng_google
from app.extensions import db
from app.models import DjiFlightKmlRoute, DjiFlightLogImport, DjiFlightRecord, EquipePiloto, OrdemServico, Solicitacao
from app.shared.access import ADMIN_PANEL_VIEW_TYPES, can_access_regiao
from app.shared.place_id import clean_place_id, resolve_google_place_id_for_address
from app.shared.uploads import get_upload_folder


DJI_LOG_ALLOWED_VIEW_TYPES = {"dev", "diretor", "admin"}
DJI_LOG_ALLOWED_IMPORT_TYPES = {"dev", "diretor", "admin"}

_FLIGHT_WINDOW_RE = re.compile(
    r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})-(\d{2}:\d{2}:\d{2}|\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})$"
)

_HEADER_ALIASES = {
    "flight_time": {"flighttime"},
    "location": {"location"},
    "place_id": {"placeid", "googleplaceid"},
    "aircraft_name": {"aircraftname"},
    "task_type": {"tasktype"},
    "sprayed_area": {"sprayedarea"},
    "total_amount": {"totalamountlkg", "totalamount"},
    "flight_duration": {"flightdurationminsec", "flightduration"},
    "crop": {"crop"},
    "pilot_name": {"pliotname", "pilotname"},
    "team_name": {"teamname"},
    "field_name": {"fieldname"},
    "serial_number": {"serialnumber"},
    "starting_battery_level": {"startingbatterylevel"},
    "ending_battery_level": {"endingbatterylevel"},
    "battery_sn": {"batterysn"},
}


def can_access_dji_logs(user) -> bool:
    return getattr(user, "tipo_usuario", None) in DJI_LOG_ALLOWED_VIEW_TYPES


def can_import_dji_logs(user) -> bool:
    return getattr(user, "tipo_usuario", None) in DJI_LOG_ALLOWED_IMPORT_TYPES


def can_access_dji_kml_route(user, route_id) -> bool:
    if can_access_dji_logs(user):
        return True

    ordem = _get_linked_os_for_kml_route(route_id)
    if not ordem or not ordem.solicitacao:
        return False

    solicitacao = ordem.solicitacao
    tipo_usuario = getattr(user, "tipo_usuario", None)

    if tipo_usuario in ADMIN_PANEL_VIEW_TYPES:
        pedido_regiao = getattr(getattr(solicitacao, "usuario", None), "regiao", None)
        return can_access_regiao(user, pedido_regiao)

    if tipo_usuario == "equipe_oceano":
        try:
            equipe_id = int((getattr(user, "codigo_setor", None) or "").strip())
        except (TypeError, ValueError):
            return False
        return equipe_id == getattr(solicitacao, "equipe_id", None)

    if tipo_usuario == "piloto" and getattr(user, "piloto_id", None):
        return (
            EquipePiloto.query
            .filter(
                EquipePiloto.equipe_id == getattr(solicitacao, "equipe_id", None),
                EquipePiloto.piloto_id == user.piloto_id,
            )
            .first()
            is not None
        )

    return False


def import_dji_log_excel(file_storage, user):
    if not file_storage or not file_storage.filename:
        raise ValueError("Selecione um arquivo Excel para importar.")

    original_filename = (file_storage.filename or "").strip()
    if not original_filename.lower().endswith(".xlsx"):
        raise ValueError("O arquivo precisa estar no formato .xlsx.")

    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("O arquivo enviado esta vazio.")

    file_sha256 = hashlib.sha256(file_bytes).hexdigest()
    previous_import = (
        DjiFlightLogImport.query
        .filter(DjiFlightLogImport.file_sha256 == file_sha256)
        .order_by(DjiFlightLogImport.uploaded_at.desc())
        .first()
    )
    if previous_import:
        when = previous_import.uploaded_at.strftime("%d/%m/%Y %H:%M")
        raise ValueError(
            f"Esse mesmo arquivo ja foi importado em {when}. Envie um Excel novo da DJI."
        )

    workbook = _load_workbook(file_bytes)
    rows = _parse_workbook_rows(workbook)
    if not rows:
        raise ValueError("Nenhum voo valido foi encontrado no Excel informado.")

    stored_filename, stored_path = _save_uploaded_excel(original_filename, file_bytes)

    import_batch = DjiFlightLogImport(
        uploaded_by_id=getattr(user, "id", None),
        original_filename=original_filename,
        stored_filename=stored_filename,
        stored_path=stored_path,
        file_sha256=file_sha256,
        total_rows=len(rows),
        imported_rows=0,
        skipped_rows=0,
        period_start=min(item["flight_start"] for item in rows),
        period_end=max(item["flight_end"] for item in rows),
    )
    db.session.add(import_batch)
    db.session.flush()

    fingerprints = [item["fingerprint"] for item in rows]
    existing_fingerprints = {
        value
        for (value,) in db.session.query(DjiFlightRecord.fingerprint)
        .filter(DjiFlightRecord.fingerprint.in_(fingerprints))
        .all()
    }

    imported_rows = 0
    skipped_rows = 0

    for item in rows:
        if item["fingerprint"] in existing_fingerprints:
            skipped_rows += 1
            continue

        existing_fingerprints.add(item["fingerprint"])
        db.session.add(
            DjiFlightRecord(
                import_id=import_batch.id,
                source_row_number=item["source_row_number"],
                fingerprint=item["fingerprint"],
                flight_window=item["flight_window"],
                flight_start=item["flight_start"],
                flight_end=item["flight_end"],
                location=item["location"],
                place_id=item["place_id"],
                aircraft_name=item["aircraft_name"],
                task_type=item["task_type"],
                sprayed_area_ha=item["sprayed_area_ha"],
                total_amount_l_kg=item["total_amount_l_kg"],
                flight_duration_seconds=item["flight_duration_seconds"],
                flight_duration_label=item["flight_duration_label"],
                crop=item["crop"],
                pilot_name=item["pilot_name"],
                team_name=item["team_name"],
                field_name=item["field_name"],
                serial_number=item["serial_number"],
                starting_battery_level=item["starting_battery_level"],
                ending_battery_level=item["ending_battery_level"],
                battery_consumed_level=item["battery_consumed_level"],
                battery_sn=item["battery_sn"],
                raw_payload=json.dumps(item["raw_payload"], ensure_ascii=False),
            )
        )
        imported_rows += 1

    import_batch.imported_rows = imported_rows
    import_batch.skipped_rows = skipped_rows

    db.session.commit()
    return import_batch


def build_dji_logs_context(args):
    data_inicio = (args.get("data_inicio") or "").strip()
    data_fim = (args.get("data_fim") or "").strip()
    piloto = (args.get("piloto") or "").strip()
    aeronave = (args.get("aeronave") or "").strip()
    equipe = (args.get("equipe") or "").strip()
    q = (args.get("q") or "").strip()
    endereco = (args.get("endereco") or "").strip()
    voo_id = (args.get("voo_id") or "").strip()
    status_rota = (args.get("status_rota") or "").strip()
    kml_q = (args.get("kml_q") or "").strip()
    kml_data_inicio = (args.get("kml_data_inicio") or "").strip()
    kml_data_fim = (args.get("kml_data_fim") or "").strip()
    kml_endereco = (args.get("kml_endereco") or "").strip()
    kml_piloto = (args.get("kml_piloto") or "").strip()
    kml_aeronave = (args.get("kml_aeronave") or "").strip()
    kml_status_os = (args.get("kml_status_os") or "").strip()
    kml_status_voo = (args.get("kml_status_voo") or "").strip()
    kml_voo_id = (args.get("kml_voo_id") or "").strip()
    page = args.get("page", 1, type=int)
    kml_page = args.get("kml_page", 1, type=int)

    filtered_query = _build_filtered_query(
        data_inicio=data_inicio,
        data_fim=data_fim,
        piloto=piloto,
        aeronave=aeronave,
        equipe=equipe,
        q=q,
        endereco=endereco,
        voo_id=voo_id,
        status_rota=status_rota,
    )

    paginacao = (
        filtered_query
        .order_by(DjiFlightRecord.flight_start.desc(), DjiFlightRecord.id.desc())
        .paginate(page=page, per_page=20, error_out=False)
    )
    registros = paginacao.items
    registro_kml_os_por_rota = _build_kml_os_map(
        [registro.route_kml for registro in registros if registro.route_kml]
    )

    totals = filtered_query.with_entities(
        func.count(DjiFlightRecord.id),
        func.coalesce(func.sum(DjiFlightRecord.sprayed_area_ha), 0.0),
        func.coalesce(func.sum(DjiFlightRecord.total_amount_l_kg), 0.0),
        func.coalesce(func.sum(DjiFlightRecord.flight_duration_seconds), 0),
        func.avg(DjiFlightRecord.battery_consumed_level),
    ).first()

    total_voos, total_area, total_volume, total_duracao, media_bateria = totals

    pilotos_disponiveis = _distinct_non_empty_values(DjiFlightRecord.pilot_name)
    aeronaves_disponiveis = _distinct_non_empty_values(DjiFlightRecord.aircraft_name)
    equipes_disponiveis = _distinct_non_empty_values(DjiFlightRecord.team_name)
    kml_pilotos_disponiveis = _distinct_non_empty_values(DjiFlightKmlRoute.pilot_name)
    kml_aeronaves_disponiveis = _distinct_non_empty_values(DjiFlightKmlRoute.aircraft_name)

    total_pilotos = filtered_query.with_entities(
        func.count(func.distinct(func.nullif(DjiFlightRecord.pilot_name, "")))
    ).scalar() or 0
    total_aeronaves = filtered_query.with_entities(
        func.count(func.distinct(func.nullif(DjiFlightRecord.aircraft_name, "")))
    ).scalar() or 0

    weekly_rows = filtered_query.with_entities(
        DjiFlightRecord.flight_start,
        DjiFlightRecord.sprayed_area_ha,
        DjiFlightRecord.total_amount_l_kg,
        DjiFlightRecord.flight_duration_seconds,
    ).all()
    monthly_rows = filtered_query.with_entities(
        DjiFlightRecord.flight_start,
        DjiFlightRecord.sprayed_area_ha,
        DjiFlightRecord.total_amount_l_kg,
        DjiFlightRecord.flight_duration_seconds,
    ).all()
    resumo_mensal = _build_monthly_summary(monthly_rows)
    comparativo_mensal = _build_monthly_comparison(resumo_mensal)

    importacoes_recentes = (
        DjiFlightLogImport.query
        .order_by(DjiFlightLogImport.uploaded_at.desc(), DjiFlightLogImport.id.desc())
        .limit(10)
        .all()
    )
    total_rotas_kml = DjiFlightKmlRoute.query.count()
    total_rotas_kml_vinculadas = (
        db.session.query(func.count(func.distinct(OrdemServico.dji_kml_route_id)))
        .filter(OrdemServico.dji_kml_route_id.isnot(None))
        .scalar()
        or 0
    )
    kml_query = _build_filtered_kml_query(
        q=kml_q,
        data_inicio=kml_data_inicio,
        data_fim=kml_data_fim,
        endereco=kml_endereco,
        piloto=kml_piloto,
        aeronave=kml_aeronave,
        status_os=kml_status_os,
        status_voo=kml_status_voo,
        voo_id=kml_voo_id,
    )
    kml_paginacao = (
        kml_query
        .order_by(DjiFlightKmlRoute.imported_at.desc(), DjiFlightKmlRoute.id.desc())
        .paginate(page=kml_page, per_page=25, error_out=False)
    )
    kml_rotas_recentes = kml_paginacao.items
    kml_os_por_rota = _build_kml_os_map(kml_rotas_recentes)

    return {
        "registros": registros,
        "paginacao": paginacao,
        "registro_kml_os_por_rota": registro_kml_os_por_rota,
        "filtros": {
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "piloto": piloto,
            "aeronave": aeronave,
            "equipe": equipe,
            "q": q,
            "endereco": endereco,
            "voo_id": voo_id,
            "status_rota": status_rota,
            "total": total_voos or 0,
        },
        "pilotos_disponiveis": pilotos_disponiveis,
        "aeronaves_disponiveis": aeronaves_disponiveis,
        "equipes_disponiveis": equipes_disponiveis,
        "kml_pilotos_disponiveis": kml_pilotos_disponiveis,
        "kml_aeronaves_disponiveis": kml_aeronaves_disponiveis,
        "kml_filtros": {
            "q": kml_q,
            "data_inicio": kml_data_inicio,
            "data_fim": kml_data_fim,
            "endereco": kml_endereco,
            "piloto": kml_piloto,
            "aeronave": kml_aeronave,
            "status_os": kml_status_os,
            "status_voo": kml_status_voo,
            "voo_id": kml_voo_id,
        },
        "importacoes_recentes": importacoes_recentes,
        "kml_rotas_recentes": kml_rotas_recentes,
        "kml_paginacao": kml_paginacao,
        "kml_os_por_rota": kml_os_por_rota,
        "total_importacoes": DjiFlightLogImport.query.count(),
        "total_rotas_kml": total_rotas_kml,
        "total_rotas_kml_vinculadas": total_rotas_kml_vinculadas,
        "total_rotas_kml_sem_os": total_rotas_kml - total_rotas_kml_vinculadas,
        "total_voos": total_voos or 0,
        "total_area": float(total_area or 0),
        "total_volume": float(total_volume or 0),
        "total_duracao": format_duration_seconds(total_duracao or 0),
        "total_pilotos": total_pilotos,
        "total_aeronaves": total_aeronaves,
        "media_bateria": round(media_bateria or 0, 1) if media_bateria is not None else None,
        "top_pilotos": _build_top_groups(filtered_query, DjiFlightRecord.pilot_name),
        "top_aeronaves": _build_top_groups(filtered_query, DjiFlightRecord.aircraft_name),
        "top_equipes": _build_top_groups(filtered_query, DjiFlightRecord.team_name),
        "resumo_semanal": _build_weekly_summary(weekly_rows),
        "resumo_mensal": resumo_mensal,
        "comparativo_mensal": comparativo_mensal,
    }


def import_dji_kml_files(files, user):
    valid_files = [file for file in (files or []) if file and file.filename]
    if not valid_files:
        raise ValueError("Selecione ao menos um arquivo KML para importar.")

    imported = 0
    skipped = 0
    linked = 0
    os_linked = 0

    for file_storage in valid_files:
        original_filename = (file_storage.filename or "").strip()
        if not original_filename.lower().endswith(".kml"):
            raise ValueError("Todos os arquivos enviados devem estar no formato .kml.")

        file_bytes = file_storage.read()
        if not file_bytes:
            continue

        file_sha256 = hashlib.sha256(file_bytes).hexdigest()
        if DjiFlightKmlRoute.query.filter(DjiFlightKmlRoute.file_sha256 == file_sha256).first():
            skipped += 1
            continue

        parsed = _parse_kml_payload(file_bytes, original_filename)
        if DjiFlightKmlRoute.query.filter(DjiFlightKmlRoute.route_code == parsed["route_code"]).first():
            skipped += 1
            continue

        stored_filename, stored_path = _save_uploaded_kml(original_filename, file_bytes)
        matched_record = (
            DjiFlightRecord.query
            .filter(DjiFlightRecord.serial_number == parsed["route_code"])
            .order_by(DjiFlightRecord.flight_start.desc(), DjiFlightRecord.id.desc())
            .first()
        )

        reverse_geocode = _reverse_geocode_kml_route(parsed["points"]) if not parsed["place_id"] else {}
        route_place_id = parsed["place_id"] or reverse_geocode.get("place_id")

        route = DjiFlightKmlRoute(
            flight_record=matched_record,
            uploaded_by_id=getattr(user, "id", None),
            route_code=parsed["route_code"],
            original_filename=original_filename,
            stored_filename=stored_filename,
            stored_path=stored_path,
            file_sha256=file_sha256,
            aircraft_name=parsed["aircraft_name"],
            pilot_name=parsed["pilot_name"],
            flight_controller_id=parsed["flight_controller_id"],
            route_timestamp=parsed["route_timestamp"],
            place_id=route_place_id,
            mode_selection=parsed["mode_selection"],
            flight_time_raw=parsed["flight_time_raw"],
            task_area=parsed["task_area"],
            spray_amount=parsed["spray_amount"],
            route_color=parsed["route_color"],
            route_width=parsed["route_width"],
            point_count=len(parsed["points"]),
            points_json=json.dumps(parsed["points"], ensure_ascii=False),
        )
        route.formatted_address = reverse_geocode.get("formatted_address")
        db.session.add(route)
        db.session.flush()

        imported += 1
        if matched_record:
            linked += 1
        if _auto_link_kml_route_to_os(route, parsed["points"]):
            os_linked += 1

    db.session.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "linked": linked,
        "os_linked": os_linked,
        "unlinked": imported - linked,
    }


def _auto_link_kml_route_to_os(route, points):
    match = _find_best_os_match_for_kml_route(route, points)
    if not match:
        return None
    ordem, _score, _details = match
    ordem.dji_kml_route_id = route.id
    _fill_solicitacao_place_id_from_route(ordem, route)
    return ordem


def _route_source_place_id(route):
    flight_record = getattr(route, "flight_record", None)
    return (
        clean_place_id(getattr(route, "place_id", None))
        or clean_place_id(getattr(flight_record, "place_id", None))
        or clean_place_id(
            _extract_place_id_from_raw_payload(getattr(flight_record, "raw_payload", None))
        )
    )


def _fill_solicitacao_place_id_from_route(ordem, route):
    solicitacao = getattr(ordem, "solicitacao", None)
    if not solicitacao or clean_place_id(solicitacao.place_id):
        return

    solicitacao.place_id = resolve_google_place_id_for_address(
        place_id=_route_source_place_id(route),
        cep=solicitacao.cep,
        logradouro=solicitacao.logradouro,
        numero=solicitacao.numero,
        bairro=solicitacao.bairro,
        cidade=solicitacao.cidade,
        uf=solicitacao.uf,
    )


def link_kml_route_to_os_by_solicitacao_id(route_id, solicitacao_id):
    route = DjiFlightKmlRoute.query.get(route_id)
    if not route:
        raise ValueError("Rota KML nao encontrada.")

    ordem = (
        OrdemServico.query
        .options(joinedload(OrdemServico.solicitacao))
        .filter(OrdemServico.solicitacao_id == solicitacao_id)
        .first()
    )
    if not ordem:
        raise ValueError("Nao existe OS de piloto para essa solicitacao.")

    if ordem.dji_kml_route_id and ordem.dji_kml_route_id != route.id:
        raise ValueError("Essa OS ja possui outra rota KML vinculada.")

    current_linked = (
        OrdemServico.query
        .filter(
            OrdemServico.dji_kml_route_id == route.id,
            OrdemServico.id != ordem.id,
        )
        .all()
    )
    for linked_ordem in current_linked:
        linked_ordem.dji_kml_route_id = None

    ordem.dji_kml_route_id = route.id
    _fill_solicitacao_place_id_from_route(ordem, route)
    db.session.commit()
    return ordem


def delete_kml_route(route_id):
    route = DjiFlightKmlRoute.query.get(route_id)
    if not route:
        raise ValueError("Rota KML nao encontrada.")

    stored_path = route.stored_path
    original_filename = route.original_filename
    linked_count = (
        OrdemServico.query
        .filter(OrdemServico.dji_kml_route_id == route.id)
        .update({OrdemServico.dji_kml_route_id: None}, synchronize_session=False)
    )

    db.session.delete(route)
    db.session.commit()

    file_removed = False
    if stored_path:
        absolute_path = os.path.join(get_upload_folder(), stored_path)
        if os.path.exists(absolute_path):
            try:
                os.remove(absolute_path)
                file_removed = True
            except OSError:
                file_removed = False

    return {
        "original_filename": original_filename,
        "linked_count": linked_count or 0,
        "file_removed": file_removed,
    }


def _extract_scalar_query_value(row):
    try:
        value = row[0]
    except (TypeError, KeyError, IndexError):
        value = row
    if isinstance(value, tuple):
        return value[0] if value else None
    return value


def auto_link_existing_kml_routes_to_os(
    *,
    limit=None,
    route_id_min=None,
    route_id_max=None,
    commit=True,
    resolve_missing_place_id=False,
    progress_callback=None,
):
    linked_route_rows = (
        db.session.query(OrdemServico.dji_kml_route_id)
        .filter(OrdemServico.dji_kml_route_id.isnot(None))
        .all()
    )
    linked_route_ids = {
        route_id
        for route_id in (_extract_scalar_query_value(row) for row in linked_route_rows)
        if route_id is not None
    }
    query = (
        db.session.query(DjiFlightKmlRoute)
        .options(joinedload(DjiFlightKmlRoute.flight_record))
        .order_by(DjiFlightKmlRoute.route_timestamp.desc().nullslast(), DjiFlightKmlRoute.id.desc())
    )
    if linked_route_ids:
        query = query.filter(~DjiFlightKmlRoute.id.in_(linked_route_ids))
    if route_id_min is not None:
        query = query.filter(DjiFlightKmlRoute.id >= route_id_min)
    if route_id_max is not None:
        query = query.filter(DjiFlightKmlRoute.id <= route_id_max)
    if limit:
        query = query.limit(limit)

    result = {
        "scanned": 0,
        "linked": 0,
        "no_match": 0,
        "errors": 0,
        "place_resolved": 0,
        "matches": [],
    }

    for route in query.all():
        result["scanned"] += 1
        try:
            points = json.loads(route.points_json or "[]")
        except (TypeError, ValueError, json.JSONDecodeError):
            result["errors"] += 1
            if progress_callback:
                progress_callback("error", route, None, None, None)
            continue

        if resolve_missing_place_id and not _clean_place_id(getattr(route, "place_id", None)):
            reverse_geocode = _reverse_geocode_kml_route(points)
            if reverse_geocode.get("place_id"):
                route.place_id = reverse_geocode["place_id"]
                result["place_resolved"] += 1
            route.formatted_address = reverse_geocode.get("formatted_address")

        match = _find_best_os_match_for_kml_route(route, points)
        if not match:
            result["no_match"] += 1
            if progress_callback:
                progress_callback("no_match", route, None, None, None)
            continue

        ordem, score, details = match
        if commit:
            ordem.dji_kml_route_id = route.id
        result["linked"] += 1
        result["matches"].append(
            {
                "route_id": route.id,
                "route_code": route.route_code,
                "ordem_id": ordem.id,
                "identificador_os": ordem.identificador_os,
                "score": score,
                "details": details,
            }
        )
        if progress_callback:
            progress_callback("linked", route, ordem, score, details)

    if commit:
        db.session.commit()

    return result


def _find_best_os_match_for_kml_route(route, points):
    candidates = _candidate_ordens_for_kml_route(route)
    scored = []
    for ordem in candidates:
        score, details = _score_os_kml_match(ordem, route, points)
        if _is_confident_os_kml_match(score, details):
            scored.append((ordem, score, details))

    if not scored:
        return None

    scored.sort(key=lambda item: item[1], reverse=True)
    best = scored[0]
    if len(scored) > 1 and (best[1] - scored[1][1]) < 15:
        return None
    return best


def _candidate_ordens_for_kml_route(route):
    query = (
        OrdemServico.query
        .join(Solicitacao, Solicitacao.id == OrdemServico.solicitacao_id)
        .options(joinedload(OrdemServico.solicitacao))
        .filter(OrdemServico.dji_kml_route_id.is_(None))
    )

    if route.route_timestamp:
        start_date = (route.route_timestamp - timedelta(days=2)).date()
        end_date = (route.route_timestamp + timedelta(days=2)).date()
        query = query.filter(
            or_(
                OrdemServico.data_aplicacao.between(start_date, end_date),
                Solicitacao.data_agendamento.between(start_date, end_date),
                func.date(OrdemServico.respondido_em).between(start_date, end_date),
            )
        )

    return (
        query
        .order_by(OrdemServico.respondido_em.desc(), OrdemServico.id.desc())
        .limit(300)
        .all()
    )


def _score_os_kml_match(ordem, route, points):
    time_score = _score_kml_os_time_match(ordem, getattr(route, "route_timestamp", None))
    aircraft_score = _score_kml_os_aircraft_match(ordem, getattr(route, "aircraft_name", None))
    pilot_score = _score_kml_os_pilot_match(ordem, getattr(route, "pilot_name", None))
    place_score = _score_kml_os_place_match(ordem, route)
    address_score = _score_kml_os_address_match(ordem, route)
    distance_meters = _kml_distance_to_solicitacao_meters(points, getattr(ordem, "solicitacao", None))
    geo_score = _score_kml_os_geo_match(distance_meters)

    score = time_score + aircraft_score + pilot_score + place_score + address_score + geo_score
    return score, {
        "time": time_score,
        "aircraft": aircraft_score,
        "pilot": pilot_score,
        "place": place_score,
        "address": address_score,
        "geo": geo_score,
        "distance_meters": distance_meters,
    }


def _is_confident_os_kml_match(score, details):
    if score < 70:
        return False
    has_exact_place = (
        details.get("place", 0) >= 65
        and (
            details.get("time", 0) >= 12
            or details.get("geo", 0) >= 5
            or details.get("aircraft", 0) >= 25
            or details.get("pilot", 0) >= 10
        )
    )
    has_address_and_time = (
        details.get("address", 0) >= 24
        and details.get("time", 0) >= 20
        and (details.get("geo", 0) >= 12 or details.get("aircraft", 0) >= 25)
    )
    has_address_and_location = details.get("address", 0) >= 24 and details.get("geo", 0) >= 22
    has_time_and_identity = details.get("time", 0) >= 20 and (
        details.get("aircraft", 0) >= 25 or details.get("geo", 0) >= 22
    )
    has_aircraft_and_location = details.get("aircraft", 0) >= 25 and details.get("geo", 0) >= 22
    return (
        has_exact_place
        or has_address_and_time
        or has_address_and_location
        or has_time_and_identity
        or has_aircraft_and_location
    )


def _score_kml_os_time_match(ordem, route_timestamp):
    if not route_timestamp:
        return 0

    windows = _ordem_time_windows(ordem)
    if not windows:
        return 0

    best = 0
    for start_dt, end_dt in windows:
        if start_dt <= route_timestamp <= end_dt:
            best = max(best, 40)
            continue
        delta_seconds = min(
            abs((route_timestamp - start_dt).total_seconds()),
            abs((route_timestamp - end_dt).total_seconds()),
        )
        delta_minutes = delta_seconds / 60
        if delta_minutes <= 60:
            best = max(best, 30)
        elif delta_minutes <= 120:
            best = max(best, 20)
        elif route_timestamp.date() == start_dt.date() or route_timestamp.date() == end_dt.date():
            best = max(best, 12)
    return best


def _ordem_time_windows(ordem):
    windows = []
    if ordem.data_aplicacao:
        if ordem.hora_inicio_aplicacao and ordem.hora_termino_aplicacao:
            start_dt = datetime.combine(ordem.data_aplicacao, ordem.hora_inicio_aplicacao)
            end_dt = datetime.combine(ordem.data_aplicacao, ordem.hora_termino_aplicacao)
            if end_dt < start_dt:
                end_dt += timedelta(days=1)
            windows.append((start_dt - timedelta(minutes=30), end_dt + timedelta(minutes=30)))
        else:
            windows.append((
                datetime.combine(ordem.data_aplicacao, datetime.min.time()),
                datetime.combine(ordem.data_aplicacao, datetime.max.time()),
            ))

    if ordem.respondido_em:
        windows.append((ordem.respondido_em - timedelta(hours=4), ordem.respondido_em + timedelta(hours=4)))

    solicitacao = getattr(ordem, "solicitacao", None)
    if solicitacao and solicitacao.data_agendamento:
        if solicitacao.hora_agendamento:
            scheduled_at = datetime.combine(solicitacao.data_agendamento, solicitacao.hora_agendamento)
            windows.append((scheduled_at - timedelta(hours=2), scheduled_at + timedelta(hours=8)))
        else:
            windows.append((
                datetime.combine(solicitacao.data_agendamento, datetime.min.time()),
                datetime.combine(solicitacao.data_agendamento, datetime.max.time()),
            ))
    return windows


def _score_kml_os_aircraft_match(ordem, aircraft_name):
    route_aircraft = _normalize_match_text(aircraft_name)
    if not route_aircraft:
        return 0

    candidates = [
        getattr(ordem, "prefixo_aeronave_pulverizacao", None),
        getattr(ordem, "prefixo_aeronave_monitoramento", None),
        getattr(ordem, "drone_denominacao", None),
        getattr(ordem, "drone_monitoramento_denominacao", None),
        getattr(ordem, "drone_numero_serie", None),
        getattr(ordem, "drone_monitoramento_numero_serie", None),
    ]
    return max((_text_match_score(route_aircraft, candidate) for candidate in candidates), default=0)


def _score_kml_os_pilot_match(ordem, pilot_name):
    route_pilot = _normalize_match_text(pilot_name)
    if not route_pilot:
        return 0
    return max(
        _text_match_score(route_pilot, getattr(ordem, "piloto", None), exact_score=15, contains_score=10),
        _text_match_score(route_pilot, getattr(ordem, "auxiliar", None), exact_score=8, contains_score=5),
    )


def _score_kml_os_place_match(ordem, route):
    solicitacao = getattr(ordem, "solicitacao", None)
    os_place_id = _clean_place_id(getattr(solicitacao, "place_id", None))
    if not os_place_id:
        return 0

    flight_record = getattr(route, "flight_record", None)
    route_place_ids = {
        place_id
        for place_id in (
            _clean_place_id(getattr(route, "place_id", None)),
            _clean_place_id(getattr(flight_record, "place_id", None)),
            _clean_place_id(_extract_place_id_from_raw_payload(getattr(flight_record, "raw_payload", None))),
        )
        if place_id
    }
    return 65 if os_place_id in route_place_ids else 0


def _score_kml_os_address_match(ordem, route):
    solicitacao = getattr(ordem, "solicitacao", None)
    route_address_text = _route_address_text(route)
    if not solicitacao or not route_address_text:
        return 0

    route_normalized = _normalize_match_text(route_address_text)
    route_tokens = _address_tokens(route_address_text)
    score = 0

    full_address = _solicitacao_address_text(solicitacao)
    full_normalized = _normalize_match_text(full_address)
    if full_normalized and (
        full_normalized in route_normalized or route_normalized in full_normalized
    ):
        score = max(score, 30)

    street_tokens = _address_tokens(getattr(solicitacao, "logradouro", None))
    number_tokens = _address_tokens(getattr(solicitacao, "numero", None))
    neighborhood_tokens = _address_tokens(getattr(solicitacao, "bairro", None))
    city_tokens = _address_tokens(getattr(solicitacao, "cidade", None))
    cep_tokens = _address_tokens(getattr(solicitacao, "cep", None))

    street_matches = bool(street_tokens and street_tokens.issubset(route_tokens))
    if street_matches:
        score = max(score, 12)
        if number_tokens and number_tokens.intersection(route_tokens):
            score = max(score, 28)
        if city_tokens and city_tokens.intersection(route_tokens):
            score = max(score, 24)
        if neighborhood_tokens and neighborhood_tokens.intersection(route_tokens):
            score = max(score, 22)
        if cep_tokens and cep_tokens.intersection(route_tokens):
            score = max(score, 26)

    return score


def _clean_place_id(value):
    value = str(value or "").strip()
    return value or None


def _extract_place_id_from_raw_payload(raw_payload):
    if not raw_payload:
        return None
    try:
        payload = json.loads(raw_payload) if isinstance(raw_payload, str) else raw_payload
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    if not isinstance(payload, dict):
        return None
    return payload.get("place_id") or payload.get("placeId") or payload.get("google_place_id")


def _route_address_text(route):
    flight_record = getattr(route, "flight_record", None)
    parts = [
        getattr(route, "location", None),
        getattr(route, "formatted_address", None),
        getattr(flight_record, "location", None),
        getattr(flight_record, "field_name", None),
    ]
    return " ".join(str(part).strip() for part in parts if part)


def _solicitacao_address_text(solicitacao):
    fields = ("logradouro", "numero", "bairro", "cidade", "uf", "cep")
    return " ".join(
        str(getattr(solicitacao, field, "") or "").strip()
        for field in fields
        if getattr(solicitacao, field, None)
    )


_ADDRESS_STOPWORDS = {
    "r",
    "rua",
    "av",
    "avenida",
    "al",
    "alameda",
    "travessa",
    "tv",
    "estrada",
    "rodovia",
    "praca",
    "pc",
    "bairro",
    "jd",
    "jardim",
    "vl",
    "vila",
    "n",
    "no",
    "numero",
    "num",
    "de",
    "da",
    "do",
    "das",
    "dos",
}


def _address_tokens(value):
    value = str(value or "").strip()
    if not value:
        return set()
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    tokens = re.findall(r"[a-z0-9]+", normalized.lower())
    return {
        token
        for token in tokens
        if token not in _ADDRESS_STOPWORDS and (token.isdigit() or len(token) > 1)
    }


def _text_match_score(normalized_left, raw_right, *, exact_score=35, contains_score=25):
    normalized_right = _normalize_match_text(raw_right)
    if not normalized_left or not normalized_right:
        return 0
    if normalized_left == normalized_right:
        return exact_score
    if normalized_left in normalized_right or normalized_right in normalized_left:
        return contains_score
    return 0


def _normalize_match_text(value):
    value = str(value or "").strip()
    if not value:
        return ""
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", normalized.lower())


def _kml_distance_to_solicitacao_meters(points, solicitacao):
    if not solicitacao:
        return None
    os_lat = _parse_coordinate(getattr(solicitacao, "latitude", None))
    os_lng = _parse_coordinate(getattr(solicitacao, "longitude", None))
    if os_lat is None or os_lng is None:
        return None

    distances = []
    for point in points or []:
        point_lat = point.get("lat")
        point_lng = point.get("lng")
        if point_lat is None or point_lng is None:
            continue
        distances.append(_haversine_distance_meters(os_lat, os_lng, point_lat, point_lng))
    return min(distances) if distances else None


def _reverse_geocode_kml_route(points):
    representative_point = _representative_kml_point(points)
    if not representative_point:
        return {}

    try:
        formatted_address, place_id = reverse_geocode_lat_lng_google(
            lat=representative_point["lat"],
            lng=representative_point["lng"],
        )
    except Exception:
        return {}

    return {
        "formatted_address": formatted_address,
        "place_id": _clean_place_id(place_id),
    }


def _representative_kml_point(points):
    valid_points = [
        point
        for point in points or []
        if point.get("lat") is not None and point.get("lng") is not None
    ]
    if not valid_points:
        return None

    return {
        "lat": sum(point["lat"] for point in valid_points) / len(valid_points),
        "lng": sum(point["lng"] for point in valid_points) / len(valid_points),
    }


def _parse_coordinate(value):
    if value is None:
        return None
    try:
        return float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


def _score_kml_os_geo_match(distance_meters):
    if distance_meters is None:
        return 0
    if distance_meters <= 150:
        return 35
    if distance_meters <= 300:
        return 30
    if distance_meters <= 750:
        return 22
    if distance_meters <= 1500:
        return 12
    if distance_meters <= 3000:
        return 5
    return 0


def get_dji_route_payload(route_id):
    route = DjiFlightKmlRoute.query.get_or_404(route_id)
    linked_os = _get_linked_os_for_kml_route(route.id)
    points = json.loads(route.points_json or "[]")
    center_point = points[0] if points else None
    start_point = points[0] if points else None
    end_point = points[-1] if points else None
    altitude_values = [float(point.get("alt", 0) or 0) for point in points if point.get("alt") is not None]
    route_distance_meters = _compute_route_distance_meters(points)
    altitude_min = min(altitude_values) if altitude_values else None
    altitude_max = max(altitude_values) if altitude_values else None
    return {
        "id": route.id,
        "route_code": route.route_code,
        "aircraft_name": route.aircraft_name,
        "pilot_name": route.pilot_name,
        "flight_controller_id": route.flight_controller_id,
        "route_timestamp": route.route_timestamp.strftime("%d/%m/%Y %H:%M:%S") if route.route_timestamp else "",
        "task_area": route.task_area,
        "spray_amount": route.spray_amount,
        "point_count": route.point_count,
        "route_color_css": _kml_color_to_css(route.route_color) or "#ff3b30",
        "route_width": route.route_width or 2,
        "center": center_point,
        "start_point": start_point,
        "end_point": end_point,
        "start_label": _format_coordinate_label(start_point),
        "end_label": _format_coordinate_label(end_point),
        "distance_meters": route_distance_meters,
        "distance_label": _format_distance_label(route_distance_meters),
        "altitude_min": altitude_min,
        "altitude_max": altitude_max,
        "altitude_label": _format_altitude_range_label(altitude_min, altitude_max),
        "linked_os": _build_linked_os_payload(linked_os),
        "points": points,
    }


def _build_kml_os_map(routes):
    route_ids = [route.id for route in routes or []]
    if not route_ids:
        return {}

    ordens = (
        OrdemServico.query
        .options(joinedload(OrdemServico.solicitacao))
        .filter(OrdemServico.dji_kml_route_id.in_(route_ids))
        .order_by(OrdemServico.respondido_em.desc(), OrdemServico.id.desc())
        .all()
    )
    os_por_rota = {}
    for ordem in ordens:
        os_por_rota.setdefault(ordem.dji_kml_route_id, ordem)
    return os_por_rota


def _get_linked_os_for_kml_route(route_id):
    return (
        OrdemServico.query
        .options(joinedload(OrdemServico.solicitacao))
        .filter(OrdemServico.dji_kml_route_id == route_id)
        .order_by(OrdemServico.respondido_em.desc(), OrdemServico.id.desc())
        .first()
    )


def _build_linked_os_payload(ordem):
    if not ordem:
        return None

    solicitacao = getattr(ordem, "solicitacao", None)
    solicitacao_id = getattr(ordem, "solicitacao_id", None)
    identificador = (getattr(ordem, "identificador_os", None) or "").strip()
    label = identificador or (f"OS #{solicitacao_id}" if solicitacao_id else f"OS interna #{ordem.id}")
    data_aplicacao = getattr(ordem, "data_aplicacao", None)

    return {
        "ordem_id": ordem.id,
        "solicitacao_id": solicitacao_id,
        "identificador_os": identificador,
        "label": label,
        "protocolo": getattr(solicitacao, "protocolo", None) or "",
        "status": getattr(solicitacao, "status", None) or "",
        "data_aplicacao": data_aplicacao.strftime("%d/%m/%Y") if data_aplicacao else "",
        "piloto": getattr(ordem, "piloto", None) or "",
    }


def build_dji_logs_excel_export(args):
    data_inicio = (args.get("data_inicio") or "").strip()
    data_fim = (args.get("data_fim") or "").strip()
    piloto = (args.get("piloto") or "").strip()
    aeronave = (args.get("aeronave") or "").strip()
    equipe = (args.get("equipe") or "").strip()
    q = (args.get("q") or "").strip()
    endereco = (args.get("endereco") or "").strip()
    voo_id = (args.get("voo_id") or "").strip()
    status_rota = (args.get("status_rota") or "").strip()

    filtered_query = _build_filtered_query(
        data_inicio=data_inicio,
        data_fim=data_fim,
        piloto=piloto,
        aeronave=aeronave,
        equipe=equipe,
        q=q,
        endereco=endereco,
        voo_id=voo_id,
        status_rota=status_rota,
    )
    registros = filtered_query.order_by(
        DjiFlightRecord.flight_start.desc(),
        DjiFlightRecord.id.desc(),
    ).all()

    totals = filtered_query.with_entities(
        func.count(DjiFlightRecord.id),
        func.coalesce(func.sum(DjiFlightRecord.sprayed_area_ha), 0.0),
        func.coalesce(func.sum(DjiFlightRecord.total_amount_l_kg), 0.0),
        func.coalesce(func.sum(DjiFlightRecord.flight_duration_seconds), 0),
        func.avg(DjiFlightRecord.battery_consumed_level),
    ).first()
    total_voos, total_area, total_volume, total_duracao, media_bateria = totals
    total_pilotos = filtered_query.with_entities(
        func.count(func.distinct(func.nullif(DjiFlightRecord.pilot_name, "")))
    ).scalar() or 0
    total_aeronaves = filtered_query.with_entities(
        func.count(func.distinct(func.nullif(DjiFlightRecord.aircraft_name, "")))
    ).scalar() or 0

    weekly_rows = filtered_query.with_entities(
        DjiFlightRecord.flight_start,
        DjiFlightRecord.sprayed_area_ha,
        DjiFlightRecord.total_amount_l_kg,
        DjiFlightRecord.flight_duration_seconds,
    ).all()
    resumo_semanal = _build_weekly_summary(weekly_rows)

    workbook = Workbook()
    ws_resumo = workbook.active
    ws_resumo.title = "Resumo"
    ws_voos = workbook.create_sheet("Voos")

    _fill_header = PatternFill("solid", fgColor="0D6EFD")
    _fill_section = PatternFill("solid", fgColor="EAF2FF")
    _fill_zebra = PatternFill("solid", fgColor="F8FBFF")
    _font_header = Font(color="FFFFFF", bold=True)
    _font_title = Font(bold=True, size=16, color="0D3B66")
    _font_section = Font(bold=True, color="0D3B66")
    _thin = Side(style="thin", color="D0D7DE")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    ws_resumo.merge_cells("A1:D1")
    ws_resumo["A1"] = "Relatorio de Logs DJI"
    ws_resumo["A1"].font = _font_title
    ws_resumo["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.row_dimensions[1].height = 26

    ws_resumo.merge_cells("A2:D2")
    filtros_texto = (
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
        f"Periodo: {data_inicio or 'inicio'} ate {data_fim or 'hoje'} | "
        f"Piloto: {piloto or 'Todos'} | Aeronave: {aeronave or 'Todas'} | Equipe: {equipe or 'Todas'}"
    )
    ws_resumo["A2"] = filtros_texto
    ws_resumo["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_resumo.row_dimensions[2].height = 34

    resumo_headers = ["Indicador", "Valor"]
    resumo_rows = [
        ("Total de voos", total_voos or 0),
        ("Area pulverizada (ha)", float(total_area or 0)),
        ("Produto aplicado (L/Kg)", float(total_volume or 0)),
        ("Tempo de voo", format_duration_seconds(total_duracao or 0)),
        ("Media de bateria consumida (%)", round(media_bateria or 0, 1) if media_bateria is not None else "-"),
        ("Pilotos unicos", total_pilotos),
        ("Aeronaves unicas", total_aeronaves),
    ]

    start_row = 4
    for col, header in enumerate(resumo_headers, start=1):
        cell = ws_resumo.cell(row=start_row, column=col, value=header)
        cell.fill = _fill_header
        cell.font = _font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border

    current_row = start_row + 1
    for index, (label, value) in enumerate(resumo_rows, start=1):
        ws_resumo.cell(row=current_row, column=1, value=label)
        ws_resumo.cell(row=current_row, column=2, value=value)
        for col in (1, 2):
            cell = ws_resumo.cell(row=current_row, column=col)
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if index % 2 == 0:
                cell.fill = _fill_zebra
        if label in {"Area pulverizada (ha)", "Produto aplicado (L/Kg)"}:
            ws_resumo.cell(row=current_row, column=2).number_format = "#,##0.000"
        if label == "Media de bateria consumida (%)":
            ws_resumo.cell(row=current_row, column=2).number_format = "0.0"
        current_row += 1

    current_row += 1
    ws_resumo.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    section_cell = ws_resumo.cell(row=current_row, column=1, value="Resumo semanal")
    section_cell.fill = _fill_section
    section_cell.font = _font_section
    section_cell.border = _border
    section_cell.alignment = Alignment(vertical="center")

    weekly_header_row = current_row + 1
    weekly_headers = ["Semana", "Voos", "Area (ha)", "Volume", "Duracao"]
    for col, header in enumerate(weekly_headers, start=1):
        cell = ws_resumo.cell(row=weekly_header_row, column=col, value=header)
        cell.fill = _fill_header
        cell.font = _font_header
        cell.border = _border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, item in enumerate(resumo_semanal, start=1):
        row = weekly_header_row + idx
        values = [item["semana_label"], item["voos"], item["area"], item["volume"], item["duracao"]]
        for col, value in enumerate(values, start=1):
            cell = ws_resumo.cell(row=row, column=col, value=value)
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if idx % 2 == 0:
                cell.fill = _fill_zebra
        ws_resumo.cell(row=row, column=3).number_format = "#,##0.000"
        ws_resumo.cell(row=row, column=4).number_format = "#,##0.000"

    ws_resumo.freeze_panes = "A4"
    ws_resumo.column_dimensions["A"].width = 34
    ws_resumo.column_dimensions["B"].width = 18
    ws_resumo.column_dimensions["C"].width = 18
    ws_resumo.column_dimensions["D"].width = 18
    ws_resumo.column_dimensions["E"].width = 18

    headers = [
        "Inicio",
        "Fim",
        "Periodo DJI",
        "Piloto",
        "Aeronave",
        "Equipe",
        "Tipo de tarefa",
        "Area (ha)",
        "Volume (L/Kg)",
        "Duracao",
        "Cultura",
        "Local resumido",
        "Local completo",
        "Serial da aeronave",
        "Bateria inicial (%)",
        "Bateria final (%)",
        "Bateria consumida (%)",
        "Bateria SN",
        "Campo",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws_voos.cell(row=1, column=col, value=header)
        cell.fill = _fill_header
        cell.font = _font_header
        cell.border = _border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, registro in enumerate(registros, start=2):
        values = [
            registro.flight_start.strftime("%d/%m/%Y %H:%M") if registro.flight_start else "",
            registro.flight_end.strftime("%d/%m/%Y %H:%M") if registro.flight_end else "",
            registro.flight_window or "",
            registro.pilot_name or "",
            registro.aircraft_name or "",
            registro.team_name or "",
            registro.task_type or "",
            float(registro.sprayed_area_ha or 0),
            float(registro.total_amount_l_kg or 0),
            registro.duration_display,
            registro.crop or "",
            registro.location_short,
            registro.location or "",
            registro.serial_number or "",
            registro.starting_battery_level,
            registro.ending_battery_level,
            registro.battery_consumption_pct,
            registro.battery_sn or "",
            registro.field_name or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws_voos.cell(row=row_index, column=col, value=value)
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = _fill_zebra
        ws_voos.cell(row=row_index, column=8).number_format = "#,##0.000"
        ws_voos.cell(row=row_index, column=9).number_format = "#,##0.000"

    ws_voos.freeze_panes = "A2"
    ws_voos.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(registros) + 1, 2)}"
    widths = {
        "A": 18, "B": 18, "C": 24, "D": 22, "E": 18, "F": 18, "G": 18,
        "H": 12, "I": 14, "J": 12, "K": 14, "L": 22, "M": 60, "N": 18,
        "O": 16, "P": 16, "Q": 18, "R": 18, "S": 18,
    }
    for col, width in widths.items():
        ws_voos.column_dimensions[col].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    nome = f"dji_logs_voo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, nome


def format_duration_seconds(total_seconds):
    total_seconds = int(total_seconds or 0)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    return f"{minutes:02d}:{seconds:02d}"


def _load_workbook(file_bytes):
    try:
        return load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Nao foi possivel abrir o Excel enviado.") from exc


def _parse_workbook_rows(workbook):
    worksheet = None
    for candidate in workbook.worksheets:
        if _normalize_header(candidate.title) == "flightrecord":
            worksheet = candidate
            break
    if worksheet is None:
        worksheet = workbook.worksheets[0]

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("A planilha enviada nao possui cabecalho.")

    header_map = _build_header_map(header_row)
    missing_fields = [
        field
        for field in (
            "flight_time",
            "aircraft_name",
            "task_type",
            "sprayed_area",
            "total_amount",
            "flight_duration",
            "pilot_name",
            "serial_number",
        )
        if field not in header_map
    ]
    if missing_fields:
        raise ValueError("O Excel nao possui todas as colunas esperadas do relatorio de voo da DJI.")

    rows = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        parsed = _parse_excel_row(row_number, row, header_map)
        if parsed is not None:
            rows.append(parsed)
    return rows


def _parse_excel_row(row_number, row_values, header_map):
    values = list(row_values)
    if not any(value is not None and str(value).strip() for value in values):
        return None

    flight_window = _stringify(values[header_map["flight_time"]])
    if not flight_window:
        raise ValueError(f"Linha {row_number}: periodo de voo nao informado.")

    flight_start, flight_end = _parse_flight_window(flight_window, row_number)

    payload = {
        "flight_time": flight_window,
        "location": _stringify(_get_row_value(values, header_map, "location")),
        "place_id": _clean_place_id(_get_row_value(values, header_map, "place_id")),
        "aircraft_name": _stringify(_get_row_value(values, header_map, "aircraft_name")),
        "task_type": _stringify(_get_row_value(values, header_map, "task_type")),
        "sprayed_area": _parse_optional_float(_get_row_value(values, header_map, "sprayed_area")),
        "total_amount": _parse_optional_float(_get_row_value(values, header_map, "total_amount")),
        "flight_duration": _stringify(_get_row_value(values, header_map, "flight_duration")),
        "crop": _stringify(_get_row_value(values, header_map, "crop")),
        "pilot_name": _stringify(_get_row_value(values, header_map, "pilot_name")),
        "team_name": _stringify(_get_row_value(values, header_map, "team_name")),
        "field_name": _stringify(_get_row_value(values, header_map, "field_name")),
        "serial_number": _stringify(_get_row_value(values, header_map, "serial_number")),
        "starting_battery_level": _parse_optional_int(_get_row_value(values, header_map, "starting_battery_level")),
        "ending_battery_level": _parse_optional_int(_get_row_value(values, header_map, "ending_battery_level")),
        "battery_sn": _stringify(_get_row_value(values, header_map, "battery_sn")),
    }

    flight_duration_seconds = _parse_duration_to_seconds(payload["flight_duration"], row_number)
    battery_consumed_level = None
    if (
        payload["starting_battery_level"] is not None
        and payload["ending_battery_level"] is not None
    ):
        battery_consumed_level = payload["starting_battery_level"] - payload["ending_battery_level"]

    fingerprint = _build_record_fingerprint(payload)

    return {
        "source_row_number": row_number,
        "fingerprint": fingerprint,
        "flight_window": flight_window,
        "flight_start": flight_start,
        "flight_end": flight_end,
        "location": payload["location"],
        "place_id": payload["place_id"],
        "aircraft_name": payload["aircraft_name"],
        "task_type": payload["task_type"],
        "sprayed_area_ha": payload["sprayed_area"] or 0.0,
        "total_amount_l_kg": payload["total_amount"] or 0.0,
        "flight_duration_seconds": flight_duration_seconds,
        "flight_duration_label": payload["flight_duration"],
        "crop": payload["crop"],
        "pilot_name": payload["pilot_name"],
        "team_name": payload["team_name"],
        "field_name": payload["field_name"],
        "serial_number": payload["serial_number"],
        "starting_battery_level": payload["starting_battery_level"],
        "ending_battery_level": payload["ending_battery_level"],
        "battery_consumed_level": battery_consumed_level,
        "battery_sn": payload["battery_sn"],
        "raw_payload": payload,
    }


def _build_header_map(header_row):
    normalized_to_index = {}
    for index, header in enumerate(header_row):
        normalized = _normalize_header(header)
        if normalized:
            normalized_to_index[normalized] = index

    header_map = {}
    for field_name, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized_to_index:
                header_map[field_name] = normalized_to_index[alias]
                break
    return header_map


def _normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _get_extended_data_value(extended_data, *aliases):
    if not extended_data:
        return None
    normalized_values = {
        _normalize_header(key): value
        for key, value in extended_data.items()
    }
    for alias in aliases:
        value = normalized_values.get(_normalize_header(alias))
        if value:
            return value
    return None


def _parse_flight_window(value, row_number):
    match = _FLIGHT_WINDOW_RE.match(value)
    if not match:
        raise ValueError(f"Linha {row_number}: periodo de voo invalido ({value}).")
    start_dt = datetime.strptime(match.group(1), "%Y-%m-%d %H:%M:%S")
    end_raw = match.group(2)

    if len(end_raw) == 8:
        end_dt = datetime.strptime(
            f"{start_dt.strftime('%Y-%m-%d')} {end_raw}",
            "%Y-%m-%d %H:%M:%S",
        )
        if end_dt < start_dt:
            end_dt += timedelta(days=1)
    else:
        end_dt = datetime.strptime(end_raw, "%Y-%m-%d %H:%M:%S")

    return start_dt, end_dt


def _parse_duration_to_seconds(value, row_number):
    if not value:
        return 0

    parts = [part.strip() for part in str(value).split(":")]
    if len(parts) == 2:
        minutes, seconds = parts
        return (int(minutes) * 60) + int(seconds)
    if len(parts) == 3:
        hours, minutes, seconds = parts
        return (int(hours) * 3600) + (int(minutes) * 60) + int(seconds)
    raise ValueError(f"Linha {row_number}: duracao de voo invalida ({value}).")


def _parse_optional_float(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_optional_int(value):
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip().replace(",", ".")))
    except ValueError:
        return None


def _stringify(value):
    if value is None:
        return ""
    return str(value).strip()


def _get_row_value(values, header_map, field_name):
    index = header_map.get(field_name)
    if index is None:
        return None
    return values[index]


def _build_record_fingerprint(payload):
    canonical_parts = [
        payload.get("flight_time") or "",
        payload.get("location") or "",
        payload.get("aircraft_name") or "",
        payload.get("task_type") or "",
        str(payload.get("sprayed_area") or ""),
        str(payload.get("total_amount") or ""),
        payload.get("flight_duration") or "",
        payload.get("crop") or "",
        payload.get("pilot_name") or "",
        payload.get("team_name") or "",
        payload.get("field_name") or "",
        payload.get("serial_number") or "",
        str(payload.get("starting_battery_level") or ""),
        str(payload.get("ending_battery_level") or ""),
        payload.get("battery_sn") or "",
    ]
    canonical = "||".join(part.strip().lower() for part in canonical_parts)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _save_uploaded_excel(original_filename, file_bytes):
    base_folder = os.path.join(get_upload_folder(), "dji-flight-logs")
    os.makedirs(base_folder, exist_ok=True)

    name_root, extension = os.path.splitext(original_filename)
    safe_root = secure_filename(name_root) or "dji_flight_logs"
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    short_hash = hashlib.sha256(file_bytes).hexdigest()[:10]
    stored_filename = f"{safe_root}_{stamp}_{short_hash}{extension or '.xlsx'}"
    absolute_path = os.path.join(base_folder, stored_filename)

    with open(absolute_path, "wb") as file_handle:
        file_handle.write(file_bytes)

    relative_path = os.path.join("dji-flight-logs", stored_filename).replace("\\", "/")
    return stored_filename, relative_path


def _save_uploaded_kml(original_filename, file_bytes):
    base_folder = os.path.join(get_upload_folder(), "dji-flight-routes")
    os.makedirs(base_folder, exist_ok=True)

    name_root, extension = os.path.splitext(original_filename)
    safe_root = secure_filename(name_root) or "dji_flight_route"
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    short_hash = hashlib.sha256(file_bytes).hexdigest()[:10]
    stored_filename = f"{safe_root}_{stamp}_{short_hash}{extension or '.kml'}"
    absolute_path = os.path.join(base_folder, stored_filename)

    with open(absolute_path, "wb") as file_handle:
        file_handle.write(file_bytes)

    relative_path = os.path.join("dji-flight-routes", stored_filename).replace("\\", "/")
    return stored_filename, relative_path


def _build_filtered_query(
    *,
    data_inicio="",
    data_fim="",
    piloto="",
    aeronave="",
    equipe="",
    q="",
    endereco="",
    voo_id="",
    status_rota="",
):
    return _apply_dji_filters(
        DjiFlightRecord.query.options(joinedload(DjiFlightRecord.route_kml)),
        data_inicio=data_inicio,
        data_fim=data_fim,
        piloto=piloto,
        aeronave=aeronave,
        equipe=equipe,
        q=q,
        endereco=endereco,
        voo_id=voo_id,
        status_rota=status_rota,
    )


def _build_filtered_kml_query(
    *,
    q="",
    data_inicio="",
    data_fim="",
    endereco="",
    piloto="",
    aeronave="",
    status_os="",
    status_voo="",
    voo_id="",
):
    query = DjiFlightKmlRoute.query.options(joinedload(DjiFlightKmlRoute.flight_record))

    if voo_id:
        try:
            query = query.filter(DjiFlightKmlRoute.flight_record_id == int(voo_id))
        except (TypeError, ValueError):
            pass

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                DjiFlightKmlRoute.route_code.ilike(like),
                DjiFlightKmlRoute.original_filename.ilike(like),
                DjiFlightKmlRoute.aircraft_name.ilike(like),
                DjiFlightKmlRoute.pilot_name.ilike(like),
                DjiFlightKmlRoute.flight_controller_id.ilike(like),
            )
        )

    if endereco:
        query = query.filter(
            DjiFlightKmlRoute.flight_record.has(
                DjiFlightRecord.location.ilike(f"%{endereco}%")
            )
        )

    if data_inicio:
        try:
            start_dt = datetime.strptime(data_inicio, "%Y-%m-%d")
            query = query.filter(DjiFlightKmlRoute.route_timestamp >= start_dt)
        except ValueError:
            pass

    if data_fim:
        try:
            end_dt = datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(DjiFlightKmlRoute.route_timestamp < end_dt)
        except ValueError:
            pass

    if piloto:
        query = query.filter(DjiFlightKmlRoute.pilot_name == piloto)

    if aeronave:
        query = query.filter(DjiFlightKmlRoute.aircraft_name == aeronave)

    if status_voo == "com_voo":
        query = query.filter(DjiFlightKmlRoute.flight_record_id.isnot(None))
    elif status_voo == "sem_voo":
        query = query.filter(DjiFlightKmlRoute.flight_record_id.is_(None))

    linked_route_ids = db.session.query(OrdemServico.dji_kml_route_id).filter(
        OrdemServico.dji_kml_route_id.isnot(None)
    )
    if status_os == "com_os":
        query = query.filter(DjiFlightKmlRoute.id.in_(linked_route_ids))
    elif status_os == "sem_os":
        query = query.filter(~DjiFlightKmlRoute.id.in_(linked_route_ids))

    return query


def _apply_dji_filters(
    query,
    *,
    data_inicio="",
    data_fim="",
    piloto="",
    aeronave="",
    equipe="",
    q="",
    endereco="",
    voo_id="",
    status_rota="",
):
    if voo_id:
        try:
            query = query.filter(DjiFlightRecord.id == int(voo_id))
        except (TypeError, ValueError):
            pass

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                DjiFlightRecord.flight_window.ilike(like),
                DjiFlightRecord.location.ilike(like),
                DjiFlightRecord.aircraft_name.ilike(like),
                DjiFlightRecord.task_type.ilike(like),
                DjiFlightRecord.crop.ilike(like),
                DjiFlightRecord.pilot_name.ilike(like),
                DjiFlightRecord.team_name.ilike(like),
                DjiFlightRecord.field_name.ilike(like),
                DjiFlightRecord.serial_number.ilike(like),
                DjiFlightRecord.battery_sn.ilike(like),
            )
        )

    if endereco:
        query = query.filter(DjiFlightRecord.location.ilike(f"%{endereco}%"))

    if data_inicio:
        try:
            start_dt = datetime.strptime(data_inicio, "%Y-%m-%d")
            query = query.filter(DjiFlightRecord.flight_start >= start_dt)
        except ValueError:
            pass

    if data_fim:
        try:
            end_dt = datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(DjiFlightRecord.flight_start < end_dt)
        except ValueError:
            pass

    if piloto:
        query = query.filter(DjiFlightRecord.pilot_name == piloto)

    if aeronave:
        query = query.filter(DjiFlightRecord.aircraft_name == aeronave)

    if equipe:
        query = query.filter(DjiFlightRecord.team_name == equipe)

    linked_record_ids = db.session.query(DjiFlightKmlRoute.flight_record_id).filter(
        DjiFlightKmlRoute.flight_record_id.isnot(None)
    )
    if status_rota == "com_kml":
        query = query.filter(DjiFlightRecord.id.in_(linked_record_ids))
    elif status_rota == "sem_kml":
        query = query.filter(~DjiFlightRecord.id.in_(linked_record_ids))

    return query


def _distinct_non_empty_values(column):
    return [
        value
        for (value,) in db.session.query(column)
        .filter(func.length(func.trim(func.coalesce(column, ""))) > 0)
        .distinct()
        .order_by(column.asc())
        .all()
    ]


def _build_top_groups(filtered_query, column, limit=8):
    return [
        {"label": label or "Nao informado", "total": total}
        for label, total in (
            filtered_query
            .with_entities(func.coalesce(column, "Nao informado"), func.count(DjiFlightRecord.id))
            .group_by(column)
            .order_by(func.count(DjiFlightRecord.id).desc(), func.coalesce(column, "Nao informado").asc())
            .limit(limit)
            .all()
        )
    ]


def _build_weekly_summary(rows):
    weekly = defaultdict(lambda: {"voos": 0, "area": 0.0, "volume": 0.0, "duracao": 0})
    for flight_start, area, volume, duration in rows:
        if not flight_start:
            continue
        week_start = (flight_start - timedelta(days=flight_start.weekday())).date()
        item = weekly[week_start]
        item["voos"] += 1
        item["area"] += float(area or 0)
        item["volume"] += float(volume or 0)
        item["duracao"] += int(duration or 0)

    summary = []
    for week_start in sorted(weekly.keys(), reverse=True):
        values = weekly[week_start]
        week_end = week_start + timedelta(days=6)
        summary.append(
            {
                "semana_label": f"{week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}",
                "voos": values["voos"],
                "area": round(values["area"], 3),
                "volume": round(values["volume"], 3),
                "duracao": format_duration_seconds(values["duracao"]),
            }
        )
    return summary[:12]


def _parse_kml_payload(file_bytes, filename):
    try:
        root = ET.fromstring(file_bytes.decode("utf-8", errors="ignore"))
    except Exception as exc:
        raise ValueError(f"Nao foi possivel ler o KML {filename}.") from exc

    placemark = next((elem for elem in root.iter() if _xml_local_name(elem.tag) == "Placemark"), None)
    if placemark is None:
        raise ValueError(f"O arquivo {filename} nao possui um Placemark valido.")

    route_code = _xml_child_text(placemark, "name")
    extended_data = {}
    for data_elem in placemark.iter():
        if _xml_local_name(data_elem.tag) != "Data":
            continue
        name = (data_elem.attrib.get("name") or "").strip()
        value = ""
        for child in data_elem:
            if _xml_local_name(child.tag) == "value":
                value = (child.text or "").strip()
                break
        if name:
            extended_data[name] = value

    coordinates_text = None
    for elem in placemark.iter():
        if _xml_local_name(elem.tag) == "coordinates":
            coordinates_text = elem.text or ""
            break
    points = _parse_kml_coordinates(coordinates_text or "")
    if not points:
        raise ValueError(f"O arquivo {filename} nao possui coordenadas de rota.")

    route_color = None
    route_width = None
    for elem in placemark.iter():
        local_name = _xml_local_name(elem.tag)
        if local_name == "color" and route_color is None:
            route_color = (elem.text or "").strip()
        if local_name == "width" and route_width is None:
            try:
                route_width = float((elem.text or "").strip())
            except ValueError:
                route_width = None

    aircraft_name, route_timestamp, route_code_from_filename = _parse_kml_filename(filename)
    final_route_code = route_code or route_code_from_filename
    if not final_route_code:
        raise ValueError(f"Nao foi possivel identificar o codigo da rota no arquivo {filename}.")

    return {
        "route_code": final_route_code,
        "aircraft_name": extended_data.get("Aircraft Name") or aircraft_name,
        "pilot_name": extended_data.get("Pilot Name") or "",
        "flight_controller_id": extended_data.get("Flight Controller ID") or "",
        "route_timestamp": route_timestamp,
        "place_id": _clean_place_id(_get_extended_data_value(
            extended_data,
            "Place ID",
            "PlaceId",
            "Google Place ID",
            "google_place_id",
        )),
        "mode_selection": extended_data.get("Mode Selection") or "",
        "flight_time_raw": extended_data.get("Flight Time") or "",
        "task_area": _parse_optional_float(extended_data.get("Task Area")),
        "spray_amount": _parse_optional_float(extended_data.get("Spray amount")),
        "route_color": route_color,
        "route_width": route_width,
        "points": points,
    }


def _parse_kml_filename(filename):
    match = re.match(r"^(?P<aircraft>.+?)_(?P<stamp>\d{14})_(?P<code>[^.]+)\.kml$", filename, re.IGNORECASE)
    if not match:
        return "", None, ""

    aircraft = match.group("aircraft").strip()
    stamp_raw = match.group("stamp")
    route_code = match.group("code").strip()
    route_timestamp = None
    try:
        route_timestamp = datetime.strptime(stamp_raw, "%Y%m%d%H%M%S")
    except ValueError:
        route_timestamp = None
    return aircraft, route_timestamp, route_code


def _parse_kml_coordinates(raw_text):
    points = []
    for chunk in (raw_text or "").strip().split():
        parts = chunk.split(",")
        if len(parts) < 2:
            continue
        try:
            lng = float(parts[0])
            lat = float(parts[1])
            alt = float(parts[2]) if len(parts) > 2 and parts[2] != "" else 0.0
        except ValueError:
            continue
        points.append({"lat": lat, "lng": lng, "alt": alt})
    return points


def _compute_route_distance_meters(points):
    if len(points or []) < 2:
        return 0.0

    total = 0.0
    for index in range(1, len(points)):
        previous = points[index - 1]
        current = points[index]
        total += _haversine_distance_meters(
            previous.get("lat"),
            previous.get("lng"),
            current.get("lat"),
            current.get("lng"),
        )
    return round(total, 2)


def _haversine_distance_meters(lat1, lng1, lat2, lng2):
    try:
        lat1 = math.radians(float(lat1))
        lng1 = math.radians(float(lng1))
        lat2 = math.radians(float(lat2))
        lng2 = math.radians(float(lng2))
    except (TypeError, ValueError):
        return 0.0

    delta_lat = lat2 - lat1
    delta_lng = lng2 - lng1
    a = (
        math.sin(delta_lat / 2) ** 2
        + math.cos(lat1) * math.cos(lat2) * math.sin(delta_lng / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    earth_radius_meters = 6_371_000
    return earth_radius_meters * c


def _format_distance_label(distance_meters):
    distance_meters = float(distance_meters or 0)
    if distance_meters >= 1000:
        return f"{distance_meters / 1000:.2f} km"
    return f"{distance_meters:.0f} m"


def _format_coordinate_label(point):
    if not point:
        return "-"
    try:
        return f"{float(point.get('lat')):.6f}, {float(point.get('lng')):.6f}"
    except (TypeError, ValueError):
        return "-"


def _format_altitude_range_label(altitude_min, altitude_max):
    if altitude_min is None or altitude_max is None:
        return "-"
    if abs(float(altitude_max) - float(altitude_min)) < 0.01:
        return f"{float(altitude_min):.1f} m"
    return f"{float(altitude_min):.1f} m a {float(altitude_max):.1f} m"


def _xml_local_name(tag):
    return str(tag).split("}", 1)[-1]


def _xml_child_text(parent, local_name):
    for child in parent:
        if _xml_local_name(child.tag) == local_name:
            return (child.text or "").strip()
    return ""


def _kml_color_to_css(raw_color):
    value = (raw_color or "").strip().lstrip("#")
    if len(value) != 8:
        return None
    alpha = value[0:2]
    blue = value[2:4]
    green = value[4:6]
    red = value[6:8]
    try:
        int(alpha, 16)
        int(blue, 16)
        int(green, 16)
        int(red, 16)
    except ValueError:
        return None
    return f"#{red}{green}{blue}"


def _build_monthly_summary(rows):
    monthly = defaultdict(lambda: {"voos": 0, "area": 0.0, "volume": 0.0, "duracao": 0})
    for flight_start, area, volume, duration in rows:
        if not flight_start:
            continue
        month_start = flight_start.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        item = monthly[month_start]
        item["voos"] += 1
        item["area"] += float(area or 0)
        item["volume"] += float(volume or 0)
        item["duracao"] += int(duration or 0)

    summary = []
    for month_start in sorted(monthly.keys()):
        values = monthly[month_start]
        summary.append(
            {
                "mes_iso": month_start.strftime("%Y-%m"),
                "mes_label": month_start.strftime("%m/%Y"),
                "voos": values["voos"],
                "area": round(values["area"], 3),
                "volume": round(values["volume"], 3),
                "duracao": format_duration_seconds(values["duracao"]),
                "duracao_seconds": values["duracao"],
            }
        )
    return summary[-12:]


def _build_monthly_comparison(resumo_mensal):
    if not resumo_mensal:
        return {
            "atual": None,
            "anterior": None,
            "voos_delta_pct": None,
            "area_delta_pct": None,
            "volume_delta_pct": None,
            "duracao_delta_pct": None,
        }

    atual = resumo_mensal[-1]
    anterior = resumo_mensal[-2] if len(resumo_mensal) > 1 else None

    return {
        "atual": atual,
        "anterior": anterior,
        "voos_delta_pct": _pct_change(
            atual["voos"],
            anterior["voos"] if anterior else None,
        ),
        "area_delta_pct": _pct_change(
            atual["area"],
            anterior["area"] if anterior else None,
        ),
        "volume_delta_pct": _pct_change(
            atual["volume"],
            anterior["volume"] if anterior else None,
        ),
        "duracao_delta_pct": _pct_change(
            atual["duracao_seconds"],
            anterior["duracao_seconds"] if anterior else None,
        ),
    }


def _pct_change(current_value, previous_value):
    if previous_value in (None, 0):
        return None
    return round(((current_value - previous_value) / previous_value) * 100, 1)
