import gc
import os
import tempfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from io import BytesIO

from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Image as RLImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.extensions import db
from app.models import Solicitacao, Usuario
from app.modules.piloto_os.exporters import (
    _download_remote_media_bytes,
    _fmt_date,
    _is_remote_media_path,
    _prepare_pdf_image_source_for_canvas,
    _try_make_local_rlimage,
    _try_make_logo,
    _try_prepare_pdf_image_for_canvas,
)
from app.modules.relatorios.service import (
    _coleta_imagens_max_export_items,
    build_relatorio_coleta_imagens_export_data,
    build_relatorio_os_export_data,
)
from app.shared.access import (
    apply_prefeitura_scope,
    apply_regiao_scope,
    apply_solicitacao_prefeitura_scope,
    apply_solicitacao_regiao_scope,
)
from app.shared.query_filters import aplicar_filtros_base

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


def _env_int(name, default, minimum=None, maximum=None):
    try:
        value = int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        value = default
    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


RELATORIO_PDF_DETALHE_MAX_ROWS = _env_int("RELATORIO_PDF_DETALHE_MAX_ROWS", 300, minimum=50, maximum=2000)
COLETA_IMAGENS_PDF_IMAGE_DPI = _env_int("RELATORIO_COLETA_IMAGENS_PDF_IMAGE_DPI", 150, minimum=120, maximum=350)
COLETA_IMAGENS_PDF_JPEG_QUALITY = _env_int("RELATORIO_COLETA_IMAGENS_PDF_JPEG_QUALITY", 78, minimum=70, maximum=95)
COLETA_IMAGENS_PDF_IMAGE_SPOOL_BYTES = _env_int(
    "RELATORIO_COLETA_IMAGENS_PDF_IMAGE_SPOOL_MB",
    1,
    minimum=1,
    maximum=8,
) * 1024 * 1024
COLETA_IMAGENS_PDF_REMOTE_PREFETCH = _env_int(
    "RELATORIO_COLETA_IMAGENS_PDF_REMOTE_PREFETCH",
    0,
    minimum=0,
    maximum=3,
)
COLETA_IMAGENS_PDF_PREFETCH_MISS = object()


def _resolve_filters(user, args):
    mes = args.get("mes", datetime.now().month, type=int)
    ano = args.get("ano", datetime.now().year, type=int)
    filtro_data = f"{ano}-{mes:02d}"

    if getattr(user, "tipo_usuario", None) == "uvis":
        uvis_id = user.id
    else:
        uvis_id = args.get("uvis_id", type=int)

    return mes, ano, filtro_data, uvis_id


def _build_pdf_export_data(user, args):
    mes, ano, filtro_data, uvis_id = _resolve_filters(user, args)
    orient = args.get("orient", default="portrait")

    base_query = aplicar_filtros_base(
        db.session.query(Solicitacao),
        filtro_data,
        uvis_id,
    )
    base_query = apply_solicitacao_prefeitura_scope(base_query, user)
    base_query = apply_solicitacao_regiao_scope(base_query, user)

    query_detalhe = aplicar_filtros_base(
        db.session.query(Solicitacao, Usuario).join(Usuario, Usuario.id == Solicitacao.usuario_id),
        filtro_data,
        uvis_id,
    )
    query_detalhe = apply_solicitacao_prefeitura_scope(query_detalhe, user)
    query_detalhe = apply_regiao_scope(query_detalhe, user, Usuario.regiao)

    query_results_total = query_detalhe.count()
    query_results = (
        query_detalhe
        .order_by(Solicitacao.data_criacao.desc())
        .limit(RELATORIO_PDF_DETALHE_MAX_ROWS)
        .all()
    )

    total_solicitacoes = base_query.count()
    total_aprovadas = base_query.filter(Solicitacao.status == "APROVADO").count()
    total_aprovadas_com_recomendacoes = base_query.filter(
        Solicitacao.status == "APROVADO COM RECOMENDAÇÕES"
    ).count()
    total_recusadas = base_query.filter(Solicitacao.status == "NEGADO").count()
    total_analise = base_query.filter(Solicitacao.status == "EM ANÁLISE").count()
    total_pendentes = base_query.filter(Solicitacao.status == "PENDENTE").count()

    dados_regiao = [
        (regiao or "Não informado", total)
        for regiao, total in (
            apply_regiao_scope(
                apply_solicitacao_prefeitura_scope(
                    aplicar_filtros_base(
                        db.session.query(Usuario.regiao, db.func.count(Solicitacao.id)).join(Usuario),
                        filtro_data,
                        uvis_id,
                    ),
                    user,
                ),
                user,
                Usuario.regiao,
            )
            .group_by(Usuario.regiao)
            .all()
        )
    ]

    dados_status = [
        (status or "Não informado", total)
        for status, total in (
            base_query
            .with_entities(Solicitacao.status, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.status)
            .all()
        )
    ]

    dados_foco = [
        (foco or "Não informado", total)
        for foco, total in (
            base_query
            .with_entities(Solicitacao.foco, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.foco)
            .all()
        )
    ]

    dados_tipo_visita = [
        (tipo or "Não informado", total)
        for tipo, total in (
            base_query
            .with_entities(Solicitacao.tipo_visita, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.tipo_visita)
            .all()
        )
    ]

    dados_tipo_imovel = [
        (tipo or "Não informado", total)
        for tipo, total in (
            base_query
            .with_entities(Solicitacao.tipo_imovel, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.tipo_imovel)
            .all()
        )
    ]

    dados_tipo_operacao = [
        (tipo or "Não informado", total)
        for tipo, total in (
            base_query
            .with_entities(Solicitacao.tipo_operacao, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.tipo_operacao)
            .all()
        )
    ]

    dados_altura_voo = [
        (altura or "Não informado", total)
        for altura, total in (
            base_query
            .with_entities(Solicitacao.altura_voo, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.altura_voo)
            .all()
        )
    ]

    dados_unidade = [
        (uvis_nome or "Não informado", total)
        for uvis_nome, total in (
            apply_regiao_scope(
                apply_solicitacao_prefeitura_scope(
                    aplicar_filtros_base(
                        db.session.query(Usuario.nome_uvis, db.func.count(Solicitacao.id))
                        .join(Usuario)
                        .filter(Usuario.tipo_usuario == "uvis"),
                        filtro_data,
                        uvis_id,
                    ),
                    user,
                ),
                user,
                Usuario.regiao,
            )
            .group_by(Usuario.nome_uvis)
            .all()
        )
    ]

    if db.engine.name == "postgresql":
        func_mes = db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM")
    else:
        func_mes = db.func.strftime("%Y-%m", Solicitacao.data_agendamento)

    dados_mensais = [
        tuple(row)
        for row in (
            apply_solicitacao_regiao_scope(
                apply_solicitacao_prefeitura_scope(
                    db.session.query(func_mes.label("mes"), db.func.count(Solicitacao.id)).filter(
                        Solicitacao.data_agendamento.isnot(None)
                    ),
                    user,
                ),
                user,
            )
            .group_by("mes")
            .order_by("mes")
            .all()
        )
    ]

    return {
        "mes": mes,
        "ano": ano,
        "orient": orient,
        "filtro_data": filtro_data,
        "uvis_id": uvis_id,
        "query_results": query_results,
        "query_results_total": query_results_total,
        "query_results_limit": RELATORIO_PDF_DETALHE_MAX_ROWS,
        "query_results_limited": query_results_total > RELATORIO_PDF_DETALHE_MAX_ROWS,
        "total_solicitacoes": total_solicitacoes,
        "total_aprovadas": total_aprovadas,
        "total_aprovadas_com_recomendacoes": total_aprovadas_com_recomendacoes,
        "total_recusadas": total_recusadas,
        "total_analise": total_analise,
        "total_pendentes": total_pendentes,
        "dados_regiao": dados_regiao,
        "dados_status": dados_status,
        "dados_foco": dados_foco,
        "dados_tipo_operacao": dados_tipo_operacao,
        "dados_tipo_visita": dados_tipo_visita,
        "dados_tipo_imovel": dados_tipo_imovel,
        "dados_altura_voo": dados_altura_voo,
        "dados_unidade": dados_unidade,
        "dados_mensais": dados_mensais,
    }


def build_relatorio_pdf_export(user, args):
    data = _build_pdf_export_data(user, args)

    status_colors = {
        "APROVADO": "#2f855a",
        "APROVADO COM RECOMENDAÇÕES": "#F7630C",
        "EM ANÁLISE": "#f3e526",
        "PENDENTE": "#718096",
        "NEGADO": "#e53e3e",
        "CANCELADO": "#343a40",
        "CONCLUÍDO": "#9B30FF",
        "CONCLUIDO": "#9B30FF",
    }

    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    caminho_pdf = tmp_pdf.name
    tmp_pdf.close()

    pagesize = landscape(A4) if data["orient"] == "landscape" else A4
    doc = SimpleDocTemplate(
        caminho_pdf,
        pagesize=pagesize,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        alignment=1,
        textColor=colors.HexColor("#0d6efd"),
        spaceAfter=10,
    )
    subtitle_style = ParagraphStyle(
        "subtitle",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        alignment=1,
        textColor=colors.HexColor("#555"),
        spaceAfter=12,
    )
    section_h = ParagraphStyle(
        "sec",
        parent=styles["Heading2"],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#0d6efd"),
        spaceBefore=10,
        spaceAfter=6,
    )
    normal = ParagraphStyle(
        "normal",
        parent=styles["Normal"],
        fontSize=9.5,
        leading=13,
    )
    cell_style = ParagraphStyle(
        "cell",
        parent=styles["BodyText"],
        fontSize=8.6,
        leading=11,
        textColor=colors.HexColor("#222"),
        wordWrap="CJK",
        splitLongWords=True,
    )

    story = []
    story.append(Paragraph(f"Relatório Mensal — {data['mes']:02d}/{data['ano']}", title_style))

    filtro_txt = f"Filtro: {data['filtro_data']}"
    if data["uvis_id"]:
        filtro_txt += f" | UVIS ID: {data['uvis_id']}"
    else:
        filtro_txt += " | UVIS: Todas"
    story.append(Paragraph(filtro_txt, subtitle_style))

    def resumo_cards():
        cards = [
            ("Total", data["total_solicitacoes"], "#0d6efd"),
            ("Aprovadas", data["total_aprovadas"], "#198754"),
            ("Aprov. c/ Recom.", data["total_aprovadas_com_recomendacoes"], "#F7630C"),
            ("Negadas", data["total_recusadas"], "#dc3545"),
            ("Em Análise", data["total_analise"], "#ffc107"),
            ("Pendentes", data["total_pendentes"], "#718096"),
        ]

        rows = []
        row = []
        for label, value, hexcolor in cards:
            box = Table(
                [
                    [Paragraph(label, ParagraphStyle("l", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#666")))],
                    [Paragraph(str(value), ParagraphStyle("v", parent=styles["Normal"], fontSize=18, leading=20, textColor=colors.HexColor(hexcolor)))],
                ],
                colWidths=[48 * mm] if data["orient"] == "portrait" else [52 * mm],
            )
            box.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8f9fa")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#e5e7eb")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            row.append(box)
            if len(row) == 3:
                rows.append(row)
                row = []

        if row:
            while len(row) < 3:
                row.append(Spacer(1, 1))
            rows.append(row)

        grid = Table(rows, colWidths=None)
        grid.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        return grid

    story.append(resumo_cards())
    story.append(Spacer(1, 10))

    def add_count_table(titulo, dados, col1="Categoria"):
        story.append(Paragraph(titulo, section_h))
        rows = [[
            Paragraph(col1, ParagraphStyle("th", parent=cell_style, textColor=colors.white, fontSize=9)),
            Paragraph("Total", ParagraphStyle("th2", parent=cell_style, textColor=colors.white, fontSize=9)),
        ]]

        for nome, total in (dados or [("Nenhum", 0)]):
            rows.append([Paragraph(str(nome), cell_style), Paragraph(str(total), cell_style)])

        tbl = Table(rows, repeatRows=1, colWidths=[140 * mm, 25 * mm] if data["orient"] == "portrait" else [190 * mm, 30 * mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9dee7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fbfdff")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 10))

    story.append(Paragraph("Resumo por Agrupamentos", section_h))
    story.append(Paragraph("Abaixo estão os agrupamentos do mês selecionado, apresentados em formato de tabela.", normal))
    story.append(Spacer(1, 6))

    add_count_table("Agrupamento — Região", data["dados_regiao"])
    add_count_table("Agrupamento — Status", data["dados_status"])
    add_count_table("Agrupamento — Foco", data["dados_foco"])
    add_count_table("Agrupamento — Tipo de Operação", data["dados_tipo_operacao"])
    add_count_table("Agrupamento — Tipo de Visita", data["dados_tipo_visita"])
    add_count_table("Agrupamento — Tipo de Imovel", data["dados_tipo_imovel"])
    add_count_table("Agrupamento — Altura do Voo", data["dados_altura_voo"])
    add_count_table("Agrupamento — Unidade (UVIS)", data["dados_unidade"])
    add_count_table("Histórico Mensal (tabela)", data["dados_mensais"], col1="Mês")

    story.append(PageBreak())
    story.append(Paragraph("Gráficos", section_h))
    story.append(Paragraph("Os gráficos abaixo representam visualmente os dados apresentados nas tabelas anteriores.", normal))
    story.append(Spacer(1, 8))

    def safe_img_from_plt(fig, width_mm=170):
        bio = BytesIO()
        fig.tight_layout()
        fig.savefig(bio, format="png", dpi=220, bbox_inches="tight")
        plt.close(fig)
        bio.seek(0)
        return RLImage(bio, width=width_mm * mm)

    if MATPLOTLIB_AVAILABLE:
        try:
            labels = [status for status, _ in data["dados_status"]]
            values = [count for _, count in data["dados_status"]]
            colors_status = [status_colors.get(status, "#bdc3c7") for status in labels]

            fig1, ax1 = plt.subplots(figsize=(6.4, 3.0))

            def autopct(percent):
                return f"{percent:.0f}%" if percent >= 6 else ""

            wedges, *_ = ax1.pie(
                values or [1],
                labels=None,
                colors=colors_status,
                autopct=autopct,
                startangle=90,
                pctdistance=0.78,
                textprops={"fontsize": 9},
            )
            centre_circle = plt.Circle((0, 0), 0.58, fc="white")
            ax1.add_artist(centre_circle)
            ax1.legend(wedges, labels, loc="center left", bbox_to_anchor=(1.02, 0.5), fontsize=9, frameon=False)
            ax1.set_title("Distribuição por Status", fontsize=11, pad=10)
            ax1.axis("equal")
            story.append(safe_img_from_plt(fig1, width_mm=170))
            story.append(Spacer(1, 10))

            u_names = [uvis for uvis, _ in data["dados_unidade"][:10]]
            u_vals = [count for _, count in data["dados_unidade"][:10]]
            fig2, ax2 = plt.subplots(figsize=(7.2, 3.0))
            ax2.barh(u_names[::-1] or ["Nenhum"], u_vals[::-1] or [0])
            ax2.set_xlabel("Total", fontsize=9)
            ax2.set_title("Top UVIS", fontsize=11, pad=10)
            ax2.tick_params(axis="both", labelsize=9)
            ax2.grid(axis="x", linestyle=":", linewidth=0.6, alpha=0.6)
            story.append(safe_img_from_plt(fig2, width_mm=180 if data["orient"] == "landscape" else 170))
            story.append(Spacer(1, 10))

            months = [mes for mes, _ in data["dados_mensais"]]
            counts = [count for _, count in data["dados_mensais"]]
            fig3, ax3 = plt.subplots(figsize=(7.2, 3.0))
            if months:
                ax3.plot(range(len(months)), counts, marker="o", linewidth=1.6)
                ax3.set_xticks(range(len(months)))
                ax3.set_xticklabels(months, rotation=45, ha="right", fontsize=9)
            ax3.set_title("Histórico Mensal", fontsize=11, pad=10)
            ax3.tick_params(axis="y", labelsize=9)
            ax3.grid(axis="y", linestyle=":", linewidth=0.6, alpha=0.6)
            story.append(safe_img_from_plt(fig3, width_mm=185 if data["orient"] == "landscape" else 170))
            story.append(Spacer(1, 8))
        except Exception:
            story.append(Paragraph("Gráficos indisponíveis (erro ao gerar).", normal))
    else:
        story.append(Paragraph("Matplotlib não disponível — gráficos foram omitidos.", normal))

    story.append(PageBreak())
    story.append(Paragraph("Registros Detalhados", section_h))
    if data.get("query_results_limited"):
        story.append(Paragraph(
            (
                f"Totais e agrupamentos consideram {data['query_results_total']} registros. "
                f"Para preservar memoria do servidor, o detalhamento abaixo exibe os "
                f"{data['query_results_limit']} registros mais recentes. Refine os filtros "
                "para obter um detalhamento completo de um periodo menor."
            ),
            normal,
        ))
    else:
        story.append(Paragraph("Listagem completa dos registros retornados pelo filtro selecionado.", normal))
    story.append(Spacer(1, 8))

    registros_header = [
        "Data",
        "Hora",
        "Unidade",
        "Região",
        "Protocolo",
        "Status",
        "Foco",
        "Tipo Operação",
        "Tipo Visita",
        "Tipo Imovel",
        "Altura Voo",
        "Observação",
    ]
    hdr_style = ParagraphStyle("hdr", parent=cell_style, textColor=colors.white, fontSize=7.8, leading=9.2)
    cell_style_small = ParagraphStyle(
        "cell_small",
        parent=cell_style,
        fontSize=7.6,
        leading=9.2,
        wordWrap="CJK",
        splitLongWords=True,
    )

    registros_rows = [[Paragraph(header, hdr_style) for header in registros_header]]
    for solicitacao, usuario in data["query_results"]:
        data_str = solicitacao.data_criacao.strftime("%d/%m/%Y") if getattr(solicitacao, "data_criacao", None) else ""
        hora_str = getattr(solicitacao, "hora_agendamento", "")
        hora_str = hora_str.strftime("%H:%M") if hasattr(hora_str, "strftime") else str(hora_str or "")

        unidade = getattr(usuario, "nome_uvis", "") or "Não informado"
        regiao = getattr(usuario, "regiao", "") or "Não informado"
        protocolo = getattr(solicitacao, "protocolo", "") or ""
        status = getattr(solicitacao, "status", "") or ""
        foco = getattr(solicitacao, "foco", "") or ""
        tipo_operacao = getattr(solicitacao, "tipo_operacao", "") or ""
        tipo_visita = getattr(solicitacao, "tipo_visita", "") or ""
        tipo_imovel = getattr(solicitacao, "tipo_imovel", "") or ""
        altura_voo = getattr(solicitacao, "altura_voo", "") or ""
        observacao = getattr(solicitacao, "observacao", "") or ""

        registros_rows.append([
            Paragraph(str(data_str), cell_style_small),
            Paragraph(str(hora_str), cell_style_small),
            Paragraph(str(unidade), cell_style_small),
            Paragraph(str(regiao), cell_style_small),
            Paragraph(str(protocolo), cell_style_small),
            Paragraph(str(status), cell_style_small),
            Paragraph(str(foco), cell_style_small),
            Paragraph(str(tipo_operacao), cell_style_small),
            Paragraph(str(tipo_visita), cell_style_small),
            Paragraph(str(tipo_imovel), cell_style_small),
            Paragraph(str(altura_voo), cell_style_small),
            Paragraph(str(observacao), cell_style_small),
        ])

    base_col_widths = [18 * mm, 14 * mm, 22 * mm, 18 * mm, 18 * mm, 20 * mm, 18 * mm, 20 * mm, 18 * mm, 18 * mm, 16 * mm, 34 * mm]
    total_w = sum(base_col_widths)
    max_w = doc.width
    col_widths = [w * (max_w / total_w) for w in base_col_widths] if total_w > max_w else base_col_widths
    chunk_size = 28 if data["orient"] == "landscape" else 24

    for index in range(0, len(registros_rows), chunk_size):
        chunk = registros_rows[index:index + chunk_size]
        tbl = Table(chunk, repeatRows=1, colWidths=col_widths, hAlign="LEFT")
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9dee7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fbfdff")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, 0), "LEFT"),
            ("ALIGN", (0, 1), (-1, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 6))
        if index + chunk_size < len(registros_rows):
            story.append(PageBreak())

    def _header_footer(canvas, doc_):
        canvas.saveState()
        _, height = pagesize
        canvas.setFillColor(colors.HexColor("#0d6efd"))
        canvas.rect(doc_.leftMargin, height - (12 * mm), doc_.width, 3, fill=1, stroke=0)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#777"))
        canvas.drawString(doc_.leftMargin, 9 * mm, f"Relatório — {data['mes']:02d}/{data['ano']} — IJASystem")
        canvas.drawRightString(doc_.leftMargin + doc_.width, 9 * mm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

    nome_arquivo = f"relatorio_OceanoAzul_{data['ano']}_{data['mes']:02d}"
    if data["uvis_id"]:
        nome_arquivo += f"_UVIS_{data['uvis_id']}"

    return caminho_pdf, f"{nome_arquivo}.pdf"


def _montar_endereco(row):
    partes_rua = []
    if row.logradouro:
        partes_rua.append(row.logradouro.strip())
    if row.numero is not None and str(row.numero).strip():
        partes_rua.append(str(row.numero).strip())

    rua_numero = ", ".join([parte for parte in partes_rua if parte]).strip()

    cidade_uf = ""
    if row.cidade and row.uf:
        cidade_uf = f"{row.cidade.strip()}/{row.uf.strip()}"
    elif row.cidade:
        cidade_uf = row.cidade.strip()
    elif row.uf:
        cidade_uf = row.uf.strip()

    bairro_cidade = " - ".join([parte for parte in [(row.bairro or "").strip(), cidade_uf] if parte]).strip()
    cep_txt = f"CEP {row.cep.strip()}" if row.cep else ""
    return " | ".join([parte for parte in [rua_numero, bairro_cidade, cep_txt] if parte])


def build_relatorio_excel_export(user, args):
    mes, ano, filtro_data, uvis_id = _resolve_filters(user, args)

    query_dados = db.session.query(
        Solicitacao.id,
        Solicitacao.status,
        Solicitacao.foco,
        Solicitacao.tipo_operacao,
        Solicitacao.tipo_visita,
        Solicitacao.tipo_imovel,
        Solicitacao.altura_voo,
        Solicitacao.data_agendamento,
        Solicitacao.hora_agendamento,
        Solicitacao.cep,
        Solicitacao.logradouro,
        Solicitacao.numero,
        Solicitacao.bairro,
        Solicitacao.cidade,
        Solicitacao.uf,
        Solicitacao.latitude,
        Solicitacao.longitude,
        Usuario.nome_uvis,
        Usuario.regiao,
    ).join(Usuario, Usuario.id == Solicitacao.usuario_id)
    query_dados = apply_solicitacao_prefeitura_scope(query_dados, user)
    query_dados = apply_regiao_scope(query_dados, user, Usuario.regiao)

    if db.engine.name == "postgresql":
        query_dados = query_dados.filter(
            Solicitacao.data_agendamento.isnot(None),
            db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM") == filtro_data,
        )
    else:
        query_dados = query_dados.filter(
            Solicitacao.data_agendamento.isnot(None),
            db.func.strftime("%Y-%m", Solicitacao.data_agendamento) == filtro_data,
        )

    if uvis_id:
        query_dados = query_dados.filter(Solicitacao.usuario_id == uvis_id)

    dados = query_dados.order_by(
        Solicitacao.data_agendamento.desc(),
        Solicitacao.hora_agendamento.desc(),
    ).all()

    nome_uvis_filtro = None
    if uvis_id:
        nome_uvis_filtro = (
            apply_regiao_scope(
                apply_prefeitura_scope(
                    db.session.query(Usuario.nome_uvis).filter(Usuario.id == uvis_id),
                    user,
                    Usuario.prefeitura_id,
                ),
                user,
                Usuario.regiao,
            )
            .scalar()
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    colunas = [
        "UVIS",
        "Região",
        "ID",
        "Status",
        "Foco",
        "Tipo Operacao",
        "Tipo Visita",
        "Tipo Imovel",
        "Altura Voo",
        "Data Agendamento",
        "Hora Agendamento",
        "ENDEREÇO DE AÇÃO",
        "Latitude",
        "Longitude",
    ]

    header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra1 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    zebra2 = PatternFill(start_color="FFF7FBFF", end_color="FFF7FBFF", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left_center = Alignment(horizontal="left", vertical="center")

    for col_num, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    ws.row_dimensions[1].height = 22

    for row_num, row in enumerate(dados, 2):
        data_agendamento_fmt = row.data_agendamento.strftime("%d/%m/%Y") if row.data_agendamento else ""
        hora_agendamento_fmt = row.hora_agendamento.strftime("%H:%M") if row.hora_agendamento else ""
        endereco_acao = _montar_endereco(row)
        values = [
            row.nome_uvis,
            row.regiao,
            row.id,
            row.status,
            row.foco,
            row.tipo_operacao,
            row.tipo_visita,
            row.tipo_imovel,
            row.altura_voo,
            data_agendamento_fmt,
            hora_agendamento_fmt,
            endereco_acao,
            row.latitude,
            row.longitude,
        ]

        ws.row_dimensions[row_num].height = 20
        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_index, value=value)
            cell.border = thin_border
            cell.fill = zebra1 if row_num % 2 == 0 else zebra2
            cell.alignment = center if col_index in (3, 9, 10, 11, 13, 14) else left_center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    larguras = {
        "A": 24,
        "B": 12,
        "C": 6,
        "D": 18,
        "E": 22,
        "F": 16,
        "G": 16,
        "H": 18,
        "I": 10,
        "J": 14,
        "K": 14,
        "L": 90,
        "M": 14,
        "N": 14,
    }
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"relatorio_OceanoAzul_{ano}_{mes:02d}"
    if uvis_id:
        safe_nome = (nome_uvis_filtro or f"ID_{uvis_id}").replace(" ", "_")
        nome_arquivo += f"_UVIS_{safe_nome}"

    return output, f"{nome_arquivo}.xlsx"


OS_THIN = Side(style="thin", color="D0D7DE")
OS_BORDER = Border(left=OS_THIN, right=OS_THIN, top=OS_THIN, bottom=OS_THIN)
OS_FILL_HEADER = PatternFill("solid", fgColor="0D6EFD")
OS_FILL_SECTION = PatternFill("solid", fgColor="EAF2FF")
OS_FILL_ZEBRA = PatternFill("solid", fgColor="FBFDFF")
OS_FONT_HEADER = Font(bold=True, color="FFFFFF")
OS_FONT_TITLE = Font(bold=True, size=16, color="0D6EFD")
OS_FONT_SUBTITLE = Font(size=10, color="555555")
OS_FONT_SECTION = Font(bold=True, color="0D6EFD")
FIELD_CONTROL_HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
FIELD_CONTROL_HEADER_FONT = Font(name="Calibri", size=11, color="1F1F1F", bold=True)
FIELD_CONTROL_BORDER_SIDE = Side(style="thin", color="D9D9D9")
FIELD_CONTROL_BORDER = Border(
    left=FIELD_CONTROL_BORDER_SIDE,
    right=FIELD_CONTROL_BORDER_SIDE,
    top=FIELD_CONTROL_BORDER_SIDE,
    bottom=FIELD_CONTROL_BORDER_SIDE,
)
FIELD_CONTROL_BODY_FONT = Font(name="Calibri", size=10, color="1F2937")


def _os_fmt_dt(value):
    if not value:
        return ""
    try:
        return value.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(value)


def _os_safe(value):
    if value is None:
        return ""
    return str(value)


def _os_excel_add_title(ws, title: str, subtitle: str = ""):
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = OS_FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    ws["A2"] = subtitle
    ws["A2"].font = OS_FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18


def _os_excel_add_section(ws, row: int, title: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = OS_FILL_SECTION
    cell.font = OS_FONT_SECTION
    cell.alignment = Alignment(vertical="center")
    cell.border = OS_BORDER
    ws.cell(row=row, column=2).border = OS_BORDER
    ws.row_dimensions[row].height = 18


def _os_excel_apply_table_style(ws, header_row: int, end_row: int, col_count: int = 2):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = OS_FILL_HEADER
        cell.font = OS_FONT_HEADER
        cell.border = OS_BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for row in range(header_row + 1, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = OS_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if (row - header_row) % 2 == 0:
                cell.fill = OS_FILL_ZEBRA


def _os_excel_write_kv(ws, start_row: int, items: list[tuple[str, object]]):
    ws.cell(row=start_row, column=1, value="Campo")
    ws.cell(row=start_row, column=2, value="Valor")

    row = start_row + 1
    for key, value in items:
        ws.cell(row=row, column=1, value=str(key))
        ws.cell(row=row, column=2, value=_os_safe(value))
        row += 1

    _os_excel_apply_table_style(ws, start_row, row - 1, col_count=2)
    return row


def _os_excel_write_table(ws, start_row: int, headers: list[str], rows: list[tuple], col_widths=None):
    for index, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=index, value=header)

    row = start_row + 1
    for values in rows:
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=_os_safe(value))
        row += 1

    _os_excel_apply_table_style(ws, start_row, row - 1, col_count=len(headers))

    if col_widths:
        for index, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

    return row


def _os_excel_auto_width(ws, max_col=2, min_w=12, max_w=60):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = 0
        for cell in ws[letter]:
            if cell.value:
                best = max(best, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, best + 2))


def _os_fmt_date(value):
    if not value:
        return ""
    try:
        return value.strftime("%d/%m/%Y")
    except Exception:
        return str(value)


def _os_fmt_time(value):
    if not value:
        return ""
    try:
        return value.strftime("%H:%M")
    except Exception:
        return str(value)


def _os_fmt_dt_seconds(value):
    if not value:
        return ""
    try:
        return value.strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        return str(value)


def _os_yes_registered(value):
    return "REGISTRADA" if value else ""


def _os_cliente_nome(solicitacao):
    prefeitura = getattr(solicitacao, "prefeitura", None)
    return getattr(prefeitura, "nome", None) or "PMSP"


def _os_localizacao_nome(solicitacao):
    usuario = getattr(solicitacao, "usuario", None)
    equipe = getattr(solicitacao, "equipe", None)
    return (
        getattr(usuario, "nome_uvis", None)
        or getattr(equipe, "nome_equipe", None)
        or getattr(solicitacao, "equipe_uvis_nome", None)
        or ""
    )


def _os_endereco_completo(solicitacao):
    parts = []
    if getattr(solicitacao, "logradouro", None):
        numero = getattr(solicitacao, "numero", None) or "S/N"
        parts.append(f"{solicitacao.logradouro}, {numero}")
    if getattr(solicitacao, "bairro", None):
        parts.append(solicitacao.bairro)
    cidade_uf = " - ".join(part for part in [getattr(solicitacao, "cidade", None), getattr(solicitacao, "uf", None)] if part)
    if cidade_uf:
        parts.append(cidade_uf)
    if getattr(solicitacao, "cep", None):
        parts.append(solicitacao.cep)
    if getattr(solicitacao, "complemento", None):
        parts.append(solicitacao.complemento)
    return " - ".join(parts)


def _os_status_export(solicitacao, ordem):
    status = (getattr(solicitacao, "status", None) or "").strip()
    if status.upper() in {"CONCLUIDO", "CONCLUÍDO"} or getattr(ordem, "respondido_em", None):
        return "Concluído"
    return status


def _os_duration_text(start_time, end_time):
    if not start_time or not end_time:
        return ""
    try:
        start_minutes = (start_time.hour * 60) + start_time.minute
        end_minutes = (end_time.hour * 60) + end_time.minute
        if end_minutes < start_minutes:
            end_minutes += 24 * 60
        minutes = max(0, end_minutes - start_minutes)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    except Exception:
        return ""


def _os_midias_text(ordem):
    imagens = getattr(ordem, "quantidade_imagens_registradas", None)
    videos = getattr(ordem, "quantidade_videos_registradas", None)
    parts = []
    if imagens not in (None, ""):
        parts.append(f"{imagens} FOTO{'S' if int(imagens or 0) != 1 else ''}")
    if videos not in (None, ""):
        parts.append(f"{videos} VIDEO{'S' if int(videos or 0) != 1 else ''}")
    return "; ".join(parts)


def _os_equipamentos_text(ordem):
    equipamentos = [
        getattr(ordem, "prefixo_aeronave_pulverizacao", None),
        getattr(ordem, "prefixo_aeronave_monitoramento", None),
    ]
    return "; ".join(item for item in equipamentos if item)


def _setup_field_control_sheet(ws, headers, widths):
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = FIELD_CONTROL_HEADER_FILL
        cell.font = FIELD_CONTROL_HEADER_FONT
        cell.border = FIELD_CONTROL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 16)


def _finalize_field_control_sheet(ws):
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).border = FIELD_CONTROL_BORDER


ATIVIDADES_HEADERS = [
    "Identificador da OS",
    "Etiqueta da Tarefa",
    "Etiqueta do Cliente",
    "Data do deslocamento",
    "Hora do deslocamento",
    "Duração do deslocamento",
    "Data agendada",
    "Hora",
    "Prazo",
    "Data/Hora limite da resposta",
    "Data/hora da conclusão da resposta",
    "Situação do prazo de resposta",
    "Data/Hora da solução",
    "Data/Hora da conclusão da solução",
    "Situação do prazo de solução",
    "Nome do cliente",
    "Nome da Localização",
    "Cidade",
    "Estado",
    "Endereço Completo",
    "Descrição",
    "Descrição da Tarefa",
    "Status",
    "Data de início",
    "Hora de início",
    "Data de finalização",
    "Hora de finalização",
    "Duração",
    "Duração estimada",
    "Descrição fechamento",
    "Avaliação",
    "Comentário da avaliação",
    "Tipo de OS",
    "Criador",
    "Tipo do criador",
    "Colaborador",
    "Colaboradores secundários",
    "Equipamentos",
    "Data de criação",
    "Hora de criação",
    "Criado em",
    "Link do relatório",
    "Latitude do colaborador no início da atividade",
    "Longitude do colaborador no início da atividade",
    "Latitude do colaborador na finalização da atividade",
    "Longitude do colaborador na finalização da atividade",
    "Data de criação da solicitação",
    "ID",
    "Duração precisa",
    "Data de abertura da solicitação",
    "Contrato de prazo",
    "Arquivado",
    "Última atualização da OS",
    "Descrição da Localização",
    "Identificador de importação",
    "Link do relatório estendido",
]


FORMULARIO_RD_HEADERS = [
    "Nome do Formulário",
    "Situação",
    "Ordem de Serviço",
    "Nome do cliente",
    "Nome da Localização",
    "Respondido em",
    "Respondido por",
    "SITUAÇÃO DA APLICAÇÃO",
    "LARVA VISUALIZADA",
    "RETORNAR NA PROXIMA SEMANA PARA MONITORAR LARVAS?",
    "DA",
    "NOME E RF DO ACE RESPONSAVEL PELA OS",
    "CRIADOURO DA OS (TIPO E VOLUME)",
    "DATA DA APLICAÇÃO",
    "HORA DO INICIO DA APLICAÇÃO",
    "HORA DO TERMINO DA APLICAÇÃO",
    "TRATAMENTO ADICIONAL REALIZADO?",
    "QUANTOS? QUAIS?",
    "DESCRIÇÃO DO PRODUTO",
    "FORMULAÇÃO DO PRODUTO",
    "DOSAGEM (g/10L)",
    "TIPO DE APLICAÇÃO",
    "QUANTIDADE DE PRODUTO ADMNISTRADA (ML)",
    "PULVERIZAÇÃO DE AREA - liq - (l/ha)",
    "PULVERIZAÇÃO DE FOCO-  liq - (TEMPO ESTIMADO DE APLICAÇÃO)(SEGUNDOS)",
    "PULVERIZAÇÃO DE FOCO - liq - (l/min)",
    "PREFIXO AERONAVE DE PULVERIZAÇÃO",
    "PREFIXO DA AERONAVE DE MONITORAMENTO",
    "QUANTIDADE DE IMAGENS REGISTRADAS",
    "PONTA DE PULVERIZAÇÃO",
    "TEMPERATURA (°C)",
    "UMIDADE RELATIVA (%)",
    "VELOCIDADE DO VENTO (km/h)",
    "MOTIVO DE NAO REALIZAÇÃO",
    "OBSERVAÇÕES",
    "PILOTO",
    "ASSINATURA PILOTO",
    "AUXILIAR",
    "PROPRIETÁRIO OU PREPOSTO",
    "ASSINATURA PROPRIETÁRIO OU PREPOSTO",
]


ATIVIDADES_WIDTHS = {
    1: 20,
    17: 40,
    18: 20,
    20: 100,
    21: 20,
    48: 39,
    49: 20,
}


FORMULARIO_RD_WIDTHS = {
    1: 20,
    5: 40,
    6: 20,
    8: 58,
    9: 20,
    12: 21,
    13: 35,
    14: 20,
    18: 85,
    19: 20,
    20: 22,
    21: 20,
    22: 30,
    23: 20,
    28: 26,
    29: 20,
    34: 50,
    35: 100,
    36: 29,
    37: 100,
    38: 26,
    39: 21,
    40: 100,
}


def _build_atividades_row(ordem):
    solicitacao = ordem.solicitacao
    duracao = _os_duration_text(
        getattr(ordem, "hora_inicio_aplicacao", None),
        getattr(ordem, "hora_termino_aplicacao", None),
    )
    usuario = getattr(solicitacao, "usuario", None)
    identificador = getattr(ordem, "identificador_os", None) or str(getattr(solicitacao, "id", ""))
    data_final = getattr(ordem, "data_aplicacao", None)
    hora_final = getattr(ordem, "hora_termino_aplicacao", None)
    if getattr(ordem, "respondido_em", None):
        data_final = data_final or ordem.respondido_em
        hora_final = hora_final or ordem.respondido_em

    return [
        identificador,
        "",
        "",
        "",
        "",
        "",
        _os_fmt_date(getattr(solicitacao, "data_agendamento", None)),
        _os_fmt_time(getattr(solicitacao, "hora_agendamento", None)),
        "",
        "",
        _os_fmt_dt_seconds(getattr(ordem, "respondido_em", None)),
        "",
        "",
        _os_fmt_dt_seconds(getattr(ordem, "respondido_em", None)),
        "",
        _os_cliente_nome(solicitacao),
        _os_localizacao_nome(solicitacao),
        getattr(solicitacao, "cidade", None) or "",
        getattr(solicitacao, "uf", None) or "",
        _os_endereco_completo(solicitacao),
        getattr(solicitacao, "observacao", None) or "",
        getattr(solicitacao, "foco", None) or "",
        _os_status_export(solicitacao, ordem),
        _os_fmt_date(getattr(ordem, "data_aplicacao", None)),
        _os_fmt_time(getattr(ordem, "hora_inicio_aplicacao", None)),
        _os_fmt_date(data_final),
        _os_fmt_time(hora_final),
        duracao,
        "",
        getattr(ordem, "observacoes", None) or getattr(ordem, "motivo_nao_realizacao", None) or "",
        "",
        "",
        getattr(solicitacao, "tipo_operacao", None) or "",
        getattr(usuario, "login", None) or getattr(usuario, "nome_uvis", None) or "",
        getattr(usuario, "tipo_usuario", None) or "",
        getattr(ordem, "piloto", None) or "",
        getattr(ordem, "auxiliar", None) or "",
        _os_equipamentos_text(ordem),
        _os_fmt_date(getattr(solicitacao, "data_criacao", None)),
        _os_fmt_time(getattr(solicitacao, "data_criacao", None)),
        _os_fmt_dt_seconds(getattr(solicitacao, "data_criacao", None)),
        "",
        "",
        "",
        "",
        "",
        _os_fmt_date(getattr(solicitacao, "data_criacao", None)),
        getattr(solicitacao, "id", None) or "",
        duracao,
        _os_fmt_date(getattr(solicitacao, "data_criacao", None)),
        "",
        "",
        _os_fmt_dt_seconds(getattr(ordem, "respondido_em", None)),
        getattr(solicitacao, "complemento", None) or getattr(solicitacao, "bairro", None) or "",
        getattr(solicitacao, "protocolo", None) or "",
        "",
    ]


def _build_formulario_rd_row(ordem):
    solicitacao = ordem.solicitacao
    identificador = getattr(ordem, "identificador_os", None) or str(getattr(solicitacao, "id", ""))

    return [
        "RD - PROJETO SP",
        _os_status_export(solicitacao, ordem),
        identificador,
        _os_cliente_nome(solicitacao),
        _os_localizacao_nome(solicitacao),
        _os_fmt_dt_seconds(getattr(ordem, "respondido_em", None)),
        getattr(ordem, "respondido_por", None) or "",
        getattr(ordem, "situacao_aplicacao", None) or "",
        getattr(ordem, "larva_visualizada", None) or "",
        getattr(ordem, "retornar_proxima_semana_monitorar_larvas", None) or "",
        getattr(ordem, "distrito_administrativo", None) or "",
        getattr(ordem, "nome_rf_ace_responsavel_os", None) or "",
        getattr(ordem, "criadouro_os_tipo_volume", None) or "",
        _os_fmt_date(getattr(ordem, "data_aplicacao", None)),
        _os_fmt_time(getattr(ordem, "hora_inicio_aplicacao", None)),
        _os_fmt_time(getattr(ordem, "hora_termino_aplicacao", None)),
        getattr(ordem, "tratamento_adicional_realizado", None) or "",
        getattr(ordem, "quantos_quais", None) or "",
        getattr(ordem, "descricao_produto", None) or "",
        getattr(ordem, "formulacao_produto", None) or "",
        getattr(ordem, "dosagem_g_10l", None) or "",
        getattr(ordem, "tipo_aplicacao", None) or "",
        getattr(ordem, "quantidade_produto_administrada_ml", None) or "",
        getattr(ordem, "pulverizacao_area_l_ha", None) or "",
        getattr(ordem, "pulverizacao_foco_tempo_estimado_segundos", None) or "",
        getattr(ordem, "pulverizacao_foco_l_min", None) or "",
        getattr(ordem, "prefixo_aeronave_pulverizacao", None) or "",
        getattr(ordem, "prefixo_aeronave_monitoramento", None) or "",
        _os_midias_text(ordem),
        getattr(ordem, "ponta_pulverizacao", None) or "",
        getattr(ordem, "temperatura_c", None) or "",
        getattr(ordem, "umidade_relativa_pct", None) or "",
        getattr(ordem, "velocidade_vento_kmh", None) or "",
        getattr(ordem, "motivo_nao_realizacao", None) or "",
        getattr(ordem, "observacoes", None) or "",
        getattr(ordem, "piloto", None) or "",
        _os_yes_registered(getattr(ordem, "assinatura_piloto", None)),
        getattr(ordem, "auxiliar", None) or "",
        getattr(ordem, "proprietario_ou_preposto", None) or "",
        _os_yes_registered(getattr(ordem, "assinatura_proprietario_ou_preposto", None)),
    ]


def _os_pdf_header_footer_factory(title: str, args=None):
    def _hf(canvas, doc):
        canvas.saveState()
        page_width, height = doc.pagesize

        logo = _try_make_logo(args or {}, width_mm=30)
        if logo:
            try:
                logo.drawOn(canvas, doc.leftMargin, height - 7 * mm - logo.drawHeight)
            except Exception:
                pass

        canvas.setFillColor(colors.HexColor("#0d6efd"))
        canvas.rect(doc.leftMargin, height - 28 * mm, doc.width, 2, fill=1, stroke=0)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#666"))
        canvas.drawRightString(doc.leftMargin + doc.width, height - 13 * mm, f"Página {canvas.getPageNumber()}")

        canvas.setFillColor(colors.HexColor("#4f81c7"))
        canvas.setFont("Helvetica-Bold", 7.2)
        canvas.drawCentredString(page_width / 2, 17 * mm, "OCEANO AZUL COMERCIO INTERNACIONAL LTDA")
        canvas.setFillColor(colors.HexColor("#a0a7b3"))
        canvas.setFont("Helvetica", 5.8)
        canvas.drawCentredString(page_width / 2, 13.5 * mm, "Alameda Rio Negro, 503 - sala 2401")
        canvas.drawCentredString(page_width / 2, 10.5 * mm, "Alphaville Centro Industrial e Empresarial - Barueri SP")
        canvas.restoreState()

    return _hf


def _os_pdf_table(title: str, rows: list[list[str]], styles, col_widths=None):
    section = ParagraphStyle(
        "os_sec",
        parent=styles["Heading2"],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#0d6efd"),
        spaceBefore=10,
        spaceAfter=6,
    )
    story = [Paragraph(title, section)]

    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9dee7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fbfdff")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    story.append(Spacer(1, 8))
    return story


def _os_display_text(value):
    text = str(value or "Não informado").strip() or "Não informado"
    replacements = {
        "Nao informado": "Não informado",
        "nao informado": "Não informado",
        "NAO INFORMADO": "Não informado",
    }
    return replacements.get(text, text)


def _os_chart_rows(rows, limit=None):
    cleaned = []
    for label, total in rows or []:
        label = _os_display_text(label)
        try:
            value = int(total or 0)
        except (TypeError, ValueError):
            value = 0
        cleaned.append((label, value))
    return cleaned[:limit] if limit else cleaned


def _os_short_label(value, max_chars=32):
    value = _os_display_text(value)
    return value if len(value) <= max_chars else f"{value[:max_chars - 3]}..."


def _os_chart_image(fig, *, width_mm=170):
    bio = BytesIO()
    fig.tight_layout()
    fig.savefig(bio, format="png", dpi=220, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    bio.seek(0)
    image_width, image_height = ImageReader(bio).getSize()
    bio.seek(0)
    draw_width = width_mm * mm
    draw_height = draw_width * (image_height / image_width) if image_width else 90 * mm
    return RLImage(bio, width=draw_width, height=draw_height)


def _os_no_data_chart(title):
    fig, ax = plt.subplots(figsize=(7.0, 2.6))
    ax.set_title(title, fontsize=11, pad=10)
    ax.text(0.5, 0.5, "Sem dados para o gráfico", ha="center", va="center", fontsize=10, color="#6c757d")
    ax.axis("off")
    return fig


def _os_pdf_chart_section(data, styles):
    section = ParagraphStyle(
        "os_charts_sec",
        parent=styles["Heading2"],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#0d6efd"),
        spaceBefore=10,
        spaceAfter=6,
    )
    note = ParagraphStyle(
        "os_charts_note",
        parent=styles["Normal"],
        fontSize=9.5,
        leading=13,
        textColor=colors.HexColor("#555"),
        spaceAfter=8,
    )

    story = [
        PageBreak(),
        Paragraph("Gráficos", section),
        Paragraph("Visualização dos mesmos indicadores apresentados na tela do relatório de OS.", note),
    ]

    if not MATPLOTLIB_AVAILABLE:
        story.append(Paragraph("Matplotlib não disponível; gráficos foram omitidos.", note))
        return story

    palette = ["#0d6efd", "#20c997", "#ffc107", "#dc3545", "#6f42c1", "#0dcaf0", "#fd7e14", "#198754", "#adb5bd", "#343a40"]

    def append_chart(fig, width_mm=170):
        story.append(_os_chart_image(fig, width_mm=width_mm))
        story.append(Spacer(1, 10))

    try:
        situacao = _os_chart_rows(data["dados_situacao_aplicacao"])
        labels = [label for label, _ in situacao]
        values = [value for _, value in situacao]
        if sum(values) > 0:
            fig1, ax1 = plt.subplots(figsize=(6.8, 3.2))

            def autopct(percent):
                return f"{percent:.0f}%" if percent >= 6 else ""

            wedges, *_ = ax1.pie(
                values,
                labels=None,
                colors=[palette[i % len(palette)] for i in range(len(values))],
                autopct=autopct,
                startangle=90,
                pctdistance=0.78,
                textprops={"fontsize": 9},
            )
            ax1.add_artist(plt.Circle((0, 0), 0.58, fc="white"))
            ax1.legend(
                wedges,
                [_os_short_label(label, 38) for label in labels],
                loc="center left",
                bbox_to_anchor=(1.02, 0.5),
                fontsize=8.5,
                frameon=False,
            )
            ax1.set_title("Situação da Aplicação", fontsize=11, pad=10)
            ax1.axis("equal")
            append_chart(fig1)
        else:
            append_chart(_os_no_data_chart("Situação da Aplicação"))

        tipo = _os_chart_rows(data["dados_tipo_aplicacao"], limit=15)
        tipo_labels = [_os_short_label(label, 36) for label, _ in tipo]
        tipo_values = [value for _, value in tipo]
        fig2, ax2 = plt.subplots(figsize=(7.4, 3.6))
        ax2.barh((tipo_labels or ["Sem dados"])[::-1], (tipo_values or [0])[::-1], color="#6f42c1")
        ax2.set_xlabel("Total", fontsize=9)
        ax2.set_title("Tipo de Aplicação", fontsize=11, pad=10)
        ax2.tick_params(axis="both", labelsize=8.5)
        ax2.grid(axis="x", linestyle=":", linewidth=0.6, alpha=0.6)
        append_chart(fig2)

        unidade = _os_chart_rows(data["dados_unidade"], limit=10)
        unidade_labels = [_os_short_label(label, 36) for label, _ in unidade]
        unidade_values = [value for _, value in unidade]
        fig3, ax3 = plt.subplots(figsize=(7.4, 3.3))
        ax3.barh((unidade_labels or ["Sem dados"])[::-1], (unidade_values or [0])[::-1], color="#20c997")
        ax3.set_xlabel("Total", fontsize=9)
        ax3.set_title("OS por Unidade", fontsize=11, pad=10)
        ax3.tick_params(axis="both", labelsize=8.5)
        ax3.grid(axis="x", linestyle=":", linewidth=0.6, alpha=0.6)
        append_chart(fig3)

        mensal = _os_chart_rows(data["dados_mensais"])
        meses = [label for label, _ in mensal]
        totais = [value for _, value in mensal]
        fig4, ax4 = plt.subplots(figsize=(7.4, 3.2))
        if meses:
            ax4.plot(range(len(meses)), totais, marker="o", linewidth=1.8, color="#0d6efd")
            ax4.fill_between(range(len(meses)), totais, color="#0d6efd", alpha=0.15)
            ax4.set_xticks(range(len(meses)))
            ax4.set_xticklabels(meses, rotation=45, ha="right", fontsize=8.5)
        else:
            ax4.plot([], [])
        ax4.set_title("Histórico Mensal de OS", fontsize=11, pad=10)
        ax4.tick_params(axis="y", labelsize=8.5)
        ax4.grid(axis="y", linestyle=":", linewidth=0.6, alpha=0.6)
        append_chart(fig4)
    except Exception:
        story.append(Paragraph("Gráficos indisponíveis por erro ao gerar as imagens.", note))

    return story


def build_relatorio_os_excel_export(user, args):
    data = build_relatorio_os_export_data(user, args, include_ordens=True, only_concluidas=True)

    workbook = Workbook()
    ws_atividades = workbook.active
    ws_atividades.title = "atividades"
    _setup_field_control_sheet(ws_atividades, ATIVIDADES_HEADERS, ATIVIDADES_WIDTHS)

    ws_form = workbook.create_sheet("formulário-rd_-_projeto_sp")
    _setup_field_control_sheet(ws_form, FORMULARIO_RD_HEADERS, FORMULARIO_RD_WIDTHS)

    for ordem in data["ordens"]:
        ws_atividades.append(_build_atividades_row(ordem))
        ws_form.append(_build_formulario_rd_row(ordem))

    _finalize_field_control_sheet(ws_atividades)
    _finalize_field_control_sheet(ws_form)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    nome = f"atividades_os_{data['ano']}_{data['mes']:02d}"
    if data["uvis_id"]:
        nome += f"_uvis_{data['uvis_id']}"
    nome += ".xlsx"
    return output, nome


def build_relatorio_os_pdf_export(user, args):
    data = build_relatorio_os_export_data(user, args)

    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    path = tmp_pdf.name
    tmp_pdf.close()

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=34 * mm,
        bottomMargin=26 * mm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "os_title",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        alignment=1,
        textColor=colors.HexColor("#0d6efd"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "os_sub",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        alignment=1,
        textColor=colors.HexColor("#555"),
        spaceAfter=12,
    )

    story = [
        Paragraph("Relatório Geral de OS", title_style),
        Paragraph(
            f"Filtro: {data['mes']:02d}/{data['ano']} | Unidade: {data['uvis_nome']} | Gerado em {_os_fmt_dt(datetime.now())}",
            subtitle_style,
        ),
    ]

    story += _os_pdf_table(
        "Indicadores",
        rows=[
            ["Indicador", "Total"],
            ["Total OS", str(data["total_os"])],
            ["Concluídas", str(data["total_concluidas"])],
            ["Larva (SIM)", str(data["total_larva_sim"])],
            ["Tratamento adicional", str(data["total_tratamento_adicional"])],
            ["Não realizadas", str(data["total_nao_realizadas"])],
        ],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )
    story += _os_pdf_table(
        "Situação da Aplicação",
        rows=[["Situação", "Total"]] + [[_os_display_text(a), str(b)] for a, b in data["dados_situacao_aplicacao"]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )
    story += _os_pdf_table(
        "Tipo de Aplicação",
        rows=[["Tipo", "Total"]] + [[_os_display_text(a), str(b)] for a, b in data["dados_tipo_aplicacao"]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )
    story += _os_pdf_table(
        "Larva Visualizada",
        rows=[["Resposta", "Total"]] + [[_os_display_text(a), str(b)] for a, b in data["dados_larva"]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )
    story += _os_pdf_table(
        "Pilotos (Top 10)",
        rows=[["Piloto", "Total"]] + [[_os_display_text(a), str(b)] for a, b in data["dados_piloto"][:10]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )
    story += _os_pdf_table(
        "OS por Unidade",
        rows=[["Unidade (UVIS)", "Total"]] + [[_os_display_text(a), str(b)] for a, b in data["dados_unidade"]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )
    story += _os_pdf_table(
        "Histórico Mensal",
        rows=[["Mês", "Total"]] + [[_os_display_text(a), str(b)] for a, b in data["dados_mensais"]],
        styles=styles,
        col_widths=[60 * mm, 110 * mm],
    )
    story += _os_pdf_chart_section(data, styles)

    header_title = f"Relatório OS - {data['mes']:02d}/{data['ano']}"
    doc.build(
        story,
        onFirstPage=_os_pdf_header_footer_factory(header_title, args),
        onLaterPages=_os_pdf_header_footer_factory(header_title, args),
    )

    nome = f"relatorio_os_{data['ano']}_{data['mes']:02d}"
    if data["uvis_id"]:
        nome += f"_uvis_{data['uvis_id']}"
    nome += ".pdf"
    return path, nome


import re
import unicodedata

MESES_PT_BR = {
    1: "janeiro",
    2: "fevereiro",
    3: "marco",
    4: "abril",
    5: "maio",
    6: "junho",
    7: "julho",
    8: "agosto",
    9: "setembro",
    10: "outubro",
    11: "novembro",
    12: "dezembro",
}


def _slug_filename(value):
    value = str(value or "").strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def _coleta_imagens_pdf_name(data):
    nome = "relatorio_coleta_imagens"

    if data.get("ano_selecionado"):
        nome += f"_{data['ano_selecionado']}"

    if data.get("mes_selecionado"):
        mes_numero = int(data["mes_selecionado"])
        nome += f"_{MESES_PT_BR.get(mes_numero, mes_numero)}"

    if data.get("regiao_selecionada"):
        nome += f"_regiao_{_slug_filename(data['regiao_selecionada'])}"

    if data.get("uvis_id_selecionado"):
        uvis_nome = data.get("uvis_nome_selecionado") or f"uvis_{data['uvis_id_selecionado']}"
        uvis_slug = _slug_filename(uvis_nome)

        if uvis_slug.startswith("uvis_"):
            nome += f"_{uvis_slug}"
        else:
            nome += f"_uvis_{uvis_slug}"

    return f"{nome}.pdf"


def _canvas_text_lines(text, max_width, font_name="Helvetica", font_size=8):
    words = _os_safe(text).split()
    if not words:
        return [""]
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if pdf_canvas.pdfmetrics.stringWidth(candidate, font_name, font_size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def _draw_wrapped_text(c, text, x, y, max_width, *, font_name="Helvetica", font_size=8, leading=10, max_lines=None):
    c.setFont(font_name, font_size)
    lines = _canvas_text_lines(text, max_width, font_name, font_size)
    if max_lines:
        lines = lines[:max_lines]
    for line in lines:
        c.drawString(x, y, line)
        y -= leading
    return y


def _draw_centered_wrapped_text(c, text, x, y, width, *, font_name="Helvetica-Bold", font_size=8, leading=10, max_lines=2):
    c.setFont(font_name, font_size)
    lines = _canvas_text_lines(text, width - 8, font_name, font_size)[:max_lines]
    for line in lines:
        c.drawCentredString(x + width / 2, y, line)
        y -= leading
    return y


def _draw_coleta_pdf_footer(c, page_width):
    c.setFillColor(colors.HexColor("#4f81c7"))
    c.setFont("Helvetica-Bold", 7.2)
    c.drawCentredString(page_width / 2, 17 * mm, "OCEANO AZUL COMERCIO INTERNACIONAL LTDA")
    c.setFillColor(colors.HexColor("#a0a7b3"))
    c.setFont("Helvetica", 5.8)
    c.drawCentredString(page_width / 2, 13.5 * mm, "Alameda Rio Negro, 503 - sala 2401")
    c.drawCentredString(page_width / 2, 10.5 * mm, "Alphaville Centro Industrial e Empresarial - Barueri SP")


def _draw_coleta_pdf_logo(c, args, x, y_top):
    logo = _try_make_logo(args, width_mm=30)
    if not logo:
        return
    try:
        logo.drawOn(c, x, y_top - logo.drawHeight)
    except Exception:
        return


def _close_coleta_pdf_image_source(image_source):
    if hasattr(image_source, "close"):
        try:
            image_source.close()
        except Exception:
            pass


def _close_coleta_pdf_image_result(image_result):
    if image_result and isinstance(image_result, tuple):
        _close_coleta_pdf_image_source(image_result[0])


def _prefetch_coleta_remote_image_for_pdf(app, rel_path):
    with app.app_context():
        image_source = _download_remote_media_bytes(rel_path)
        if image_source is None:
            return None
        return _prepare_pdf_image_source_for_canvas(
            image_source,
            width_mm=148,
            max_height_mm=96,
            image_dpi=COLETA_IMAGENS_PDF_IMAGE_DPI,
            jpeg_quality=COLETA_IMAGENS_PDF_JPEG_QUALITY,
            spool_max_size=COLETA_IMAGENS_PDF_IMAGE_SPOOL_BYTES,
        )


def _prepare_coleta_pdf_image(rel_path, prefetched_image_result=None):
    if prefetched_image_result is COLETA_IMAGENS_PDF_PREFETCH_MISS:
        return None

    if prefetched_image_result is not None:
        return prefetched_image_result

    return _try_prepare_pdf_image_for_canvas(
        rel_path,
        width_mm=148,
        max_height_mm=96,
        image_dpi=COLETA_IMAGENS_PDF_IMAGE_DPI,
        jpeg_quality=COLETA_IMAGENS_PDF_JPEG_QUALITY,
        spool_max_size=COLETA_IMAGENS_PDF_IMAGE_SPOOL_BYTES,
    )


def _draw_coleta_pdf_page(c, item, data, args, *, is_first=False, prefetched_image_result=None):
    page_width, page_height = A4
    left = 21 * mm
    right = 21 * mm
    content_width = page_width - left - right

    _draw_coleta_pdf_logo(c, args, left, page_height - 21 * mm)

    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 12.2)
    c.drawCentredString(page_width / 2, page_height - 43 * mm, "Relatorio de Coleta de Imagens - Operacao Dengue PMSP")

    if is_first and data.get("export_limit_aplicado"):
        c.setFillColor(colors.HexColor("#111827"))
        _draw_wrapped_text(
            c,
            (
                f"Exportacao limitada aos primeiros {data.get('total_levantamentos_exportados')} "
                f"levantamentos de {data.get('total_levantamentos')} para preservar a memoria do servidor. "
                "Refine os filtros para exportar um periodo menor."
            ),
            left,
            page_height - 50 * mm,
            content_width,
            font_size=8,
            leading=10,
            max_lines=3,
        )

    summary_width = min(content_width - (20 * mm), 128 * mm)
    summary_x = left + (content_width - summary_width) / 2
    summary_y = page_height - 59 * mm
    summary_h = 15 * mm
    col_w = summary_width / 3
    c.setStrokeColor(colors.black)
    c.rect(summary_x, summary_y - summary_h, summary_width, summary_h, stroke=1, fill=0)
    c.line(summary_x + col_w, summary_y, summary_x + col_w, summary_y - summary_h)
    c.line(summary_x + (2 * col_w), summary_y, summary_x + (2 * col_w), summary_y - summary_h)
    c.setFillColor(colors.HexColor("#111827"))
    _draw_centered_wrapped_text(c, f"UVIS: {item['uvis_nome']}", summary_x, summary_y - 6 * mm, col_w, font_size=8.4)
    _draw_centered_wrapped_text(c, f"REGIAO: {item['regiao_nome']}", summary_x + col_w, summary_y - 6 * mm, col_w, font_size=8.4)
    _draw_centered_wrapped_text(c, f"PERIODO: {data['periodo_label']}", summary_x + 2 * col_w, summary_y - 6 * mm, col_w, font_size=8.4)

    block_width = min(content_width, 162 * mm)
    block_x = left + (content_width - block_width) / 2
    block_h = 154 * mm
    block_top = page_height - 78 * mm
    c.setStrokeColor(colors.black)
    c.rect(block_x, block_top - block_h, block_width, block_h, stroke=1, fill=0)

    image_area_h = 103 * mm
    image_area_top = block_top
    image_area_bottom = block_top - image_area_h
    image_result = _prepare_coleta_pdf_image(
        item.get("imagem_principal_path"),
        prefetched_image_result=prefetched_image_result,
    )
    if image_result:
        image_source, draw_w, draw_h = image_result
        image_x = block_x + (block_width - draw_w) / 2
        image_y = image_area_bottom + (image_area_h - draw_h) / 2
        image_reader = None
        try:
            image_reader = ImageReader(image_source)
            c.drawImage(image_reader, image_x, image_y, width=draw_w, height=draw_h, preserveAspectRatio=True, mask="auto")
        finally:
            image_reader = None
            _close_coleta_pdf_image_source(image_source)
    else:
        c.setFillColor(colors.HexColor("#4b5563"))
        c.setFont("Helvetica", 9.4)
        c.drawCentredString(block_x + block_width / 2, image_area_bottom + image_area_h / 2, "Imagem principal indisponivel no momento da exportacao.")

    details_top = image_area_bottom - 5 * mm
    label_w = 44 * mm
    row_h = 7 * mm
    rows = [
        ("ENDERECO:", item.get("endereco")),
        ("CEP:", item.get("cep")),
        ("COORDENADAS:", item.get("coordenadas")),
        ("DATA DE COLETA DE IMG:", item.get("data_coleta_label")),
        ("FOCO:", item.get("foco")),
    ]
    y_row = details_top
    for label, value in rows:
        c.setFillColor(colors.HexColor("#111827"))
        c.setFont("Helvetica-Bold", 7.4)
        c.drawString(block_x + 8, y_row, label)
        _draw_wrapped_text(
            c,
            value,
            block_x + label_w,
            y_row,
            block_width - label_w - 12,
            font_size=7.4,
            leading=8,
            max_lines=2 if label == "ENDERECO:" else 1,
        )
        y_row -= row_h if label != "ENDERECO:" else 11 * mm

    _draw_coleta_pdf_footer(c, page_width)


def _build_relatorio_coleta_imagens_pdf_export_streamed(data, args):
    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    path = tmp_pdf.name
    tmp_pdf.close()

    c = pdf_canvas.Canvas(path, pagesize=A4)
    if not data["levantamentos"]:
        item = {
            "uvis_nome": data["uvis_nome_selecionado"],
            "regiao_nome": data["regiao_nome_selecionada"],
            "imagem_principal_path": None,
            "endereco": "Nenhum levantamento com foto principal foi encontrado para os filtros selecionados.",
            "cep": "",
            "coordenadas": "",
            "data_coleta_label": "",
            "foco": "",
        }
        _draw_coleta_pdf_page(c, item, data, args, is_first=True)
        c.save()
        gc.collect()
        return path, "relatorio_coleta_imagens.pdf"

    levantamentos = data["levantamentos"]
    prefetch_workers = min(COLETA_IMAGENS_PDF_REMOTE_PREFETCH, max(0, len(levantamentos) - 1))
    prefetch_executor = None
    prefetch_futures = {}

    def schedule_prefetch(item_index):
        if prefetch_executor is None or item_index >= len(levantamentos) or item_index in prefetch_futures:
            return
        rel_path = levantamentos[item_index].get("imagem_principal_path")
        if not _is_remote_media_path(rel_path):
            return
        prefetch_futures[item_index] = prefetch_executor.submit(
            _prefetch_coleta_remote_image_for_pdf,
            current_app._get_current_object(),
            rel_path,
        )

    try:
        if prefetch_workers > 0:
            prefetch_executor = ThreadPoolExecutor(max_workers=prefetch_workers)
            for next_index in range(1, min(len(levantamentos), 1 + prefetch_workers)):
                schedule_prefetch(next_index)

        for index, item in enumerate(levantamentos):
            if index:
                c.showPage()

            prefetched_image_result = None
            future = prefetch_futures.pop(index, None)
            if future is not None:
                try:
                    prefetched_image_result = future.result()
                except Exception:
                    current_app.logger.exception("Erro ao pre-baixar imagem remota para PDF de coleta.")
                    prefetched_image_result = COLETA_IMAGENS_PDF_PREFETCH_MISS
                if prefetched_image_result is None:
                    prefetched_image_result = COLETA_IMAGENS_PDF_PREFETCH_MISS

            schedule_prefetch(index + prefetch_workers)
            _draw_coleta_pdf_page(
                c,
                item,
                data,
                args,
                is_first=(index == 0),
                prefetched_image_result=prefetched_image_result,
            )
    finally:
        for future in prefetch_futures.values():
            if future.done():
                try:
                    _close_coleta_pdf_image_result(future.result())
                except Exception:
                    pass
            else:
                future.cancel()
        if prefetch_executor is not None:
            prefetch_executor.shutdown(wait=False, cancel_futures=True)

    c.save()
    gc.collect()
    return path, _coleta_imagens_pdf_name(data)


def build_relatorio_coleta_imagens_pdf_export(user, args):
    data = build_relatorio_coleta_imagens_export_data(
        user,
        args,
        max_items=_coleta_imagens_max_export_items(),
    )
    return _build_relatorio_coleta_imagens_pdf_export_streamed(data, args)

    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    path = tmp_pdf.name
    tmp_pdf.close()

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=21 * mm,
        rightMargin=21 * mm,
        topMargin=18 * mm,
        bottomMargin=20 * mm,
    )
    styles = getSampleStyleSheet()

    page_inner_width = min(doc.width, 162 * mm)
    summary_width = min(doc.width - (20 * mm), 128 * mm)
    detail_label_width = 44 * mm

    title_style = ParagraphStyle(
        "coleta_title",
        parent=styles["Title"],
        fontSize=14.4,
        leading=19,
        alignment=1,
        textColor=colors.HexColor("#111827"),
        spaceAfter=0,
    )
    top_info_style = ParagraphStyle(
        "coleta_top_info",
        parent=styles["BodyText"],
        fontSize=9.2,
        leading=12.2,
        alignment=1,
        textColor=colors.HexColor("#111827"),
    )
    detail_label_style = ParagraphStyle(
        "coleta_detail_label",
        parent=styles["BodyText"],
        fontSize=7.4,
        leading=10.2,
        textColor=colors.HexColor("#111827"),
    )
    detail_value_style = ParagraphStyle(
        "coleta_detail_value",
        parent=styles["BodyText"],
        fontSize=7.4,
        leading=10.2,
        textColor=colors.HexColor("#111827"),
    )
    footer_company_style = ParagraphStyle(
        "coleta_footer_company",
        parent=styles["Normal"],
        fontSize=7.2,
        leading=8.4,
        alignment=1,
        textColor=colors.HexColor("#4f81c7"),
    )
    footer_text_style = ParagraphStyle(
        "coleta_footer_text",
        parent=styles["Normal"],
        fontSize=5.8,
        leading=7,
        alignment=1,
        textColor=colors.HexColor("#a0a7b3"),
    )
    empty_state_style = ParagraphStyle(
        "coleta_empty",
        parent=styles["Normal"],
        fontSize=9.4,
        leading=13,
        alignment=1,
        textColor=colors.HexColor("#4b5563"),
    )

    story = []

    if not data["levantamentos"]:
        logo = _try_make_logo(args, width_mm=30)
        if logo:
            logo_box = Table([[logo]], colWidths=[doc.width])
            logo_box.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            story.append(logo_box)
            story.append(Spacer(1, 24))

        story.append(Paragraph("Relatorio de Coleta de Imagens - Operacao Dengue PMSP", title_style))
        story.append(Spacer(1, 24))

        filtro_box = Table([[
            Paragraph(f"<b>UVIS:</b> {_os_safe(data['uvis_nome_selecionado'])}", top_info_style),
            Paragraph(f"<b>REGIAO:</b> {_os_safe(data['regiao_nome_selecionada'])}", top_info_style),
            Paragraph(f"<b>PERIODO:</b> {_os_safe(data['periodo_label'])}", top_info_style),
        ]], colWidths=[summary_width / 3, summary_width / 3, summary_width / 3])
        filtro_box.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 5.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5.5),
        ]))
        filtro_wrap = Table([[filtro_box]], colWidths=[doc.width])
        filtro_wrap.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(filtro_wrap)
        story.append(Spacer(1, 22))

        vazio = Table(
            [[Paragraph("Nenhum levantamento com foto principal foi encontrado para os filtros selecionados.", empty_state_style)]],
            colWidths=[page_inner_width],
            rowHeights=[102 * mm],
        )
        vazio.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        vazio_wrap = Table([[vazio]], colWidths=[doc.width])
        vazio_wrap.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(vazio_wrap)
        story.append(Spacer(1, 28))
        story.append(Paragraph("OCEANO AZUL COMERCIO INTERNACIONAL LTDA", footer_company_style))
        story.append(Paragraph("Alameda Rio Negro, 503 - sala 2401", footer_text_style))
        story.append(Paragraph("Alphaville Centro Industrial e Empresarial - Barueri SP", footer_text_style))

        doc.build(story)
        return path, "relatorio_coleta_imagens.pdf"

    for index, item in enumerate(data["levantamentos"]):
        if index:
            story.append(PageBreak())

        logo = _try_make_logo(args, width_mm=30)
        if logo:
            logo_box = Table([[logo]], colWidths=[doc.width])
            logo_box.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            story.append(logo_box)
            story.append(Spacer(1, 24))

        story.append(Paragraph("Relatorio de Coleta de Imagens - Operacao Dengue PMSP", title_style))
        story.append(Spacer(1, 24))
        if index == 0 and data.get("export_limit_aplicado"):
            story.append(Paragraph(
                (
                    f"Exportacao limitada aos primeiros {data.get('total_levantamentos_exportados')} "
                    f"levantamentos de {data.get('total_levantamentos')} para preservar a memoria do servidor. "
                    "Refine os filtros para exportar um periodo menor."
                ),
                top_info_style,
            ))
            story.append(Spacer(1, 10))

        resumo_topo = Table([[
            Paragraph(f"<b>UVIS:</b> {_os_safe(item['uvis_nome'])}", top_info_style),
            Paragraph(f"<b>REGIAO:</b> {_os_safe(item['regiao_nome'])}", top_info_style),
            Paragraph(f"<b>PERIODO:</b> {_os_safe(data['periodo_label'])}", top_info_style),
        ]], colWidths=[summary_width / 3, summary_width / 3, summary_width / 3])
        resumo_topo.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 5.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5.5),
        ]))
        resumo_wrap = Table([[resumo_topo]], colWidths=[doc.width])
        resumo_wrap.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(resumo_wrap)
        story.append(Spacer(1, 22))

        imagem_principal = _try_make_local_rlimage(
            item["imagem_principal_path"],
            width_mm=148,
            max_height_mm=96,
        )

        if imagem_principal:
            imagem_box = Table([[imagem_principal]], colWidths=[page_inner_width])
            imagem_box.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 14),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]))
        else:
            imagem_box = Table(
                [[Paragraph("Imagem principal indisponivel no momento da exportacao.", empty_state_style)]],
                colWidths=[page_inner_width],
                rowHeights=[96 * mm],
            )
            imagem_box.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 14),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]))

        detalhes = Table(
            [
                [Paragraph("<b>ENDERECO:</b>", detail_label_style), Paragraph(_os_safe(item["endereco"]), detail_value_style)],
                [Paragraph("<b>CEP:</b>", detail_label_style), Paragraph(_os_safe(item["cep"]), detail_value_style)],
                [Paragraph("<b>COORDENADAS:</b>", detail_label_style), Paragraph(_os_safe(item["coordenadas"]), detail_value_style)],
                [Paragraph("<b>DATA DE COLETA DE IMG:</b>", detail_label_style), Paragraph(_os_safe(item["data_coleta_label"]), detail_value_style)],
                [Paragraph("<b>FOCO:</b>", detail_label_style), Paragraph(_os_safe(item["foco"]), detail_value_style)],
            ],
            colWidths=[detail_label_width, page_inner_width - detail_label_width],
        )
        detalhes.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        bloco_principal = Table(
            [[imagem_box], [detalhes]],
            colWidths=[page_inner_width],
        )
        bloco_principal.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (0, 0), 16),
            ("TOPPADDING", (0, 1), (0, 1), 10),
            ("BOTTOMPADDING", (0, 1), (0, 1), 12),
        ]))
        bloco_wrap = Table([[bloco_principal]], colWidths=[doc.width])
        bloco_wrap.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(bloco_wrap)
        story.append(Spacer(1, 28))
        story.append(Paragraph("OCEANO AZUL COMERCIO INTERNACIONAL LTDA", footer_company_style))
        story.append(Paragraph("Alameda Rio Negro, 503 - sala 2401", footer_text_style))
        story.append(Paragraph("Alphaville Centro Industrial e Empresarial - Barueri SP", footer_text_style))

    doc.build(story)

    nome = "relatorio_coleta_imagens"
    if data["ano_selecionado"]:
        nome += f"_{data['ano_selecionado']}"
    if data["mes_selecionado"]:
        nome += f"_{int(data['mes_selecionado']):02d}"
    if data["regiao_selecionada"]:
        nome += f"_regiao_{data['regiao_selecionada'].replace(' ', '_').lower()}"
    if data["uvis_id_selecionado"]:
        nome += f"_uvis_{data['uvis_id_selecionado']}"
    nome += ".pdf"
    return path, nome
