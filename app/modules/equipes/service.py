from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models import Equipe, EquipePiloto, Pilotos, Usuario
from app.shared.access import apply_prefeitura_scope, normalize_role
from app.shared.query_filters import id_search_clause


REGIOES = ("CENTRO", "CENTRO-OESTE", "LESTE", "NORTE", "OESTE", "SUL", "SUDESTE")
REGIOES_VALIDAS = set(REGIOES) | {"SULDESTE"}
ATIVA_TRUE_VALUES = {"1", "true", "sim", "yes", "on"}
ATIVA_FALSE_VALUES = {"0", "false", "nao", "no"}
EQUIPE_OCEANO_USER_TYPE = "equipe_oceano"


def build_regioes_list():
    return sorted(REGIOES)


def regiao_valida(regiao: str) -> bool:
    return not regiao or regiao in REGIOES_VALIDAS


def parse_optional_int(value):
    try:
        return int(value) if value else None
    except (TypeError, ValueError):
        return None


def is_truthy(value: str) -> bool:
    return (value or "").strip().lower() in ATIVA_TRUE_VALUES


def get_pilotos_ordered(user=None):
    query = Pilotos.query
    if user is not None:
        query = apply_prefeitura_scope(query, user, Pilotos.prefeitura_id)
    return query.order_by(Pilotos.nome_piloto.asc()).all()


def login_em_uso(login: str, exclude_user_id=None):
    query = Usuario.query.filter(db.func.lower(Usuario.login) == (login or "").strip().lower())
    if exclude_user_id:
        query = query.filter(Usuario.id != exclude_user_id)
    return db.session.query(query.exists()).scalar()


def suggested_equipe_login(equipe):
    nome = (getattr(equipe, "nome_equipe", None) or f"Ploa {equipe.id}").strip().lower()
    base = "".join(ch if ch.isalnum() else "." for ch in nome)
    base = ".".join(part for part in base.split(".") if part)
    return (base or f"ploa.{equipe.id}")[:50]


def get_equipe_account(equipe_id: int):
    return (
        Usuario.query
        .filter(
            Usuario.tipo_usuario == EQUIPE_OCEANO_USER_TYPE,
            Usuario.codigo_setor == str(equipe_id),
        )
        .first()
    )


def build_equipe_accounts_map(equipes):
    equipe_ids = [str(equipe.id) for equipe in equipes]
    if not equipe_ids:
        return {}

    contas = (
        Usuario.query
        .filter(
            Usuario.tipo_usuario == EQUIPE_OCEANO_USER_TYPE,
            Usuario.codigo_setor.in_(equipe_ids),
        )
        .all()
    )
    return {int(conta.codigo_setor): conta for conta in contas if str(conta.codigo_setor or "").isdigit()}


def validate_equipe_account_form(login: str, senha: str, senha2: str, current_account=None):
    errors = {}
    login = (login or "").strip()
    senha = senha or ""
    senha2 = senha2 or ""

    if not login:
        errors["login"] = "Informe o login da equipe."
    elif len(login) < 4:
        errors["login"] = "O login deve ter pelo menos 4 caracteres."
    elif len(login) > 50:
        errors["login"] = "O login deve ter no maximo 50 caracteres."
    elif login_em_uso(login, exclude_user_id=(current_account.id if current_account else None)):
        errors["login"] = "Este login ja esta em uso."

    if not current_account and not senha:
        errors["senha"] = "Informe a senha da equipe."
    elif senha and len(senha) < 6:
        errors["senha"] = "A senha deve ter pelo menos 6 caracteres."

    if senha or senha2:
        if senha != senha2:
            errors["senha2"] = "As senhas nao conferem."

    return errors


def upsert_equipe_account(equipe, login: str, senha: str):
    account = get_equipe_account(equipe.id)
    if account is None:
        account = Usuario(
            nome_uvis=equipe.nome_equipe or f"Ploa {equipe.id}",
            regiao=equipe.regiao,
            prefeitura_id=equipe.prefeitura_id,
            codigo_setor=str(equipe.id),
            login=login,
            tipo_usuario=EQUIPE_OCEANO_USER_TYPE,
            piloto_id=None,
        )
        db.session.add(account)

    account.nome_uvis = equipe.nome_equipe or f"Ploa {equipe.id}"
    account.regiao = equipe.regiao
    account.prefeitura_id = equipe.prefeitura_id
    account.codigo_setor = str(equipe.id)
    account.login = login
    account.tipo_usuario = EQUIPE_OCEANO_USER_TYPE
    account.piloto_id = None
    if senha:
        account.set_senha(senha)
    return account


def find_piloto_conflict(piloto_id: int, exclude_equipe_id=None, user=None):
    query = EquipePiloto.query.filter(EquipePiloto.piloto_id == piloto_id)
    if exclude_equipe_id is not None:
        query = query.filter(EquipePiloto.equipe_id != exclude_equipe_id)

    vinculo = query.first()
    if not vinculo:
        return None, None

    equipe_query = Equipe.query
    if user is not None:
        equipe_query = apply_prefeitura_scope(equipe_query, user, Equipe.prefeitura_id)
    equipe = equipe_query.filter(Equipe.id == vinculo.equipe_id).first()
    return vinculo, equipe


def build_equipes_query(
    tipo: str,
    regiao: str,
    ativa: str,
    piloto_id: str,
    auxiliar_id: str,
    q: str,
    sort: str,
    user_regiao: str,
    user=None,
):
    tipo = normalize_role(tipo)
    regiao = (regiao or "").strip().upper()
    ativa = (ativa or "").strip().lower()
    user_regiao = (user_regiao or "").strip().upper()

    if tipo == "uvis":
        regiao = user_regiao
        ativa = "1"

    query = Equipe.query.options(
        db.selectinload(Equipe.membros).selectinload(EquipePiloto.piloto),
        db.selectinload(Equipe.equipamentos),
    )
    if user is not None:
        query = apply_prefeitura_scope(query, user, Equipe.prefeitura_id)

    if tipo == "uvis":
        if not regiao:
            query = query.filter(db.false())
        else:
            query = query.filter(Equipe.regiao.ilike(regiao))
            query = query.filter(Equipe.ativa.is_(True))
    elif tipo not in ["dev", "admin", "visualizar", "operario", "operador", "prefeitura_admin"]:
        if user_regiao:
            query = query.filter(Equipe.regiao.ilike(user_regiao))
            regiao = user_regiao
        else:
            query = query.filter(db.false())

    if regiao:
        query = query.filter(Equipe.regiao.ilike(regiao))

    if tipo != "uvis":
        if ativa in ATIVA_TRUE_VALUES:
            query = query.filter(Equipe.ativa.is_(True))
        elif ativa in ATIVA_FALSE_VALUES:
            query = query.filter(Equipe.ativa.is_(False))

    piloto_id_int = parse_optional_int(piloto_id)
    if piloto_id_int:
        query = query.filter(
            Equipe.membros.any(
                db.and_(
                    EquipePiloto.papel == "piloto",
                    EquipePiloto.piloto_id == piloto_id_int,
                )
            )
        )

    auxiliar_id_int = parse_optional_int(auxiliar_id)
    if auxiliar_id_int:
        query = query.filter(
            Equipe.membros.any(
                db.and_(
                    EquipePiloto.papel == "auxiliar",
                    EquipePiloto.piloto_id == auxiliar_id_int,
                )
            )
        )

    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                id_search_clause(Equipe.id, q),
                Equipe.nome_equipe.ilike(like),
                Equipe.descricao.ilike(like),
                Equipe.regiao.ilike(like),
                Equipe.membros.any(
                    EquipePiloto.piloto.has(
                        db.or_(
                            id_search_clause(Pilotos.id, q),
                            Pilotos.nome_piloto.ilike(like),
                            Pilotos.telefone.ilike(like),
                        )
                    )
                ),
            )
        )

    if sort == "nome_desc":
        return query.order_by(Equipe.nome_equipe.desc()), regiao, ativa
    if sort == "id_desc":
        return query.order_by(Equipe.id.desc()), regiao, ativa
    if sort == "id_asc":
        return query.order_by(Equipe.id.asc()), regiao, ativa
    if sort == "criada_desc":
        return query.order_by(Equipe.criada_em.desc()), regiao, ativa
    if sort == "criada_asc":
        return query.order_by(Equipe.criada_em.asc()), regiao, ativa
    return query.order_by(Equipe.nome_equipe.asc()), regiao, ativa


def build_equipes_filters(q, regiao, ativa, piloto_id, auxiliar_id, sort, page, per_page, total, total_pages, locked_regiao, locked_ativa):
    return {
        "q": q,
        "regiao": regiao,
        "ativa": ativa,
        "piloto_id": piloto_id,
        "auxiliar_id": auxiliar_id,
        "sort": sort,
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
        "locked_regiao": locked_regiao,
        "locked_ativa": locked_ativa,
    }


def _format_drones(equipe):
    drones = [
        equipamento
        for equipamento in (equipe.equipamentos or [])
        if getattr(equipamento, "tipo_equipamento", None) == "drones"
    ]

    parts = []
    for drone in drones:
        nome = (
            getattr(drone, "renomacao", None)
            or getattr(drone, "modelo", None)
            or f"Drone {getattr(drone, 'id', '')}"
        ).strip()
        modelo = (getattr(drone, "modelo", None) or "").strip()

        piece = nome
        if modelo and modelo.lower() != nome.lower():
            piece = f"{nome} ({modelo})"

        parts.append(piece)

    return "; ".join(parts)


def build_equipes_export(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Equipes"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet["A1"] = "Relatorio de Equipes"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sheet["A2"].font = Font(color="6B7280")

    start_row = 4
    headers = ["ID", "Equipe", "Regiao", "Ativa", "Piloto Titular", "Auxiliar", "Drones", "Criada em", "Descricao"]

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for row_idx, equipe in enumerate(rows, start=start_row + 1):
        values = [
            equipe.id,
            equipe.nome_equipe,
            equipe.regiao or "",
            "SIM" if equipe.ativa else "NAO",
            equipe.piloto_titular.nome_piloto if equipe.piloto_titular else "",
            equipe.piloto_auxiliar.nome_piloto if equipe.piloto_auxiliar else "",
            _format_drones(equipe),
            equipe.criada_em.strftime("%d/%m/%Y %H:%M") if equipe.criada_em else "",
            equipe.descricao or "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = center_align if col_idx in (1, 4) else text_align

    last_row = start_row + len(rows)
    last_col = len(headers)
    sheet.freeze_panes = sheet["A5"]
    sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(last_col)}{max(last_row, start_row)}"
    sheet.row_dimensions[start_row].height = 22

    max_widths = {1: 8, 2: 28, 3: 14, 4: 10, 5: 26, 6: 26, 7: 60, 8: 18, 9: 50}
    for col_idx in range(1, last_col + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(start_row + 1, last_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))

        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_widths.get(col_idx, 40))

    zebra_fill = PatternFill("solid", fgColor="F9FAFB")
    for row_idx in range(start_row + 1, last_row + 1):
        if (row_idx - (start_row + 1)) % 2 == 1:
            for col_idx in range(1, last_col + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = zebra_fill

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"equipes_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    return output, filename
