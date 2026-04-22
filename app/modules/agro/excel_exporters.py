from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.agro.service import (
    AGRO_REPORT_MONTHS,
    build_agro_dre_gerencial_report,
    build_agro_fluxo_caixa_report,
    build_financeiro_agro_caixa_diario_query,
)

THIN = Side(style="thin", color="D7E4DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_TITLE = PatternFill("solid", fgColor="1F6F43")
FILL_SUBTITLE = PatternFill("solid", fgColor="2F855A")
FILL_SECTION = PatternFill("solid", fgColor="E7F4EC")
FILL_HEADER = PatternFill("solid", fgColor="25603A")
FILL_ZEBRA = PatternFill("solid", fgColor="F7FBF8")
FILL_CARD_GREEN = PatternFill("solid", fgColor="DFF3E6")
FILL_CARD_AMBER = PatternFill("solid", fgColor="FFF3D6")
FILL_CARD_RED = PatternFill("solid", fgColor="FCE3E3")
FILL_CARD_BLUE = PatternFill("solid", fgColor="DDEEFF")
FILL_RESULT = PatternFill("solid", fgColor="EAF7EF")

FONT_TITLE = Font(bold=True, size=18, color="FFFFFF")
FONT_SUBTITLE = Font(size=10, color="F4FFF8")
FONT_SECTION = Font(bold=True, size=11, color="1F5130")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_CARD_LABEL = Font(bold=True, size=10, color="385449")
FONT_CARD_VALUE = Font(bold=True, size=16, color="173425")
FONT_CARD_VALUE_NEG = Font(bold=True, size=16, color="A61B1B")
FONT_STRONG = Font(bold=True)

CURRENCY_FMT = 'R$ #,##0.00;[Red]-R$ #,##0.00'

EXPENSE_KINDS = {
    "comissao_principal_realizada",
    "comissao_cooperativa_realizada",
    "despesas_manuais_realizadas",
    "impostos_realizados",
    "retencoes_realizadas",
    "despesa_total_realizada",
}

DRE_LINE_ORDER = {
    "Receita bruta": 1,
    "Comissao principal": 2,
    "Comissao cooperativa": 3,
    "Despesas manuais": 4,
    "Impostos": 5,
    "Retencoes": 6,
}

DRE_LINE_FILL = {
    "Receita bruta": FILL_CARD_GREEN,
    "Comissao principal": FILL_CARD_AMBER,
    "Comissao cooperativa": FILL_CARD_AMBER,
    "Despesas manuais": FILL_CARD_AMBER,
    "Impostos": FILL_CARD_RED,
    "Retencoes": FILL_CARD_RED,
}


def _safe_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def _safe_number(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def _as_decimal(value) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value))


def _fmt_date(value) -> str:
    if not value:
        return ""
    return value.strftime("%d/%m/%Y")


def _fmt_datetime(value) -> str:
    if not value:
        return ""
    return value.strftime("%d/%m/%Y %H:%M")


def _apply_currency(cell):
    cell.number_format = CURRENCY_FMT
    cell.alignment = Alignment(horizontal="right", vertical="center")


def _apply_title_block(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str,
    year: int,
    max_col: int,
):
    end_col = get_column_letter(max_col)
    ws.merge_cells(f"A1:{end_col}1")
    ws.merge_cells(f"A2:{end_col}2")
    ws.merge_cells(f"A3:{end_col}3")

    for cell_ref, value, fill, font in (
        ("A1", title, FILL_TITLE, FONT_TITLE),
        ("A2", subtitle, FILL_SUBTITLE, FONT_SUBTITLE),
        ("A3", f"Ano base: {year} | Gerado em {_fmt_datetime(datetime.now())}", FILL_SECTION, FONT_SECTION),
    ):
        cell = ws[cell_ref]
        cell.value = value
        cell.fill = fill
        cell.font = font
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18


def _write_card(
    ws: Worksheet,
    *,
    row: int,
    start_col: int,
    end_col: int,
    label: str,
    value,
    fill: PatternFill,
    is_currency: bool = True,
):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    ws.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 2, end_column=end_col)

    label_cell = ws.cell(row=row, column=start_col, value=label)
    value_cell = ws.cell(row=row + 1, column=start_col, value=_safe_number(value))

    for current_row in range(row, row + 3):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = fill
            cell.border = BORDER

    label_cell.font = FONT_CARD_LABEL
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    value_cell.font = FONT_CARD_VALUE_NEG if isinstance(value, Decimal) and value < 0 else FONT_CARD_VALUE
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    if is_currency:
        _apply_currency(value_cell)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 24
    ws.row_dimensions[row + 2].height = 14


def _add_section_title(ws: Worksheet, row: int, title: str, max_col: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = FILL_SECTION
    cell.font = FONT_SECTION
    cell.border = BORDER
    cell.alignment = Alignment(vertical="center")
    for col in range(2, max_col + 1):
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 20


def _write_table_header(ws: Worksheet, row: int, headers: list[str]):
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=index, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22


def _style_data_row(
    ws: Worksheet,
    *,
    row: int,
    col_count: int,
    zebra: bool = False,
    fill: PatternFill | None = None,
):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill is not None:
            cell.fill = fill
        elif zebra:
            cell.fill = FILL_ZEBRA


def _set_column_widths(ws: Worksheet, widths: dict[str, float]):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _auto_width(ws: Worksheet, *, min_width: int = 12, max_width: int = 42):
    best: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value in (None, ""):
                continue
            letter = get_column_letter(cell.column)
            best[letter] = max(best.get(letter, 0), len(str(cell.value)))

    for letter, size in best.items():
            current = ws.column_dimensions[letter].width or 0
            target = max(min_width, min(max_width, size + 2))
            if current < target:
                ws.column_dimensions[letter].width = target


def _build_fluxo_daily_rows(user, report: dict) -> list[dict]:
    caixa_map = {}
    for caixa in build_financeiro_agro_caixa_diario_query(user).all():
        if caixa.data_caixa and caixa.data_caixa.year == report["ano"]:
            caixa_map[caixa.data_caixa] = caixa

    movimentos_por_data: dict = {}
    for item in report["lancamentos"]:
        data_caixa = item.get("data_realizada")
        if not data_caixa or data_caixa.year != report["ano"]:
            continue
        bucket = movimentos_por_data.setdefault(
            data_caixa,
            {
                "movimentos": [],
                "total_entradas": Decimal("0"),
                "total_saidas": Decimal("0"),
            },
        )
        bucket["movimentos"].append(item)
        bucket["total_entradas"] += _as_decimal(item.get("entrada_realizada"))
        bucket["total_saidas"] += _as_decimal(item.get("saida_realizada"))

    rows = []
    saldo_referencia = Decimal("0")
    for data_caixa in sorted(set(caixa_map.keys()) | set(movimentos_por_data.keys())):
        caixa = caixa_map.get(data_caixa)
        bucket = movimentos_por_data.get(data_caixa)
        movimentos = (bucket or {}).get("movimentos", [])
        total_entradas = (bucket or {}).get("total_entradas", Decimal("0"))
        total_saidas = (bucket or {}).get("total_saidas", Decimal("0"))

        if caixa is not None:
            saldo_abertura = _as_decimal(caixa.saldo_abertura_decimal)
            if not movimentos:
                total_entradas = _as_decimal(caixa.total_entradas_decimal)
                total_saidas = _as_decimal(caixa.total_saidas_decimal)
                saldo_fechamento = _as_decimal(caixa.saldo_fechamento_decimal)
            else:
                saldo_fechamento = saldo_abertura + total_entradas - total_saidas
            status_caixa = caixa.status
            aberto_por = caixa.aberto_por_nome or ""
            fechado_por = caixa.fechado_por_nome or ""
            observacoes_abertura = caixa.observacoes_abertura or ""
            observacoes_fechamento = caixa.observacoes_fechamento or ""
        elif movimentos:
            primeiro = movimentos[0]
            ultimo = movimentos[-1]
            saldo_abertura = (
                _as_decimal(primeiro.get("saldo_realizado_acumulado"))
                - _as_decimal(primeiro.get("entrada_realizada"))
                + _as_decimal(primeiro.get("saida_realizada"))
            )
            saldo_fechamento = _as_decimal(ultimo.get("saldo_realizado_acumulado"))
            status_caixa = "SEM_ABERTURA"
            aberto_por = ""
            fechado_por = ""
            observacoes_abertura = ""
            observacoes_fechamento = ""
        else:
            saldo_abertura = saldo_referencia
            saldo_fechamento = saldo_abertura
            status_caixa = "SEM_MOVIMENTO"
            aberto_por = ""
            fechado_por = ""
            observacoes_abertura = ""
            observacoes_fechamento = ""

        resultado_dia = total_entradas - total_saidas
        saldo_referencia = saldo_fechamento
        rows.append(
            {
                "data_caixa": data_caixa,
                "status_caixa": status_caixa,
                "saldo_abertura": saldo_abertura,
                "total_entradas": total_entradas,
                "total_saidas": total_saidas,
                "resultado_dia": resultado_dia,
                "saldo_fechamento": saldo_fechamento,
                "qtd_movimentos": len(movimentos),
                "aberto_por": aberto_por,
                "fechado_por": fechado_por,
                "observacoes_abertura": observacoes_abertura,
                "observacoes_fechamento": observacoes_fechamento,
            }
        )
    return rows


def _build_fluxo_dre_rows(report: dict, *, use_emission_date: bool = False) -> list[dict]:
    grouped = {}
    receita_total = _as_decimal(report["totais"].get("receita_bruta_realizada"))

    def add_row(
        linha_dre: str,
        origem: str,
        categoria: str,
        subcategoria: str,
        month: int,
        value,
        leitura: str,
    ):
        key = (linha_dre, origem, categoria, subcategoria, leitura)
        if key not in grouped:
            grouped[key] = {
                "linha_dre": linha_dre,
                "origem": origem,
                "categoria": categoria,
                "subcategoria": subcategoria,
                "leitura": leitura,
                "meses": [Decimal("0") for _ in range(12)],
                "total": Decimal("0"),
            }
        grouped[key]["meses"][month - 1] += _as_decimal(value)
        grouped[key]["total"] += _as_decimal(value)

    for item in report["lancamentos"]:
        data_base = item.get("data_emissao") if use_emission_date else item.get("data")
        if use_emission_date and data_base is None:
            data_base = item.get("data")
        if not data_base:
            continue
        month = data_base.month

        if _as_decimal(item.get("entrada_realizada")):
            if item.get("origem") == "CONTRATO":
                add_row(
                    "Receita bruta",
                    "Recebivel de contrato",
                    "Contratos do Agro",
                    "Recebivel de contrato",
                    month,
                    item.get("entrada_realizada"),
                    "Receita realizada por contrato do Agro.",
                )
            else:
                add_row(
                    "Receita bruta",
                    "Entrada manual",
                    item.get("categoria_grupo") or item.get("categoria") or "Entrada manual",
                    item.get("subcategoria") or "Sem subcategoria",
                    month,
                    item.get("entrada_realizada"),
                "Entrada manual classificada para alimentar o faturamento gerencial.",
                )

        if _as_decimal(item.get("comissao_principal_realizada")):
            add_row(
                "Comissao principal",
                "Contrato",
                "Comissoes de contrato",
                "Comissao principal",
                month,
                _as_decimal(item.get("comissao_principal_realizada")) * Decimal("-1"),
                "Comissao operacional calculada a partir dos contratos do Agro.",
            )

        if _as_decimal(item.get("comissao_cooperativa_realizada")):
            add_row(
                "Comissao cooperativa",
                "Contrato",
                "Comissoes de contrato",
                "Comissao cooperativa",
                month,
                _as_decimal(item.get("comissao_cooperativa_realizada")) * Decimal("-1"),
                "Comissao cooperativa calculada a partir dos contratos do Agro.",
            )

        saida_realizada = _as_decimal(item.get("saida_realizada"))
        if item.get("origem") == "SAIDA_MANUAL" and saida_realizada:
            tipo_saida = item.get("tipo_saida") or ""
            linha_dre = "Despesas manuais"
            leitura = "Saida manual operacional classificada para o DRE."
            if tipo_saida == "IMPOSTO":
                linha_dre = "Impostos"
                leitura = "Saida manual classificada como imposto."
            elif tipo_saida == "RETENCAO":
                linha_dre = "Retencoes"
                leitura = "Saida manual classificada como retencao."
            add_row(
                linha_dre,
                "Saida manual",
                item.get("categoria_grupo") or item.get("categoria") or "Saida manual",
                item.get("subcategoria") or "Sem subcategoria",
                month,
                saida_realizada * Decimal("-1"),
                leitura,
            )

    rows = []
    for row in grouped.values():
        participacao = Decimal("0")
        if receita_total not in (Decimal("0"), 0):
            participacao = abs(row["total"]) / receita_total
        row["participacao"] = participacao
        rows.append(row)

    return sorted(
        rows,
        key=lambda item: (
            DRE_LINE_ORDER.get(item["linha_dre"], 99),
            item["categoria"],
            item["subcategoria"],
            item["origem"],
        ),
    )


def _build_fluxo_summary_sheet(workbook: Workbook, report: dict):
    ws = workbook.active
    ws.title = "Resumo Fluxo"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    _apply_title_block(
        ws,
        title="Fluxo de Caixa Agro",
        subtitle="Visao executiva do caixa, com leitura anual e consolidado mensal.",
        year=report["ano"],
        max_col=12,
    )

    mensal = report["mensal"]
    ultimo_mes = mensal[-1] if mensal else {}
    totais = report["totais"]

    _write_card(ws, row=5, start_col=1, end_col=3, label="Receita realizada", value=totais["receita_bruta_realizada"], fill=FILL_CARD_GREEN)
    _write_card(ws, row=5, start_col=4, end_col=6, label="Saida realizada", value=totais["despesa_total_realizada"], fill=FILL_CARD_AMBER)
    _write_card(ws, row=5, start_col=7, end_col=9, label="Resultado realizado", value=totais["resultado_realizado"], fill=FILL_CARD_BLUE)
    _write_card(
        ws,
        row=5,
        start_col=10,
        end_col=12,
        label="Saldo acumulado",
        value=ultimo_mes.get("saldo_acumulado_realizado", Decimal("0")),
        fill=FILL_CARD_RED if ultimo_mes.get("saldo_acumulado_realizado", Decimal("0")) < 0 else FILL_CARD_GREEN,
    )

    row = 9
    _add_section_title(ws, row, "Consolidado mensal", 12)
    row += 1

    headers = [
        "Indicador",
        "Jan",
        "Fev",
        "Mar",
        "Abr",
        "Mai",
        "Jun",
        "Jul",
        "Ago",
        "Set",
        "Out",
        "Nov",
        "Dez",
    ]
    _write_table_header(ws, row, headers)
    row += 1

    table_rows = [
        ("Receita prevista", "receita_bruta_prevista"),
        ("Receita realizada", "receita_bruta_realizada"),
        ("Despesa prevista", "despesa_total_prevista"),
        ("Despesa realizada", "despesa_total_realizada"),
        ("Resultado previsto", "resultado_previsto"),
        ("Resultado realizado", "resultado_realizado"),
        ("Saldo acumulado previsto", "saldo_acumulado_previsto"),
        ("Saldo acumulado realizado", "saldo_acumulado_realizado"),
    ]

    for index, (label, key) in enumerate(table_rows):
        ws.cell(row=row, column=1, value=label).font = FONT_STRONG
        for month_index, mes in enumerate(mensal, start=2):
            cell = ws.cell(row=row, column=month_index, value=_safe_number(mes.get(key, Decimal("0"))))
            _apply_currency(cell)
        fill = FILL_RESULT if "Resultado" in label or "Saldo" in label else None
        _style_data_row(ws, row=row, col_count=len(headers), zebra=index % 2 == 1, fill=fill)
        row += 1

    row += 1
    _add_section_title(ws, row, "Totais anuais por categoria", 6)
    row += 1
    _write_table_header(ws, row, ["Categoria", "Previsto", "Realizado", "Diferenca", "Participacao real", "Leitura"])
    row += 1

    categorias = [
        ("Receita bruta", totais["receita_bruta_prevista"], totais["receita_bruta_realizada"]),
        ("Comissao principal", totais["comissao_principal_prevista"], totais["comissao_principal_realizada"]),
        ("Comissao cooperativa", totais["comissao_cooperativa_prevista"], totais["comissao_cooperativa_realizada"]),
        ("Despesas manuais", totais["despesas_manuais_previstas"], totais["despesas_manuais_realizadas"]),
        ("Impostos", totais["impostos_previstos"], totais["impostos_realizados"]),
        ("Retencoes", totais["retencoes_previstas"], totais["retencoes_realizadas"]),
        ("Despesa total", totais["despesa_total_prevista"], totais["despesa_total_realizada"]),
        ("Resultado", totais["resultado_previsto"], totais["resultado_realizado"]),
    ]
    receita_realizada = totais["receita_bruta_realizada"] or Decimal("0")
    for index, (label, previsto, realizado) in enumerate(categorias):
        ws.cell(row=row, column=1, value=label)
        for col, value in ((2, previsto), (3, realizado), (4, realizado - previsto)):
            cell = ws.cell(row=row, column=col, value=_safe_number(value))
            _apply_currency(cell)

        participacao = Decimal("0")
        if receita_realizada not in (Decimal("0"), 0):
            participacao = realizado / receita_realizada
        pct_cell = ws.cell(row=row, column=5, value=float(participacao))
        pct_cell.number_format = "0.00%"
        pct_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(
            row=row,
            column=6,
            value="Acima do previsto" if realizado > previsto else ("Abaixo do previsto" if realizado < previsto else "Em linha"),
        )
        _style_data_row(ws, row=row, col_count=6, zebra=index % 2 == 1)
        row += 1

    _set_column_widths(
        ws,
        {
            "A": 28,
            "B": 15,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 15,
            "K": 15,
            "L": 15,
            "M": 15,
        },
    )


def _build_fluxo_daily_sheet(workbook: Workbook, user, report: dict):
    ws = workbook.create_sheet("Caixa Diario")
    ws.freeze_panes = "A5"

    _apply_title_block(
        ws,
        title="Fluxo de Caixa Agro - Dia a Dia",
        subtitle="Livro-caixa diario com saldo de abertura, entradas, saidas e fechamento por dia.",
        year=report["ano"],
        max_col=12,
    )

    headers = [
        "Data",
        "Status do caixa",
        "Saldo abertura",
        "Entradas do dia",
        "Saidas do dia",
        "Resultado do dia",
        "Saldo fechamento",
        "Qtd. movimentos",
        "Aberto por",
        "Fechado por",
        "Obs. abertura",
        "Obs. fechamento",
    ]
    _write_table_header(ws, 4, headers)

    rows = _build_fluxo_daily_rows(user, report)
    for row_index, item in enumerate(rows, start=5):
        values = [
            _fmt_date(item["data_caixa"]),
            item["status_caixa"],
            item["saldo_abertura"],
            item["total_entradas"],
            item["total_saidas"],
            item["resultado_dia"],
            item["saldo_fechamento"],
            item["qtd_movimentos"],
            item["aberto_por"],
            item["fechado_por"],
            item["observacoes_abertura"],
            item["observacoes_fechamento"],
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=_safe_number(value))
            if 3 <= col_index <= 7:
                _apply_currency(cell)
            if col_index == 8:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        fill = FILL_RESULT
        if item["status_caixa"] == "ABERTO":
            fill = FILL_CARD_AMBER
        elif item["status_caixa"] == "SEM_ABERTURA":
            fill = FILL_CARD_RED
        elif item["status_caixa"] == "FECHADO":
            fill = FILL_CARD_GREEN
        _style_data_row(ws, row=row_index, col_count=len(headers), zebra=(row_index - 5) % 2 == 1, fill=fill)

    last_row = max(len(rows) + 4, 5)
    ws.auto_filter.ref = f"A4:L{last_row}"
    _set_column_widths(
        ws,
        {
            "A": 14,
            "B": 18,
            "C": 16,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 16,
            "H": 14,
            "I": 22,
            "J": 22,
            "K": 30,
            "L": 30,
        },
    )


def _build_fluxo_mensal_sheet(workbook: Workbook, report: dict):
    ws = workbook.create_sheet("Fluxo Mensal")
    ws.freeze_panes = "A6"

    _apply_title_block(
        ws,
        title="Fluxo de Caixa Agro - Detalhamento Mensal",
        subtitle="Meses no topo e indicadores na lateral para leitura mensal do caixa.",
        year=report["ano"],
        max_col=15,
    )

    headers = ["Indicador", "Leitura", *[month_name for _, month_name in AGRO_REPORT_MONTHS], "Total anual"]
    _write_table_header(ws, 5, headers)

    table_rows = [
        ("Receita prevista", "Entrada planejada para o caixa.", "receita_bruta_prevista"),
        ("Receita realizada", "Entrada que efetivamente caiu no caixa.", "receita_bruta_realizada"),
        ("Comissao principal prevista", "Comissao prevista vinculada aos contratos.", "comissao_principal_prevista"),
        ("Comissao principal realizada", "Comissao principal efetivamente realizada.", "comissao_principal_realizada"),
        ("Comissao cooperativa prevista", "Parcela prevista da cooperativa.", "comissao_cooperativa_prevista"),
        ("Comissao cooperativa realizada", "Parcela realizada da cooperativa.", "comissao_cooperativa_realizada"),
        ("Despesas manuais previstas", "Saidas operacionais previstas.", "despesas_manuais_previstas"),
        ("Despesas manuais realizadas", "Saidas operacionais pagas.", "despesas_manuais_realizadas"),
        ("Impostos previstos", "Tributos previstos no periodo.", "impostos_previstos"),
        ("Impostos realizados", "Tributos efetivamente pagos.", "impostos_realizados"),
        ("Retencoes previstas", "Retencoes previstas no periodo.", "retencoes_previstas"),
        ("Retencoes realizadas", "Retencoes efetivamente registradas.", "retencoes_realizadas"),
        ("Despesa total prevista", "Soma das saidas previstas.", "despesa_total_prevista"),
        ("Despesa total realizada", "Soma das saidas realizadas.", "despesa_total_realizada"),
        ("Resultado previsto", "Receita prevista menos saidas previstas.", "resultado_previsto"),
        ("Resultado realizado", "Receita realizada menos saidas realizadas.", "resultado_realizado"),
        ("Saldo acumulado previsto", "Posicao acumulada projetada do caixa.", "saldo_acumulado_previsto"),
        ("Saldo acumulado realizado", "Posicao acumulada real do caixa.", "saldo_acumulado_realizado"),
    ]

    for row_index, (label, leitura, key) in enumerate(table_rows, start=6):
        ws.cell(row=row_index, column=1, value=label).font = FONT_STRONG
        ws.cell(row=row_index, column=2, value=leitura)
        for month_index, mes in enumerate(report["mensal"], start=3):
            value = mes.get(key, Decimal("0"))
            cell = ws.cell(row=row_index, column=month_index, value=_safe_number(value))
            _apply_currency(cell)

        if key in {"saldo_acumulado_previsto", "saldo_acumulado_realizado"}:
            total = report["mensal"][-1].get(key, Decimal("0")) if report["mensal"] else Decimal("0")
        else:
            total = sum((mes.get(key, Decimal("0")) for mes in report["mensal"]), Decimal("0"))

        total_cell = ws.cell(row=row_index, column=15, value=_safe_number(total))
        _apply_currency(total_cell)
        fill = FILL_RESULT if key in {"resultado_previsto", "resultado_realizado", "saldo_acumulado_previsto", "saldo_acumulado_realizado"} else None
        _style_data_row(ws, row=row_index, col_count=len(headers), zebra=(row_index - 6) % 2 == 1, fill=fill)

    last_row = 5 + len(table_rows)
    ws.auto_filter.ref = f"A5:O{last_row}"

    _set_column_widths(
        ws,
        {
            "A": 28,
            "B": 36,
            "C": 14,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 14,
            "L": 14,
            "M": 14,
            "N": 14,
            "O": 18,
        },
    )


def _build_fluxo_dre_sheet(
    workbook: Workbook,
    report: dict,
    *,
    sheet_name: str = "Categorias DRE",
    title: str = "Fluxo de Caixa Agro - Leitura por Categoria",
    subtitle: str = "Abertura gerencial das linhas do DRE com categoria e subcategoria do financeiro.",
    use_emission_date: bool = False,
):
    ws = workbook.create_sheet(sheet_name)
    ws.freeze_panes = "A5"

    _apply_title_block(
        ws,
        title=title,
        subtitle=subtitle,
        year=report["ano"],
        max_col=19,
    )

    headers = [
        "Linha DRE",
        "Origem",
        "Categoria",
        "Subcategoria",
        "Jan",
        "Fev",
        "Mar",
        "Abr",
        "Mai",
        "Jun",
        "Jul",
        "Ago",
        "Set",
        "Out",
        "Nov",
        "Dez",
        "Total",
        "Participacao",
        "Leitura gerencial",
    ]
    _write_table_header(ws, 4, headers)

    rows = _build_fluxo_dre_rows(report, use_emission_date=use_emission_date)
    for row_index, item in enumerate(rows, start=5):
        ws.cell(row=row_index, column=1, value=item["linha_dre"]).font = FONT_STRONG
        ws.cell(row=row_index, column=2, value=item["origem"])
        ws.cell(row=row_index, column=3, value=item["categoria"])
        ws.cell(row=row_index, column=4, value=item["subcategoria"])
        for month_index, value in enumerate(item["meses"], start=5):
            cell = ws.cell(row=row_index, column=month_index, value=_safe_number(value))
            _apply_currency(cell)
        total_cell = ws.cell(row=row_index, column=17, value=_safe_number(item["total"]))
        _apply_currency(total_cell)
        pct_cell = ws.cell(row=row_index, column=18, value=float(item["participacao"]))
        pct_cell.number_format = "0.00%"
        pct_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_index, column=19, value=item["leitura"])

        fill = DRE_LINE_FILL.get(item["linha_dre"])
        _style_data_row(ws, row=row_index, col_count=len(headers), zebra=(row_index - 5) % 2 == 1, fill=fill)

    last_row = max(len(rows) + 4, 5)
    ws.auto_filter.ref = f"A4:S{last_row}"
    _set_column_widths(
        ws,
        {
            "A": 24,
            "B": 18,
            "C": 24,
            "D": 26,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 14,
            "L": 14,
            "M": 14,
            "N": 14,
            "O": 14,
            "P": 14,
            "Q": 16,
            "R": 14,
            "S": 40,
        },
    )


def _build_fluxo_lancamentos_sheet(workbook: Workbook, report: dict):
    ws = workbook.create_sheet("Lancamentos")
    ws.freeze_panes = "A5"

    _apply_title_block(
        ws,
        title="Base do Fluxo de Caixa Agro",
        subtitle="Todos os lancamentos que compoem o caixa anual, com previsto e realizado.",
        year=report["ano"],
        max_col=23,
    )

    headers = [
        "Origem",
        "NF / Doc",
        "Data emissao",
        "Data caixa",
        "Cliente / referencia",
        "Contrato",
        "Cultura",
        "Categoria",
        "Descricao",
        "Detalhamento fiscal",
        "Favorecido",
        "Forma",
        "Status",
        "Entrada prevista",
        "Entrada realizada",
        "Saida prevista",
        "Saida realizada",
        "Resultado previsto",
        "Resultado realizado",
        "Saldo previsto",
        "Saldo realizado",
        "Observacoes",
        "Tipo de saida",
    ]
    _write_table_header(ws, 4, headers)

    origin_fill = {
        "CONTRATO": FILL_CARD_GREEN,
        "ENTRADA_MANUAL": FILL_CARD_BLUE,
        "SAIDA_MANUAL": FILL_CARD_AMBER,
    }

    for row_index, item in enumerate(report["lancamentos"], start=5):
        values = [
            item["origem"],
            item.get("documento_referencia") or "",
            _fmt_date(item.get("data_emissao")),
            _fmt_date(item["data"]),
            item.get("cliente_nome") or "",
            item.get("contrato_id") or "",
            item.get("cultura") or "",
            item.get("categoria") or "",
            item.get("descricao") or "",
            item.get("detalhamento_imposto") or "",
            item.get("favorecido") or "",
            item.get("forma_recebimento") or "",
            item.get("status") or "",
            item.get("entrada_prevista") or Decimal("0"),
            item.get("entrada_realizada") or Decimal("0"),
            item.get("saida_prevista") or Decimal("0"),
            item.get("saida_realizada") or Decimal("0"),
            item.get("resultado_previsto") or Decimal("0"),
            item.get("resultado_realizado") or Decimal("0"),
            item.get("saldo_previsto_acumulado") or Decimal("0"),
            item.get("saldo_realizado_acumulado") or Decimal("0"),
            item.get("observacoes") or "",
            item.get("tipo_saida") or "",
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=_safe_number(value))
            if 14 <= col_index <= 21:
                _apply_currency(cell)

        fill = origin_fill.get(item["origem"])
        _style_data_row(ws, row=row_index, col_count=len(headers), zebra=(row_index - 5) % 2 == 1, fill=fill)

    last_row = max(len(report["lancamentos"]) + 4, 5)
    ws.auto_filter.ref = f"A4:W{last_row}"
    _set_column_widths(
        ws,
        {
            "A": 18,
            "B": 16,
            "C": 14,
            "D": 14,
            "E": 28,
            "F": 12,
            "G": 16,
            "H": 22,
            "I": 34,
            "J": 28,
            "K": 24,
            "L": 18,
            "M": 16,
            "N": 16,
            "O": 16,
            "P": 16,
            "Q": 16,
            "R": 16,
            "S": 16,
            "T": 16,
            "U": 36,
            "V": 18,
            "W": 18,
        },
    )


def build_agro_fluxo_caixa_excel_export(user, ano: int | None = None):
    report = build_agro_fluxo_caixa_report(user, ano=ano)
    workbook = Workbook()
    _build_fluxo_summary_sheet(workbook, report)
    _build_fluxo_daily_sheet(workbook, user, report)
    _build_fluxo_mensal_sheet(workbook, report)
    _build_fluxo_dre_sheet(workbook, report)
    _build_fluxo_dre_sheet(
        workbook,
        report,
        sheet_name="DRE por emissao",
        title="Fluxo de Caixa Agro - Categorias por Emissao",
        subtitle="Mesma leitura da aba Categorias DRE, mas agrupando pelo mes da data de emissao.",
        use_emission_date=True,
    )
    _build_fluxo_lancamentos_sheet(workbook, report)

    for ws in workbook.worksheets:
        _auto_width(ws)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, f"agro_fluxo_caixa_{report['ano']}.xlsx"


def _build_dre_summary_sheet(workbook: Workbook, report: dict):
    ws = workbook.active
    ws.title = "Resumo DRE"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    _apply_title_block(
        ws,
        title="DRE Gerencial Agro",
        subtitle="Resumo executivo do resultado mensal e acumulado da operacao Agro.",
        year=report["ano"],
        max_col=12,
    )

    despesas_operacionais = (
        report["comissao_principal_total"]
        + report["comissao_cooperativa_total"]
        + report["despesas_manuais_total"]
    )
    encargos = report["impostos_total"] + report["retencoes_total"]

    _write_card(ws, row=5, start_col=1, end_col=3, label="Receita bruta", value=report["faturamento_total"], fill=FILL_CARD_GREEN)
    _write_card(ws, row=5, start_col=4, end_col=6, label="Despesas operacionais", value=despesas_operacionais, fill=FILL_CARD_AMBER)
    _write_card(ws, row=5, start_col=7, end_col=9, label="Impostos e retencoes", value=encargos, fill=FILL_CARD_RED)
    _write_card(
        ws,
        row=5,
        start_col=10,
        end_col=12,
        label="Resultado anual",
        value=report["resultado_total"],
        fill=FILL_CARD_BLUE if report["resultado_total"] >= 0 else FILL_CARD_RED,
    )

    row = 9
    _add_section_title(ws, row, "Indicadores anuais", 6)
    row += 1
    _write_table_header(ws, row, ["Indicador", "Valor", "Participacao na receita", "Leitura", "Observacao", "Fechamento"])
    row += 1

    indicadores = [
        ("Receita bruta", report["faturamento_total"], "Base de faturamento do Agro", "Entrada", "Contratos e entradas manuais", "Anual"),
        ("Comissao principal", report["comissao_principal_total"], "Saida operacional direta", "Custo", "Relacionada aos contratos", "Anual"),
        ("Comissao cooperativa", report["comissao_cooperativa_total"], "Saida operacional indireta", "Custo", "Parcela de cooperativa", "Anual"),
        ("Despesas manuais", report["despesas_manuais_total"], "Saida fora do contrato", "Custo", "Combustivel, equipe, terceiros etc.", "Anual"),
        ("Impostos", report["impostos_total"], "Encargo fiscal", "Fiscal", "Saidas tributarias", "Anual"),
        ("Retencoes", report["retencoes_total"], "Encargo financeiro", "Fiscal", "Valores retidos", "Anual"),
        ("Resultado", report["resultado_total"], "Lucro / prejuizo", "Resultado", "Receita menos saidas", "Anual"),
        ("Saldo acumulado final", report["saldo_total"], "Posicao acumulada", "Caixa", "Fechamento do ano", "Dezembro"),
    ]
    receita = report["faturamento_total"] or Decimal("0")
    for index, (label, value, leitura, tipo, observacao, fechamento) in enumerate(indicadores):
        ws.cell(row=row, column=1, value=label)
        money_cell = ws.cell(row=row, column=2, value=_safe_number(value))
        _apply_currency(money_cell)

        participacao = Decimal("0")
        if receita not in (Decimal("0"), 0):
            participacao = value / receita
        pct_cell = ws.cell(row=row, column=3, value=float(participacao))
        pct_cell.number_format = "0.00%"
        pct_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=4, value=leitura)
        ws.cell(row=row, column=5, value=observacao)
        ws.cell(row=row, column=6, value=fechamento)
        _style_data_row(ws, row=row, col_count=6, zebra=index % 2 == 1)
        row += 1

    row += 1
    _add_section_title(ws, row, "Leitura mensal resumida", 15)
    row += 1
    headers = ["Linha do DRE", "Descricao", *report["meses"], "Total"]
    _write_table_header(ws, row, headers)
    row += 1

    for index, linha in enumerate(report["linhas"]):
        ws.cell(row=row, column=1, value=linha["label"]).font = FONT_STRONG
        ws.cell(row=row, column=2, value=linha["description"])
        expense_line = linha["kind"] in EXPENSE_KINDS
        for month_index, valor in enumerate(linha["valores"], start=3):
            cell = ws.cell(row=row, column=month_index, value=_safe_number(valor))
            _apply_currency(cell)
            if expense_line:
                cell.font = Font(color="A61B1B", bold=linha["kind"] in {"despesa_total_realizada"})
        total_cell = ws.cell(row=row, column=15, value=_safe_number(linha["total"]))
        _apply_currency(total_cell)
        if expense_line:
            total_cell.font = Font(color="A61B1B", bold=True)
        fill = FILL_RESULT if linha["kind"] in {"resultado_realizado", "saldo_acumulado_realizado"} else None
        _style_data_row(ws, row=row, col_count=15, zebra=index % 2 == 1, fill=fill)
        row += 1

    _set_column_widths(
        ws,
        {
            "A": 28,
            "B": 38,
            "C": 14,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 14,
            "L": 14,
            "M": 14,
            "N": 14,
            "O": 16,
        },
    )


def _build_dre_mensal_sheet(workbook: Workbook, report: dict):
    ws = workbook.create_sheet("DRE Mensal")
    ws.freeze_panes = "A5"

    _apply_title_block(
        ws,
        title="DRE Gerencial Agro - Demonstrativo",
        subtitle="Matriz mensal das linhas gerenciais utilizadas na leitura do resultado.",
        year=report["ano"],
        max_col=16,
    )

    headers = ["Linha do DRE", "Descricao", *report["meses"], "Total"]
    _write_table_header(ws, 4, headers)

    for row_index, linha in enumerate(report["linhas"], start=5):
        ws.cell(row=row_index, column=1, value=linha["label"]).font = FONT_STRONG
        ws.cell(row=row_index, column=2, value=linha["description"])
        expense_line = linha["kind"] in EXPENSE_KINDS
        for month_index, valor in enumerate(linha["valores"], start=3):
            cell = ws.cell(row=row_index, column=month_index, value=_safe_number(valor))
            _apply_currency(cell)
            if expense_line:
                cell.font = Font(color="A61B1B")
        total_cell = ws.cell(row=row_index, column=15, value=_safe_number(linha["total"]))
        _apply_currency(total_cell)
        total_cell.font = Font(bold=True, color="A61B1B" if expense_line else "1F5130")
        fill = FILL_RESULT if linha["kind"] in {"resultado_realizado", "saldo_acumulado_realizado"} else None
        _style_data_row(ws, row=row_index, col_count=15, zebra=(row_index - 5) % 2 == 1, fill=fill)

    last_row = max(len(report["linhas"]) + 4, 5)
    ws.auto_filter.ref = f"A4:O{last_row}"
    _set_column_widths(
        ws,
        {
            "A": 28,
            "B": 40,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 15,
            "K": 15,
            "L": 15,
            "M": 15,
            "N": 15,
            "O": 18,
        },
    )


def _build_dre_base_sheet(workbook: Workbook, fluxo_report: dict):
    ws = workbook.create_sheet("Base DRE")
    ws.freeze_panes = "A5"

    _apply_title_block(
        ws,
        title="Base do DRE Gerencial Agro",
        subtitle="Lancamentos realizados que sustentam a leitura do DRE no periodo.",
        year=fluxo_report["ano"],
        max_col=15,
    )

    headers = [
        "Data caixa",
        "Data emissao",
        "NF / Doc",
        "Origem",
        "Cliente / referencia",
        "Categoria",
        "Descricao",
        "Detalhamento fiscal",
        "Favorecido",
        "Status",
        "Entrada realizada",
        "Saida realizada",
        "Resultado realizado",
        "Saldo realizado",
        "Observacoes",
        "Contrato",
    ]
    _write_table_header(ws, 4, headers)

    for row_index, item in enumerate(fluxo_report["lancamentos"], start=5):
        values = [
            _fmt_date(item["data"]),
            _fmt_date(item.get("data_emissao")),
            item.get("documento_referencia") or "",
            item.get("origem") or "",
            item.get("cliente_nome") or "",
            item.get("categoria") or "",
            item.get("descricao") or "",
            item.get("detalhamento_imposto") or "",
            item.get("favorecido") or "",
            item.get("status") or "",
            item.get("entrada_realizada") or Decimal("0"),
            item.get("saida_realizada") or Decimal("0"),
            item.get("resultado_realizado") or Decimal("0"),
            item.get("saldo_realizado_acumulado") or Decimal("0"),
            item.get("observacoes") or "",
            item.get("contrato_id") or "",
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=_safe_number(value))
            if 11 <= col_index <= 14:
                _apply_currency(cell)
        _style_data_row(ws, row=row_index, col_count=len(headers), zebra=(row_index - 5) % 2 == 1)

    last_row = max(len(fluxo_report["lancamentos"]) + 4, 5)
    ws.auto_filter.ref = f"A4:P{last_row}"
    _set_column_widths(
        ws,
        {
            "A": 14,
            "B": 14,
            "C": 16,
            "D": 18,
            "E": 28,
            "F": 22,
            "G": 34,
            "H": 28,
            "I": 24,
            "J": 16,
            "K": 16,
            "L": 16,
            "M": 16,
            "N": 36,
            "O": 36,
            "P": 12,
        },
    )


def build_agro_dre_gerencial_excel_export(user, ano: int | None = None):
    report = build_agro_dre_gerencial_report(user, ano=ano)
    fluxo_report = build_agro_fluxo_caixa_report(user, ano=report["ano"])

    workbook = Workbook()
    _build_dre_summary_sheet(workbook, report)
    _build_dre_mensal_sheet(workbook, report)
    _build_dre_base_sheet(workbook, fluxo_report)

    for ws in workbook.worksheets:
        _auto_width(ws)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, f"agro_dre_gerencial_{report['ano']}.xlsx"
