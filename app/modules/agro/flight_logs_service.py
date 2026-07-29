import hashlib
import json
import os
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models import AgroFlightKmlRoute, AgroFlightLogImport, AgroFlightRecord, OrdemServicoAgro
from app.modules.agro.service import can_access_agro_panel, can_edit_agro_panel
from app.modules.dji_flight_logs.service import (
    _build_weekly_summary,
    _compute_route_distance_meters,
    _format_altitude_range_label,
    _format_coordinate_label,
    _format_distance_label,
    _kml_color_to_css,
    _load_workbook,
    _normalize_match_text,
    _parse_kml_payload,
    _parse_workbook_rows,
    _text_match_score,
    format_duration_seconds,
)
from app.shared.uploads import get_upload_folder


def can_access_agro_flight_logs(user) -> bool:
    return can_access_agro_panel(user)


def can_import_agro_flight_logs(user) -> bool:
    return can_edit_agro_panel(user)


def can_access_agro_kml_route(user, route_id) -> bool:
    if can_access_agro_panel(user):
        return True

    if getattr(user, "tipo_usuario", None) != "piloto_agro":
        return False

    piloto = getattr(user, "piloto_agro", None)
    equipe_id = getattr(piloto, "equipe_agro_id", None)
    if not equipe_id:
        return False

    return (
        OrdemServicoAgro.query
        .filter(
            OrdemServicoAgro.agro_kml_route_id == route_id,
            OrdemServicoAgro.equipe_agro_id == equipe_id,
        )
        .first()
        is not None
    )


def import_agro_log_excel(file_storage, user):
    if not file_storage or not file_storage.filename:
        raise ValueError("Selecione um arquivo Excel para importar.")

    original_filename = (file_storage.filename or "").strip()
    if not original_filename.lower().endswith(".xlsx"):
        raise ValueError("O arquivo precisa estar no formato .xlsx.")

    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("O arquivo enviado esta vazio.")

    file_sha256 = hashlib.sha256(file_bytes).hexdigest()
    previous_import = (
        AgroFlightLogImport.query
        .filter(AgroFlightLogImport.file_sha256 == file_sha256)
        .order_by(AgroFlightLogImport.uploaded_at.desc())
        .first()
    )
    if previous_import:
        when = previous_import.uploaded_at.strftime("%d/%m/%Y %H:%M")
        raise ValueError(f"Esse mesmo arquivo Agro ja foi importado em {when}.")

    workbook = _load_workbook(file_bytes)
    rows = _parse_workbook_rows(workbook)
    if not rows:
        raise ValueError("Nenhum voo valido foi encontrado no Excel informado.")

    stored_filename, stored_path = _save_uploaded_agro_excel(original_filename, file_bytes)
    import_batch = AgroFlightLogImport(
        uploaded_by_id=getattr(user, "id", None),
        original_filename=original_filename,
        stored_filename=stored_filename,
        stored_path=stored_path,
        file_sha256=file_sha256,
        total_rows=len(rows),
        imported_rows=0,
        skipped_rows=0,
        period_start=min(item["flight_start"] for item in rows),
        period_end=max(item["flight_end"] for item in rows),
    )
    db.session.add(import_batch)
    db.session.flush()

    fingerprints = [item["fingerprint"] for item in rows]
    existing_fingerprints = {
        value
        for (value,) in db.session.query(AgroFlightRecord.fingerprint)
        .filter(AgroFlightRecord.fingerprint.in_(fingerprints))
        .all()
    }

    imported_rows = 0
    skipped_rows = 0
    for item in rows:
        if item["fingerprint"] in existing_fingerprints:
            skipped_rows += 1
            continue

        existing_fingerprints.add(item["fingerprint"])
        db.session.add(
            AgroFlightRecord(
                import_id=import_batch.id,
                source_row_number=item["source_row_number"],
                fingerprint=item["fingerprint"],
                flight_window=item["flight_window"],
                flight_start=item["flight_start"],
                flight_end=item["flight_end"],
                location=item["location"],
                aircraft_name=item["aircraft_name"],
                task_type=item["task_type"],
                sprayed_area_ha=item["sprayed_area_ha"],
                total_amount_l_kg=item["total_amount_l_kg"],
                flight_duration_seconds=item["flight_duration_seconds"],
                flight_duration_label=item["flight_duration_label"],
                crop=item["crop"],
                pilot_name=item["pilot_name"],
                team_name=item["team_name"],
                field_name=item["field_name"],
                serial_number=item["serial_number"],
                starting_battery_level=item["starting_battery_level"],
                ending_battery_level=item["ending_battery_level"],
                battery_consumed_level=item["battery_consumed_level"],
                battery_sn=item["battery_sn"],
                raw_payload=json.dumps(item["raw_payload"], ensure_ascii=False),
            )
        )
        imported_rows += 1

    import_batch.imported_rows = imported_rows
    import_batch.skipped_rows = skipped_rows
    db.session.commit()
    return import_batch


def build_agro_flight_logs_context(args):
    data_inicio = (args.get("data_inicio") or "").strip()
    data_fim = (args.get("data_fim") or "").strip()
    piloto = (args.get("piloto") or "").strip()
    aeronave = (args.get("aeronave") or "").strip()
    equipe = (args.get("equipe") or "").strip()
    q = (args.get("q") or "").strip()
    endereco = (args.get("endereco") or "").strip()
    voo_id = (args.get("voo_id") or "").strip()
    status_rota = (args.get("status_rota") or "").strip()
    kml_q = (args.get("kml_q") or "").strip()
    kml_data_inicio = (args.get("kml_data_inicio") or "").strip()
    kml_data_fim = (args.get("kml_data_fim") or "").strip()
    kml_piloto = (args.get("kml_piloto") or "").strip()
    kml_aeronave = (args.get("kml_aeronave") or "").strip()
    kml_status_os = (args.get("kml_status_os") or "").strip()
    kml_status_voo = (args.get("kml_status_voo") or "").strip()
    kml_voo_id = (args.get("kml_voo_id") or "").strip()
    page = args.get("page", 1, type=int)
    kml_page = args.get("kml_page", 1, type=int)

    filtered_query = _build_filtered_agro_record_query(
        data_inicio=data_inicio,
        data_fim=data_fim,
        piloto=piloto,
        aeronave=aeronave,
        equipe=equipe,
        q=q,
        endereco=endereco,
        voo_id=voo_id,
        status_rota=status_rota,
    )
    paginacao = (
        filtered_query
        .order_by(AgroFlightRecord.flight_start.desc(), AgroFlightRecord.id.desc())
        .paginate(page=page, per_page=20, error_out=False)
    )

    totals = filtered_query.with_entities(
        func.count(AgroFlightRecord.id),
        func.coalesce(func.sum(AgroFlightRecord.sprayed_area_ha), 0.0),
        func.coalesce(func.sum(AgroFlightRecord.total_amount_l_kg), 0.0),
        func.coalesce(func.sum(AgroFlightRecord.flight_duration_seconds), 0),
        func.avg(AgroFlightRecord.battery_consumed_level),
    ).first()
    total_voos, total_area, total_volume, total_duracao, media_bateria = totals

    kml_query = _build_agro_kml_query(
        q=kml_q,
        data_inicio=kml_data_inicio,
        data_fim=kml_data_fim,
        piloto=kml_piloto,
        aeronave=kml_aeronave,
        status_os=kml_status_os,
        status_voo=kml_status_voo,
        voo_id=kml_voo_id,
    )
    kml_paginacao = (
        kml_query
        .order_by(AgroFlightKmlRoute.imported_at.desc(), AgroFlightKmlRoute.id.desc())
        .paginate(page=kml_page, per_page=25, error_out=False)
    )
    kml_rotas = kml_paginacao.items

    total_rotas_kml = AgroFlightKmlRoute.query.count()
    total_rotas_kml_vinculadas = (
        db.session.query(func.count(func.distinct(OrdemServicoAgro.agro_kml_route_id)))
        .filter(OrdemServicoAgro.agro_kml_route_id.isnot(None))
        .scalar()
        or 0
    )

    weekly_rows = filtered_query.with_entities(
        AgroFlightRecord.flight_start,
        AgroFlightRecord.sprayed_area_ha,
        AgroFlightRecord.total_amount_l_kg,
        AgroFlightRecord.flight_duration_seconds,
    ).all()

    return {
        "registros": paginacao.items,
        "paginacao": paginacao,
        "registro_kml_os_por_rota": build_agro_kml_os_map(
            [registro.route_kml for registro in paginacao.items if registro.route_kml]
        ),
        "filtros": {
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "piloto": piloto,
            "aeronave": aeronave,
            "equipe": equipe,
            "q": q,
            "endereco": endereco,
            "voo_id": voo_id,
            "status_rota": status_rota,
            "total": total_voos or 0,
        },
        "kml_filtros": {
            "q": kml_q,
            "data_inicio": kml_data_inicio,
            "data_fim": kml_data_fim,
            "piloto": kml_piloto,
            "aeronave": kml_aeronave,
            "status_os": kml_status_os,
            "status_voo": kml_status_voo,
            "voo_id": kml_voo_id,
        },
        "pilotos_disponiveis": _distinct_non_empty_values(AgroFlightRecord.pilot_name),
        "aeronaves_disponiveis": _distinct_non_empty_values(AgroFlightRecord.aircraft_name),
        "equipes_disponiveis": _distinct_non_empty_values(AgroFlightRecord.team_name),
        "kml_pilotos_disponiveis": _distinct_non_empty_values(AgroFlightKmlRoute.pilot_name),
        "kml_aeronaves_disponiveis": _distinct_non_empty_values(AgroFlightKmlRoute.aircraft_name),
        "importacoes_recentes": (
            AgroFlightLogImport.query
            .order_by(AgroFlightLogImport.uploaded_at.desc(), AgroFlightLogImport.id.desc())
            .limit(8)
            .all()
        ),
        "kml_rotas": kml_rotas,
        "kml_paginacao": kml_paginacao,
        "kml_os_por_rota": build_agro_kml_os_map(kml_rotas),
        "total_importacoes": AgroFlightLogImport.query.count(),
        "total_rotas_kml": total_rotas_kml,
        "total_rotas_kml_vinculadas": total_rotas_kml_vinculadas,
        "total_rotas_kml_sem_os": total_rotas_kml - total_rotas_kml_vinculadas,
        "total_voos": total_voos or 0,
        "total_area": float(total_area or 0),
        "total_volume": float(total_volume or 0),
        "total_duracao": format_duration_seconds(total_duracao or 0),
        "media_bateria": round(media_bateria or 0, 1) if media_bateria is not None else None,
        "top_pilotos": _build_top_groups(filtered_query, AgroFlightRecord.pilot_name, limit=5),
        "resumo_semanal": _build_weekly_summary(weekly_rows)[:6],
    }


def import_agro_kml_files(files, user):
    valid_files = [file for file in (files or []) if file and file.filename]
    if not valid_files:
        raise ValueError("Selecione ao menos um arquivo KML para importar.")

    imported = 0
    skipped = 0
    linked = 0
    os_linked = 0
    existing_linked = 0

    for file_storage in valid_files:
        original_filename = (file_storage.filename or "").strip()
        if not original_filename.lower().endswith(".kml"):
            raise ValueError("Todos os arquivos enviados devem estar no formato .kml.")

        file_bytes = file_storage.read()
        if not file_bytes:
            continue

        file_sha256 = hashlib.sha256(file_bytes).hexdigest()
        parsed = _parse_kml_payload(file_bytes, original_filename)
        existing_route = AgroFlightKmlRoute.query.filter(
            AgroFlightKmlRoute.file_sha256 == file_sha256
        ).first()
        if existing_route is None:
            existing_route = AgroFlightKmlRoute.query.filter(
                AgroFlightKmlRoute.route_code == parsed["route_code"]
            ).first()

        if existing_route:
            skipped += 1
            if _auto_link_agro_kml_route_to_os(existing_route, _route_points(existing_route)):
                existing_linked += 1
            continue

        stored_filename, stored_path = _save_uploaded_agro_kml(original_filename, file_bytes)
        matched_record = (
            AgroFlightRecord.query
            .filter(AgroFlightRecord.serial_number == parsed["route_code"])
            .order_by(AgroFlightRecord.flight_start.desc(), AgroFlightRecord.id.desc())
            .first()
        )
        route = AgroFlightKmlRoute(
            flight_record_id=matched_record.id if matched_record else None,
            uploaded_by_id=getattr(user, "id", None),
            route_code=parsed["route_code"],
            original_filename=original_filename,
            stored_filename=stored_filename,
            stored_path=stored_path,
            file_sha256=file_sha256,
            aircraft_name=parsed["aircraft_name"],
            pilot_name=parsed["pilot_name"],
            flight_controller_id=parsed["flight_controller_id"],
            route_timestamp=parsed["route_timestamp"],
            mode_selection=parsed["mode_selection"],
            flight_time_raw=parsed["flight_time_raw"],
            task_area=parsed["task_area"],
            spray_amount=parsed["spray_amount"],
            route_color=parsed["route_color"],
            route_width=parsed["route_width"],
            point_count=len(parsed["points"]),
            points_json=json.dumps(parsed["points"], ensure_ascii=False),
        )
        db.session.add(route)
        db.session.flush()

        imported += 1
        if matched_record:
            linked += 1
        if _auto_link_agro_kml_route_to_os(route, parsed["points"]):
            os_linked += 1

    db.session.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "linked": linked,
        "os_linked": os_linked,
        "existing_linked": existing_linked,
        "unlinked": imported - linked,
    }


def build_agro_logs_excel_export(args):
    filtered_query = _build_filtered_agro_record_query(
        data_inicio=(args.get("data_inicio") or "").strip(),
        data_fim=(args.get("data_fim") or "").strip(),
        piloto=(args.get("piloto") or "").strip(),
        aeronave=(args.get("aeronave") or "").strip(),
        equipe=(args.get("equipe") or "").strip(),
        q=(args.get("q") or "").strip(),
        endereco=(args.get("endereco") or "").strip(),
        voo_id=(args.get("voo_id") or "").strip(),
        status_rota=(args.get("status_rota") or "").strip(),
    )
    registros = filtered_query.order_by(AgroFlightRecord.flight_start.desc(), AgroFlightRecord.id.desc()).all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Voos Agro"
    fill_header = PatternFill("solid", fgColor="198754")
    font_header = Font(color="FFFFFF", bold=True)
    headers = [
        "Inicio", "Fim", "Periodo DJI", "Piloto", "Aeronave", "Equipe",
        "Tipo de tarefa", "Area (ha)", "Volume (L/Kg)", "Duracao",
        "Cultura", "Local", "Serial da aeronave", "Rota KML",
    ]
    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, registro in enumerate(registros, start=2):
        values = [
            registro.flight_start.strftime("%d/%m/%Y %H:%M") if registro.flight_start else "",
            registro.flight_end.strftime("%d/%m/%Y %H:%M") if registro.flight_end else "",
            registro.flight_window or "",
            registro.pilot_name or "",
            registro.aircraft_name or "",
            registro.team_name or "",
            registro.task_type or "",
            float(registro.sprayed_area_ha or 0),
            float(registro.total_amount_l_kg or 0),
            registro.duration_display,
            registro.crop or "",
            registro.location or "",
            registro.serial_number or "",
            registro.route_kml.route_code if registro.route_kml else "",
        ]
        for col, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=col, value=value)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(registros) + 1, 2)}"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, f"agro_logs_voo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def link_agro_kml_route_to_os(route_id, os_ref):
    route = AgroFlightKmlRoute.query.get(route_id)
    if not route:
        raise ValueError("Rota KML Agro nao encontrada.")

    os_ref = str(os_ref or "").strip()
    if not os_ref:
        raise ValueError("Informe o ID ou identificador da OS Agro.")

    query = OrdemServicoAgro.query
    if os_ref.isdigit():
        query = query.filter(or_(OrdemServicoAgro.id == int(os_ref), OrdemServicoAgro.identificador_os == os_ref))
    else:
        query = query.filter(func.lower(OrdemServicoAgro.identificador_os) == os_ref.lower())
    ordem = query.first()
    if not ordem:
        raise ValueError("OS Agro nao encontrada.")

    if ordem.agro_kml_route_id and ordem.agro_kml_route_id != route.id:
        raise ValueError("Essa OS Agro ja possui outra rota KML vinculada.")

    current_linked = (
        OrdemServicoAgro.query
        .filter(OrdemServicoAgro.agro_kml_route_id == route.id, OrdemServicoAgro.id != ordem.id)
        .all()
    )
    for linked_ordem in current_linked:
        linked_ordem.agro_kml_route_id = None

    ordem.agro_kml_route_id = route.id
    db.session.commit()
    return ordem


def unlink_agro_kml_route_from_os(route_id):
    linked_count = (
        OrdemServicoAgro.query
        .filter(OrdemServicoAgro.agro_kml_route_id == route_id)
        .update({OrdemServicoAgro.agro_kml_route_id: None}, synchronize_session=False)
    )
    db.session.commit()
    return linked_count or 0


def get_agro_route_payload(route_id):
    route = AgroFlightKmlRoute.query.get_or_404(route_id)
    linked_os = _get_linked_agro_os_for_kml_route(route.id)
    points = _route_points(route)
    start_point = points[0] if points else None
    end_point = points[-1] if points else None
    altitude_values = [float(point.get("alt", 0) or 0) for point in points if point.get("alt") is not None]
    altitude_min = min(altitude_values) if altitude_values else None
    altitude_max = max(altitude_values) if altitude_values else None
    distance_meters = _compute_route_distance_meters(points)
    return {
        "id": route.id,
        "route_code": route.route_code,
        "aircraft_name": route.aircraft_name,
        "pilot_name": route.pilot_name,
        "flight_controller_id": route.flight_controller_id,
        "route_timestamp": route.route_timestamp.strftime("%d/%m/%Y %H:%M:%S") if route.route_timestamp else "",
        "task_area": route.task_area,
        "spray_amount": route.spray_amount,
        "point_count": route.point_count,
        "route_color_css": _kml_color_to_css(route.route_color) or "#198754",
        "route_width": route.route_width or 2,
        "start_label": _format_coordinate_label(start_point),
        "end_label": _format_coordinate_label(end_point),
        "distance_meters": distance_meters,
        "distance_label": _format_distance_label(distance_meters),
        "altitude_label": _format_altitude_range_label(altitude_min, altitude_max),
        "linked_os": _build_linked_agro_os_payload(linked_os),
        "points": points,
    }


def build_agro_kml_os_map(routes):
    route_ids = [route.id for route in routes or [] if route]
    if not route_ids:
        return {}

    ordens = (
        OrdemServicoAgro.query
        .options(joinedload(OrdemServicoAgro.contrato), joinedload(OrdemServicoAgro.equipe))
        .filter(OrdemServicoAgro.agro_kml_route_id.in_(route_ids))
        .order_by(OrdemServicoAgro.data_aplicacao.desc().nullslast(), OrdemServicoAgro.id.desc())
        .all()
    )
    os_por_rota = {}
    for ordem in ordens:
        os_por_rota.setdefault(ordem.agro_kml_route_id, ordem)
    return os_por_rota


def _build_filtered_agro_record_query(*, data_inicio="", data_fim="", piloto="", aeronave="", equipe="", q="", endereco="", voo_id="", status_rota=""):
    query = AgroFlightRecord.query.options(joinedload(AgroFlightRecord.route_kml))

    if voo_id:
        try:
            query = query.filter(AgroFlightRecord.id == int(voo_id))
        except (TypeError, ValueError):
            pass
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            AgroFlightRecord.flight_window.ilike(like),
            AgroFlightRecord.location.ilike(like),
            AgroFlightRecord.aircraft_name.ilike(like),
            AgroFlightRecord.task_type.ilike(like),
            AgroFlightRecord.crop.ilike(like),
            AgroFlightRecord.pilot_name.ilike(like),
            AgroFlightRecord.team_name.ilike(like),
            AgroFlightRecord.field_name.ilike(like),
            AgroFlightRecord.serial_number.ilike(like),
            AgroFlightRecord.battery_sn.ilike(like),
        ))
    if endereco:
        query = query.filter(AgroFlightRecord.location.ilike(f"%{endereco}%"))
    if data_inicio:
        try:
            query = query.filter(AgroFlightRecord.flight_start >= datetime.strptime(data_inicio, "%Y-%m-%d"))
        except ValueError:
            pass
    if data_fim:
        try:
            query = query.filter(AgroFlightRecord.flight_start < datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    if piloto:
        query = query.filter(AgroFlightRecord.pilot_name == piloto)
    if aeronave:
        query = query.filter(AgroFlightRecord.aircraft_name == aeronave)
    if equipe:
        query = query.filter(AgroFlightRecord.team_name == equipe)

    linked_record_ids = db.session.query(AgroFlightKmlRoute.flight_record_id).filter(
        AgroFlightKmlRoute.flight_record_id.isnot(None)
    )
    if status_rota == "com_kml":
        query = query.filter(AgroFlightRecord.id.in_(linked_record_ids))
    elif status_rota == "sem_kml":
        query = query.filter(~AgroFlightRecord.id.in_(linked_record_ids))
    return query


def _build_agro_kml_query(*, q="", data_inicio="", data_fim="", piloto="", aeronave="", status_os="", status_voo="", voo_id=""):
    query = AgroFlightKmlRoute.query.options(joinedload(AgroFlightKmlRoute.flight_record))
    if voo_id:
        try:
            query = query.filter(AgroFlightKmlRoute.flight_record_id == int(voo_id))
        except (TypeError, ValueError):
            pass
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            AgroFlightKmlRoute.route_code.ilike(like),
            AgroFlightKmlRoute.original_filename.ilike(like),
            AgroFlightKmlRoute.aircraft_name.ilike(like),
            AgroFlightKmlRoute.pilot_name.ilike(like),
            AgroFlightKmlRoute.flight_controller_id.ilike(like),
        ))
    if data_inicio:
        try:
            query = query.filter(AgroFlightKmlRoute.route_timestamp >= datetime.strptime(data_inicio, "%Y-%m-%d"))
        except ValueError:
            pass
    if data_fim:
        try:
            query = query.filter(AgroFlightKmlRoute.route_timestamp < datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    if piloto:
        query = query.filter(AgroFlightKmlRoute.pilot_name == piloto)
    if aeronave:
        query = query.filter(AgroFlightKmlRoute.aircraft_name == aeronave)
    if status_voo == "com_voo":
        query = query.filter(AgroFlightKmlRoute.flight_record_id.isnot(None))
    elif status_voo == "sem_voo":
        query = query.filter(AgroFlightKmlRoute.flight_record_id.is_(None))

    linked_route_ids = db.session.query(OrdemServicoAgro.agro_kml_route_id).filter(
        OrdemServicoAgro.agro_kml_route_id.isnot(None)
    )
    if status_os == "com_os":
        query = query.filter(AgroFlightKmlRoute.id.in_(linked_route_ids))
    elif status_os == "sem_os":
        query = query.filter(~AgroFlightKmlRoute.id.in_(linked_route_ids))
    return query


def _auto_link_agro_kml_route_to_os(route, points):
    match = _find_best_agro_os_match_for_kml_route(route, points)
    if not match:
        return None
    ordem, _score, _details = match
    ordem.agro_kml_route_id = route.id
    return ordem


def _find_best_agro_os_match_for_kml_route(route, points):
    candidates = _candidate_agro_ordens_for_kml_route(route)
    scored = []
    for ordem in candidates:
        score, details = _score_agro_os_kml_match(ordem, route)
        if _is_confident_agro_os_kml_match(score, details):
            scored.append((ordem, score, details))
    if not scored:
        return None
    scored.sort(key=lambda item: item[1], reverse=True)
    best = scored[0]
    if len(scored) > 1 and (best[1] - scored[1][1]) < 12:
        return None
    return best


def _candidate_agro_ordens_for_kml_route(route):
    query = (
        OrdemServicoAgro.query
        .options(
            joinedload(OrdemServicoAgro.equipe),
            joinedload(OrdemServicoAgro.piloto),
            joinedload(OrdemServicoAgro.drone_pulverizacao),
            joinedload(OrdemServicoAgro.drone_mapeamento),
        )
        .filter(OrdemServicoAgro.agro_kml_route_id.is_(None))
    )
    if route.route_timestamp:
        start_date = (route.route_timestamp - timedelta(days=2)).date()
        end_date = (route.route_timestamp + timedelta(days=2)).date()
        query = query.filter(OrdemServicoAgro.data_aplicacao.between(start_date, end_date))
    return query.order_by(OrdemServicoAgro.data_aplicacao.desc().nullslast(), OrdemServicoAgro.id.desc()).limit(300).all()


def _score_agro_os_kml_match(ordem, route):
    time_score = _score_agro_time_match(ordem, route.route_timestamp)
    aircraft_score = _score_agro_aircraft_match(ordem, route.aircraft_name)
    pilot_score = _score_agro_pilot_match(ordem, route.pilot_name)
    service_score = _score_agro_service_match(ordem, route)
    score = time_score + aircraft_score + pilot_score + service_score
    return score, {"time": time_score, "aircraft": aircraft_score, "pilot": pilot_score, "service": service_score}


def _is_confident_agro_os_kml_match(score, details):
    return score >= 60 and details["time"] >= 25 and (details["aircraft"] >= 25 or details["pilot"] >= 12)


def _score_agro_time_match(ordem, route_timestamp):
    if not route_timestamp or not ordem.data_aplicacao:
        return 0
    delta_days = abs((route_timestamp.date() - ordem.data_aplicacao).days)
    if delta_days == 0:
        return 35
    if delta_days == 1:
        return 25
    if delta_days == 2:
        return 15
    return 0


def _score_agro_aircraft_match(ordem, aircraft_name):
    route_aircraft = _normalize_match_text(aircraft_name)
    if not route_aircraft:
        return 0
    candidates = [
        ordem.drone_pulverizacao_identificacao,
        ordem.drone_mapeamento_identificacao,
        ordem.drone_pulverizacao_modelo,
        ordem.drone_mapeamento_modelo,
        getattr(getattr(ordem, "drone_pulverizacao", None), "numero_serie", None),
        getattr(getattr(ordem, "drone_mapeamento", None), "numero_serie", None),
    ]
    return max((_text_match_score(route_aircraft, candidate) for candidate in candidates), default=0)


def _score_agro_pilot_match(ordem, pilot_name):
    route_pilot = _normalize_match_text(pilot_name)
    if not route_pilot:
        return 0
    candidates = [getattr(getattr(ordem, "piloto", None), "nome", None)]
    if ordem.equipe and ordem.equipe.pilotos:
        candidates.extend(piloto.nome for piloto in ordem.equipe.pilotos if getattr(piloto, "ativo", True))
    return max((_text_match_score(route_pilot, candidate, exact_score=18, contains_score=12) for candidate in candidates), default=0)


def _score_agro_service_match(ordem, route):
    task_area = float(route.task_area or 0)
    os_area = float(ordem.area_total_ha or 0)
    if task_area <= 0 or os_area <= 0:
        return 0
    diff_ratio = abs(task_area - os_area) / max(task_area, os_area)
    if diff_ratio <= 0.10:
        return 12
    if diff_ratio <= 0.25:
        return 8
    return 0


def _get_linked_agro_os_for_kml_route(route_id):
    return (
        OrdemServicoAgro.query
        .options(joinedload(OrdemServicoAgro.contrato), joinedload(OrdemServicoAgro.equipe))
        .filter(OrdemServicoAgro.agro_kml_route_id == route_id)
        .order_by(OrdemServicoAgro.data_aplicacao.desc().nullslast(), OrdemServicoAgro.id.desc())
        .first()
    )


def _build_linked_agro_os_payload(ordem):
    if not ordem:
        return None
    data_aplicacao = getattr(ordem, "data_aplicacao", None)
    return {
        "ordem_id": ordem.id,
        "identificador_os": ordem.identificador_os or "",
        "label": ordem.identificador_os or f"OS Agro #{ordem.id}",
        "protocolo": ordem.protocolo or "",
        "status": ordem.status or "",
        "cliente": ordem.cliente_nome or "",
        "propriedade": ordem.propriedade_nome or "",
        "data_aplicacao": data_aplicacao.strftime("%d/%m/%Y") if data_aplicacao else "",
    }


def _distinct_non_empty_values(column):
    return [
        value
        for (value,) in db.session.query(column)
        .filter(func.length(func.trim(func.coalesce(column, ""))) > 0)
        .distinct()
        .order_by(column.asc())
        .all()
    ]


def _build_top_groups(filtered_query, column, limit=5):
    return [
        {"label": label or "Nao informado", "total": total}
        for label, total in (
            filtered_query
            .with_entities(func.coalesce(column, "Nao informado"), func.count(AgroFlightRecord.id))
            .group_by(column)
            .order_by(func.count(AgroFlightRecord.id).desc(), func.coalesce(column, "Nao informado").asc())
            .limit(limit)
            .all()
        )
    ]


def _save_uploaded_agro_excel(original_filename, file_bytes):
    return _save_uploaded_agro_file(original_filename, file_bytes, "agro-flight-logs", ".xlsx", "agro_flight_logs")


def _save_uploaded_agro_kml(original_filename, file_bytes):
    return _save_uploaded_agro_file(original_filename, file_bytes, "agro-flight-routes", ".kml", "agro_flight_route")


def _save_uploaded_agro_file(original_filename, file_bytes, subfolder, default_extension, fallback_name):
    base_folder = os.path.join(get_upload_folder(), subfolder)
    os.makedirs(base_folder, exist_ok=True)
    name_root, extension = os.path.splitext(original_filename)
    safe_root = secure_filename(name_root) or fallback_name
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    short_hash = hashlib.sha256(file_bytes).hexdigest()[:10]
    stored_filename = f"{safe_root}_{stamp}_{short_hash}{extension or default_extension}"
    absolute_path = os.path.join(base_folder, stored_filename)
    with open(absolute_path, "wb") as file_handle:
        file_handle.write(file_bytes)
    return stored_filename, os.path.join(subfolder, stored_filename).replace("\\", "/")


def _route_points(route):
    try:
        return json.loads(route.points_json or "[]")
    except (TypeError, ValueError, json.JSONDecodeError):
        return []
