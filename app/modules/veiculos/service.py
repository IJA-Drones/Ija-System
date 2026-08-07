import json
import os
import re
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

from flask import current_app, make_response, send_file
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import case
from sqlalchemy.orm import joinedload, selectinload
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models import (
    Abastecimento,
    AuditoriaUsuario,
    Equipe,
    EquipePiloto,
    LimpezaVeiculo,
    LimpezaVeiculoAlertaCiencia,
    LogVeiculo,
    OrdemServico,
    Pilotos,
    Solicitacao,
    Usuario,
    Veiculos,
)
from app.shared.access import apply_prefeitura_scope, normalize_role
from app.shared.query_filters import id_search_clause
from app.shared.skybox import (
    SkyboxError,
    build_veiculo_media_remote_path,
    is_skybox_path,
    skybox_enabled,
    upload_file_to_skybox,
)


EQUIPE_OCEANO_USER_TYPE = "equipe_oceano"
UTC_TZ = ZoneInfo("UTC")
BRAZIL_TZ = ZoneInfo("America/Sao_Paulo")
MAX_KM_POR_TURNO = 500
LIMPEZA_ALERTA_OPERACIONAL_DIAS = 14
LIMPEZA_ALERTA_ADMIN_DIAS = 21
VEICULO_LOG_DELETE_AUDIT_ENDPOINT = "main.deletar_log_veiculo.snapshot"
VEICULOS_ALLOWED_TYPES = (
    "dev",
    "diretor",
    "admin",
    "visualizar",
    "operario",
    "operador",
    "uvis",
    "piloto",
    EQUIPE_OCEANO_USER_TYPE,
    "prefeitura_admin",
)


def _now_brazil():
    return datetime.now(BRAZIL_TZ).replace(tzinfo=None)


def _now_utc():
    return datetime.now(UTC_TZ).replace(tzinfo=None)
VEICULOS_LOGS_ALLOWED_TYPES = (
    "dev",
    "diretor",
    "admin",
    "visualizar",
    "operario",
    "operador",
    "prefeitura_admin",
    EQUIPE_OCEANO_USER_TYPE,
)
class VeiculoTurnoError(Exception):
    def __init__(self, message, category="warning"):
        super().__init__(message)
        self.category = category


def list_veiculos(tipo_usuario, args, user=None):
    tipo_usuario = normalize_role(tipo_usuario)
    if tipo_usuario not in VEICULOS_ALLOWED_TYPES:
        raise PermissionError

    q = (args.get("q") or "").strip()
    operacao = (args.get("operacao") or "").strip().upper()
    frota = (args.get("frota") or "").strip().upper()
    status = (args.get("status") or "").strip()

    query = Veiculos.query
    if user is not None and tipo_usuario != EQUIPE_OCEANO_USER_TYPE:
        query = apply_prefeitura_scope(query, user, Veiculos.prefeitura_id)

    if tipo_usuario == EQUIPE_OCEANO_USER_TYPE:
        equipe = _equipe_oceano_logada(user)
        if not equipe:
            raise PermissionError
        query = query.filter(_veiculo_equipe_operacional_filter(equipe, user))

    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                id_search_clause(Veiculos.id, q),
                Veiculos.modelo.ilike(like),
                Veiculos.placa.ilike(like),
                Veiculos.equipe.has(Equipe.nome_equipe.ilike(like)),
            )
        )

    if operacao:
        query = query.filter(Veiculos.operacao == operacao)

    if frota:
        query = query.filter(Veiculos.frota == frota)

    if status:
        query = query.filter(Veiculos.status == status)

    veiculos = query.order_by(Veiculos.criado_em.desc()).all()
    equipes = list_equipes_choices(user=user)

    return {
        "veiculos": veiculos,
        "is_admin": tipo_usuario in {"dev", "diretor", "admin"},
        "can_manage": tipo_usuario in {"dev", "diretor", "admin", "operario", "operador", "prefeitura_admin"},
        "equipes": equipes,
        "equipes_por_id": {item["value"]: item for item in equipes},
        "filters": {
            "q": q,
            "operacao": operacao,
            "frota": frota,
            "status": status,
            "total": len(veiculos),
        },
        "ultimos_logs": _build_ultimos_logs(veiculos),
    }


def list_equipes_choices(user=None):
    query = (
        Equipe.query
        .options(selectinload(Equipe.membros).joinedload(EquipePiloto.piloto))
        .filter(Equipe.ativa.is_(True))
    )
    if user is not None:
        query = apply_prefeitura_scope(query, user, Equipe.prefeitura_id)

    equipes = query.order_by(Equipe.nome_equipe.asc()).all()
    options = []
    for equipe in equipes:
        piloto_titular = next(
            (
                membro.piloto
                for membro in sorted(equipe.membros or [], key=lambda membro: membro.id or 0)
                if (membro.papel or "").lower() == "piloto"
                and membro.piloto
                and (membro.piloto.nome_piloto or "").strip()
            ),
            None,
        )
        options.append(
            {
                "value": str(equipe.id),
                "label": equipe.nome_equipe,
                "piloto_label": piloto_titular.nome_piloto if piloto_titular else "Sem piloto vinculado",
            }
        )
    return options


def validate_veiculo_form(form_data, *, equipes=None, existing_veiculo=None):
    errors = {}
    equipes = equipes or []

    modelo = (form_data.get("modelo") or "").strip()
    ano_raw = (form_data.get("ano_fabricacao") or "").strip()
    frota = (form_data.get("frota") or "").strip().upper()
    operacao = (form_data.get("operacao") or "").strip().upper()
    placa = (form_data.get("placa") or "").strip().upper()
    equipe_id_raw = (form_data.get("equipe_id") or "").strip()
    km_atual_raw = (form_data.get("km_atual") or "").strip()
    km_prox_raw = (form_data.get("km_prox_revisao") or "").strip()
    status = (form_data.get("status") or "Ativo").strip()
    revisao_marcada_raw = (form_data.get("revisao_marcada_em") or "").strip()
    revisao_obs = (form_data.get("revisao_obs") or "").strip()

    form = {
        "modelo": modelo,
        "ano_fabricacao": ano_raw,
        "frota": frota,
        "operacao": operacao,
        "placa": placa,
        "equipe_id": equipe_id_raw,
        "km_atual": km_atual_raw,
        "km_prox_revisao": km_prox_raw,
        "status": status,
        "revisao_marcada_em": revisao_marcada_raw,
        "revisao_obs": revisao_obs,
    }

    if not modelo:
        errors["modelo"] = "Informe o modelo."
    if not ano_raw:
        errors["ano_fabricacao"] = "Informe o ano."
    if frota not in ("PROPRIA", "ALUGADA"):
        errors["frota"] = "Selecione PROPRIA ou ALUGADA."
    if not operacao:
        errors["operacao"] = (
            "Informe a operação (ex: PMSP / AGRO)."
            if existing_veiculo is None
            else "Informe a operação."
        )
    if not placa:
        errors["placa"] = "Informe a placa."

    valid_equipe_ids = {item["value"] for item in equipes}
    equipe_id = None
    if equipe_id_raw:
        if equipe_id_raw not in valid_equipe_ids:
            errors["equipe_id"] = "Selecione uma equipe válida."
        else:
            equipe_id = int(equipe_id_raw)

    ano_fabricacao = None
    if ano_raw:
        try:
            ano_fabricacao = int(ano_raw)
            if ano_fabricacao < 1900 or ano_fabricacao > 2100:
                errors["ano_fabricacao"] = "Ano inválido."
        except ValueError:
            errors["ano_fabricacao"] = "Ano inválido."

    km_atual = (existing_veiculo.km_atual if existing_veiculo is not None else 0) or 0
    if km_atual_raw:
        try:
            km_atual = _parse_km_form(km_atual_raw, "KM atual")
            if km_atual < 0:
                errors["km_atual"] = (
                    "KM atual não pode ser negativo."
                    if existing_veiculo is None
                    else "KM atual inválido."
                )
        except ValueError:
            errors["km_atual"] = (
                "KM atual inválido."
                if existing_veiculo is not None
                else "KM atual inválido."
            )

    km_prox_revisao = None
    if km_prox_raw:
        try:
            km_prox_revisao = float(km_prox_raw.replace(",", "."))
            if existing_veiculo is None and km_prox_revisao < 0:
                errors["km_prox_revisao"] = "Próx revisão inválida."
        except ValueError:
            errors["km_prox_revisao"] = "Próx revisão inválida."

    revisao_marcada_em = None
    if revisao_marcada_raw:
        try:
            revisao_marcada_em = datetime.strptime(revisao_marcada_raw, "%Y-%m-%dT%H:%M")
        except ValueError:
            errors["revisao_marcada_em"] = "Data/hora inválida."

    if placa and not errors.get("placa"):
        query = Veiculos.query.filter_by(placa=placa)
        if existing_veiculo is not None:
            query = query.filter(Veiculos.id != existing_veiculo.id)
        if query.first():
            errors["placa"] = "Já existe um veículo com essa placa."

    cleaned = {
        "modelo": modelo,
        "ano_fabricacao": ano_fabricacao,
        "frota": frota,
        "operacao": operacao,
        "placa": placa,
        "responsavel": None,
        "equipe_id": equipe_id,
        "km_atual": km_atual,
        "km_prox_revisao": km_prox_revisao,
        "status": status,
        "revisao_marcada_em": revisao_marcada_em,
        "revisao_obs": revisao_obs or None,
    }

    return form, cleaned, errors


def create_veiculo(cleaned, *, prefeitura_id=None):
    prefeitura_id = _resolve_prefeitura_id_veiculo(cleaned["equipe_id"], prefeitura_id)
    novo = Veiculos(
        tipo_equipamento="veiculos",
        status=cleaned["status"],
        modelo=cleaned["modelo"],
        ano_fabricacao=cleaned["ano_fabricacao"],
        renomacao=cleaned["placa"],
        categoria=None,
        numero_serie=None,
        ultima_manutencao=None,
        frota=cleaned["frota"],
        operacao=cleaned["operacao"],
        placa=cleaned["placa"],
        responsavel=cleaned["responsavel"],
        km_atual=cleaned["km_atual"],
        km_prox_revisao=cleaned["km_prox_revisao"],
        revisao_marcada_em=cleaned["revisao_marcada_em"],
        revisao_obs=cleaned["revisao_obs"],
        equipe_id=cleaned["equipe_id"],
        prefeitura_id=prefeitura_id,
    )
    db.session.add(novo)
    db.session.commit()
    return novo


def update_veiculo(veiculo, cleaned):
    veiculo.modelo = cleaned["modelo"]
    veiculo.ano_fabricacao = cleaned["ano_fabricacao"]
    veiculo.frota = cleaned["frota"]
    veiculo.operacao = cleaned["operacao"]
    veiculo.placa = cleaned["placa"]
    veiculo.responsavel = cleaned["responsavel"]
    veiculo.equipe_id = cleaned["equipe_id"]
    veiculo.prefeitura_id = _resolve_prefeitura_id_veiculo(cleaned["equipe_id"], veiculo.prefeitura_id)
    veiculo.km_atual = cleaned["km_atual"]
    veiculo.km_prox_revisao = cleaned["km_prox_revisao"]
    veiculo.status = cleaned["status"]
    veiculo.revisao_marcada_em = cleaned["revisao_marcada_em"]
    veiculo.revisao_obs = cleaned["revisao_obs"]
    veiculo.renomacao = cleaned["placa"]
    db.session.commit()
    return veiculo


def update_veiculos_equipes(user, form_data):
    tipo_usuario = normalize_role(getattr(user, "tipo_usuario", None))
    if tipo_usuario not in {"dev", "diretor", "admin", "operario", "operador", "prefeitura_admin"}:
        raise PermissionError

    getlist = getattr(form_data, "getlist", None)
    raw_ids = getlist("veiculo_ids") if getlist else form_data.get("veiculo_ids", [])
    if isinstance(raw_ids, str):
        raw_ids = [raw_ids]

    veiculo_ids = []
    for raw_id in raw_ids:
        try:
            veiculo_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue

    if not veiculo_ids:
        raise VeiculoTurnoError("Nenhum veiculo foi enviado para atualizacao.", "warning")

    valid_equipe_ids = {item["value"] for item in list_equipes_choices(user=user)}
    updates = {}
    for veiculo_id in veiculo_ids:
        raw_equipe_id = (form_data.get(f"equipe_id_{veiculo_id}") or "").strip()
        if raw_equipe_id and raw_equipe_id not in valid_equipe_ids:
            raise VeiculoTurnoError("Uma das equipes selecionadas nao e valida para o seu acesso.", "danger")
        updates[veiculo_id] = int(raw_equipe_id) if raw_equipe_id else None

    query = apply_prefeitura_scope(Veiculos.query, user, Veiculos.prefeitura_id)
    veiculos = query.filter(Veiculos.id.in_(veiculo_ids)).all()
    veiculos_por_id = {veiculo.id: veiculo for veiculo in veiculos}
    if len(veiculos_por_id) != len(set(veiculo_ids)):
        raise PermissionError

    alterados = 0
    for veiculo_id, equipe_id in updates.items():
        veiculo = veiculos_por_id[veiculo_id]
        if veiculo.equipe_id != equipe_id or veiculo.responsavel:
            alterados += 1
        veiculo.equipe_id = equipe_id
        veiculo.responsavel = None
        veiculo.prefeitura_id = _resolve_prefeitura_id_veiculo(equipe_id, veiculo.prefeitura_id)

    db.session.commit()
    if alterados == 1:
        return "Equipe responsavel atualizada em 1 veiculo."
    return f"Equipe responsavel atualizada em {alterados} veiculos."


def delete_veiculo(veiculo):
    db.session.delete(veiculo)
    db.session.commit()


def build_veiculo_form(veiculo):
    return {
        "modelo": veiculo.modelo or "",
        "ano_fabricacao": str(veiculo.ano_fabricacao or ""),
        "frota": veiculo.frota or "",
        "operacao": veiculo.operacao or "",
        "placa": veiculo.placa or "",
        "equipe_id": str(veiculo.equipe_id or ""),
        "km_atual": str(veiculo.km_atual or ""),
        "km_prox_revisao": str(veiculo.km_prox_revisao or "") if veiculo.km_prox_revisao is not None else "",
        "status": veiculo.status or "Ativo",
        "revisao_marcada_em": (
            veiculo.revisao_marcada_em.strftime("%Y-%m-%dT%H:%M")
            if veiculo.revisao_marcada_em else ""
        ),
        "revisao_obs": veiculo.revisao_obs or "",
    }


def build_piloto_veiculos_context(user):
    if getattr(user, "tipo_usuario", None) == EQUIPE_OCEANO_USER_TYPE:
        equipe = _equipe_oceano_logada(user)
        if not equipe:
            return {
                "piloto_vinculado": False,
                "veiculos": [],
                "turnos_abertos": {},
                "km_inicial_referencias": {},
                "agora_brasilia": _now_brazil(),
            }

        veiculos = (
            Veiculos.query
            .filter(_veiculo_equipe_operacional_filter(equipe, user))
            .order_by(Veiculos.operacao.asc(), Veiculos.modelo.asc())
            .all()
        )
        return {
            "piloto_vinculado": True,
            "veiculos": veiculos,
            "turnos_abertos": _build_turnos_abertos_veiculos(veiculos, user),
            "km_inicial_referencias": _build_km_inicial_referencias(veiculos),
            "agora_brasilia": _now_brazil(),
        }

    if not getattr(user, "piloto_id", None):
        return {
            "piloto_vinculado": False,
            "veiculos": [],
            "turnos_abertos": {},
            "km_inicial_referencias": {},
            "agora_brasilia": _now_brazil(),
        }

    equipe_ids = _equipe_ids_do_piloto(user)
    if not equipe_ids:
        return {
            "piloto_vinculado": False,
            "veiculos": [],
            "turnos_abertos": {},
            "km_inicial_referencias": {},
            "agora_brasilia": _now_brazil(),
        }
    veiculos = (
        Veiculos.query
        .filter(Veiculos.equipe_id.in_(equipe_ids))
        .order_by(Veiculos.operacao.asc(), Veiculos.modelo.asc())
        .all()
    )

    return {
        "piloto_vinculado": True,
        "veiculos": veiculos,
        "turnos_abertos": _build_turnos_abertos_veiculos(veiculos, user),
        "km_inicial_referencias": _build_km_inicial_referencias(veiculos),
        "agora_brasilia": _now_brazil(),
    }


def can_access_limpeza_alertas_operacionais(user):
    return (
        getattr(user, "tipo_usuario", None) in {"piloto", EQUIPE_OCEANO_USER_TYPE}
        and bool(getattr(user, "trabalha_oceano_azul", False))
    )


def can_access_limpeza_alertas_admin(user):
    tipo_usuario = normalize_role(getattr(user, "tipo_usuario", None))
    return (
        tipo_usuario in {"dev", "diretor", "admin", "operario", "operador", "visualizar", "prefeitura_admin"}
        and bool(getattr(user, "trabalha_oceano_azul", False))
    )


def count_limpeza_alertas_operacionais(user):
    if not can_access_limpeza_alertas_operacionais(user):
        return 0
    return sum(1 for alerta in build_limpeza_alertas_operacionais_context(user)["alertas"] if not alerta["ciencia"])


def count_limpeza_alertas_admin(user):
    if not can_access_limpeza_alertas_admin(user):
        return 0
    return len(build_limpeza_alertas_admin_context(user)["alertas"])


def build_limpeza_alertas_operacionais_context(user):
    if not can_access_limpeza_alertas_operacionais(user):
        raise PermissionError

    veiculos = _veiculos_operacionais_do_usuario(user)
    alertas = _build_alertas_limpeza_veiculos(
        veiculos,
        prazo_dias=LIMPEZA_ALERTA_OPERACIONAL_DIAS,
        usuario=user,
    )

    return {
        "alertas": alertas,
        "total_pendentes": sum(1 for alerta in alertas if not alerta["ciencia"]),
        "prazo_dias": LIMPEZA_ALERTA_OPERACIONAL_DIAS,
    }


def confirmar_alerta_limpeza_operacional(user, veiculo_id):
    if not can_access_limpeza_alertas_operacionais(user):
        raise PermissionError

    alertas = _build_alertas_limpeza_veiculos(
        _veiculos_operacionais_do_usuario(user),
        prazo_dias=LIMPEZA_ALERTA_OPERACIONAL_DIAS,
        usuario=user,
    )
    alerta = next((item for item in alertas if item["veiculo"].id == veiculo_id), None)
    if not alerta:
        raise VeiculoTurnoError("Este alerta nao esta mais ativo para o seu usuario.", "warning")

    agora = _now_brazil()
    ciencia = alerta["ciencia"]
    if ciencia is None:
        ciencia = LimpezaVeiculoAlertaCiencia(
            veiculo_id=veiculo_id,
            usuario_id=getattr(user, "id", None),
            piloto_id=getattr(user, "piloto_id", None),
            equipe_id=_actor_equipe_id_from_user(user),
            referencia_limpeza_em=alerta["referencia_em"],
            prazo_dias=LIMPEZA_ALERTA_OPERACIONAL_DIAS,
            criado_em=agora,
            atualizado_em=agora,
        )
        db.session.add(ciencia)

    ciencia.reconhecido_em = agora
    ciencia.atualizado_em = agora
    db.session.commit()
    return "Ciencia do alerta de limpeza registrada."


def build_limpeza_alertas_admin_context(user):
    if not can_access_limpeza_alertas_admin(user):
        raise PermissionError

    query = Veiculos.query.options(joinedload(Veiculos.equipe)).filter(
        db.func.lower(db.func.coalesce(Veiculos.status, "")) == "ativo",
        db.func.upper(db.func.coalesce(Veiculos.operacao, "")) != "AGRO",
    )
    query = apply_prefeitura_scope(query, user, Veiculos.prefeitura_id)
    veiculos = query.order_by(Veiculos.operacao.asc(), Veiculos.placa.asc(), Veiculos.id.asc()).all()

    alertas = _build_alertas_limpeza_veiculos(
        veiculos,
        prazo_dias=LIMPEZA_ALERTA_ADMIN_DIAS,
    )
    _attach_ciencias_operacionais(alertas)

    total_atores = sum(len(alerta["atores"]) for alerta in alertas)
    total_cientes = sum(
        1
        for alerta in alertas
        for ator in alerta["atores"]
        if ator.get("ciencia") is not None
    )

    return {
        "alertas": alertas,
        "total_alertas": len(alertas),
        "total_atores": total_atores,
        "total_cientes": total_cientes,
        "prazo_dias": LIMPEZA_ALERTA_ADMIN_DIAS,
        "prazo_operacional_dias": LIMPEZA_ALERTA_OPERACIONAL_DIAS,
    }


def _build_turnos_abertos_veiculos(veiculos, user):
    turnos_abertos = {}
    veiculo_ids = [veiculo.id for veiculo in veiculos]

    if veiculo_ids:
        query = (
            LogVeiculo.query
            .options(
                selectinload(LogVeiculo.abastecimentos_detalhados),
                selectinload(LogVeiculo.limpezas_detalhadas),
            )
            .filter(LogVeiculo.veiculo_id.in_(veiculo_ids), LogVeiculo.km_final.is_(None))
        )
        query = _apply_log_actor_scope(query, user)
        logs_abertos = query.order_by(LogVeiculo.veiculo_id.asc(), LogVeiculo.data_registro.desc()).all()
        for log in logs_abertos:
            if log.veiculo_id not in turnos_abertos:
                turnos_abertos[log.veiculo_id] = log

    return turnos_abertos


def _build_km_inicial_referencias(veiculos):
    referencias = {}
    if not veiculos:
        return referencias

    veiculo_ids = [veiculo.id for veiculo in veiculos]
    logs_fechados = (
        LogVeiculo.query
        .filter(LogVeiculo.veiculo_id.in_(veiculo_ids), LogVeiculo.km_final.isnot(None))
        .order_by(LogVeiculo.veiculo_id.asc(), LogVeiculo.data_registro.desc(), LogVeiculo.id.desc())
        .all()
    )
    for log in logs_fechados:
        if log.veiculo_id not in referencias:
            referencias[log.veiculo_id] = {
                "km": log.km_final,
                "origem": "ultimo_fechamento",
                "data": log.data_registro,
                "log_id": log.id,
            }

    for veiculo in veiculos:
        referencias.setdefault(
            veiculo.id,
            {
                "km": veiculo.km_atual or 0,
                "origem": "km_atual",
                "data": None,
                "log_id": None,
            },
        )

    return referencias


def _buscar_ultimo_fechamento_veiculo(veiculo_id):
    return (
        LogVeiculo.query
        .filter(LogVeiculo.veiculo_id == veiculo_id, LogVeiculo.km_final.isnot(None))
        .order_by(LogVeiculo.data_registro.desc(), LogVeiculo.id.desc())
        .first()
    )


def iniciar_turno_piloto(user, veiculo_id, form_data, files_data, root_path):
    veiculo = _veiculo_do_operacional_logado(veiculo_id, user=user)
    piloto_id = getattr(user, "piloto_id", None) if getattr(user, "tipo_usuario", None) != EQUIPE_OCEANO_USER_TYPE else None
    equipe_id = veiculo.equipe_id if getattr(user, "tipo_usuario", None) != EQUIPE_OCEANO_USER_TYPE else _parse_equipe_oceano_id(user)
    ultimo_fechamento = _buscar_ultimo_fechamento_veiculo(veiculo.id)

    try:
        km_inicial_enviado = _parse_km_form(form_data.get("km_inicial"), "Kilometragem inicial")
    except ValueError as exc:
        raise VeiculoTurnoError("Kilometragem inicial invalida.", "warning") from exc

    assinatura_b64 = form_data.get("assinatura_b64")
    foto_painel = files_data.get("foto_painel")
    km_inicial = ultimo_fechamento.km_final if ultimo_fechamento is not None else (veiculo.km_atual or 0)

    if km_inicial_enviado is not None and abs(km_inicial_enviado - km_inicial) > 0.0001:
        origem_km_inicial = (
            "o ultimo KM final registrado para este veiculo"
            if ultimo_fechamento is not None
            else "o KM atual cadastrado para este veiculo"
        )
        raise VeiculoTurnoError(
            f"KM inicial travado em {km_inicial:.0f} km, conforme {origem_km_inicial}.",
            "danger",
        )

    if km_inicial is None or not assinatura_b64 or not foto_painel or not foto_painel.filename:
        raise VeiculoTurnoError(
            "Foto do painel e assinatura sao obrigatorias.",
            "warning",
        )

    km_atual_veiculo = veiculo.km_atual or 0
    if ultimo_fechamento is None and km_atual_veiculo > 0 and abs(km_inicial - km_atual_veiculo) > 0.0001:
        raise VeiculoTurnoError("KM inicial deve ser igual ao KM atual do veiculo.", "danger")

    turno_aberto = _buscar_turno_aberto_usuario(veiculo.id, user)
    if turno_aberto:
        raise VeiculoTurnoError(
            "Ja existe um turno aberto para este veiculo. Finalize-o antes de iniciar outro.",
            "warning",
        )

    novo_log = LogVeiculo(
        veiculo_id=veiculo.id,
        piloto_id=piloto_id,
        equipe_id=equipe_id,
        km_inicial=km_inicial,
        km_final=None,
        check_diario=True,
        assinatura_piloto=assinatura_b64,
        data_registro=_now_brazil(),
    )
    novo_log.foto_painel_path = _salvar_upload_veiculo(
        foto_painel,
        root_path,
        "paineis",
        "painel_inicial",
        veiculo.placa,
        copiar_skybox=True,
    )

    db.session.add(novo_log)
    veiculo.km_atual = km_inicial

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        error_text = str(getattr(exc, "orig", exc)).lower()
        if "km_final" in error_text:
            raise VeiculoTurnoError(
                "Erro de banco: o campo km_final ainda nao aceita vazio. Rode a migracao para permitir NULL.",
                "danger",
            ) from exc
        raise

    return f"Turno de {veiculo.modelo} iniciado com sucesso!"


def registrar_abastecimento_turno_piloto(user, veiculo_id, form_data, files_data, root_path):
    veiculo = _veiculo_do_operacional_logado(veiculo_id, user=user)
    log = _buscar_turno_aberto_usuario(veiculo.id, user, incluir_abastecimentos=True)

    if not log:
        raise VeiculoTurnoError(
            "Nenhum turno aberto encontrado para registrar abastecimento.",
            "warning",
        )

    try:
        km_registro = _parse_km_form(form_data.get("km_abastecimento"), "KM do abastecimento")
        litros = _parse_decimal_form(form_data.get("litros"))
        valor_total = _parse_decimal_form(form_data.get("valor_abastecimento"))
    except ValueError as exc:
        raise VeiculoTurnoError("Os dados do abastecimento estao invalidos.", "warning") from exc

    tipo_abastecimento = (form_data.get("tipo_abastecimento") or "").strip()
    foto_nf = files_data.get("foto_nf")
    foto_painel = files_data.get("foto_painel_abastecimento")

    if (
        km_registro is None
        or litros is None
        or valor_total is None
        or not tipo_abastecimento
        or not foto_nf
        or not foto_nf.filename
        or not foto_painel
        or not foto_painel.filename
    ):
        raise VeiculoTurnoError(
            "KM, tipo, litros, valor total, foto do painel e foto da nota sao obrigatorios no abastecimento.",
            "warning",
        )

    if len(tipo_abastecimento) > 100:
        raise VeiculoTurnoError(
            "O tipo de abastecimento deve ter no maximo 100 caracteres.",
            "warning",
        )

    _validar_limite_km_turno(log.km_inicial or 0, km_registro, "KM do abastecimento")

    novo_abastecimento = Abastecimento(
        log_veiculo_id=log.id,
        data_hora=_now_brazil(),
        km_registro=km_registro,
        tipo_abastecimento=tipo_abastecimento,
        litros=litros,
        valor_total=valor_total,
        foto_nf_path=_salvar_upload_veiculo(
            foto_nf,
            root_path,
            "notas",
            "nf",
            veiculo.placa,
            copiar_skybox=True,
        ),
        foto_painel_path=_salvar_upload_veiculo(
            foto_painel,
            root_path,
            "paineis",
            "painel_abastecimento",
            veiculo.placa,
            copiar_skybox=True,
        ),
    )

    db.session.add(novo_abastecimento)
    db.session.commit()
    return "Abastecimento registrado com sucesso!"


def registrar_limpeza_turno_piloto(user, veiculo_id, form_data):
    veiculo = _veiculo_do_operacional_logado(veiculo_id, user=user)
    log = _buscar_turno_aberto_usuario(veiculo.id, user, incluir_abastecimentos=True)

    if not log:
        raise VeiculoTurnoError(
            "Nenhum turno aberto encontrado para registrar limpeza.",
            "warning",
        )

    limpeza_realizada = _bool_from_form(form_data.get("limpeza_realizada"), default=True)
    tipo_limpeza = (form_data.get("tipo_limpeza") or "").strip().lower()
    observacao = (form_data.get("observacao_limpeza") or "").strip() or None
    data_hora_limpeza = _parse_datetime_local_form(form_data.get("data_hora_limpeza"))

    try:
        valor_total = _parse_decimal_form(form_data.get("valor_limpeza"))
    except ValueError as exc:
        raise VeiculoTurnoError("Valor da limpeza invalido.", "warning") from exc

    tipos_validos = {"completa", "ducha"}
    if tipo_limpeza not in tipos_validos:
        raise VeiculoTurnoError("Selecione se a limpeza foi completa ou apenas ducha.", "warning")

    if data_hora_limpeza is None:
        raise VeiculoTurnoError("Informe a data e hora da limpeza.", "warning")

    if valor_total is None:
        raise VeiculoTurnoError("Informe o valor da limpeza.", "warning")

    if valor_total < 0:
        raise VeiculoTurnoError("Valor da limpeza nao pode ser negativo.", "warning")

    nova_limpeza = LimpezaVeiculo(
        log_veiculo_id=log.id,
        veiculo_id=veiculo.id,
        piloto_id=log.piloto_id,
        equipe_id=log.equipe_id,
        data_registro=_now_brazil(),
        data_hora=data_hora_limpeza,
        limpeza_realizada=limpeza_realizada,
        tipo_limpeza=tipo_limpeza,
        valor_total=valor_total,
        observacao=observacao,
    )

    db.session.add(nova_limpeza)
    db.session.commit()
    return "Limpeza registrada com sucesso!"


def encerrar_turno_piloto(user, veiculo_id, form_data, files_data=None, root_path=None):
    veiculo = _veiculo_do_operacional_logado(veiculo_id, user=user)

    try:
        km_final = _parse_km_form(form_data.get("km_final"), "Kilometragem final")
    except ValueError as exc:
        raise VeiculoTurnoError("Kilometragem final invalida.", "warning") from exc

    qtd_fazendas_enderecos = _parse_optional_int(form_data.get("qtd_fazendas_enderecos"))
    observacao = (form_data.get("observacao") or "").strip() or None
    files_data = files_data or {}
    foto_painel_final = files_data.get("foto_painel_final") if hasattr(files_data, "get") else None

    if km_final is None:
        raise VeiculoTurnoError(
            "Informe a kilometragem final para encerrar o turno.",
            "warning",
        )

    if not foto_painel_final or not foto_painel_final.filename:
        raise VeiculoTurnoError(
            "A foto do painel no fechamento do turno e obrigatoria.",
            "warning",
        )

    log = _buscar_turno_aberto_usuario(veiculo_id, user, incluir_abastecimentos=True)
    if not log:
        raise VeiculoTurnoError("Nenhum turno aberto encontrado.", "warning")

    km_inicial_turno = log.km_inicial or 0
    if km_final < km_inicial_turno:
        raise VeiculoTurnoError(
            f"KM final nao pode ser menor que o KM inicial do turno ({km_inicial_turno:.0f}).",
            "danger",
        )
    maior_km_abastecimento = max(
        [
            abastecimento.km_registro
            for abastecimento in (log.abastecimentos_detalhados or [])
            if abastecimento.km_registro is not None
        ],
        default=None,
    )
    if maior_km_abastecimento is not None and km_final < maior_km_abastecimento:
        raise VeiculoTurnoError(
            f"KM final nao pode ser menor que o KM do abastecimento ({maior_km_abastecimento:.0f}).",
            "danger",
        )
    _validar_limite_km_turno(km_inicial_turno, km_final, "KM final")

    log.qtd_fazendas_enderecos = qtd_fazendas_enderecos
    log.km_final = km_final
    log.observacao = observacao
    log.foto_painel_final_path = _salvar_upload_veiculo(
        foto_painel_final,
        root_path,
        "paineis",
        "painel_fechamento",
        veiculo.placa,
        copiar_skybox=True,
    )
    if log.veiculo:
        log.veiculo.km_atual = km_final

    db.session.commit()
    return "Turno encerrado com sucesso!"


def _piloto_nome_logado(user, *, strict=False):
    nome_piloto = (getattr(user, "nome_uvis", None) or "").strip()
    if strict and (not nome_piloto or not getattr(user, "piloto_id", None)):
        raise PermissionError
    return nome_piloto


def _parse_equipe_oceano_id(user):
    raw = (getattr(user, "codigo_setor", None) or "").strip()
    if not raw:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _equipe_oceano_logada(user):
    if getattr(user, "tipo_usuario", None) != EQUIPE_OCEANO_USER_TYPE:
        return None

    equipe_id = _parse_equipe_oceano_id(user)
    if not equipe_id:
        return None

    query = Equipe.query.filter(Equipe.id == equipe_id, Equipe.ativa.is_(True))
    query = apply_prefeitura_scope(query, user, Equipe.prefeitura_id)
    return query.first()


def _resolve_prefeitura_id_veiculo(equipe_id, fallback_prefeitura_id=None):
    if not equipe_id:
        return fallback_prefeitura_id

    equipe = Equipe.query.filter(Equipe.id == equipe_id).first()
    if equipe and equipe.prefeitura_id is not None:
        return equipe.prefeitura_id
    return fallback_prefeitura_id


def _veiculo_equipe_operacional_filter(equipe, user):
    return Veiculos.equipe_id == equipe.id


def _equipe_ids_do_piloto(user):
    piloto_id = getattr(user, "piloto_id", None)
    if not piloto_id:
        return []

    query = (
        db.session.query(EquipePiloto.equipe_id)
        .join(Equipe, Equipe.id == EquipePiloto.equipe_id)
        .filter(
            EquipePiloto.piloto_id == piloto_id,
            Equipe.ativa.is_(True),
        )
    )
    prefeitura_id = getattr(user, "prefeitura_id", None)
    if prefeitura_id is not None:
        query = query.filter(Equipe.prefeitura_id == prefeitura_id)

    rows = query.all()
    return [row.equipe_id for row in rows if row.equipe_id]


def _veiculos_operacionais_do_usuario(user):
    if getattr(user, "tipo_usuario", None) == EQUIPE_OCEANO_USER_TYPE:
        equipe = _equipe_oceano_logada(user)
        if not equipe:
            return []
        return (
            Veiculos.query
            .options(joinedload(Veiculos.equipe))
            .filter(
                db.func.lower(db.func.coalesce(Veiculos.status, "")) == "ativo",
                _veiculo_equipe_operacional_filter(equipe, user),
            )
            .order_by(Veiculos.operacao.asc(), Veiculos.modelo.asc(), Veiculos.id.asc())
            .all()
        )

    if not getattr(user, "piloto_id", None):
        return []

    equipe_ids = _equipe_ids_do_piloto(user)
    if not equipe_ids:
        return []

    return (
        Veiculos.query
        .options(joinedload(Veiculos.equipe))
        .filter(
            db.func.lower(db.func.coalesce(Veiculos.status, "")) == "ativo",
            Veiculos.equipe_id.in_(equipe_ids),
        )
        .order_by(Veiculos.operacao.asc(), Veiculos.modelo.asc(), Veiculos.id.asc())
        .all()
    )


def _actor_equipe_id_from_user(user):
    if getattr(user, "tipo_usuario", None) == EQUIPE_OCEANO_USER_TYPE:
        return _parse_equipe_oceano_id(user)
    return None


def _build_alertas_limpeza_veiculos(veiculos, *, prazo_dias, usuario=None):
    veiculos = list(veiculos or [])
    if not veiculos:
        return []

    hoje = _now_brazil()
    veiculo_ids = [veiculo.id for veiculo in veiculos]
    referencias = _latest_limpeza_realizada_por_veiculo(veiculo_ids)
    ciencias = {}

    if usuario is not None:
        referencias_validas = [
            _referencia_limpeza_alerta(veiculo, referencias.get(veiculo.id))
            for veiculo in veiculos
        ]
        referencias_validas = [ref for ref in referencias_validas if ref is not None]
        if referencias_validas:
            rows = (
                LimpezaVeiculoAlertaCiencia.query
                .filter(
                    LimpezaVeiculoAlertaCiencia.usuario_id == getattr(usuario, "id", None),
                    LimpezaVeiculoAlertaCiencia.veiculo_id.in_(veiculo_ids),
                    LimpezaVeiculoAlertaCiencia.prazo_dias == prazo_dias,
                    LimpezaVeiculoAlertaCiencia.referencia_limpeza_em.in_(referencias_validas),
                )
                .all()
            )
            ciencias = {
                (row.veiculo_id, row.referencia_limpeza_em): row
                for row in rows
            }

    alertas = []
    for veiculo in veiculos:
        ultima_limpeza = referencias.get(veiculo.id)
        referencia_em = _referencia_limpeza_alerta(veiculo, ultima_limpeza)
        if referencia_em is None:
            continue

        dias_desde = max((hoje.date() - referencia_em.date()).days, 0)
        if dias_desde < prazo_dias:
            continue

        alertas.append(
            {
                "veiculo": veiculo,
                "ultima_limpeza_em": ultima_limpeza,
                "referencia_em": referencia_em,
                "referencia_tipo": "limpeza" if ultima_limpeza is not None else "cadastro",
                "dias_desde": dias_desde,
                "vencido_em": referencia_em + timedelta(days=prazo_dias),
                "prazo_dias": prazo_dias,
                "ciencia": ciencias.get((veiculo.id, referencia_em)),
                "atores": [],
            }
        )

    return sorted(alertas, key=lambda item: (-item["dias_desde"], item["veiculo"].placa or ""))


def _latest_limpeza_realizada_por_veiculo(veiculo_ids):
    if not veiculo_ids:
        return {}

    rows = (
        db.session.query(
            LimpezaVeiculo.veiculo_id,
            db.func.max(LimpezaVeiculo.data_hora).label("ultima_limpeza_em"),
        )
        .filter(
            LimpezaVeiculo.veiculo_id.in_(veiculo_ids),
            LimpezaVeiculo.limpeza_realizada.is_(True),
        )
        .group_by(LimpezaVeiculo.veiculo_id)
        .all()
    )
    return {row.veiculo_id: row.ultima_limpeza_em for row in rows}


def _referencia_limpeza_alerta(veiculo, ultima_limpeza):
    return ultima_limpeza or getattr(veiculo, "criado_em", None)


def _attach_ciencias_operacionais(alertas):
    if not alertas:
        return

    veiculo_ids = [alerta["veiculo"].id for alerta in alertas]
    atores_por_veiculo = _build_limpeza_alerta_atores_por_veiculo(veiculo_ids)
    usuario_ids = {
        ator["usuario"].id
        for atores in atores_por_veiculo.values()
        for ator in atores
        if ator.get("usuario") is not None
    }
    refs = [alerta["referencia_em"] for alerta in alertas if alerta.get("referencia_em") is not None]

    ciencias = {}
    if usuario_ids and refs:
        rows = (
            LimpezaVeiculoAlertaCiencia.query
            .options(joinedload(LimpezaVeiculoAlertaCiencia.usuario))
            .filter(
                LimpezaVeiculoAlertaCiencia.veiculo_id.in_(veiculo_ids),
                LimpezaVeiculoAlertaCiencia.usuario_id.in_(usuario_ids),
                LimpezaVeiculoAlertaCiencia.prazo_dias == LIMPEZA_ALERTA_OPERACIONAL_DIAS,
                LimpezaVeiculoAlertaCiencia.referencia_limpeza_em.in_(refs),
            )
            .all()
        )
        ciencias = {
            (row.veiculo_id, row.usuario_id, row.referencia_limpeza_em): row
            for row in rows
        }

    for alerta in alertas:
        veiculo = alerta["veiculo"]
        atores = []
        for ator in atores_por_veiculo.get(veiculo.id, []):
            usuario = ator.get("usuario")
            ciencia = None
            if usuario is not None:
                ciencia = ciencias.get((veiculo.id, usuario.id, alerta["referencia_em"]))
            item = dict(ator)
            item["ciencia"] = ciencia
            atores.append(item)
        alerta["atores"] = atores


def _build_limpeza_alerta_atores_por_veiculo(veiculo_ids):
    veiculos = (
        Veiculos.query
        .options(joinedload(Veiculos.equipe))
        .filter(Veiculos.id.in_(veiculo_ids))
        .all()
    )
    por_veiculo = {veiculo_id: [] for veiculo_id in veiculo_ids}
    piloto_ids_por_veiculo = defaultdict(set)
    equipe_ids = {veiculo.equipe_id for veiculo in veiculos if veiculo.equipe_id}

    if equipe_ids:
        membros = (
            EquipePiloto.query
            .filter(EquipePiloto.equipe_id.in_(equipe_ids))
            .all()
        )
        for membro in membros:
            for veiculo in veiculos:
                if veiculo.equipe_id == membro.equipe_id and membro.piloto_id:
                    piloto_ids_por_veiculo[veiculo.id].add(membro.piloto_id)

    all_piloto_ids = {
        piloto_id
        for ids in piloto_ids_por_veiculo.values()
        for piloto_id in ids
        if piloto_id
    }
    usuarios_por_piloto = defaultdict(list)
    if all_piloto_ids:
        usuarios_pilotos = (
            Usuario.query
            .filter(
                Usuario.piloto_id.in_(all_piloto_ids),
                Usuario.trabalha_oceano_azul.is_(True),
            )
            .all()
        )
        for usuario in usuarios_pilotos:
            usuarios_por_piloto[usuario.piloto_id].append(usuario)

    usuarios_equipe = []
    if equipe_ids:
        usuarios_equipe = (
            Usuario.query
            .filter(
                Usuario.tipo_usuario == EQUIPE_OCEANO_USER_TYPE,
                Usuario.trabalha_oceano_azul.is_(True),
            )
            .all()
        )
    usuarios_equipe_por_id = defaultdict(list)
    for usuario in usuarios_equipe:
        equipe_id = None
        try:
            equipe_id = int((usuario.codigo_setor or "").strip())
        except (TypeError, ValueError):
            equipe_id = None
        if equipe_id in equipe_ids:
            usuarios_equipe_por_id[equipe_id].append(usuario)

    for veiculo in veiculos:
        seen = set()
        atores = []

        if veiculo.equipe_id:
            for usuario in usuarios_equipe_por_id.get(veiculo.equipe_id, []):
                if usuario.id in seen:
                    continue
                seen.add(usuario.id)
                atores.append(
                    {
                        "usuario": usuario,
                        "tipo": "Equipe",
                        "nome": getattr(veiculo.equipe, "nome_equipe", None) or usuario.nome_uvis or usuario.login,
                    }
                )

        for piloto_id in sorted(piloto_ids_por_veiculo.get(veiculo.id, set())):
            for usuario in usuarios_por_piloto.get(piloto_id, []):
                if usuario.id in seen:
                    continue
                seen.add(usuario.id)
                atores.append(
                    {
                        "usuario": usuario,
                        "tipo": "Piloto",
                        "nome": usuario.nome_uvis or usuario.login,
                    }
                )

        por_veiculo[veiculo.id] = atores

    return por_veiculo


def _parse_decimal_form(raw_value):
    raw_value = (raw_value or "").strip()
    if not raw_value:
        return None

    if "," in raw_value and "." in raw_value:
        raw_value = raw_value.replace(".", "").replace(",", ".")
    else:
        raw_value = raw_value.replace(",", ".")

    return float(raw_value)


def _parse_datetime_local_form(raw_value):
    raw_value = (raw_value or "").strip()
    if not raw_value:
        return None

    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(raw_value, fmt)
        except ValueError:
            continue
    return None


def _bool_from_form(value, default=True):
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "on", "sim", "ok", "realizada"}


def _parse_km_form(raw_value, label="KM"):
    raw_value = (raw_value or "").strip().replace(" ", "")
    if not raw_value:
        return None

    if re.fullmatch(r"\d+", raw_value):
        return float(raw_value)

    br_milhares = re.fullmatch(r"(\d{1,3}(?:\.\d{3})+)(?:,(\d+))?", raw_value)
    if br_milhares:
        decimal = br_milhares.group(2)
        normalized = br_milhares.group(1).replace(".", "")
        if decimal:
            normalized = f"{normalized}.{decimal}"
        return float(normalized)

    us_milhares = re.fullmatch(r"(\d{1,3}(?:,\d{3})+)(?:\.(\d+))?", raw_value)
    if us_milhares:
        decimal = us_milhares.group(2)
        normalized = us_milhares.group(1).replace(",", "")
        if decimal:
            normalized = f"{normalized}.{decimal}"
        return float(normalized)

    decimal_simples = re.fullmatch(r"(\d+)[,.](\d+)", raw_value)
    if decimal_simples:
        return float(f"{decimal_simples.group(1)}.{decimal_simples.group(2)}")

    raise ValueError(f"{label} deve ser informado como numero valido.")


def _validar_limite_km_turno(km_referencia, km_informado, label):
    km_referencia = km_referencia or 0
    if km_informado is None:
        return

    km_rodado = km_informado - km_referencia
    if km_rodado > MAX_KM_POR_TURNO:
        raise VeiculoTurnoError(
            (
                f"{label} ultrapassa o limite de {MAX_KM_POR_TURNO} km por turno. "
                f"Conferir o painel: referencia {km_referencia:.0f} km, informado {km_informado:.0f} km."
            ),
            "danger",
        )


def _parse_optional_int(raw_value):
    raw_value = (raw_value or "").strip()
    if not raw_value:
        return None
    try:
        return int(raw_value)
    except ValueError:
        return None


def _veiculo_do_operacional_logado(veiculo_id, *, user=None):
    query = Veiculos.query.filter(Veiculos.id == veiculo_id)
    veiculo = query.first_or_404()

    if getattr(user, "tipo_usuario", None) == EQUIPE_OCEANO_USER_TYPE:
        equipe = _equipe_oceano_logada(user)
        if not equipe or veiculo.equipe_id != equipe.id:
            raise PermissionError
        return veiculo

    _piloto_nome_logado(user, strict=True)
    equipe_ids = _equipe_ids_do_piloto(user)
    equipe_ok = bool(veiculo.equipe_id and veiculo.equipe_id in equipe_ids)
    if not equipe_ok:
        raise PermissionError
    return veiculo


def _apply_log_actor_scope(query, user):
    if getattr(user, "tipo_usuario", None) == EQUIPE_OCEANO_USER_TYPE:
        equipe_id = _parse_equipe_oceano_id(user)
        if not equipe_id:
            return query.filter(db.false())
        return query.filter(LogVeiculo.equipe_id == equipe_id)
    return query.filter(LogVeiculo.piloto_id == getattr(user, "piloto_id", None))


def _buscar_turno_aberto_usuario(veiculo_id, user, incluir_abastecimentos=False):
    query = LogVeiculo.query.filter(LogVeiculo.veiculo_id == veiculo_id, LogVeiculo.km_final.is_(None))
    query = _apply_log_actor_scope(query, user)
    if incluir_abastecimentos:
        query = query.options(
            selectinload(LogVeiculo.abastecimentos_detalhados),
            selectinload(LogVeiculo.limpezas_detalhadas),
        )
    return query.order_by(LogVeiculo.data_registro.desc()).first()


def _upload_veiculo_media_para_skybox(arquivo, placa, subpasta, nome, *, tipo="arquivo", dia=None):
    remote_path = build_veiculo_media_remote_path(placa, subpasta, nome, day=dia)
    try:
        marker = upload_file_to_skybox(arquivo, remote_path)
        current_app.logger.info(
            "Foto de veiculo enviada ao Skybox: placa=%s tipo=%s caminho=%s",
            placa,
            tipo,
            remote_path,
        )
        return marker
    except SkyboxError as exc:
        raise VeiculoTurnoError(
            f"Falha ao enviar {tipo} para o Skybox: {exc}",
            "danger",
        ) from exc


def build_veiculo_media_skybox_path(media_path, placa):
    if not media_path:
        return None
    if is_skybox_path(media_path):
        return media_path

    placa = (placa or "").strip()
    if not placa:
        return None

    parts = [part for part in str(media_path or "").replace("\\", "/").split("/") if part]
    if len(parts) < 4 or parts[-4] != "uploads" or parts[-3] != "veiculos":
        return None

    return build_veiculo_media_remote_path(placa, parts[-2], parts[-1], day=_dia_media_veiculo(parts[-1]))


def _dia_media_veiculo(filename):
    match = re.search(r"_(\d{4}-\d{2}-\d{2})_\d{2}-\d{2}-\d{2}-\d{3}\.", str(filename or ""))
    if match:
        return match.group(1)
    return None


def get_veiculo_log_for_media(user, log_id):
    tipo_usuario = normalize_role(getattr(user, "tipo_usuario", None))
    if tipo_usuario not in VEICULOS_LOGS_ALLOWED_TYPES:
        raise PermissionError

    return _build_veiculos_logs_query(user=user).filter(LogVeiculo.id == log_id).first()


def get_abastecimento_for_media(user, abastecimento_id):
    abastecimento = (
        Abastecimento.query
        .options(joinedload(Abastecimento.log_pai).joinedload(LogVeiculo.veiculo))
        .filter(Abastecimento.id == abastecimento_id)
        .first()
    )
    if not abastecimento:
        return None

    if not get_veiculo_log_for_media(user, abastecimento.log_veiculo_id):
        return None

    return abastecimento


def _salvar_upload_veiculo(arquivo, root_path, subpasta, prefixo, placa, *, copiar_skybox=False):
    if not arquivo or not arquivo.filename:
        return None

    ext = os.path.splitext(secure_filename(arquivo.filename))[1] or ".jpg"
    agora = _now_brazil()
    stamp = f"{agora:%Y-%m-%d_%H-%M-%S}-{agora.microsecond // 1000:03d}"
    nome = secure_filename(f"{prefixo}_{placa}_{stamp}{ext}")

    if copiar_skybox and skybox_enabled():
        return _upload_veiculo_media_para_skybox(
            arquivo,
            placa,
            subpasta,
            nome,
            tipo="imagem",
            dia=f"{agora:%Y-%m-%d}",
        )

    pasta_base = os.path.join(root_path, "static", "uploads", "veiculos")
    pasta_destino = os.path.join(pasta_base, subpasta)
    os.makedirs(pasta_destino, exist_ok=True)
    arquivo.save(os.path.join(pasta_destino, nome))

    return f"uploads/veiculos/{subpasta}/{nome}"


def build_veiculos_export_response(tipo_usuario, args, user=None):
    tipo_usuario = normalize_role(tipo_usuario)
    if tipo_usuario not in VEICULOS_ALLOWED_TYPES:
        raise PermissionError

    veiculos = list_veiculos(tipo_usuario, args, user=user)["veiculos"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Ve\u00edculos"

    fill_title = PatternFill("solid", fgColor="FFD966")
    fill_header = PatternFill("solid", fgColor="FFF2CC")
    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")
    fill_red = PatternFill("solid", fgColor="FFC7CE")
    fill_none = PatternFill()

    font_bold = Font(bold=True)
    font_title = Font(bold=True, size=12)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="000000")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "MODELO",
        "ANO",
        "FROTA",
        "OPERA\u00c7\u00c3O",
        "PLACA",
        "EQUIPE RESPONSAVEL",
        "KM ATUAL",
        "PROX REVISAO",
        "OBS",
    ]

    def write_section(title, rows, start_row):
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=8)
        cell = ws.cell(row=start_row, column=2, value=title)
        cell.font = font_title
        cell.alignment = align_center
        cell.fill = fill_title

        for c in range(2, 9):
            ws.cell(row=start_row, column=c).fill = fill_title
            ws.cell(row=start_row, column=c).border = border_thin

        header_row = start_row + 1
        for col_idx, header in enumerate(headers, start=1):
            current = ws.cell(row=header_row, column=col_idx, value=header)
            current.font = font_bold
            current.alignment = align_center
            current.fill = fill_header
            current.border = border_thin

        row = header_row + 1
        for veiculo in rows:
            falt = veiculo.km_restante_revisao

            obs = ""
            if veiculo.revisao_marcada_em:
                obs = "MARCADO " + veiculo.revisao_marcada_em.strftime("%d/%m %H:%M")
            elif veiculo.revisao_obs:
                obs = veiculo.revisao_obs

            data = [
                veiculo.modelo or "",
                veiculo.ano_fabricacao or "",
                veiculo.frota or "",
                veiculo.operacao or "",
                veiculo.placa or "",
                veiculo.equipe.nome_equipe if veiculo.equipe else "",
                float(veiculo.km_atual or 0),
                float(veiculo.km_prox_revisao) if veiculo.km_prox_revisao is not None else "",
                obs,
            ]

            for col_idx, value in enumerate(data, start=1):
                current = ws.cell(row=row, column=col_idx, value=value)
                current.border = border_thin
                current.alignment = align_left if col_idx in (1, 5, 6, 9) else align_center

                if col_idx in (7, 8) and isinstance(value, (int, float)):
                    current.number_format = "#,##0.00"

                if col_idx == 7 and isinstance(value, (int, float)):
                    current.fill = fill_green

                if col_idx == 8:
                    if value == "" or falt is None:
                        current.fill = fill_none
                    elif falt < 0:
                        current.fill = fill_red
                    elif falt <= 2000:
                        current.fill = fill_yellow
                    else:
                        current.fill = fill_green

            row += 1

        return row + 2

    by_op = {}
    for veiculo in veiculos:
        operacao = (veiculo.operacao or "OUTROS").upper()
        by_op.setdefault(operacao, []).append(veiculo)

    ops_order = []
    for current in ("PMSP", "AGRO"):
        if current in by_op:
            ops_order.append(current)
    for current in sorted(by_op.keys()):
        if current not in ops_order:
            ops_order.append(current)

    current_row = 2
    for operacao in ops_order:
        current_row = write_section(f"VEICULOS {operacao}", by_op[operacao], current_row)

    col_widths = {
        1: 16,
        2: 8,
        3: 12,
        4: 12,
        5: 14,
        6: 18,
        7: 14,
        8: 14,
        9: 26,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    response = make_response(file_stream.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = (
        f'attachment; filename="veiculos_{_now_brazil().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    return response


def _build_ultimos_logs(veiculos):
    ultimos_logs = {}
    if not veiculos:
        return ultimos_logs

    veiculo_ids = [veiculo.id for veiculo in veiculos]
    logs = (
        LogVeiculo.query
        .options(
            selectinload(LogVeiculo.abastecimentos_detalhados),
            selectinload(LogVeiculo.limpezas_detalhadas),
        )
        .filter(LogVeiculo.veiculo_id.in_(veiculo_ids))
        .order_by(LogVeiculo.veiculo_id.asc(), LogVeiculo.data_registro.desc())
        .all()
    )
    for log in logs:
        if log.veiculo_id not in ultimos_logs:
            ultimos_logs[log.veiculo_id] = log
    return ultimos_logs


def _build_veiculos_timeline_from_logs(logs):
    historico = {}
    dias_por_veiculo = defaultdict(dict)

    for log in logs:
        veiculo = log.veiculo
        item = historico.setdefault(
            log.veiculo_id,
            {
                "veiculo": veiculo,
                "logs": [],
                "dias": [],
                "total_logs": 0,
                "total_km": 0,
                "total_gasto": 0,
                "total_gasto_veiculo": 0,
                "total_gasto_gerador": 0,
                "total_limpeza": 0,
                "total_abastecimentos": 0,
                "total_abastecimentos_veiculo": 0,
                "total_abastecimentos_gerador": 0,
                "total_limpezas": 0,
                "total_limpezas_realizadas": {"quantidade": 0, "valor_total": 0},
            },
        )
        if item["veiculo"] is None and veiculo is not None:
            item["veiculo"] = veiculo

        km_rodado = _km_rodado_log_veiculo(log)
        gasto = log.total_valor_abastecido or 0
        gasto_limpeza = log.total_valor_limpeza or 0
        limpezas_realizadas = [
            limpeza
            for limpeza in (log.limpezas_detalhadas or [])
            if limpeza.limpeza_realizada
        ]
        qtd_limpezas_realizadas = len(limpezas_realizadas)
        valor_limpezas_realizadas = sum(float(limpeza.valor_total or 0) for limpeza in limpezas_realizadas)
        gasto_por_tipo = _sum_abastecimentos_por_tipo(log)
        data_inicio = log.data_registro
        movimentacao = log.ultima_movimentacao_em or data_inicio
        dia = data_inicio.date() if data_inicio else (movimentacao.date() if movimentacao else None)
        log_info = {
            "id": log.id,
            "data": movimentacao,
            "data_inicio": data_inicio,
            "operador": _operador_log_veiculo(log),
            "km_inicial": log.km_inicial or 0,
            "km_final": log.km_final if log.km_final is not None else log.ultimo_km_registrado,
            "km_rodado": km_rodado,
            "gasto": gasto,
            "gasto_veiculo": gasto_por_tipo["veiculo"],
            "gasto_gerador": gasto_por_tipo["gerador"],
            "gasto_limpeza": gasto_limpeza,
            "status": "Aberto" if log.km_final is None else "Encerrado",
            "abastecimentos": log.qtd_abastecimentos,
            "abastecimentos_veiculo": gasto_por_tipo["qtd_veiculo"],
            "abastecimentos_gerador": gasto_por_tipo["qtd_gerador"],
            "limpezas": log.qtd_limpezas,
            "qtd_fazendas_enderecos": log.qtd_fazendas_enderecos,
            "observacao": log.observacao,
            "foto_painel_path": log.foto_painel_path,
            "foto_painel_final_path": log.foto_painel_final_path,
            "assinatura_piloto": log.assinatura_piloto,
            "eventos_abastecimento": _eventos_abastecimento_log_veiculo(log),
            "eventos_limpeza": _eventos_limpeza_log_veiculo(log),
        }

        item["logs"].append(log_info)
        item["total_logs"] += 1
        item["total_km"] += km_rodado
        item["total_gasto"] += gasto
        item["total_gasto_veiculo"] += gasto_por_tipo["veiculo"]
        item["total_gasto_gerador"] += gasto_por_tipo["gerador"]
        item["total_limpeza"] += gasto_limpeza
        item["total_abastecimentos"] += log.qtd_abastecimentos
        item["total_abastecimentos_veiculo"] += gasto_por_tipo["qtd_veiculo"]
        item["total_abastecimentos_gerador"] += gasto_por_tipo["qtd_gerador"]
        item["total_limpezas"] += log.qtd_limpezas
        item["total_limpezas_realizadas"]["quantidade"] += qtd_limpezas_realizadas
        item["total_limpezas_realizadas"]["valor_total"] += valor_limpezas_realizadas

        if dia is not None:
            dia_item = dias_por_veiculo[log.veiculo_id].setdefault(
                dia,
                {
                    "dia": dia,
                    "km_rodado": 0,
                    "gasto": 0,
                    "gasto_veiculo": 0,
                    "gasto_gerador": 0,
                    "gasto_limpeza": 0,
                    "logs": 0,
                    "abastecimentos": 0,
                    "abastecimentos_veiculo": 0,
                    "abastecimentos_gerador": 0,
                    "limpezas": 0,
                    "limpezas_realizadas": 0,
                    "valor_limpezas_realizadas": 0,
                    "turnos": [],
                },
            )
            dia_item["km_rodado"] += km_rodado
            dia_item["gasto"] += gasto
            dia_item["gasto_veiculo"] += gasto_por_tipo["veiculo"]
            dia_item["gasto_gerador"] += gasto_por_tipo["gerador"]
            dia_item["gasto_limpeza"] += gasto_limpeza
            dia_item["logs"] += 1
            dia_item["abastecimentos"] += log.qtd_abastecimentos
            dia_item["abastecimentos_veiculo"] += gasto_por_tipo["qtd_veiculo"]
            dia_item["abastecimentos_gerador"] += gasto_por_tipo["qtd_gerador"]
            dia_item["limpezas"] += log.qtd_limpezas
            dia_item["limpezas_realizadas"] += qtd_limpezas_realizadas
            dia_item["valor_limpezas_realizadas"] += valor_limpezas_realizadas
            dia_item["turnos"].append(log_info)

    for veiculo_id, dias in dias_por_veiculo.items():
        historico[veiculo_id]["dias"] = [
            dias[dia]
            for dia in sorted(dias.keys(), reverse=True)
        ]

    return sorted(
        historico.values(),
        key=lambda item: (
            (getattr(item["veiculo"], "modelo", "") or "").upper(),
            (getattr(item["veiculo"], "placa", "") or "").upper(),
        ),
    )


def _can_view_retorno_automatico_audit(tipo_usuario, user=None):
    return normalize_role(tipo_usuario) in VEICULOS_LOGS_ALLOWED_TYPES


def _attach_retornos_automaticos_turnos(timeline, logs, *, user=None):
    logs_fechados = [
        log
        for log in (logs or [])
        if log.km_final is not None and log.equipe_id and log.data_registro
    ]
    if not logs_fechados:
        return

    equipe_ids = {log.equipe_id for log in logs_fechados if log.equipe_id}
    dias = {log.data_registro.date() for log in logs_fechados if log.data_registro}
    if not equipe_ids or not dias:
        return

    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe),
            joinedload(Solicitacao.ordem_servico).joinedload(OrdemServico.equipe),
        )
        .outerjoin(OrdemServico, OrdemServico.solicitacao_id == Solicitacao.id)
        .filter(
            db.or_(
                Solicitacao.gerada_automaticamente.is_(True),
                Solicitacao.origem_retorno_id.isnot(None),
            ),
            db.or_(
                Solicitacao.equipe_id.in_(equipe_ids),
                OrdemServico.equipe_id.in_(equipe_ids),
            ),
            db.or_(
                Solicitacao.data_agendamento.in_(dias),
                OrdemServico.data_aplicacao.in_(dias),
            ),
        )
    )
    query = apply_prefeitura_scope(query, user, Solicitacao.prefeitura_id)

    retornos_por_equipe_dia = defaultdict(list)
    for solicitacao in query.order_by(Solicitacao.data_agendamento.asc(), Solicitacao.id.asc()).all():
        ordem = solicitacao.ordem_servico
        equipe_id = (ordem.equipe_id if ordem else None) or solicitacao.equipe_id
        data_ref = (ordem.data_aplicacao if ordem else None) or solicitacao.data_agendamento
        if not equipe_id or not data_ref:
            continue

        data_ref_dia = data_ref.date() if hasattr(data_ref, "date") else data_ref
        if equipe_id not in equipe_ids or data_ref_dia not in dias:
            continue

        retornos_por_equipe_dia[(equipe_id, data_ref_dia)].append(
            _serialize_retorno_automatico_solicitacao(solicitacao)
        )

    if not retornos_por_equipe_dia:
        return

    retornos_por_log = {
        log.id: retornos_por_equipe_dia.get((log.equipe_id, log.data_registro.date()), [])
        for log in logs_fechados
    }

    for turno in timeline.get("logs", []):
        retornos = retornos_por_log.get(turno.get("id"), [])
        turno["retornos_automaticos"] = retornos
        turno["retornos_automaticos_count"] = len(retornos)

    for dia in timeline.get("dias", []):
        total = 0
        for turno in dia.get("turnos", []):
            retornos = retornos_por_log.get(turno.get("id"), [])
            turno["retornos_automaticos"] = retornos
            turno["retornos_automaticos_count"] = len(retornos)
            total += len(retornos)
        dia["retornos_automaticos_count"] = total


def _serialize_retorno_automatico_solicitacao(solicitacao):
    ordem = solicitacao.ordem_servico
    equipe = (ordem.equipe if ordem else None) or solicitacao.equipe
    data_ref = (ordem.data_aplicacao if ordem else None) or solicitacao.data_agendamento
    endereco = (
        f"{solicitacao.logradouro or ''}, {solicitacao.numero or 'S/N'} - "
        f"{solicitacao.bairro or ''} - {solicitacao.cidade or ''}/{solicitacao.uf or ''}"
    )
    if solicitacao.complemento:
        endereco = f"{endereco} - {solicitacao.complemento}"

    return {
        "id": solicitacao.id,
        "origem_id": solicitacao.origem_retorno_id,
        "protocolo": solicitacao.protocolo or "",
        "identificador_os": (ordem.identificador_os if ordem else "") or "",
        "status": solicitacao.status or "",
        "data": data_ref,
        "endereco": endereco,
        "bairro": solicitacao.bairro or "",
        "uvis": (solicitacao.usuario.nome_uvis if solicitacao.usuario else "") or "-",
        "equipe_nome": (equipe.nome_equipe if equipe else "") or "-",
    }


def _abastecimento_tipo_key(tipo_abastecimento):
    tipo = (tipo_abastecimento or "").strip().lower()
    return "gerador" if "gerador" in tipo else "veiculo"


def _sum_abastecimentos_por_tipo(log):
    resumo = {
        "veiculo": 0,
        "gerador": 0,
        "qtd_veiculo": 0,
        "qtd_gerador": 0,
    }
    for abastecimento in log.abastecimentos_detalhados or []:
        tipo_key = _abastecimento_tipo_key(abastecimento.tipo_abastecimento)
        resumo[tipo_key] += abastecimento.valor_total or 0
        resumo[f"qtd_{tipo_key}"] += 1
    return resumo


def _build_veiculos_summary_from_logs_query(
    *,
    user=None,
    q="",
    data_inicio="",
    data_fim="",
    limpeza_realizada="",
    tipo_limpeza="",
    data_limpeza_inicio="",
    data_limpeza_fim="",
    valor_limpeza_min="",
    valor_limpeza_max="",
):
    log_ids = (
        _build_veiculos_logs_query(
            user=user,
            q=q,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limpeza_realizada=limpeza_realizada,
            tipo_limpeza=tipo_limpeza,
            data_limpeza_inicio=data_limpeza_inicio,
            data_limpeza_fim=data_limpeza_fim,
            valor_limpeza_min=valor_limpeza_min,
            valor_limpeza_max=valor_limpeza_max,
            include_options=False,
            include_order=False,
        )
        .with_entities(LogVeiculo.id)
        .subquery()
    )
    tipo_lower = db.func.lower(db.func.coalesce(Abastecimento.tipo_abastecimento, ""))
    is_abastecimento = Abastecimento.id.isnot(None)
    is_gerador = db.and_(is_abastecimento, tipo_lower.like("%gerador%"))
    is_veiculo = db.and_(is_abastecimento, db.not_(tipo_lower.like("%gerador%")))
    log_summary = (
        db.session.query(
            LogVeiculo.id.label("log_id"),
            LogVeiculo.veiculo_id.label("veiculo_id"),
            LogVeiculo.km_inicial.label("km_inicial"),
            db.func.coalesce(
                LogVeiculo.km_final,
                db.func.max(Abastecimento.km_registro),
                LogVeiculo.km_inicial,
            ).label("km_referencia"),
            db.func.count(Abastecimento.id).label("qtd_abastecimentos"),
            db.func.coalesce(db.func.sum(Abastecimento.valor_total), 0).label("total_gasto"),
            db.func.coalesce(
                db.func.sum(case((is_veiculo, Abastecimento.valor_total), else_=0)),
                0,
            ).label("total_gasto_veiculo"),
            db.func.coalesce(
                db.func.sum(case((is_gerador, Abastecimento.valor_total), else_=0)),
                0,
            ).label("total_gasto_gerador"),
            db.func.coalesce(db.func.sum(case((is_veiculo, 1), else_=0)), 0).label("qtd_abastecimentos_veiculo"),
            db.func.coalesce(db.func.sum(case((is_gerador, 1), else_=0)), 0).label("qtd_abastecimentos_gerador"),
        )
        .join(log_ids, log_ids.c.id == LogVeiculo.id)
        .outerjoin(Abastecimento, Abastecimento.log_veiculo_id == LogVeiculo.id)
        .group_by(LogVeiculo.id, LogVeiculo.veiculo_id, LogVeiculo.km_inicial, LogVeiculo.km_final)
        .subquery()
    )
    km_rodado = case(
        (log_summary.c.km_referencia < log_summary.c.km_inicial, 0),
        else_=log_summary.c.km_referencia - log_summary.c.km_inicial,
    )
    summary_rows = (
        db.session.query(
            log_summary.c.veiculo_id,
            db.func.count(log_summary.c.log_id).label("total_logs"),
            db.func.coalesce(db.func.sum(km_rodado), 0).label("total_km"),
            db.func.coalesce(db.func.sum(log_summary.c.total_gasto), 0).label("total_gasto"),
            db.func.coalesce(db.func.sum(log_summary.c.total_gasto_veiculo), 0).label("total_gasto_veiculo"),
            db.func.coalesce(db.func.sum(log_summary.c.total_gasto_gerador), 0).label("total_gasto_gerador"),
            db.func.coalesce(db.func.sum(log_summary.c.qtd_abastecimentos), 0).label("total_abastecimentos"),
            db.func.coalesce(db.func.sum(log_summary.c.qtd_abastecimentos_veiculo), 0).label("total_abastecimentos_veiculo"),
            db.func.coalesce(db.func.sum(log_summary.c.qtd_abastecimentos_gerador), 0).label("total_abastecimentos_gerador"),
        )
        .group_by(log_summary.c.veiculo_id)
        .all()
    )

    limpeza_summary_conditions = _limpeza_filter_conditions(
        limpeza_realizada=limpeza_realizada,
        tipo_limpeza=tipo_limpeza,
        data_limpeza_inicio=data_limpeza_inicio,
        data_limpeza_fim=data_limpeza_fim,
        valor_limpeza_min=valor_limpeza_min,
        valor_limpeza_max=valor_limpeza_max,
    )
    limpeza_rows_query = (
        db.session.query(
            LogVeiculo.veiculo_id.label("veiculo_id"),
            db.func.coalesce(db.func.sum(LimpezaVeiculo.valor_total), 0).label("total_limpeza"),
            db.func.count(LimpezaVeiculo.id).label("total_limpezas"),
        )
        .join(log_ids, log_ids.c.id == LogVeiculo.id)
        .join(LimpezaVeiculo, LimpezaVeiculo.log_veiculo_id == LogVeiculo.id)
    )
    if limpeza_summary_conditions:
        limpeza_rows_query = limpeza_rows_query.filter(*limpeza_summary_conditions)
    limpeza_rows = (
        limpeza_rows_query
        .group_by(LogVeiculo.veiculo_id)
        .all()
    )
    limpezas_por_veiculo = {
        row.veiculo_id: {
            "total_limpeza": row.total_limpeza or 0,
            "total_limpezas": int(row.total_limpezas or 0),
        }
        for row in limpeza_rows
    }

    if not summary_rows:
        return []

    veiculo_ids = [row.veiculo_id for row in summary_rows]
    veiculos = {
        veiculo.id: veiculo
        for veiculo in (
            Veiculos.query
            .options(joinedload(Veiculos.equipe))
            .filter(Veiculos.id.in_(veiculo_ids))
            .all()
        )
    }
    items = []
    for row in summary_rows:
        veiculo = veiculos.get(row.veiculo_id)
        limpeza_summary = limpezas_por_veiculo.get(row.veiculo_id, {})
        items.append(
            {
                "veiculo": veiculo,
                "logs": [],
                "dias": [],
                "total_logs": int(row.total_logs or 0),
                "total_km": row.total_km or 0,
                "total_gasto": row.total_gasto or 0,
                "total_gasto_veiculo": row.total_gasto_veiculo or 0,
                "total_gasto_gerador": row.total_gasto_gerador or 0,
                "total_limpeza": limpeza_summary.get("total_limpeza", 0),
                "total_abastecimentos": int(row.total_abastecimentos or 0),
                "total_abastecimentos_veiculo": int(row.total_abastecimentos_veiculo or 0),
                "total_abastecimentos_gerador": int(row.total_abastecimentos_gerador or 0),
                "total_limpezas": limpeza_summary.get("total_limpezas", 0),
            }
        )

    return sorted(
        items,
        key=lambda item: (
            (getattr(item["veiculo"], "modelo", "") or "").upper(),
            (getattr(item["veiculo"], "placa", "") or "").upper(),
        ),
    )


def _build_veiculo_km_conferencia(logs):
    logs_ordenados = sorted(
        logs,
        key=lambda log: (
            log.data_registro or datetime.min,
            log.id or 0,
        ),
    )
    linhas = []
    anterior = None

    for log in logs_ordenados:
        diferenca_anterior = None
        status = "primeiro"
        status_label = "Primeiro turno"

        if anterior is not None:
            if anterior.km_final is None:
                status = "anterior_aberto"
                status_label = "Anterior aberto"
            else:
                diferenca_anterior = (log.km_inicial or 0) - (anterior.km_final or 0)
                if abs(diferenca_anterior) < 0.0001:
                    status = "ok"
                    status_label = "OK"
                elif diferenca_anterior > 0:
                    status = "perdido"
                    status_label = "KM perdido"
                else:
                    status = "sobreposto"
                    status_label = "KM sobreposto"

        linhas.append(
            {
                "id": log.id,
                "data_inicio": log.data_registro,
                "operador": _operador_log_veiculo(log),
                "km_inicial": log.km_inicial or 0,
                "km_final": log.km_final,
                "ultimo_km_registrado": log.ultimo_km_registrado,
                "diferenca_anterior": diferenca_anterior,
                "status": status,
                "status_label": status_label,
                "log_anterior_id": anterior.id if anterior is not None else None,
            }
        )
        anterior = log

    return linhas


def _km_rodado_log_veiculo(log):
    km_inicial = log.km_inicial or 0
    km_final = log.km_final if log.km_final is not None else log.ultimo_km_registrado
    return max((km_final or 0) - km_inicial, 0)


def _operador_log_veiculo(log):
    if log.piloto:
        return log.piloto.nome_piloto
    if log.equipe:
        return log.equipe.nome_equipe
    return "-"


def _eventos_abastecimento_log_veiculo(log):
    eventos = []
    for abastecimento in log.abastecimentos_ordenados:
        eventos.append(
            {
                "id": abastecimento.id,
                "data": abastecimento.data_hora,
                "km": abastecimento.km_registro or 0,
                "tipo": abastecimento.tipo_abastecimento or "Abastecimento",
                "tipo_key": _abastecimento_tipo_key(abastecimento.tipo_abastecimento),
                "litros": abastecimento.litros or 0,
                "valor": abastecimento.valor_total or 0,
                "foto_painel_path": abastecimento.foto_painel_path,
                "foto_nf_path": abastecimento.foto_nf_path,
            }
        )
    return eventos


def _eventos_limpeza_log_veiculo(log):
    eventos = []
    for limpeza in log.limpezas_ordenadas:
        eventos.append(
            {
                "id": limpeza.id,
                "data": limpeza.data_hora,
                "data_registro": limpeza.data_registro,
                "realizada": bool(limpeza.limpeza_realizada),
                "tipo": limpeza.tipo_limpeza or "",
                "tipo_label": _limpeza_tipo_label(limpeza.tipo_limpeza),
                "valor": float(limpeza.valor_total or 0),
                "observacao": limpeza.observacao or "",
            }
        )
    return eventos


def _limpeza_tipo_label(tipo_limpeza):
    tipo = (tipo_limpeza or "").strip().lower()
    if tipo == "completa":
        return "Completa"
    if tipo == "ducha":
        return "Apenas ducha"
    return "Nao informado"


def _limpezas_tipos_resumo(log):
    tipos = []
    for item in log.limpezas_ordenadas:
        label = _limpeza_tipo_label(item.tipo_limpeza)
        if label and label not in tipos:
            tipos.append(label)
    return ", ".join(tipos)


def _limpezas_registros_resumo(log):
    return " | ".join(
        item.data_registro.strftime("%d/%m/%Y %H:%M")
        for item in log.limpezas_ordenadas
        if item.data_registro
    )


def _limpezas_datas_resumo(log):
    return " | ".join(
        item.data_hora.strftime("%d/%m/%Y %H:%M")
        for item in log.limpezas_ordenadas
        if item.data_hora
    )


def _ultima_movimentacao_log_subquery():
    movimentos = (
        db.session.query(
            Abastecimento.log_veiculo_id.label("log_id"),
            Abastecimento.data_hora.label("data_hora"),
        )
        .union_all(
            db.session.query(
                LimpezaVeiculo.log_veiculo_id.label("log_id"),
                LimpezaVeiculo.data_hora.label("data_hora"),
            )
        )
        .subquery()
    )

    return (
        db.session.query(
            movimentos.c.log_id,
            db.func.max(movimentos.c.data_hora).label("ultima_movimentacao_em"),
        )
        .group_by(movimentos.c.log_id)
        .subquery()
    )


def list_veiculos_logs(tipo_usuario, args, user=None):
    tipo_usuario = normalize_role(tipo_usuario)
    if tipo_usuario not in VEICULOS_LOGS_ALLOWED_TYPES:
        raise PermissionError

    q = (args.get("q") or "").strip()
    data_inicio = (args.get("data_inicio") or "").strip()
    data_fim = (args.get("data_fim") or "").strip()
    limpeza_realizada = (args.get("limpeza_realizada") or "").strip()
    tipo_limpeza = (args.get("tipo_limpeza") or "").strip().lower()
    data_limpeza_inicio = (args.get("data_limpeza_inicio") or "").strip()
    data_limpeza_fim = (args.get("data_limpeza_fim") or "").strip()
    valor_limpeza_min = (args.get("valor_limpeza_min") or "").strip()
    valor_limpeza_max = (args.get("valor_limpeza_max") or "").strip()
    page = args.get("page", 1, type=int)

    query = _build_veiculos_logs_query(
        user=user,
        q=q,
        data_inicio=data_inicio,
        data_fim=data_fim,
        limpeza_realizada=limpeza_realizada,
        tipo_limpeza=tipo_limpeza,
        data_limpeza_inicio=data_limpeza_inicio,
        data_limpeza_fim=data_limpeza_fim,
        valor_limpeza_min=valor_limpeza_min,
        valor_limpeza_max=valor_limpeza_max,
    )
    total_logs = _build_veiculos_logs_query(
        user=user,
        q=q,
        data_inicio=data_inicio,
        data_fim=data_fim,
        limpeza_realizada=limpeza_realizada,
        tipo_limpeza=tipo_limpeza,
        data_limpeza_inicio=data_limpeza_inicio,
        data_limpeza_fim=data_limpeza_fim,
        valor_limpeza_min=valor_limpeza_min,
        valor_limpeza_max=valor_limpeza_max,
        include_options=False,
        include_order=False,
    ).count()
    paginacao = query.paginate(page=page, per_page=20, error_out=False)
    logs = paginacao.items

    return {
        "logs": logs,
        "paginacao": paginacao,
        "total_logs": total_logs,
        "total_abastecido": _sum_veiculos_logs_abastecido(
            user=user,
            q=q,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limpeza_realizada=limpeza_realizada,
            tipo_limpeza=tipo_limpeza,
            data_limpeza_inicio=data_limpeza_inicio,
            data_limpeza_fim=data_limpeza_fim,
            valor_limpeza_min=valor_limpeza_min,
            valor_limpeza_max=valor_limpeza_max,
        ),
        "total_limpezas_realizadas": _sum_veiculos_logs_limpezas_realizadas(
            user=user,
            q=q,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limpeza_realizada=limpeza_realizada,
            tipo_limpeza=tipo_limpeza,
            data_limpeza_inicio=data_limpeza_inicio,
            data_limpeza_fim=data_limpeza_fim,
            valor_limpeza_min=valor_limpeza_min,
            valor_limpeza_max=valor_limpeza_max,
        ),
        "filters": {
            "q": q,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "limpeza_realizada": limpeza_realizada,
            "tipo_limpeza": tipo_limpeza,
            "data_limpeza_inicio": data_limpeza_inicio,
            "data_limpeza_fim": data_limpeza_fim,
            "valor_limpeza_min": valor_limpeza_min,
            "valor_limpeza_max": valor_limpeza_max,
        },
        "can_edit_logs": tipo_usuario in {"dev", "diretor", "admin", "operario", "operador", "prefeitura_admin"},
        "can_delete_logs": tipo_usuario == "admin",
        "can_view_deleted_logs": tipo_usuario == "dev",
        "veiculos_timeline": _build_veiculos_summary_from_logs_query(
            user=user,
            q=q,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limpeza_realizada=limpeza_realizada,
            tipo_limpeza=tipo_limpeza,
            data_limpeza_inicio=data_limpeza_inicio,
            data_limpeza_fim=data_limpeza_fim,
            valor_limpeza_min=valor_limpeza_min,
            valor_limpeza_max=valor_limpeza_max,
        ),
    }


def list_veiculos_limpezas(tipo_usuario, args, user=None):
    tipo_usuario = normalize_role(tipo_usuario)
    if tipo_usuario not in VEICULOS_LOGS_ALLOWED_TYPES:
        raise PermissionError

    q = (args.get("q") or "").strip()
    limpeza_realizada = (args.get("limpeza_realizada") or "").strip()
    tipo_limpeza = (args.get("tipo_limpeza") or "").strip().lower()
    operacao = (args.get("operacao") or "").strip().upper()
    data_limpeza_inicio = (args.get("data_limpeza_inicio") or "").strip()
    data_limpeza_fim = (args.get("data_limpeza_fim") or "").strip()
    data_registro_inicio = (args.get("data_registro_inicio") or "").strip()
    data_registro_fim = (args.get("data_registro_fim") or "").strip()
    page = args.get("page", 1, type=int)

    query = _build_veiculos_limpezas_query(
        user=user,
        q=q,
        limpeza_realizada=limpeza_realizada,
        tipo_limpeza=tipo_limpeza,
        operacao=operacao,
        data_limpeza_inicio=data_limpeza_inicio,
        data_limpeza_fim=data_limpeza_fim,
        data_registro_inicio=data_registro_inicio,
        data_registro_fim=data_registro_fim,
    )
    total_query = _build_veiculos_limpezas_query(
        user=user,
        q=q,
        limpeza_realizada=limpeza_realizada,
        tipo_limpeza=tipo_limpeza,
        operacao=operacao,
        data_limpeza_inicio=data_limpeza_inicio,
        data_limpeza_fim=data_limpeza_fim,
        data_registro_inicio=data_registro_inicio,
        data_registro_fim=data_registro_fim,
        include_options=False,
        include_order=False,
    )
    resumo = _build_veiculos_limpezas_resumo(total_query)
    paginacao = query.paginate(page=page, per_page=25, error_out=False)

    return {
        "limpezas": paginacao.items,
        "paginacao": paginacao,
        "resumo": resumo,
        "filters": {
            "q": q,
            "limpeza_realizada": limpeza_realizada,
            "tipo_limpeza": tipo_limpeza,
            "operacao": operacao,
            "data_limpeza_inicio": data_limpeza_inicio,
            "data_limpeza_fim": data_limpeza_fim,
            "data_registro_inicio": data_registro_inicio,
            "data_registro_fim": data_registro_fim,
        },
    }


def build_veiculo_logs_detalhe_context(tipo_usuario, veiculo_id, args, user=None):
    tipo_usuario = normalize_role(tipo_usuario)
    if tipo_usuario not in VEICULOS_LOGS_ALLOWED_TYPES:
        raise PermissionError

    data_inicio = (args.get("data_inicio") or "").strip()
    data_fim = (args.get("data_fim") or "").strip()

    veiculo = _get_veiculo_logs_scoped(veiculo_id, user)
    logs = (
        _build_veiculos_logs_query(user=user, data_inicio=data_inicio, data_fim=data_fim)
        .filter(LogVeiculo.veiculo_id == veiculo_id)
        .all()
    )
    timeline_items = _build_veiculos_timeline_from_logs(logs)
    timeline = timeline_items[0] if timeline_items else {
        "veiculo": veiculo,
        "logs": [],
        "dias": [],
        "total_logs": 0,
        "total_km": 0,
        "total_gasto": 0,
        "total_gasto_veiculo": 0,
        "total_gasto_gerador": 0,
        "total_limpeza": 0,
        "total_abastecimentos": 0,
        "total_abastecimentos_veiculo": 0,
        "total_abastecimentos_gerador": 0,
        "total_limpezas": 0,
        "total_limpezas_realizadas": {"quantidade": 0, "valor_total": 0},
    }
    can_view_retorno_automatico_audit = _can_view_retorno_automatico_audit(tipo_usuario, user)
    if can_view_retorno_automatico_audit:
        _attach_retornos_automaticos_turnos(timeline, logs, user=user)

    return {
        "veiculo": veiculo,
        "timeline": timeline,
        "km_conferencia": _build_veiculo_km_conferencia(logs),
        "filters": {"data_inicio": data_inicio, "data_fim": data_fim},
        "can_edit_logs": tipo_usuario in {"dev", "diretor", "admin", "operario", "operador", "prefeitura_admin"},
        "can_delete_logs": tipo_usuario == "admin",
        "can_view_deleted_logs": tipo_usuario == "dev",
        "can_view_retorno_automatico_audit": can_view_retorno_automatico_audit,
    }


def build_veiculos_deleted_logs_context(tipo_usuario, args):
    tipo_usuario = normalize_role(tipo_usuario)
    if tipo_usuario != "dev":
        raise PermissionError

    q = (args.get("q") or "").strip()
    data_inicio = (args.get("data_inicio") or "").strip()
    data_fim = (args.get("data_fim") or "").strip()
    page = args.get("page", 1, type=int)

    query = AuditoriaUsuario.query.filter(
        AuditoriaUsuario.endpoint == VEICULO_LOG_DELETE_AUDIT_ENDPOINT,
        AuditoriaUsuario.tipo_evento == "EXCLUSAO",
    )

    if q:
        like = f"%{q.lower()}%"
        query = query.filter(
            db.or_(
                id_search_clause(AuditoriaUsuario.id, q),
                id_search_clause(AuditoriaUsuario.usuario_id, q),
                db.func.lower(db.func.coalesce(AuditoriaUsuario.usuario_nome, "")).like(like),
                db.func.lower(db.func.coalesce(AuditoriaUsuario.usuario_login, "")).like(like),
                db.func.lower(db.func.coalesce(AuditoriaUsuario.query_string, "")).like(like),
            )
        )

    if data_inicio:
        try:
            query = query.filter(AuditoriaUsuario.criado_em >= datetime.strptime(data_inicio, "%Y-%m-%d"))
        except ValueError:
            pass

    if data_fim:
        try:
            query = query.filter(
                AuditoriaUsuario.criado_em <= datetime.strptime(data_fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            )
        except ValueError:
            pass

    paginacao = query.order_by(AuditoriaUsuario.criado_em.desc(), AuditoriaUsuario.id.desc()).paginate(
        page=page,
        per_page=25,
        error_out=False,
    )

    return {
        "logs_excluidos": [_build_deleted_log_history_item(log) for log in paginacao.items],
        "paginacao": paginacao,
        "filters": {"q": q, "data_inicio": data_inicio, "data_fim": data_fim},
    }


def _build_deleted_log_history_item(audit_log):
    try:
        snapshot = json.loads(audit_log.query_string or "{}")
    except (TypeError, ValueError):
        snapshot = {}

    return {
        "audit_log": audit_log,
        "snapshot": snapshot,
        "veiculo": snapshot.get("veiculo") or {},
        "turno": snapshot.get("turno") or {},
        "operador": snapshot.get("operador") or {},
        "abastecimentos": snapshot.get("abastecimentos") or [],
        "totais": snapshot.get("totais") or {},
    }


def _get_veiculo_logs_scoped(veiculo_id, user):
    query = Veiculos.query.filter(Veiculos.id == veiculo_id)
    if user is not None and getattr(user, "tipo_usuario", None) == EQUIPE_OCEANO_USER_TYPE:
        equipe = _equipe_oceano_logada(user)
        if not equipe:
            raise PermissionError
        query = query.filter(_veiculo_equipe_operacional_filter(equipe, user))
    elif user is not None:
        query = apply_prefeitura_scope(query, user, Veiculos.prefeitura_id)

    veiculo = query.first()
    if veiculo is None:
        raise PermissionError
    return veiculo


def update_veiculo_log_km(user, log_id, form_data):
    tipo_usuario = normalize_role(getattr(user, "tipo_usuario", None))
    if tipo_usuario not in {"dev", "diretor", "admin", "operario", "operador", "prefeitura_admin"}:
        raise PermissionError

    log = (
        _build_veiculos_logs_query(user=user)
        .filter(LogVeiculo.id == log_id)
        .first()
    )
    if log is None:
        raise PermissionError

    km_inicial = _parse_log_km_field(form_data, "km_inicial", "KM inicial", required=True)
    km_final = _parse_log_km_field(form_data, "km_final", "KM final")
    if km_final is not None and km_final < km_inicial:
        raise VeiculoTurnoError("KM final nao pode ser menor que o KM inicial.")

    abastecimentos = list(log.abastecimentos_detalhados or [])
    abastecimento_updates = []
    for abastecimento in abastecimentos:
        km_field_name = f"abastecimento_{abastecimento.id}_km"
        km_registro = _parse_log_km_field(
            form_data,
            km_field_name,
            f"KM do abastecimento #{abastecimento.id}",
            required=True,
        )
        valor_field_name = f"abastecimento_{abastecimento.id}_valor"
        valor_total = abastecimento.valor_total
        if valor_field_name in form_data:
            valor_total = _parse_log_decimal_field(
                form_data,
                valor_field_name,
                f"Valor do abastecimento #{abastecimento.id}",
                required=True,
            )
        abastecimento_updates.append((abastecimento, km_registro, valor_total))

    maior_km_abastecimento = max([km for _item, km, _valor in abastecimento_updates], default=None)
    if km_final is not None and maior_km_abastecimento is not None and km_final < maior_km_abastecimento:
        raise VeiculoTurnoError("KM final nao pode ser menor que o maior KM de abastecimento.")

    log.km_inicial = km_inicial
    log.km_final = km_final
    for abastecimento, km_registro, valor_total in abastecimento_updates:
        abastecimento.km_registro = km_registro
        abastecimento.valor_total = valor_total

    _recalcular_km_atual_veiculo(log.veiculo_id)
    db.session.commit()
    return f"Log #{log.id} corrigido com sucesso."


def delete_veiculo_log(user, log_id, request_info=None):
    tipo_usuario = normalize_role(getattr(user, "tipo_usuario", None))
    if tipo_usuario != "admin":
        raise PermissionError

    log = (
        _build_veiculos_logs_query(user=user)
        .filter(LogVeiculo.id == log_id)
        .first()
    )
    if log is None:
        raise PermissionError

    veiculo_id = log.veiculo_id
    _record_veiculo_log_delete_audit(user, log, request_info=request_info)
    db.session.delete(log)
    db.session.flush()
    _recalcular_km_atual_veiculo(veiculo_id)
    db.session.commit()
    return f"Log #{log_id} removido com sucesso."


def _serialize_datetime(value):
    return value.isoformat() if value else None


def _build_veiculo_log_delete_snapshot(log):
    abastecimentos = []
    for item in log.abastecimentos_ordenados:
        abastecimentos.append(
            {
                "id": item.id,
                "data_hora": _serialize_datetime(item.data_hora),
                "km_registro": item.km_registro,
                "tipo_abastecimento": item.tipo_abastecimento,
                "litros": item.litros,
                "valor_total": item.valor_total,
                "foto_nf_path": item.foto_nf_path,
                "foto_painel_path": item.foto_painel_path,
            }
        )

    limpezas = []
    for item in log.limpezas_ordenadas:
        limpezas.append(
            {
                "id": item.id,
                "data_registro": _serialize_datetime(item.data_registro),
                "data_hora": _serialize_datetime(item.data_hora),
                "limpeza_realizada": bool(item.limpeza_realizada),
                "tipo_limpeza": item.tipo_limpeza,
                "valor_total": float(item.valor_total or 0),
                "observacao": item.observacao,
            }
        )

    return {
        "log_id": log.id,
        "veiculo": {
            "id": log.veiculo.id if log.veiculo else log.veiculo_id,
            "modelo": log.veiculo.modelo if log.veiculo else None,
            "placa": log.veiculo.placa if log.veiculo else None,
            "responsavel": log.veiculo.responsavel if log.veiculo else None,
            "prefeitura_id": getattr(log.veiculo, "prefeitura_id", None) if log.veiculo else None,
        },
        "operador": {
            "piloto_id": log.piloto_id,
            "piloto_nome": log.piloto.nome_piloto if log.piloto else None,
            "equipe_id": log.equipe_id,
            "equipe_nome": log.equipe.nome_equipe if log.equipe else None,
        },
        "turno": {
            "data_registro": _serialize_datetime(log.data_registro),
            "km_inicial": log.km_inicial,
            "km_final": log.km_final,
            "ultimo_km_registrado": log.ultimo_km_registrado,
            "km_rodado": _km_rodado_log_veiculo(log),
            "check_diario": bool(log.check_diario),
            "qtd_fazendas_enderecos": log.qtd_fazendas_enderecos,
            "observacao": log.observacao,
            "foto_painel_path": log.foto_painel_path,
            "foto_painel_final_path": log.foto_painel_final_path,
            "assinatura_presente": bool(log.assinatura_piloto),
        },
        "totais": {
            "qtd_abastecimentos": log.qtd_abastecimentos,
            "total_litros_abastecidos": log.total_litros_abastecidos,
            "total_valor_abastecido": log.total_valor_abastecido,
            "qtd_limpezas": log.qtd_limpezas,
            "total_valor_limpeza": log.total_valor_limpeza,
        },
        "abastecimentos": abastecimentos,
        "limpezas": limpezas,
    }


def _record_veiculo_log_delete_audit(user, log, request_info=None):
    request_info = request_info or {}
    snapshot = _build_veiculo_log_delete_snapshot(log)
    usuario_nome = (
        getattr(user, "nome_uvis", None)
        or getattr(user, "login", None)
        or "Usuario sem nome"
    )
    db.session.add(
        AuditoriaUsuario(
            usuario_id=getattr(user, "id", None),
            usuario_nome=usuario_nome[:100],
            usuario_login=getattr(user, "login", None),
            tipo_usuario=getattr(user, "tipo_usuario", None),
            metodo="POST",
            tipo_evento="EXCLUSAO",
            endpoint=VEICULO_LOG_DELETE_AUDIT_ENDPOINT,
            path=(request_info.get("path") or f"/veiculos/logs/{log.id}/deletar")[:255],
            query_string=json.dumps(snapshot, ensure_ascii=False, default=str),
            status_code=200,
            ip=request_info.get("ip"),
            user_agent=request_info.get("user_agent"),
            referrer=(request_info.get("referrer") or None),
            criado_em=_now_utc(),
        )
    )


def _parse_log_km_field(form_data, field_name, label, *, required=False):
    raw_value = (form_data.get(field_name) or "").strip()
    if not raw_value:
        if required:
            raise VeiculoTurnoError(f"Informe {label}.")
        return None

    try:
        value = _parse_km_form(raw_value, label)
    except (TypeError, ValueError):
        raise VeiculoTurnoError(f"{label} deve ser informado em KM inteiro, sem virgula decimal.")

    if value is None:
        if required:
            raise VeiculoTurnoError(f"Informe {label}.")
        return None
    if value < 0:
        raise VeiculoTurnoError(f"{label} nao pode ser negativo.")
    return value


def _parse_log_decimal_field(form_data, field_name, label, *, required=False):
    raw_value = (form_data.get(field_name) or "").strip()
    if not raw_value:
        if required:
            raise VeiculoTurnoError(f"Informe {label}.")
        return None

    try:
        value = _parse_decimal_form(raw_value)
    except (TypeError, ValueError):
        raise VeiculoTurnoError(f"{label} deve ser informado com valor numerico valido.")

    if value is None:
        if required:
            raise VeiculoTurnoError(f"Informe {label}.")
        return None
    if value < 0:
        raise VeiculoTurnoError(f"{label} nao pode ser negativo.")
    return value


def _recalcular_km_atual_veiculo(veiculo_id):
    veiculo = db.session.get(Veiculos, veiculo_id)
    if veiculo is None:
        return None

    ultimo_log_fechado = (
        LogVeiculo.query
        .filter(LogVeiculo.veiculo_id == veiculo_id, LogVeiculo.km_final.isnot(None))
        .order_by(LogVeiculo.data_registro.desc(), LogVeiculo.id.desc())
        .first()
    )
    if ultimo_log_fechado is not None:
        veiculo.km_atual = ultimo_log_fechado.km_final or 0
        return veiculo.km_atual

    ultimo_log = (
        LogVeiculo.query
        .filter(LogVeiculo.veiculo_id == veiculo_id)
        .order_by(LogVeiculo.data_registro.desc(), LogVeiculo.id.desc())
        .first()
    )
    veiculo.km_atual = (ultimo_log.km_inicial if ultimo_log is not None else 0) or 0
    return veiculo.km_atual


def build_veiculos_logs_export(tipo_usuario, args, user=None):
    tipo_usuario = normalize_role(tipo_usuario)
    if tipo_usuario not in VEICULOS_LOGS_ALLOWED_TYPES:
        raise PermissionError

    q = (args.get("q") or "").strip()
    data_inicio = (args.get("data_inicio") or "").strip()
    data_fim = (args.get("data_fim") or "").strip()
    limpeza_realizada = (args.get("limpeza_realizada") or "").strip()
    tipo_limpeza = (args.get("tipo_limpeza") or "").strip().lower()
    data_limpeza_inicio = (args.get("data_limpeza_inicio") or "").strip()
    data_limpeza_fim = (args.get("data_limpeza_fim") or "").strip()
    valor_limpeza_min = (args.get("valor_limpeza_min") or "").strip()
    valor_limpeza_max = (args.get("valor_limpeza_max") or "").strip()

    logs = _build_veiculos_logs_query(
        user=user,
        q=q,
        data_inicio=data_inicio,
        data_fim=data_fim,
        limpeza_realizada=limpeza_realizada,
        tipo_limpeza=tipo_limpeza,
        data_limpeza_inicio=data_limpeza_inicio,
        data_limpeza_fim=data_limpeza_fim,
        valor_limpeza_min=valor_limpeza_min,
        valor_limpeza_max=valor_limpeza_max,
    ).all()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row=1, height=28):
        ws.row_dimensions[row].height = height
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

    def auto_width(ws, max_col, min_w=10, max_w=48):
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))

    def center_cols(ws, cols, start_row, end_row):
        for col in cols:
            for row in range(start_row, end_row + 1):
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalhamento"

    headers = [
        "Ultima Movimentacao",
        "Veiculo",
        "Placa",
        "Responsavel",
        "Operador / Equipe",
        "Check Diario",
        "KM Inicial",
        "KM Final",
        "KM Rodado",
        "Abasteceu",
        "Qtd. Abastecimentos",
        "Tipos de Abastecimento",
        "Litros",
        "Valor Abastecimento (R$)",
        "Valor por Litro (R$)",
        "Custo por KM (R$)",
        "Limpeza Realizada",
        "Qtd. Limpezas",
        "Data(s) Registro Limpeza",
        "Data(s) Limpeza",
        "Tipo(s) Limpeza",
        "Valor Limpeza (R$)",
        "Assinatura",
        "Observacao",
    ]
    ws.append(headers)
    style_header(ws, row=1)
    ws.freeze_panes = "A2"

    for log in logs:
        ws.append([
            log.ultima_movimentacao_em.strftime("%d/%m/%Y %H:%M") if log.ultima_movimentacao_em else "",
            (log.veiculo.modelo if log.veiculo else "") or "",
            (log.veiculo.placa if log.veiculo else "") or "",
            (log.veiculo.responsavel if log.veiculo else "") or "",
            (log.piloto.nome_piloto if log.piloto else None)
            or (log.equipe.nome_equipe if log.equipe else "")
            or "",
            "SIM" if log.check_diario else "N\u00c3O",
            float(log.km_inicial or 0),
            "" if log.km_final is None else float(log.km_final),
            None,
            "SIM" if log.teve_abastecimento else "N\u00c3O",
            int(log.qtd_abastecimentos or 0),
            log.tipos_abastecimento_resumo or "",
            float(log.total_litros_abastecidos or 0),
            float(log.total_valor_abastecido or 0),
            None,
            None,
            "SIM" if log.teve_limpeza else "N\u00c3O",
            int(log.qtd_limpezas or 0),
            _limpezas_registros_resumo(log),
            _limpezas_datas_resumo(log),
            _limpezas_tipos_resumo(log),
            float(log.total_valor_limpeza or 0),
            "SIM" if log.assinatura_piloto else "N\u00c3O",
            log.observacao or "",
        ])

    last_row = ws.max_row
    last_col = ws.max_column
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

    col_data = 1
    col_check = 6
    col_km_ini = 7
    col_km_fim = 8
    col_km_rod = 9
    col_abast = 10
    col_qtd_ab = 11
    col_litros = 13
    col_valor = 14
    col_val_litro = 15
    col_custo_km = 16
    col_limpeza = 17
    col_qtd_limpeza = 18
    col_valor_limpeza = 22
    col_ass = 23

    for row in range(2, last_row + 1):
        ws.cell(row, col_km_rod).value = (
            f'=IF({get_column_letter(col_km_fim)}{row}="","",'
            f'{get_column_letter(col_km_fim)}{row}-{get_column_letter(col_km_ini)}{row})'
        )
        ws.cell(row, col_val_litro).value = (
            f'=IF(OR({get_column_letter(col_litros)}{row}="",{get_column_letter(col_litros)}{row}=0),"",'
            f'{get_column_letter(col_valor)}{row}/{get_column_letter(col_litros)}{row})'
        )
        ws.cell(row, col_custo_km).value = (
            f'=IF(OR({get_column_letter(col_km_rod)}{row}="",{get_column_letter(col_km_rod)}{row}=0),"",'
            f'{get_column_letter(col_valor)}{row}/{get_column_letter(col_km_rod)}{row})'
        )

        ws.cell(row, col_km_ini).number_format = "#,##0.00"
        ws.cell(row, col_km_fim).number_format = "#,##0.00"
        ws.cell(row, col_km_rod).number_format = "#,##0.00"
        ws.cell(row, col_qtd_ab).number_format = "0"
        ws.cell(row, col_litros).number_format = "#,##0.00"
        ws.cell(row, col_valor).number_format = '"R$" #,##0.00'
        ws.cell(row, col_val_litro).number_format = '"R$" #,##0.00'
        ws.cell(row, col_custo_km).number_format = '"R$" #,##0.00'
        ws.cell(row, col_qtd_limpeza).number_format = "0"
        ws.cell(row, col_valor_limpeza).number_format = '"R$" #,##0.00'

        for col in range(1, last_col + 1):
            current = ws.cell(row, col)
            current.border = border
            current.alignment = Alignment(vertical="center", wrap_text=True)

    center_cols(ws, cols=[col_data, col_check, col_abast, col_limpeza, col_ass], start_row=2, end_row=last_row)

    stripe_fill = PatternFill("solid", fgColor="F2F2F2")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    highlight_fill = PatternFill("solid", fgColor="FFF2CC")
    highlight_cols = {col_litros, col_valor, col_val_litro, col_custo_km, col_valor_limpeza}

    for row in range(2, last_row + 1):
        row_fill = stripe_fill if (row % 2 == 0) else white_fill
        for col in range(1, last_col + 1):
            current = ws.cell(row, col)
            current.fill = row_fill
            if col in highlight_cols and current.value not in (None, ""):
                current.fill = highlight_fill

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(color="006100", bold=True)
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006", bold=True)

    for col in (col_check, col_abast, col_limpeza, col_ass):
        col_letter = get_column_letter(col)
        cell_range = f"{col_letter}2:{col_letter}{last_row}"
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'{col_letter}2="SIM"'], fill=green_fill, font=green_font),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'{col_letter}2="N\u00c3O"'], fill=red_fill, font=red_font),
        )

    auto_width(ws, last_col, min_w=10, max_w=48)

    ws2 = wb.create_sheet("Resumo (M\u00e9dias)")

    title_fill = PatternFill("solid", fgColor="0B2F4F")
    card_fill = PatternFill("solid", fgColor="E7EFF8")
    info_fill = PatternFill("solid", fgColor="F8F8F8")
    kpi_fill = PatternFill("solid", fgColor="D9E1F2")

    title_font = Font(color="FFFFFF", bold=True, size=14)
    section_font = Font(color="1F4E79", bold=True, size=12)
    label_font = Font(color="1F4E79", bold=True)
    small_font = Font(color="404040", size=10)
    big_font = Font(color="1F4E79", bold=True, size=16)

    has_data = last_row >= 2
    rng_km_rod = f"Detalhamento!{get_column_letter(col_km_rod)}2:{get_column_letter(col_km_rod)}{last_row}"
    rng_valor = f"Detalhamento!{get_column_letter(col_valor)}2:{get_column_letter(col_valor)}{last_row}"
    rng_litros = f"Detalhamento!{get_column_letter(col_litros)}2:{get_column_letter(col_litros)}{last_row}"
    rng_vl = f"Detalhamento!{get_column_letter(col_val_litro)}2:{get_column_letter(col_val_litro)}{last_row}"
    rng_ckm = f"Detalhamento!{get_column_letter(col_custo_km)}2:{get_column_letter(col_custo_km)}{last_row}"
    rng_qtd_ab = f"Detalhamento!{get_column_letter(col_qtd_ab)}2:{get_column_letter(col_qtd_ab)}{last_row}"
    rng_valor_limpeza = f"Detalhamento!{get_column_letter(col_valor_limpeza)}2:{get_column_letter(col_valor_limpeza)}{last_row}"

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "RESUMO - M\u00c9DIAS E INDICADORES (FROTA)"
    ws2["A1"].fill = title_fill
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:H2")
    ws2["A2"] = "Painel de custos, consumo e produtividade com base nos registros da aba Detalhamento."
    ws2["A2"].font = small_font
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 18

    ws2.merge_cells("A4:C4")
    ws2.merge_cells("A5:C6")
    ws2["A4"] = "Total de Registros"
    ws2["A5"] = (f"={max(0, last_row - 1)}" if has_data else "")
    ws2["A4"].font = label_font
    ws2["A5"].font = big_font
    ws2["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A4"].fill = kpi_fill
    ws2["A5"].fill = card_fill

    ws2.merge_cells("D4:F4")
    ws2.merge_cells("D5:F6")
    ws2["D4"] = "Total Abastecido (R$)"
    ws2["D5"] = (f"=SUM({rng_valor})" if has_data else "")
    ws2["D4"].font = label_font
    ws2["D5"].font = big_font
    ws2["D4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["D5"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["D4"].fill = kpi_fill
    ws2["D5"].fill = card_fill
    ws2["D5"].number_format = '"R$" #,##0.00'

    ws2.merge_cells("G4:H4")
    ws2.merge_cells("G5:H6")
    ws2["G4"] = "Total de Litros"
    ws2["G5"] = (f"=SUM({rng_litros})" if has_data else "")
    ws2["G4"].font = label_font
    ws2["G5"].font = big_font
    ws2["G4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["G5"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["G4"].fill = kpi_fill
    ws2["G5"].fill = card_fill
    ws2["G5"].number_format = "#,##0.00"

    for row in range(4, 7):
        for col in range(1, 9):
            ws2.cell(row, col).border = border
            ws2.cell(row, col).alignment = Alignment(wrap_text=True, vertical="center")

    ws2["A8"] = "M\u00c9DIAS PRINCIPAIS"
    ws2["A8"].font = section_font
    ws2.merge_cells("A8:H8")

    ws2["A9"] = "M\u00e9trica"
    ws2["D9"] = "Resultado"
    ws2["F9"] = "Como interpretar"
    ws2.merge_cells("A9:C9")
    ws2.merge_cells("D9:E9")
    ws2.merge_cells("F9:H9")

    for cell in ws2["A9:H9"][0]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws2.row_dimensions[9].height = 22

    metrics = [
        ("M\u00e9dia de KM Rodado", f'=AVERAGEIF({rng_km_rod},">0")', "M\u00e9dia de km por registro."),
        ("M\u00e9dia Valor Abastecimento (R$)", f'=AVERAGEIF({rng_valor},">0")', "Valor m\u00e9dio por abastecimento."),
        ("M\u00e9dia Valor por Litro (R$)", f'=AVERAGEIF({rng_vl},">0")', "Pre\u00e7o m\u00e9dio pago por litro."),
        ("M\u00e9dia Custo por KM (R$)", f'=AVERAGEIF({rng_ckm},">0")', "Custo m\u00e9dio por km."),
        ("Qtd. de Abastecimentos", f"=SUM({rng_qtd_ab})", "Quantidade total registrada."),
        ("Total Limpeza (R$)", f"=SUM({rng_valor_limpeza})", "Valor total registrado em limpezas."),
    ]

    base_row = 10
    for offset, (name, formula, tip) in enumerate(metrics):
        row = base_row + offset
        ws2.merge_cells(f"A{row}:C{row}")
        ws2.merge_cells(f"D{row}:E{row}")
        ws2.merge_cells(f"F{row}:H{row}")

        ws2[f"A{row}"] = name
        ws2[f"D{row}"] = (formula if has_data else "")
        ws2[f"F{row}"] = tip
        ws2[f"A{row}"].font = Font(bold=True, color="1F4E79")
        ws2[f"F{row}"].font = small_font

        for col in range(1, 9):
            current = ws2.cell(row, col)
            current.border = border
            current.alignment = Alignment(vertical="center", wrap_text=True)
            current.fill = info_fill if (offset % 2 == 0) else PatternFill("solid", fgColor="FFFFFF")

    if has_data:
        ws2[f"D{base_row}"].number_format = "#,##0.00"
        ws2[f"D{base_row + 1}"].number_format = '"R$" #,##0.00'
        ws2[f"D{base_row + 2}"].number_format = '"R$" #,##0.00'
        ws2[f"D{base_row + 3}"].number_format = '"R$" #,##0.00'
        ws2[f"D{base_row + 4}"].number_format = "0"
        ws2[f"D{base_row + 5}"].number_format = '"R$" #,##0.00'

    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")

    alert_row = base_row + len(metrics) + 2
    ws2[f"A{alert_row}"] = "ALERTAS (LEITURA R\u00c1PIDA)"
    ws2[f"A{alert_row}"].font = section_font
    ws2.merge_cells(f"A{alert_row}:H{alert_row}")

    limiar_ckm = 1.50
    ws2.merge_cells(f"A{alert_row + 1}:E{alert_row + 1}")
    ws2.merge_cells(f"F{alert_row + 1}:H{alert_row + 1}")
    ws2[f"A{alert_row + 1}"] = f"Custo por KM acima de R$ {limiar_ckm:.2f}"
    ws2[f"F{alert_row + 1}"] = (
        f'=IF(AVERAGEIF({rng_ckm},">0")>{limiar_ckm},"ATEN\u00c7\u00c3O: ALTO","OK")'
        if has_data else ""
    )
    ws2[f"A{alert_row + 1}"].font = Font(bold=True, color="1F4E79")
    ws2[f"F{alert_row + 1}"].font = Font(bold=True)

    for col in range(1, 9):
        current = ws2.cell(alert_row + 1, col)
        current.border = border
        current.alignment = Alignment(vertical="center", wrap_text=True)
        current.fill = warn_fill

    alert_range = f"F{alert_row + 1}:H{alert_row + 1}"
    ws2.conditional_formatting.add(
        alert_range,
        FormulaRule(
            formula=[f'F{alert_row + 1}="OK"'],
            fill=ok_fill,
            font=Font(color="006100", bold=True),
        ),
    )
    ws2.conditional_formatting.add(
        alert_range,
        FormulaRule(
            formula=[f'F{alert_row + 1}<>"OK"'],
            fill=bad_fill,
            font=Font(color="9C0006", bold=True),
        ),
    )

    ws2.freeze_panes = "A10"
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 22
    ws2.column_dimensions["G"].width = 14
    ws2.column_dimensions["H"].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"logs_veiculos_{_now_brazil().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _parse_date_filter(raw_value, *, end_of_day=False):
    raw_value = (raw_value or "").strip()
    if not raw_value:
        return None

    try:
        parsed = datetime.strptime(raw_value, "%Y-%m-%d")
    except ValueError:
        return None

    if end_of_day:
        parsed = parsed.replace(hour=23, minute=59, second=59, microsecond=999999)
    return parsed


def _parse_decimal_filter(raw_value):
    try:
        return _parse_decimal_form(raw_value)
    except ValueError:
        return None


def _limpeza_filter_conditions(
    *,
    limpeza_realizada="",
    tipo_limpeza="",
    data_limpeza_inicio="",
    data_limpeza_fim="",
    valor_limpeza_min="",
    valor_limpeza_max="",
    force_realizada=None,
):
    conditions = []
    status = (limpeza_realizada or "").strip()
    tipo = (tipo_limpeza or "").strip().lower()

    if force_realizada is not None:
        conditions.append(LimpezaVeiculo.limpeza_realizada.is_(force_realizada))
    elif status in {"1", "com_limpeza"}:
        conditions.append(LimpezaVeiculo.limpeza_realizada.is_(True))
    elif status == "0":
        conditions.append(LimpezaVeiculo.limpeza_realizada.is_(False))

    if tipo in {"completa", "ducha"}:
        conditions.append(LimpezaVeiculo.tipo_limpeza == tipo)

    dt_inicio = _parse_date_filter(data_limpeza_inicio)
    if dt_inicio is not None:
        conditions.append(LimpezaVeiculo.data_hora >= dt_inicio)

    dt_fim = _parse_date_filter(data_limpeza_fim, end_of_day=True)
    if dt_fim is not None:
        conditions.append(LimpezaVeiculo.data_hora <= dt_fim)

    valor_min = _parse_decimal_filter(valor_limpeza_min)
    if valor_min is not None:
        conditions.append(LimpezaVeiculo.valor_total >= valor_min)

    valor_max = _parse_decimal_filter(valor_limpeza_max)
    if valor_max is not None:
        conditions.append(LimpezaVeiculo.valor_total <= valor_max)

    return conditions


def _has_limpeza_filters(
    *,
    limpeza_realizada="",
    tipo_limpeza="",
    data_limpeza_inicio="",
    data_limpeza_fim="",
    valor_limpeza_min="",
    valor_limpeza_max="",
):
    status = (limpeza_realizada or "").strip()
    tipo = (tipo_limpeza or "").strip().lower()
    return (
        status in {"com_limpeza", "1", "0", "sem_limpeza"}
        or tipo in {"completa", "ducha"}
        or bool((data_limpeza_inicio or "").strip())
        or bool((data_limpeza_fim or "").strip())
        or bool((valor_limpeza_min or "").strip())
        or bool((valor_limpeza_max or "").strip())
    )


def _limpeza_exists_expression(*, conditions=None):
    return (
        db.session.query(LimpezaVeiculo.id)
        .filter(
            LimpezaVeiculo.log_veiculo_id == LogVeiculo.id,
            *(conditions or []),
        )
        .exists()
    )


def _sum_veiculos_logs_abastecido(
    *,
    user=None,
    q="",
    data_inicio="",
    data_fim="",
    limpeza_realizada="",
    tipo_limpeza="",
    data_limpeza_inicio="",
    data_limpeza_fim="",
    valor_limpeza_min="",
    valor_limpeza_max="",
):
    log_ids = (
        _build_veiculos_logs_query(
            user=user,
            q=q,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limpeza_realizada=limpeza_realizada,
            tipo_limpeza=tipo_limpeza,
            data_limpeza_inicio=data_limpeza_inicio,
            data_limpeza_fim=data_limpeza_fim,
            valor_limpeza_min=valor_limpeza_min,
            valor_limpeza_max=valor_limpeza_max,
            include_options=False,
            include_order=False,
        )
        .with_entities(LogVeiculo.id)
        .subquery()
    )
    total = (
        db.session.query(db.func.coalesce(db.func.sum(Abastecimento.valor_total), 0))
        .join(log_ids, log_ids.c.id == Abastecimento.log_veiculo_id)
        .scalar()
    )
    return total or 0


def _sum_veiculos_logs_limpezas_realizadas(
    *,
    user=None,
    q="",
    data_inicio="",
    data_fim="",
    limpeza_realizada="",
    tipo_limpeza="",
    data_limpeza_inicio="",
    data_limpeza_fim="",
    valor_limpeza_min="",
    valor_limpeza_max="",
):
    status = (limpeza_realizada or "").strip()
    if status in {"0", "sem_limpeza"}:
        return {"quantidade": 0, "valor_total": 0}

    log_ids = (
        _build_veiculos_logs_query(
            user=user,
            q=q,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limpeza_realizada=limpeza_realizada,
            tipo_limpeza=tipo_limpeza,
            data_limpeza_inicio=data_limpeza_inicio,
            data_limpeza_fim=data_limpeza_fim,
            valor_limpeza_min=valor_limpeza_min,
            valor_limpeza_max=valor_limpeza_max,
            include_options=False,
            include_order=False,
        )
        .with_entities(LogVeiculo.id)
        .subquery()
    )
    conditions = _limpeza_filter_conditions(
        tipo_limpeza=tipo_limpeza,
        data_limpeza_inicio=data_limpeza_inicio,
        data_limpeza_fim=data_limpeza_fim,
        valor_limpeza_min=valor_limpeza_min,
        valor_limpeza_max=valor_limpeza_max,
        force_realizada=True,
    )
    row = (
        db.session.query(
            db.func.count(LimpezaVeiculo.id).label("quantidade"),
            db.func.coalesce(db.func.sum(LimpezaVeiculo.valor_total), 0).label("valor_total"),
        )
        .join(log_ids, log_ids.c.id == LimpezaVeiculo.log_veiculo_id)
        .filter(*conditions)
        .first()
    )
    return {
        "quantidade": int((row.quantidade if row else 0) or 0),
        "valor_total": (row.valor_total if row else 0) or 0,
    }


def _build_veiculos_logs_query(
    *,
    user=None,
    q="",
    data_inicio="",
    data_fim="",
    limpeza_realizada="",
    tipo_limpeza="",
    data_limpeza_inicio="",
    data_limpeza_fim="",
    valor_limpeza_min="",
    valor_limpeza_max="",
    include_options=True,
    include_order=True,
):
    ultima_movimentacao_subq = _ultima_movimentacao_log_subquery()
    ultima_movimentacao_expr = db.func.coalesce(
        ultima_movimentacao_subq.c.ultima_movimentacao_em,
        LogVeiculo.data_registro,
    )

    query = (
        LogVeiculo.query
        .outerjoin(ultima_movimentacao_subq, ultima_movimentacao_subq.c.log_id == LogVeiculo.id)
        .join(Veiculos, LogVeiculo.veiculo_id == Veiculos.id)
        .outerjoin(Pilotos, LogVeiculo.piloto_id == Pilotos.id)
        .outerjoin(Equipe, LogVeiculo.equipe_id == Equipe.id)
    )
    if include_options:
        query = query.options(
            joinedload(LogVeiculo.veiculo),
            joinedload(LogVeiculo.piloto),
            joinedload(LogVeiculo.equipe),
            selectinload(LogVeiculo.abastecimentos_detalhados),
            selectinload(LogVeiculo.limpezas_detalhadas),
        )
    if user is not None:
        if getattr(user, "tipo_usuario", None) == EQUIPE_OCEANO_USER_TYPE:
            equipe = _equipe_oceano_logada(user)
            if not equipe:
                query = query.filter(db.false())
            else:
                query = query.filter(
                    db.or_(
                        LogVeiculo.equipe_id == equipe.id,
                        _veiculo_equipe_operacional_filter(equipe, user),
                    )
                )
        else:
            query = apply_prefeitura_scope(query, user, Veiculos.prefeitura_id)

    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                id_search_clause(LogVeiculo.id, q),
                id_search_clause(Veiculos.id, q),
                id_search_clause(Pilotos.id, q),
                id_search_clause(Equipe.id, q),
                Veiculos.modelo.ilike(like),
                Veiculos.placa.ilike(like),
                Veiculos.responsavel.ilike(like),
                Pilotos.nome_piloto.ilike(like),
                Equipe.nome_equipe.ilike(like),
            )
        )

    if data_inicio:
        try:
            dt_ini = datetime.strptime(data_inicio, "%Y-%m-%d")
            query = query.filter(ultima_movimentacao_expr >= dt_ini)
        except ValueError:
            pass

    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(ultima_movimentacao_expr <= dt_fim)
        except ValueError:
            pass

    if limpeza_realizada in {"0", "sem_limpeza"}:
        query = query.filter(
            ~_limpeza_exists_expression(
                conditions=_limpeza_filter_conditions(
                    tipo_limpeza=tipo_limpeza,
                    data_limpeza_inicio=data_limpeza_inicio,
                    data_limpeza_fim=data_limpeza_fim,
                    valor_limpeza_min=valor_limpeza_min,
                    valor_limpeza_max=valor_limpeza_max,
                    force_realizada=True,
                )
            )
        )
    elif _has_limpeza_filters(
        limpeza_realizada=limpeza_realizada,
        tipo_limpeza=tipo_limpeza,
        data_limpeza_inicio=data_limpeza_inicio,
        data_limpeza_fim=data_limpeza_fim,
        valor_limpeza_min=valor_limpeza_min,
        valor_limpeza_max=valor_limpeza_max,
    ):
        query = query.filter(
            _limpeza_exists_expression(
                conditions=_limpeza_filter_conditions(
                    limpeza_realizada=limpeza_realizada,
                    tipo_limpeza=tipo_limpeza,
                    data_limpeza_inicio=data_limpeza_inicio,
                    data_limpeza_fim=data_limpeza_fim,
                    valor_limpeza_min=valor_limpeza_min,
                    valor_limpeza_max=valor_limpeza_max,
                )
            )
        )

    if include_order:
        query = query.order_by(ultima_movimentacao_expr.desc(), LogVeiculo.id.desc())

    return query


def _build_veiculos_limpezas_query(
    *,
    user=None,
    q="",
    limpeza_realizada="",
    tipo_limpeza="",
    operacao="",
    data_limpeza_inicio="",
    data_limpeza_fim="",
    data_registro_inicio="",
    data_registro_fim="",
    include_options=True,
    include_order=True,
):
    query = (
        LimpezaVeiculo.query
        .join(Veiculos, LimpezaVeiculo.veiculo_id == Veiculos.id)
        .outerjoin(LogVeiculo, LimpezaVeiculo.log_veiculo_id == LogVeiculo.id)
        .outerjoin(Pilotos, LimpezaVeiculo.piloto_id == Pilotos.id)
        .outerjoin(Equipe, LimpezaVeiculo.equipe_id == Equipe.id)
    )
    if include_options:
        query = query.options(
            joinedload(LimpezaVeiculo.veiculo).joinedload(Veiculos.equipe),
            joinedload(LimpezaVeiculo.log_pai),
            joinedload(LimpezaVeiculo.piloto),
            joinedload(LimpezaVeiculo.equipe),
        )

    if user is not None:
        if getattr(user, "tipo_usuario", None) == EQUIPE_OCEANO_USER_TYPE:
            equipe = _equipe_oceano_logada(user)
            if not equipe:
                query = query.filter(db.false())
            else:
                query = query.filter(
                    db.or_(
                        LimpezaVeiculo.equipe_id == equipe.id,
                        _veiculo_equipe_operacional_filter(equipe, user),
                    )
                )
        else:
            query = apply_prefeitura_scope(query, user, Veiculos.prefeitura_id)

    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                id_search_clause(LimpezaVeiculo.id, q),
                id_search_clause(Veiculos.id, q),
                id_search_clause(LogVeiculo.id, q),
                id_search_clause(Pilotos.id, q),
                id_search_clause(Equipe.id, q),
                Veiculos.modelo.ilike(like),
                Veiculos.placa.ilike(like),
                Veiculos.responsavel.ilike(like),
                Pilotos.nome_piloto.ilike(like),
                Equipe.nome_equipe.ilike(like),
                LimpezaVeiculo.observacao.ilike(like),
            )
        )

    if limpeza_realizada in {"1", "realizada", "sim"}:
        query = query.filter(LimpezaVeiculo.limpeza_realizada.is_(True))
    elif limpeza_realizada in {"0", "nao_realizada", "nao"}:
        query = query.filter(LimpezaVeiculo.limpeza_realizada.is_(False))

    if tipo_limpeza in {"completa", "ducha"}:
        query = query.filter(LimpezaVeiculo.tipo_limpeza == tipo_limpeza)

    if operacao:
        query = query.filter(db.func.upper(db.func.coalesce(Veiculos.operacao, "")) == operacao)

    dt_limpeza_inicio = _parse_date_filter(data_limpeza_inicio)
    if dt_limpeza_inicio is not None:
        query = query.filter(LimpezaVeiculo.data_hora >= dt_limpeza_inicio)

    dt_limpeza_fim = _parse_date_filter(data_limpeza_fim, end_of_day=True)
    if dt_limpeza_fim is not None:
        query = query.filter(LimpezaVeiculo.data_hora <= dt_limpeza_fim)

    dt_registro_inicio = _parse_date_filter(data_registro_inicio)
    if dt_registro_inicio is not None:
        query = query.filter(LimpezaVeiculo.data_registro >= dt_registro_inicio)

    dt_registro_fim = _parse_date_filter(data_registro_fim, end_of_day=True)
    if dt_registro_fim is not None:
        query = query.filter(LimpezaVeiculo.data_registro <= dt_registro_fim)

    if include_order:
        query = query.order_by(
            LimpezaVeiculo.data_hora.desc(),
            LimpezaVeiculo.data_registro.desc(),
            LimpezaVeiculo.id.desc(),
        )

    return query


def _build_veiculos_limpezas_resumo(query):
    subq = query.with_entities(LimpezaVeiculo.id).subquery()
    row = (
        db.session.query(
            db.func.count(LimpezaVeiculo.id).label("total"),
            db.func.coalesce(db.func.sum(LimpezaVeiculo.valor_total), 0).label("valor_total"),
            db.func.coalesce(db.func.sum(case((LimpezaVeiculo.limpeza_realizada.is_(True), 1), else_=0)), 0).label("realizadas"),
            db.func.coalesce(db.func.sum(case((LimpezaVeiculo.limpeza_realizada.is_(False), 1), else_=0)), 0).label("nao_realizadas"),
        )
        .join(subq, subq.c.id == LimpezaVeiculo.id)
        .first()
    )
    return {
        "total": int((row.total if row else 0) or 0),
        "valor_total": (row.valor_total if row else 0) or 0,
        "realizadas": int((row.realizadas if row else 0) or 0),
        "nao_realizadas": int((row.nao_realizadas if row else 0) or 0),
    }
