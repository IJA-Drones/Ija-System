from datetime import date, datetime
from io import BytesIO
import os
import re
import uuid

from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import and_, case, func, or_
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models import Equipe, OrdemServico, OrdemServicoEquipeUvis, Solicitacao, Usuario
from app.shared.access import (
    ADMIN_PANEL_EDIT_TYPES,
    ADMIN_PANEL_VIEW_TYPES,
    apply_prefeitura_scope,
    apply_regiao_scope,
    apply_solicitacao_prefeitura_scope,
)
from app.shared.os_history_filters import apply_retorno_automatico_filter
from app.shared.query_filters import id_search_clause
from app.shared.uploads import allowed_file, get_upload_folder

APPROVAL_STATUSES = {"APROVADO", "APROVADO COM RECOMENDAÇÕES"}
HISTORICO_OS_ANDAMENTO_STATUSES = (
    "APROVADO",
    "APROVADA",
    "APROVADO COM RECOMENDACOES",
    "APROVADA COM RECOMENDACOES",
    "APROVADO COM RECOMENDAÇÕES",
    "APROVADA COM RECOMENDAÇÕES",
)
HISTORICO_OS_CONCLUIDAS_STATUSES = ("CONCLUIDO", "CONCLUÍDO")


def _parse_filter_date(value: str):
    raw_value = (value or "").strip()
    if not raw_value:
        return None
    try:
        return datetime.strptime(raw_value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _apply_data_agendamento_range(query, filtro_data_ini: str = "", filtro_data_fim: str = ""):
    data_ini = _parse_filter_date(filtro_data_ini)
    data_fim = _parse_filter_date(filtro_data_fim)

    if data_ini and data_fim and data_ini > data_fim:
        data_ini, data_fim = data_fim, data_ini

    if data_ini:
        query = query.filter(Solicitacao.data_agendamento >= data_ini)

    if data_fim:
        query = query.filter(Solicitacao.data_agendamento <= data_fim)

    return query


def _apply_endereco_filter(query, filtro_endereco: str = ""):
    termo = (filtro_endereco or "").strip()
    if not termo:
        return query

    tokens = [token for token in re.split(r"[\s,;/\-]+", termo) if token]
    if not tokens:
        return query

    for token in tokens:
        like = f"%{token}%"
        query = query.filter(
            or_(
                Solicitacao.logradouro.ilike(like),
                Solicitacao.numero.ilike(like),
                Solicitacao.complemento.ilike(like),
                Solicitacao.bairro.ilike(like),
                Solicitacao.cidade.ilike(like),
                Solicitacao.uf.ilike(like),
                Solicitacao.cep.ilike(like),
            )
        )

    return query


def can_access_admin_panel(user) -> bool:
    return getattr(user, "tipo_usuario", None) in ADMIN_PANEL_VIEW_TYPES


def can_edit_admin_panel(user) -> bool:
    return getattr(user, "tipo_usuario", None) in ADMIN_PANEL_EDIT_TYPES


def get_google_maps_key():
    return current_app.config.get("KEY_API_GOOGLE_MAPS") or os.getenv("KEY_API_GOOGLE_MAPS")


def build_uvis_select(user=None):
    query = Usuario.query.filter_by(tipo_usuario="uvis")
    if user is not None:
        query = apply_prefeitura_scope(query, user, Usuario.prefeitura_id)
        query = apply_regiao_scope(query, user, Usuario.regiao)
    return query.order_by(Usuario.nome_uvis.asc()).all()


def build_active_teams(user=None):
    query = Equipe.query.filter(Equipe.ativa.is_(True))
    if user is not None:
        query = apply_prefeitura_scope(query, user, Equipe.prefeitura_id)
        query = apply_regiao_scope(query, user, Equipe.regiao)
    return query.order_by(Equipe.regiao.asc(), Equipe.nome_equipe.asc()).all()


def build_status_order():
    return case(
        {
            "PENDENTE": 1,
            "EM ANALISE": 2,
            "EM ANÁLISE": 2,
            "APROVADO COM RECOMENDACOES": 3,
            "APROVADO COM RECOMENDAÇÕES": 3,
            "APROVADO": 4,
            "NEGADO": 5,
            "CONCLUIDO": 6,
            "CONCLUÍDO": 6,
        },
        value=Solicitacao.status,
        else_=99,
    )


def build_admin_dashboard_query(
    user,
    filtro_status: str,
    filtro_unidade: str,
    filtro_regiao: str,
    filtro_apoio_cet: str,
    filtro_protocolo: str,
    filtro_endereco: str = "",
    filtro_tipo_visita: str = "",
    filtro_tipo_imovel: str = "",
    filtro_foco: str = "",
    filtro_data_ini: str = "",
    filtro_data_fim: str = "",
    filtro_retorno_automatico: str = "",
):
    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe),
        )
        .join(Usuario)
        .filter(Solicitacao.status != "CANCELADO")
    )
    query = apply_solicitacao_prefeitura_scope(query, user)
    query = apply_regiao_scope(query, user, Usuario.regiao)

    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_unidade:
        query = query.filter(
            or_(
                id_search_clause(Usuario.id, filtro_unidade),
                Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"),
            )
        )

    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))

    if filtro_apoio_cet == "SIM":
        query = query.filter(Solicitacao.apoio_cet.is_(True))
    elif filtro_apoio_cet == "NAO":
        query = query.filter(Solicitacao.apoio_cet.is_(False))

    if filtro_protocolo:
        query = query.filter(
            or_(
                id_search_clause(Solicitacao.id, filtro_protocolo, prefixes=("id", "os")),
                Solicitacao.protocolo.ilike(f"%{filtro_protocolo}%"),
                Solicitacao.ordem_servico.has(
                    OrdemServico.identificador_os.ilike(f"%{filtro_protocolo}%")
                ),
                Solicitacao.ordem_servico_equipe_uvis.has(
                    OrdemServicoEquipeUvis.identificador_os.ilike(f"%{filtro_protocolo}%")
                ),
            )
        )

    query = _apply_endereco_filter(query, filtro_endereco)

    if filtro_tipo_visita:
        query = query.filter(Solicitacao.tipo_visita == filtro_tipo_visita)

    if filtro_tipo_imovel:
        query = query.filter(Solicitacao.tipo_imovel == filtro_tipo_imovel)

    if filtro_foco:
        query = query.filter(Solicitacao.foco == filtro_foco)

    query = apply_retorno_automatico_filter(query, filtro_retorno_automatico)

    return _apply_data_agendamento_range(query, filtro_data_ini, filtro_data_fim)


def build_admin_canceladas_query(
    user,
    filtro_unidade: str,
    filtro_regiao: str,
    filtro_foco: str,
    filtro_protocolo: str,
    filtro_endereco: str = "",
    filtro_tipo_visita: str = "",
    filtro_tipo_imovel: str = "",
    filtro_data_ini: str = "",
    filtro_data_fim: str = "",
):
    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe),
        )
        .join(Usuario)
        .filter(Solicitacao.status == "CANCELADO")
    )
    query = apply_solicitacao_prefeitura_scope(query, user)
    query = apply_regiao_scope(query, user, Usuario.regiao)

    if filtro_unidade:
        query = query.filter(
            or_(
                id_search_clause(Usuario.id, filtro_unidade),
                Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"),
            )
        )

    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))

    if filtro_foco:
        query = query.filter(Solicitacao.foco == filtro_foco)

    if filtro_tipo_visita:
        query = query.filter(Solicitacao.tipo_visita == filtro_tipo_visita)

    if filtro_tipo_imovel:
        query = query.filter(Solicitacao.tipo_imovel == filtro_tipo_imovel)

    if filtro_protocolo:
        query = query.filter(
            or_(
                id_search_clause(Solicitacao.id, filtro_protocolo, prefixes=("id", "os")),
                Solicitacao.protocolo.ilike(f"%{filtro_protocolo}%"),
            )
        )

    query = _apply_endereco_filter(query, filtro_endereco)

    return _apply_data_agendamento_range(query, filtro_data_ini, filtro_data_fim)


def _has_equipe_uvis_os():
    return or_(
        Solicitacao.ordem_servico_equipe_uvis.has(),
        and_(
            Solicitacao.equipe_uvis_nome.isnot(None),
            func.trim(Solicitacao.equipe_uvis_nome) != "",
        ),
    )


def _apply_historico_equipe_filter(query, filtro_tipo_os: str, filtro_equipe: str):
    filtro_equipe = (filtro_equipe or "").strip()
    if not filtro_equipe:
        return query

    if filtro_tipo_os == "equipe_uvis":
        return query.filter(
            or_(
                Solicitacao.equipe_uvis_nome == filtro_equipe,
                Solicitacao.ordem_servico_equipe_uvis.has(
                    OrdemServicoEquipeUvis.equipe_uvis_nome == filtro_equipe
                ),
            )
        )

    try:
        equipe_id = int(filtro_equipe)
    except (TypeError, ValueError):
        return query.filter(
            or_(
                Solicitacao.equipe.has(Equipe.nome_equipe.ilike(f"%{filtro_equipe}%")),
                Solicitacao.ordem_servico.has(
                    OrdemServico.equipe.has(Equipe.nome_equipe.ilike(f"%{filtro_equipe}%"))
                ),
            )
        )

    return query.filter(
        or_(
            Solicitacao.equipe_id == equipe_id,
            Solicitacao.ordem_servico.has(OrdemServico.equipe_id == equipe_id),
        )
    )


def build_admin_historico_os_query(user, filtros, filtro_tipo_os: str, filtro_equipe: str = ""):
    query = build_admin_dashboard_query(
        user,
        filtro_status="",
        filtro_unidade=filtros["unidade"],
        filtro_regiao=filtros["regiao"],
        filtro_apoio_cet=filtros["apoio_cet"],
        filtro_protocolo=filtros["protocolo"],
        filtro_endereco=filtros["endereco"],
        filtro_tipo_visita=filtros["tipo_visita"],
        filtro_tipo_imovel=filtros["tipo_imovel"],
        filtro_foco=filtros["foco"],
        filtro_data_ini=filtros["data_ini"],
        filtro_data_fim=filtros["data_fim"],
        filtro_retorno_automatico=filtros["retorno_automatico"],
    ).options(
        db.selectinload(Solicitacao.ordem_servico).selectinload(OrdemServico.equipe),
        db.selectinload(Solicitacao.ordem_servico_equipe_uvis),
    )

    filtro_status_os = filtros["status"]
    if filtro_tipo_os == "equipe_uvis":
        query = query.filter(_has_equipe_uvis_os())
        if filtro_status_os == "EM_ANDAMENTO":
            query = query.filter(~Solicitacao.status.in_(HISTORICO_OS_CONCLUIDAS_STATUSES))
        elif filtro_status_os == "CONCLUIDAS":
            query = query.filter(Solicitacao.status.in_(HISTORICO_OS_CONCLUIDAS_STATUSES))
    else:
        if filtro_status_os == "EM_ANDAMENTO":
            query = query.filter(
                and_(
                    Solicitacao.status.in_(HISTORICO_OS_ANDAMENTO_STATUSES),
                    Solicitacao.equipe_id.isnot(None),
                )
            )
        elif filtro_status_os == "CONCLUIDAS":
            query = query.filter(Solicitacao.status.in_(HISTORICO_OS_CONCLUIDAS_STATUSES))
        else:
            query = query.filter(
                or_(
                    Solicitacao.status.in_(HISTORICO_OS_CONCLUIDAS_STATUSES),
                    and_(
                        Solicitacao.status.in_(HISTORICO_OS_ANDAMENTO_STATUSES),
                        Solicitacao.equipe_id.isnot(None),
                    ),
                )
            )

    return _apply_historico_equipe_filter(query, filtro_tipo_os, filtro_equipe)


def build_equipe_uvis_names_select(user=None):
    base_query = (
        Solicitacao.query
        .join(Usuario)
        .filter(Solicitacao.status != "CANCELADO")
    )
    if user is not None:
        base_query = apply_solicitacao_prefeitura_scope(base_query, user)
        base_query = apply_regiao_scope(base_query, user, Usuario.regiao)

    names = set()
    solicitacao_names = (
        base_query
        .with_entities(Solicitacao.equipe_uvis_nome)
        .filter(
            Solicitacao.equipe_uvis_nome.isnot(None),
            func.trim(Solicitacao.equipe_uvis_nome) != "",
        )
        .distinct()
        .all()
    )
    ordem_names = (
        base_query
        .join(OrdemServicoEquipeUvis, OrdemServicoEquipeUvis.solicitacao_id == Solicitacao.id)
        .with_entities(OrdemServicoEquipeUvis.equipe_uvis_nome)
        .filter(
            OrdemServicoEquipeUvis.equipe_uvis_nome.isnot(None),
            func.trim(OrdemServicoEquipeUvis.equipe_uvis_nome) != "",
        )
        .distinct()
        .all()
    )

    for (name,) in solicitacao_names + ordem_names:
        clean_name = (name or "").strip()
        if clean_name:
            names.add(clean_name)

    return sorted(names, key=str.lower)


def build_admin_export_query(
    user,
    filtro_status: str,
    filtro_unidade: str,
    filtro_regiao: str,
    filtro_apoio_cet: str,
    filtro_protocolo: str,
    filtro_endereco: str = "",
    filtro_tipo_visita: str = "",
    filtro_tipo_imovel: str = "",
    filtro_foco: str = "",
    filtro_data_ini: str = "",
    filtro_data_fim: str = "",
    filtro_retorno_automatico: str = "",
):
    query = (
        Solicitacao.query
        .join(Usuario)
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe),
        )
    )
    query = apply_solicitacao_prefeitura_scope(query, user)
    query = apply_regiao_scope(query, user, Usuario.regiao)
    query = query.filter(Solicitacao.status != "CANCELADO")

    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_unidade:
        query = query.filter(
            or_(
                id_search_clause(Usuario.id, filtro_unidade),
                Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"),
            )
        )

    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))

    if filtro_apoio_cet == "SIM":
        query = query.filter(Solicitacao.apoio_cet.is_(True))
    elif filtro_apoio_cet == "NAO":
        query = query.filter(Solicitacao.apoio_cet.is_(False))

    if filtro_protocolo:
        query = query.filter(
            or_(
                id_search_clause(Solicitacao.id, filtro_protocolo, prefixes=("id", "os")),
                Solicitacao.protocolo.ilike(f"%{filtro_protocolo}%"),
            )
        )

    query = _apply_endereco_filter(query, filtro_endereco)

    if filtro_tipo_visita:
        query = query.filter(Solicitacao.tipo_visita == filtro_tipo_visita)

    if filtro_tipo_imovel:
        query = query.filter(Solicitacao.tipo_imovel == filtro_tipo_imovel)

    if filtro_foco:
        query = query.filter(Solicitacao.foco == filtro_foco)

    query = apply_retorno_automatico_filter(query, filtro_retorno_automatico)

    query = _apply_data_agendamento_range(query, filtro_data_ini, filtro_data_fim)

    return query.order_by(Solicitacao.data_criacao.desc())


def build_admin_dashboard_export(
    user,
    filtro_status: str,
    filtro_unidade: str,
    filtro_regiao: str,
    filtro_apoio_cet: str,
    filtro_protocolo: str,
    filtro_endereco: str = "",
    filtro_tipo_visita: str = "",
    filtro_tipo_imovel: str = "",
    filtro_foco: str = "",
    filtro_data_ini: str = "",
    filtro_data_fim: str = "",
    filtro_retorno_automatico: str = "",
):
    pedidos = build_admin_export_query(
        user=user,
        filtro_status=filtro_status,
        filtro_unidade=filtro_unidade,
        filtro_regiao=filtro_regiao,
        filtro_apoio_cet=filtro_apoio_cet,
        filtro_protocolo=filtro_protocolo,
        filtro_endereco=filtro_endereco,
        filtro_tipo_visita=filtro_tipo_visita,
        filtro_tipo_imovel=filtro_tipo_imovel,
        filtro_foco=filtro_foco,
        filtro_data_ini=filtro_data_ini,
        filtro_data_fim=filtro_data_fim,
        filtro_retorno_automatico=filtro_retorno_automatico,
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio de Solicitacoes"

    headers = [
        "ID",
        "Unidade",
        "Regiao",
        "Equipe Responsavel",
        "Data Agendada",
        "Hora",
        "Endereco Completo",
        "CEP",
        "Coordenadas",
        "Foco",
        "Tipo Operacao",
        "Tipo Visita",
        "Tipo Imovel",
        "Altura",
        "Apoio CET?",
        "Observacao",
        "Status",
        "Protocolo",
        "Justificativa",
    ]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_number, pedido in enumerate(pedidos, start=2):
        uvis_nome = pedido.usuario.nome_uvis if pedido.usuario else "Nao informado"
        uvis_regiao = pedido.usuario.regiao if pedido.usuario else "Nao informado"

        equipe_nome = ""
        if getattr(pedido, "equipe", None):
            equipe_nome = pedido.equipe.nome_equipe or ""
        elif getattr(pedido, "equipe_id", None):
            equipe_nome = f"ID #{pedido.equipe_id}"

        endereco_completo = (
            f"{pedido.logradouro or ''}, {pedido.numero or ''} - "
            f"{pedido.bairro or ''} - "
            f"{(pedido.cidade or '')}/{(pedido.uf or '')}"
        )
        if pedido.complemento:
            endereco_completo += f" - {pedido.complemento}"

        coordenadas = ""
        if pedido.latitude or pedido.longitude:
            coordenadas = f"{pedido.latitude or ''}, {pedido.longitude or ''}"

        data_formatada = ""
        if pedido.data_agendamento:
            if isinstance(pedido.data_agendamento, (date, datetime)):
                data_formatada = pedido.data_agendamento.strftime("%d/%m/%Y")
            else:
                data_formatada = str(pedido.data_agendamento)

        row = [
            pedido.id,
            uvis_nome,
            uvis_regiao,
            equipe_nome,
            data_formatada,
            str(pedido.hora_agendamento or ""),
            endereco_completo,
            pedido.cep or "",
            coordenadas,
            pedido.foco,
            pedido.tipo_operacao or "",
            pedido.tipo_visita or "",
            pedido.tipo_imovel or "",
            pedido.altura_voo or "",
            "SIM" if pedido.apoio_cet else "NAO",
            pedido.observacao or "",
            pedido.status,
            pedido.protocolo or "",
            pedido.justificativa or "",
        ]

        for column, value in enumerate(row, start=1):
            cell = ws.cell(row=row_number, column=column, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _format_date_br(value):
    if not value:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%d/%m/%Y")
    return str(value)


def _format_datetime_br(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def _format_time(value):
    if not value:
        return ""
    return value.strftime("%H:%M") if hasattr(value, "strftime") else str(value)


def _historico_os_equipe_nome(pedido, filtro_tipo_os: str):
    if filtro_tipo_os == "equipe_uvis":
        ordem_uvis = getattr(pedido, "ordem_servico_equipe_uvis", None)
        if ordem_uvis and ordem_uvis.equipe_uvis_nome:
            return ordem_uvis.equipe_uvis_nome
        return pedido.equipe_uvis_nome or ""
    ordem = getattr(pedido, "ordem_servico", None)
    if ordem and getattr(ordem, "equipe", None):
        return ordem.equipe.nome_equipe or ""
    if getattr(pedido, "equipe", None):
        return pedido.equipe.nome_equipe or ""
    return ""


def _historico_os_identificador(pedido, filtro_tipo_os: str):
    if filtro_tipo_os == "equipe_uvis":
        ordem_uvis = getattr(pedido, "ordem_servico_equipe_uvis", None)
        return ordem_uvis.identificador_os if ordem_uvis else ""
    ordem = getattr(pedido, "ordem_servico", None)
    return ordem.identificador_os if ordem else ""


def _historico_os_situacao(pedido, filtro_tipo_os: str):
    ordem = (
        getattr(pedido, "ordem_servico_equipe_uvis", None)
        if filtro_tipo_os == "equipe_uvis"
        else getattr(pedido, "ordem_servico", None)
    )
    return ordem.situacao_aplicacao if ordem else ""


def _historico_os_respondido_por(pedido, filtro_tipo_os: str):
    ordem = (
        getattr(pedido, "ordem_servico_equipe_uvis", None)
        if filtro_tipo_os == "equipe_uvis"
        else getattr(pedido, "ordem_servico", None)
    )
    return ordem.respondido_por if ordem else ""


def _historico_os_respondido_em(pedido, filtro_tipo_os: str):
    ordem = (
        getattr(pedido, "ordem_servico_equipe_uvis", None)
        if filtro_tipo_os == "equipe_uvis"
        else getattr(pedido, "ordem_servico", None)
    )
    return _format_datetime_br(ordem.respondido_em) if ordem else ""


def build_admin_historico_os_export(user, filtros, filtro_tipo_os: str, filtro_equipe: str = ""):
    pedidos = (
        build_admin_historico_os_query(user, filtros, filtro_tipo_os, filtro_equipe)
        .order_by(build_status_order(), Solicitacao.data_criacao.desc(), Solicitacao.id.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historico OS"

    headers = [
        "ID",
        "Tipo OS",
        "Identificador OS",
        "Unidade",
        "Regiao",
        "Equipe/PLOA",
        "Status",
        "Situacao",
        "Data Agendada",
        "Hora Agendada",
        "Endereco Completo",
        "Latitude",
        "Longitude",
        "Foco",
        "Tipo Operacao",
        "Tipo Visita",
        "Tipo Imovel",
        "Apoio CET?",
        "Protocolo",
        "Respondido Por",
        "Respondido Em",
        "Observacao",
        "Justificativa",
    ]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    tipo_label = "OS Equipe UVIS" if filtro_tipo_os == "equipe_uvis" else "OS Piloto"
    for row_number, pedido in enumerate(pedidos, start=2):
        usuario = getattr(pedido, "usuario", None)
        endereco_completo = (
            f"{pedido.logradouro or ''}, {pedido.numero or 'S/N'} - "
            f"{pedido.bairro or ''} - "
            f"{(pedido.cidade or '')}/{(pedido.uf or '')} - {pedido.cep or ''}"
        )
        if pedido.complemento:
            endereco_completo += f" - {pedido.complemento}"

        row = [
            pedido.id,
            tipo_label,
            _historico_os_identificador(pedido, filtro_tipo_os),
            usuario.nome_uvis if usuario else "",
            usuario.regiao if usuario else "",
            _historico_os_equipe_nome(pedido, filtro_tipo_os),
            pedido.status or "",
            _historico_os_situacao(pedido, filtro_tipo_os),
            _format_date_br(pedido.data_agendamento),
            _format_time(pedido.hora_agendamento),
            endereco_completo,
            pedido.latitude or "",
            pedido.longitude or "",
            pedido.foco or "",
            pedido.tipo_operacao or "",
            pedido.tipo_visita or "",
            pedido.tipo_imovel or "",
            "SIM" if pedido.apoio_cet else "NAO",
            pedido.protocolo or "",
            _historico_os_respondido_por(pedido, filtro_tipo_os),
            _historico_os_respondido_em(pedido, filtro_tipo_os),
            pedido.observacao or "",
            pedido.justificativa or "",
        ]

        for column, value in enumerate(row, start=1):
            cell = ws.cell(row=row_number, column=column, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 55)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    tipo_nome = "equipe_uvis" if filtro_tipo_os == "equipe_uvis" else "piloto"
    filename = f"historico_os_{tipo_nome}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename


def apply_admin_update_fields(pedido, form, *, user=None):
    pedido.protocolo = form.get("protocolo")
    pedido.status = form.get("status")
    pedido.justificativa = form.get("justificativa")
    pedido.latitude = form.get("latitude")
    pedido.longitude = form.get("longitude")

    equipe_id = form.get("equipe_id")
    if equipe_id in (None, "", "null", "undefined"):
        pedido.equipe_id = None
        equipe_nome = None
    else:
        try:
            equipe_id_int = int(equipe_id)
        except (TypeError, ValueError) as exc:
            raise ValueError("Equipe invalida.") from exc

        equipe_query = Equipe.query.filter(Equipe.id == equipe_id_int)
        if user is not None:
            equipe_query = apply_prefeitura_scope(equipe_query, user, Equipe.prefeitura_id)
            equipe_query = apply_regiao_scope(equipe_query, user, Equipe.regiao)
        equipe = equipe_query.first()
        if not equipe:
            raise ValueError("Equipe selecionada nao existe.")

        pedido.equipe_id = equipe_id_int
        equipe_nome = equipe.nome_equipe

    if pedido.status in APPROVAL_STATUSES and not pedido.equipe_id:
        raise ValueError("Para aprovar, selecione uma equipe responsavel.")

    return equipe_nome


def save_admin_attachment(pedido, uploaded_file):
    if not uploaded_file or not uploaded_file.filename:
        return None

    if not allowed_file(uploaded_file.filename):
        raise ValueError("Formato de arquivo nao permitido.")

    original_filename = secure_filename(uploaded_file.filename)
    ext = original_filename.rsplit(".", 1)[1].lower()
    unique_name = f"sol_{pedido.id}_{uuid.uuid4().hex}.{ext}"
    upload_folder = get_upload_folder()
    file_path = os.path.join(upload_folder, unique_name)

    uploaded_file.save(file_path)

    pedido.anexo_path = f"upload-files/{unique_name}"
    pedido.anexo_nome = original_filename
    return original_filename
