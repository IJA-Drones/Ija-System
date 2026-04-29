import json
import os
from datetime import date, datetime, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

from sqlalchemy import false
from flask import current_app, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload

from app.extensions import db
from app.models import Baterias, Drones, Equipe, EquipePiloto, Notificacao, Solicitacao, Usuario, Veiculos
from app.shared.access import (
    apply_prefeitura_scope,
    apply_regiao_scope,
    apply_solicitacao_prefeitura_scope,
    apply_solicitacao_regiao_scope,
)


AGENDA_VIEW_TYPES = {"admin", "operario", "visualizar", "regional", "prefeitura_admin"}
AGENDA_EXPORT_TYPES = {"admin", "operario", "visualizar", "regional", "prefeitura_admin", "uvis"}
NOTIFICATION_GLOBAL_VIEW_TYPES = {"admin", "operario", "visualizar", "prefeitura_admin"}
AGENDA_ROUTE_STATUSES = (
    "APROVADO",
    "APROVADO COM RECOMENDACOES",
    "APROVADO COM RECOMENDAÇÕES",
)
TZ_BR = ZoneInfo("America/Sao_Paulo")
AUTO_ALERT_PREVIEW_LIMIT = 4
AUTO_ALERT_BATTERY_WARNING_CYCLES = 200
AUTO_ALERT_BATTERY_CRITICAL_CYCLES = 250
AUTO_ALERT_VEHICLE_REVIEW_WARNING_KM = 500
AUTO_ALERT_DRONE_MAINTENANCE_STALE_DAYS = 90
MANUTENCAO_STATUS = "Em Manutenção"


def _parse_filter_date(value):
    raw_value = (value or "").strip()
    if not raw_value:
        return None
    try:
        return datetime.strptime(raw_value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _apply_agenda_date_range(query, data_ini=None, data_fim=None):
    dt_ini = _parse_filter_date(data_ini)
    dt_fim = _parse_filter_date(data_fim)

    if dt_ini and dt_fim and dt_ini > dt_fim:
        dt_ini, dt_fim = dt_fim, dt_ini

    if dt_ini:
        query = query.filter(Solicitacao.data_agendamento >= dt_ini)

    if dt_fim:
        query = query.filter(Solicitacao.data_agendamento <= dt_fim)

    return query


def _has_agenda_date_range(data_ini=None, data_fim=None):
    return _parse_filter_date(data_ini) is not None or _parse_filter_date(data_fim) is not None


def agenda_status_color(status):
    if status == "APROVADO":
        return "#198754"
    if status == "APROVADO COM RECOMENDAÇÕES":
        return "#ffa023"
    if status == "NEGADO":
        return "#dc3545"
    if status == "EM ANÁLISE":
        return "#e9fa05"
    return "#0d6efd"


def agora_brasilia_naive():
    return datetime.now(TZ_BR).replace(tzinfo=None)


def can_view_all_agenda(user):
    return getattr(user, "tipo_usuario", None) in AGENDA_VIEW_TYPES


def can_export_agenda(user):
    return getattr(user, "tipo_usuario", None) in AGENDA_EXPORT_TYPES


def can_view_all_notifications(user):
    return getattr(user, "tipo_usuario", None) in NOTIFICATION_GLOBAL_VIEW_TYPES


def is_piloto_agenda_user(user):
    return getattr(user, "tipo_usuario", None) == "piloto"


def _build_piloto_equipes_query(user):
    piloto_id = getattr(user, "piloto_id", None)
    if not piloto_id:
        return None

    return (
        db.session.query(EquipePiloto.equipe_id)
        .join(Equipe, Equipe.id == EquipePiloto.equipe_id)
        .filter(
            EquipePiloto.piloto_id == piloto_id,
            EquipePiloto.equipe_id.isnot(None),
            Equipe.ativa.is_(True),
        )
        .distinct()
    )


def apply_agenda_user_scope(query, user):
    query = apply_solicitacao_prefeitura_scope(query, user)
    query = apply_solicitacao_regiao_scope(query, user)

    if is_piloto_agenda_user(user):
        equipes_query = _build_piloto_equipes_query(user)
        if equipes_query is None:
            return query.filter(false())
        return query.filter(Solicitacao.equipe_id.in_(equipes_query))

    if not can_view_all_agenda(user):
        return query.filter(Solicitacao.usuario_id == user.id)

    return query


def get_agenda_google_maps_key():
    return current_app.config.get("Maps_KEY_FRONT") or os.getenv("KEY_API_GOOGLE_MAPS") or ""


def build_agenda_query(
    user,
    *,
    filtro_status=None,
    filtro_uvis_id=None,
    filtro_tipo_visita=None,
    filtro_tipo_imovel=None,
    filtro_foco=None,
    data_ini=None,
    data_fim=None,
    mes=None,
    ano=None,
):
    query = (
        Solicitacao.query
        .options(joinedload(Solicitacao.usuario))
        .filter(Solicitacao.status != "CANCELADO")
    )
    query = apply_agenda_user_scope(query, user)

    if is_piloto_agenda_user(user):
        filtro_uvis_id = None
    elif not can_view_all_agenda(user):
        filtro_uvis_id = None
    elif filtro_uvis_id:
        query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)

    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_tipo_visita:
        query = query.filter(Solicitacao.tipo_visita == filtro_tipo_visita)

    if filtro_tipo_imovel:
        query = query.filter(Solicitacao.tipo_imovel == filtro_tipo_imovel)

    if filtro_foco:
        query = query.filter(Solicitacao.foco == filtro_foco)

    query = _apply_agenda_date_range(query, data_ini, data_fim)

    if mes and ano and not _has_agenda_date_range(data_ini, data_fim):
        filtro_mesano = f"{ano}-{mes:02d}"
        if db.engine.name == "postgresql":
            query = query.filter(db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM") == filtro_mesano)
        else:
            query = query.filter(db.func.strftime("%Y-%m", Solicitacao.data_agendamento) == filtro_mesano)

    return query


def build_agenda_eventos(solicitacoes):
    agenda_eventos = []

    for solicitacao in solicitacoes:
        if not solicitacao.data_agendamento:
            continue

        data = solicitacao.data_agendamento.strftime("%Y-%m-%d")
        hora = solicitacao.hora_agendamento.strftime("%H:%M") if solicitacao.hora_agendamento else "00:00"
        uvis_nome = (solicitacao.usuario.nome_uvis if solicitacao.usuario else "UVIS") or "UVIS"

        lat = None
        lng = None
        try:
            if solicitacao.latitude is not None and str(solicitacao.latitude).strip() != "":
                lat = float(str(solicitacao.latitude).replace(",", "."))
            if solicitacao.longitude is not None and str(solicitacao.longitude).strip() != "":
                lng = float(str(solicitacao.longitude).replace(",", "."))
        except Exception:
            lat = None
            lng = None

        logradouro = (solicitacao.logradouro or "").strip()
        numero = solicitacao.numero or "S/N"
        bairro = (solicitacao.bairro or "").strip()
        cidade = (getattr(solicitacao, "cidade", "") or "").strip()
        uf = (getattr(solicitacao, "uf", "") or "").strip()
        cep = (solicitacao.cep or "").strip()

        partes = []
        if logradouro:
            partes.append(f"{logradouro}, {numero}")
        elif numero:
            partes.append(str(numero))

        if bairro:
            partes.append(bairro)

        if cidade or uf:
            partes.append(f"{cidade}/{uf}".strip("/"))

        endereco_txt = " - ".join([parte for parte in partes if parte and parte != "S/N"])
        if cep:
            endereco_txt = (endereco_txt + f" (CEP {cep})").strip()

        agenda_eventos.append(
            {
                "id": str(solicitacao.id),
                "title": f"{solicitacao.foco} - {uvis_nome}",
                "start": f"{data}T{hora}",
                "color": agenda_status_color(solicitacao.status),
                "extendedProps": {
                    "foco": solicitacao.foco,
                    "tipo_visita": solicitacao.tipo_visita,
                    "tipo_imovel": solicitacao.tipo_imovel,
                    "tipo_operacao": solicitacao.tipo_operacao,
                    "altura_voo": solicitacao.altura_voo,
                    "apoio_cet": "Sim" if solicitacao.apoio_cet else "Não",
                    "observacao": (solicitacao.observacao or "").strip(),
                    "protocolo": solicitacao.protocolo,
                    "uvis": uvis_nome,
                    "hora": hora,
                    "status": solicitacao.status,
                    "lat": lat,
                    "lng": lng,
                    "endereco": endereco_txt,
                },
            }
        )

    return agenda_eventos


def build_agenda_uvis_disponiveis(user):
    if not can_view_all_agenda(user):
        return []

    return (
        apply_regiao_scope(
            apply_prefeitura_scope(
                db.session.query(Usuario.id, Usuario.nome_uvis).filter(Usuario.tipo_usuario == "uvis"),
                user,
                Usuario.prefeitura_id,
            ),
            user,
            Usuario.regiao,
        )
        .order_by(Usuario.nome_uvis)
        .all()
    )


def build_agenda_anos_disponiveis(user):
    if db.engine.name == "postgresql":
        func_ano = db.func.to_char(Solicitacao.data_agendamento, "YYYY")
    else:
        func_ano = db.func.strftime("%Y", Solicitacao.data_agendamento)

    anos_raw = (
        apply_solicitacao_prefeitura_scope(
            db.session.query(func_ano),
            user,
        )
        .filter(Solicitacao.data_agendamento.isnot(None))
        .distinct()
        .order_by(func_ano.desc())
        .all()
    )
    return [int(item[0]) for item in anos_raw if item and item[0]] or [datetime.now().year]


def build_agenda_context(user, args):
    filtro_status = (args.get("status") or "").strip() or None
    filtro_uvis_id = args.get("uvis_id", type=int)
    filtro_tipo_visita = (args.get("tipo_visita") or "").strip() or None
    filtro_tipo_imovel = (args.get("tipo_imovel") or "").strip() or None
    filtro_foco = (args.get("foco") or "").strip() or None
    filtro_data_ini = (args.get("data_ini") or "").strip() or None
    filtro_data_fim = (args.get("data_fim") or "").strip() or None
    mes = args.get("mes", datetime.now().month, type=int)
    ano = args.get("ano", datetime.now().year, type=int)
    d = (args.get("d") or "").strip()
    initial_date = d or filtro_data_ini or datetime.now().strftime("%Y-%m-%d")

    solicitacoes = build_agenda_query(
        user,
        filtro_status=filtro_status,
        filtro_uvis_id=filtro_uvis_id,
        filtro_tipo_visita=filtro_tipo_visita,
        filtro_tipo_imovel=filtro_tipo_imovel,
        filtro_foco=filtro_foco,
        data_ini=filtro_data_ini,
        data_fim=filtro_data_fim,
        mes=mes,
        ano=ano,
    ).all()

    return {
        "eventos_json": json.dumps(build_agenda_eventos(solicitacoes), ensure_ascii=False),
        "filtros": {
            "uvis_id": filtro_uvis_id if can_view_all_agenda(user) else None,
            "status": filtro_status,
            "tipo_visita": filtro_tipo_visita,
            "tipo_imovel": filtro_tipo_imovel,
            "foco": filtro_foco,
            "data_ini": filtro_data_ini,
            "data_fim": filtro_data_fim,
            "mes": mes,
            "ano": ano,
        },
        "status_opcoes": [
            "PENDENTE",
            "EM ANÁLISE",
            "APROVADO",
            "APROVADO COM RECOMENDAÇÕES",
            "NEGADO",
        ],
        "uvis_disponiveis": build_agenda_uvis_disponiveis(user),
        "anos_disponiveis": build_agenda_anos_disponiveis(user),
        "initial_date": initial_date,
        "pode_filtrar_uvis": can_view_all_agenda(user),
        "google_maps_key": get_agenda_google_maps_key(),
    }


def build_agenda_rotas_payload(user, args):
    dia = (args.get("dia") or "").strip()
    if not dia:
        raise ValueError("Par\u00e2metro 'dia' \u00e9 obrigat\u00f3rio (YYYY-MM-DD).")

    try:
        dia_date = datetime.strptime(dia, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("Formato inv\u00e1lido para 'dia'. Use YYYY-MM-DD.") from exc

    filtro_uvis_id = args.get("uvis_id", type=int)

    query = apply_agenda_user_scope(
        Solicitacao.query.options(joinedload(Solicitacao.usuario)),
        user,
    )
    if can_view_all_agenda(user) and filtro_uvis_id:
        query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)

    query = query.filter(Solicitacao.status.in_(AGENDA_ROUTE_STATUSES))
    query = query.filter(Solicitacao.data_agendamento == dia_date)

    eventos = query.order_by(Solicitacao.hora_agendamento.asc()).all()

    pontos = []
    total_com_coords = 0
    for evento in eventos:
        lat = evento.latitude
        lng = evento.longitude

        try:
            if isinstance(lat, str):
                lat = float(lat.replace(",", "."))
            else:
                lat = float(lat) if lat is not None else None

            if isinstance(lng, str):
                lng = float(lng.replace(",", "."))
            else:
                lng = float(lng) if lng is not None else None
        except Exception:
            lat = None
            lng = None

        if lat is None or lng is None:
            continue

        total_com_coords += 1
        pontos.append(
            {
                "id": evento.id,
                "lat": lat,
                "lng": lng,
                "hora": evento.hora_agendamento.strftime("%H:%M") if evento.hora_agendamento else "00:00",
                "uvis": evento.usuario.nome_uvis if evento.usuario else "",
                "foco": evento.foco or "",
                "status": evento.status,
                "endereco": f"{evento.logradouro or ''}, {evento.numero or 'S/N'} - {evento.bairro or ''}".strip(),
            }
        )

    return {
        "ok": True,
        "dia": dia,
        "total_eventos": len(eventos),
        "total_com_coordenadas": total_com_coords,
        "pontos": pontos,
    }


def build_agenda_export(user, args):
    export_all = args.get("all") == "1"

    filtro_status = None if export_all else (args.get("status") or None)
    filtro_uvis_id = None if export_all else args.get("uvis_id", type=int)
    filtro_tipo_visita = None if export_all else (args.get("tipo_visita") or None)
    filtro_tipo_imovel = None if export_all else (args.get("tipo_imovel") or None)
    filtro_foco = None if export_all else (args.get("foco") or None)
    filtro_data_ini = None if export_all else (args.get("data_ini") or None)
    filtro_data_fim = None if export_all else (args.get("data_fim") or None)
    mes = None if export_all else args.get("mes", type=int)
    ano = None if export_all else args.get("ano", type=int)

    # A exportação precisa seguir o mesmo escopo da tela da agenda.
    query = apply_agenda_user_scope(
        Solicitacao.query.options(joinedload(Solicitacao.usuario)),
        user,
    )

    if filtro_uvis_id and can_view_all_agenda(user):
        query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)
    if filtro_tipo_visita:
        query = query.filter(Solicitacao.tipo_visita == filtro_tipo_visita)
    if filtro_tipo_imovel:
        query = query.filter(Solicitacao.tipo_imovel == filtro_tipo_imovel)
    if filtro_foco:
        query = query.filter(Solicitacao.foco == filtro_foco)
    query = _apply_agenda_date_range(query, filtro_data_ini, filtro_data_fim)
    if mes and ano and not _has_agenda_date_range(filtro_data_ini, filtro_data_fim):
        filtro_mesano = f"{ano}-{mes:02d}"
        if db.engine.name == "postgresql":
            query = query.filter(db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM") == filtro_mesano)
        else:
            query = query.filter(db.func.strftime("%Y-%m", Solicitacao.data_agendamento) == filtro_mesano)

    eventos = query.order_by(
        Solicitacao.data_agendamento.desc(),
        Solicitacao.hora_agendamento.desc(),
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda"

    headers = [
        "DATA",
        "HORÁRIO",
        "REGIÃO",
        "UVIS",
        "CET",
        "ENDEREÇO DA AÇÃO",
        "CEP",
        "TIPO DE VISITA",
        "TIPO DE IMOVEL",
        "FOCO DA AÇÃO",
        "COORDENADA GEOGRÁFICA",
        "Altura dos Voos",
        "Protocolo DECA",
        "Status",
    ]
    ws.append(headers)

    for evento in eventos:
        endereco_completo = (
            f"{evento.logradouro or ''}, {getattr(evento, 'numero', '')} - "
            f"{evento.bairro or ''} - "
            f"{(evento.cidade or '')}/{(evento.uf or '')} - "
            f"{evento.cep or ''}"
        )
        if getattr(evento, "complemento", None):
            endereco_completo += f" - {evento.complemento}"

        cet_txt = "SIM" if getattr(evento, "apoio_cet", None) else "NÃO"
        data_str = evento.data_agendamento.strftime("%d/%m/%Y") if evento.data_agendamento else ""
        hora_str = evento.hora_agendamento.strftime("%H:%M") if evento.hora_agendamento else ""
        uvis_nome = evento.usuario.nome_uvis if getattr(evento, "usuario", None) else ""
        regiao = evento.usuario.regiao if getattr(evento, "usuario", None) else ""
        lat = getattr(evento, "latitude", "") or ""
        lon = getattr(evento, "longitude", "") or ""
        coordenada = f"{lat},{lon}" if (lat or lon) else ""
        protocolo_deca = getattr(evento, "protocolo_deca", None) or getattr(evento, "protocolo", "") or ""

        ws.append(
            [
                data_str,
                hora_str,
                regiao,
                uvis_nome,
                cet_txt,
                endereco_completo,
                getattr(evento, "cep", "") or "",
                getattr(evento, "tipo_visita", "") or "",
                getattr(evento, "tipo_imovel", "") or "",
                getattr(evento, "foco", "") or "",
                coordenada,
                getattr(evento, "altura_voo", "") or "",
                protocolo_deca,
                getattr(evento, "status", "") or "",
            ]
        )

    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = wrap if cell.row > 1 else center

    for col in range(1, ws.max_column + 1):
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 60)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return bio, ("agenda_tudo.xlsx" if export_all else "agenda_exportada.xlsx")


def criar_notificacao(usuario_id, titulo, mensagem="", link=None, commit=True):
    notificacao = Notificacao(
        usuario_id=usuario_id,
        titulo=titulo,
        mensagem=mensagem or "",
        link=link,
        criada_em=agora_brasilia_naive(),
    )
    db.session.add(notificacao)
    if commit:
        db.session.commit()
    return notificacao


def garantir_notificacoes_do_dia(usuario_id):
    hoje = date.today()

    agendamentos = (
        Solicitacao.query
        .options(joinedload(Solicitacao.usuario))
        .filter_by(usuario_id=usuario_id)
        .filter(Solicitacao.data_agendamento == hoje)
        .all()
    )

    for solicitacao in agendamentos:
        hora_fmt = solicitacao.hora_agendamento.strftime("%H:%M") if solicitacao.hora_agendamento else "00:00"
        link = url_for("main.agenda", sid=solicitacao.id, d=hoje.isoformat())

        ja_existe = (
            Notificacao.query
            .filter_by(usuario_id=usuario_id, link=link)
            .first()
        )
        if ja_existe:
            continue

        criar_notificacao(
            usuario_id=usuario_id,
            titulo="Agendamento para hoje",
            mensagem=f"Você tem um agendamento hoje às {hora_fmt} (Foco: {solicitacao.foco}).",
            link=link,
        )


def sincronizar_notificacoes_automaticas(user):
    if not getattr(user, "is_authenticated", False):
        return

    if can_view_all_notifications(user):
        _sincronizar_alertas_operacionais(user)
        return

    garantir_notificacoes_do_dia(user.id)


def _sincronizar_alertas_operacionais(user):
    usuario_id = getattr(user, "id", None)
    if not usuario_id:
        return

    try:
        _sincronizar_alerta_baterias_ciclos(user, usuario_id)
        _sincronizar_alerta_revisoes_veiculos(user, usuario_id)
        _sincronizar_alerta_drones_manutencao(user, usuario_id)
        _sincronizar_alerta_manutencao_desatualizada_drones(user, usuario_id)
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Erro ao sincronizar alertas automaticos operacionais.")


def _sincronizar_alerta_baterias_ciclos(user, usuario_id):
    baterias = (
        apply_prefeitura_scope(Baterias.query, user, Baterias.prefeitura_id)
        .filter(Baterias.ciclo >= AUTO_ALERT_BATTERY_WARNING_CYCLES)
        .filter(db.func.lower(db.func.coalesce(Baterias.status, "")) != "inativo")
        .order_by(Baterias.ciclo.desc(), Baterias.renomacao.asc(), Baterias.id.asc())
        .all()
    )

    link = url_for("main.listar_baterias", alert="ciclos")
    if not baterias:
        _sincronizar_notificacao_por_link(
            usuario_id,
            link=link,
            titulo="Alerta automático: baterias com ciclo alto",
            mensagem=None,
        )
        return

    qtd_criticas = sum(
        1
        for bateria in baterias
        if int(bateria.ciclo or 0) >= AUTO_ALERT_BATTERY_CRITICAL_CYCLES
    )
    preview = _build_preview_text(
        [
            f"{_bateria_label(bateria)} ({int(bateria.ciclo or 0)}c)"
            for bateria in baterias
        ]
    )
    mensagem = (
        f"{len(baterias)} bateria(s) acima do limite de atenção "
        f"(>= {AUTO_ALERT_BATTERY_WARNING_CYCLES} ciclos)"
    )
    if qtd_criticas:
        mensagem += f"; {qtd_criticas} em nível crítico (>= {AUTO_ALERT_BATTERY_CRITICAL_CYCLES})"
    mensagem += f": {preview}."

    _sincronizar_notificacao_por_link(
        usuario_id,
        link=link,
        titulo="Alerta automático: baterias com ciclo alto",
        mensagem=mensagem,
    )


def _sincronizar_alerta_revisoes_veiculos(user, usuario_id):
    veiculos_base = (
        apply_prefeitura_scope(Veiculos.query, user, Veiculos.prefeitura_id)
        .filter(db.func.lower(db.func.coalesce(Veiculos.status, "")) == "ativo")
        .filter(Veiculos.km_prox_revisao.isnot(None))
        .order_by(Veiculos.km_prox_revisao.asc(), Veiculos.placa.asc(), Veiculos.id.asc())
        .all()
    )
    veiculos = [
        item
        for item in veiculos_base
        if item.km_restante_revisao is not None and item.km_restante_revisao <= AUTO_ALERT_VEHICLE_REVIEW_WARNING_KM
    ]

    link = url_for("main.listar_veiculos", alert="revisao")
    if not veiculos:
        _sincronizar_notificacao_por_link(
            usuario_id,
            link=link,
            titulo="Alerta automático: revisões de veículo",
            mensagem=None,
        )
        return

    qtd_vencidos = sum(1 for item in veiculos if (item.km_restante_revisao or 0) < 0)
    preview = _build_preview_text([_build_veiculo_review_label(item) for item in veiculos])
    mensagem = (
        f"{len(veiculos)} veículo(s) com revisão vencida ou próxima "
        f"(até {AUTO_ALERT_VEHICLE_REVIEW_WARNING_KM:.0f} km)"
    )
    if qtd_vencidos:
        mensagem += f"; {qtd_vencidos} já vencido(s)"
    mensagem += f": {preview}."

    _sincronizar_notificacao_por_link(
        usuario_id,
        link=link,
        titulo="Alerta automático: revisões de veículo",
        mensagem=mensagem,
    )


def _sincronizar_alerta_drones_manutencao(user, usuario_id):
    drones = (
        apply_prefeitura_scope(Drones.query, user, Drones.prefeitura_id)
        .filter(Drones.status == MANUTENCAO_STATUS)
        .order_by(Drones.renomacao.asc(), Drones.id.asc())
        .all()
    )

    base_link = url_for("main.equipamentos_manutencao", alert="drones")
    active_links = set()

    for drone in drones:
        label = _drone_label(drone)
        link = url_for("main.equipamentos_manutencao", alert="drones", item=drone.id)
        active_links.add(link)
        _sincronizar_notificacao_por_link(
            usuario_id,
            link=link,
            titulo=f"Alerta automático: drone em manutenção - {label}",
            mensagem=f"{label} indisponível para operação por manutenção.",
        )

    _apagar_notificacoes_por_prefixo_link(
        usuario_id,
        link_prefix=base_link,
        keep_links=active_links,
        allowed_title_prefixes=(
            "Alerta automático: drones em manutenção",
            "Alerta automático: drone em manutenção",
            "Alerta automatico: drones em manutencao",
            "Alerta automatico: drone em manutencao",
        ),
    )


def _sincronizar_alerta_manutencao_desatualizada_drones(user, usuario_id):
    data_limite = date.today() - timedelta(days=AUTO_ALERT_DRONE_MAINTENANCE_STALE_DAYS)
    drones = (
        apply_prefeitura_scope(Drones.query, user, Drones.prefeitura_id)
        .filter(db.func.lower(db.func.coalesce(Drones.status, "")) == "ativo")
        .filter(Drones.ultima_manutencao.isnot(None))
        .filter(Drones.ultima_manutencao < data_limite)
        .order_by(Drones.ultima_manutencao.asc(), Drones.renomacao.asc(), Drones.id.asc())
        .all()
    )

    link = url_for("main.listar_drones", alert="manutencao")
    if not drones:
        _sincronizar_notificacao_por_link(
            usuario_id,
            link=link,
            titulo="Alerta automático: manutenção de drones desatualizada",
            mensagem=None,
        )
        return

    preview = _build_preview_text(
        [
            f"{_drone_label(drone)} ({_days_since(drone.ultima_manutencao)} dias)"
            for drone in drones
        ]
    )
    mensagem = (
        f"{len(drones)} drone(s) com última manutenção acima de "
        f"{AUTO_ALERT_DRONE_MAINTENANCE_STALE_DAYS} dias: {preview}."
    )

    _sincronizar_notificacao_por_link(
        usuario_id,
        link=link,
        titulo="Alerta automático: manutenção de drones desatualizada",
        mensagem=mensagem,
    )


def _sincronizar_notificacao_por_link(usuario_id, *, link, titulo, mensagem):
    if not usuario_id or not link:
        return

    existentes = (
        Notificacao.query
        .filter(
            Notificacao.usuario_id == usuario_id,
            Notificacao.link == link,
        )
        .order_by(Notificacao.id.desc())
        .all()
    )

    agora = agora_brasilia_naive()
    principal = existentes[0] if existentes else None

    if mensagem:
        if principal:
            principal.titulo = (titulo or "")[:140]
            principal.mensagem = mensagem
            principal.criada_em = agora
            principal.lida_em = None
            principal.apagada_em = None
        else:
            criar_notificacao(
                usuario_id=usuario_id,
                titulo=titulo,
                mensagem=mensagem,
                link=link,
                commit=False,
            )
        for extra in existentes[1:]:
            if extra.apagada_em is None:
                extra.apagada_em = agora
        return

    for notificacao in existentes:
        if notificacao.apagada_em is None:
            notificacao.apagada_em = agora


def _apagar_notificacoes_por_prefixo_link(
    usuario_id,
    *,
    link_prefix,
    keep_links=None,
    allowed_title_prefixes=None,
):
    if not usuario_id or not link_prefix:
        return

    keep_links = {link for link in (keep_links or set()) if link}
    title_prefixes = tuple(
        (prefix or "").strip().lower()
        for prefix in (allowed_title_prefixes or ())
        if (prefix or "").strip()
    )
    existentes = (
        Notificacao.query
        .filter(
            Notificacao.usuario_id == usuario_id,
            Notificacao.link.isnot(None),
            Notificacao.link.like(f"{link_prefix}%"),
        )
        .all()
    )

    agora = agora_brasilia_naive()
    for notificacao in existentes:
        if notificacao.link in keep_links:
            continue
        titulo = (notificacao.titulo or "").strip().lower()
        if title_prefixes and not any(titulo.startswith(prefix) for prefix in title_prefixes):
            continue
        if notificacao.apagada_em is None:
            notificacao.apagada_em = agora


def _build_preview_text(values):
    cleaned = [str(value).strip() for value in values if str(value or "").strip()]
    if not cleaned:
        return "Sem detalhes"
    if len(cleaned) <= AUTO_ALERT_PREVIEW_LIMIT:
        return ", ".join(cleaned)
    extras = len(cleaned) - AUTO_ALERT_PREVIEW_LIMIT
    return f"{', '.join(cleaned[:AUTO_ALERT_PREVIEW_LIMIT])} e +{extras}"


def _bateria_label(bateria):
    nome = (bateria.renomacao or bateria.modelo or "").strip()
    if nome:
        return nome
    return f"Bateria {bateria.id}"


def _drone_label(drone):
    nome = (drone.renomacao or drone.modelo or "").strip()
    if nome:
        return nome
    return f"Drone {drone.id}"


def _days_since(value):
    if value is None:
        return 0
    return max((date.today() - value).days, 0)


def _build_veiculo_review_label(veiculo):
    placa = (veiculo.placa or veiculo.renomacao or veiculo.modelo or "").strip() or f"Veículo {veiculo.id}"
    faltante = veiculo.km_restante_revisao
    if faltante is None:
        return placa
    if faltante < 0:
        return f"{placa} (vencido há {abs(int(faltante))} km)"
    return f"{placa} (faltam {int(faltante)} km)"


def get_notificacao_or_404(user, notif_id):
    if can_view_all_notifications(user):
        return Notificacao.query.get_or_404(notif_id)

    return (
        Notificacao.query
        .filter_by(id=notif_id, usuario_id=user.id)
        .first_or_404()
    )


def list_notificacoes(user):
    sincronizar_notificacoes_automaticas(user)

    base = Notificacao.query.filter(Notificacao.apagada_em.is_(None))
    if can_view_all_notifications(user):
        return base.order_by(Notificacao.criada_em.desc()).all()

    return (
        base
        .filter_by(usuario_id=user.id)
        .order_by(Notificacao.criada_em.desc())
        .all()
    )


def mark_notificacao_as_read(notificacao):
    if notificacao.lida_em is None:
        notificacao.lida_em = agora_brasilia_naive()
        db.session.commit()


def soft_delete_notificacao(notificacao):
    notificacao.apagada_em = agora_brasilia_naive()
    db.session.commit()


def clear_notificacoes(user):
    agora = agora_brasilia_naive()
    query = Notificacao.query.filter(Notificacao.apagada_em.is_(None))

    if not can_view_all_notifications(user):
        query = query.filter_by(usuario_id=user.id)

    query.update({"apagada_em": agora}, synchronize_session=False)
    db.session.commit()
