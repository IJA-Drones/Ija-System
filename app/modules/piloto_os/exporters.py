import base64
import json
import os
import re
import tempfile
from datetime import datetime
from io import BytesIO
from urllib.parse import quote, urlencode

import requests
from flask import current_app
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image as RLImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import joinedload

from app.extensions import db
from app.models import Solicitacao

REMOTE_MEDIA_PREFIXES = ("webdav://", "skybox://")
REMOTE_MEDIA_TIMEOUT = (30, 300)
PDF_IMAGE_DPI = 350
PDF_IMAGE_JPEG_QUALITY = 95


def _fmt_dt(value):
    if not value:
        return ""
    try:
        return value.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(value)


def _fmt_date(value):
    if not value:
        return ""
    try:
        return value.strftime("%d/%m/%Y")
    except Exception:
        return str(value)


def _fmt_time(value):
    if not value:
        return ""
    try:
        return value.strftime("%H:%M")
    except Exception:
        return str(value)


def _safe(value):
    if value is None:
        return ""
    return str(value)


def _parse_json_object(value):
    if not value:
        return None
    if isinstance(value, dict):
        return value
    try:
        data = json.loads(value)
    except Exception:
        return None
    return data if isinstance(data, dict) else None


def _build_planejado_dosagem_items(ordem):
    data = _parse_json_object(getattr(ordem, "calculo_dosagem_planejado", None))
    if not data:
        return []

    medicao = data.get("medicao") or {}
    resultado = data.get("resultado") or {}
    valor_base = medicao.get("valor_base")
    unidade_base = medicao.get("unidade_base") or ""

    items = []
    if getattr(ordem, "calculo_dosagem_planejado_em", None):
        items.append(("Calculado em", _fmt_dt(ordem.calculo_dosagem_planejado_em)))
    items.extend([
        ("Cenario planejado", data.get("cenario_label") or ""),
        (
            medicao.get("medida_label") or "Medida base",
            f"{valor_base} {unidade_base}".strip() if valor_base not in (None, "") else "",
        ),
        ("Metodo de medicao", medicao.get("modo_label") or ""),
        ("Resumo da medicao", medicao.get("resumo") or ""),
        ("Carga BTI planejada (g)", resultado.get("carga_bti_g") or ""),
        ("Calda planejada", resultado.get("calda_total_label") or resultado.get("calda_total_ml") or ""),
        ("Tempo planejado", resultado.get("tempo_aplicacao_label") or ""),
        ("Tipo aplicacao planejada", resultado.get("tipo_aplicacao") or ""),
        ("Ponta planejada", resultado.get("ponta_pulverizacao") or ""),
        ("Numero de bicos", resultado.get("numero_bicos") or ""),
        ("Vazao bicos (ml/min)", resultado.get("vazao_bicos_ml_min") or ""),
        ("Dose BTI (g/min)", resultado.get("dose_bti_g_min") or ""),
        ("Pressao (bar)", resultado.get("pressao_bar") or ""),
        ("Faixa aplicacao (m)", resultado.get("faixa_aplicacao_m") or ""),
        ("Tamanho gotas DMV", resultado.get("tamanho_gotas_dmv") or ""),
    ])

    if resultado.get("pulverizacao_area_l_ha") not in (None, ""):
        items.append(("Pulverizacao planejada (l/ha)", resultado.get("pulverizacao_area_l_ha")))

    return [(label, value) for label, value in items if value not in (None, "")]


def _normalize_coord(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return text.replace(",", ".")


def _get_logo_path(args):
    logo_mode = (args.get("logo") or "light").strip().lower()
    filename = "img/logo_oceano_azul_dark.png" if logo_mode == "dark" else "img/logo_oceano_azul_light.png"
    return os.path.join(current_app.root_path, "static", filename)


def _try_make_logo(args, width_mm=34):
    try:
        path = _get_logo_path(args)
        if not os.path.exists(path):
            return None
        image = RLImage(path)
        image.drawWidth = width_mm * mm
        image.drawHeight = (width_mm * 0.55) * mm
        return image
    except Exception:
        return None


def _setting(*names):
    for name in names:
        value = current_app.config.get(name) or os.getenv(name)
        if value:
            return value
    return None


def _is_remote_media_path(value):
    return str(value or "").startswith(REMOTE_MEDIA_PREFIXES)


def _remote_media_path_from_marker(value):
    text = str(value or "").strip().replace("\\", "/")
    for prefix in REMOTE_MEDIA_PREFIXES:
        if text.startswith(prefix):
            text = text[len(prefix):]
            break

    parts = [part.strip() for part in text.split("/") if part.strip()]
    if not parts or any(part in {".", ".."} for part in parts):
        return None
    return "/".join(parts)


def _download_remote_media_bytes(marker):
    base_url = (_setting("WEBDAV_URL", "SKYBOX_WEBDAV_URL") or "").strip().rstrip("/")
    username = (_setting("WEBDAV_USER", "SKYBOX_USERNAME") or "").strip()
    password = _setting("WEBDAV_PASS", "SKYBOX_APP_PASSWORD") or ""
    remote_path = _remote_media_path_from_marker(marker)
    if not base_url or not username or not password or not remote_path:
        return None

    encoded_path = "/".join(quote(part, safe="") for part in remote_path.split("/") if part)
    try:
        response = requests.get(
            f"{base_url}/{encoded_path}",
            auth=(username, password),
            timeout=REMOTE_MEDIA_TIMEOUT,
        )
        if response.status_code != 200:
            current_app.logger.warning(
                "Falha ao baixar imagem remota da OS para PDF (%s): HTTP %s",
                remote_path,
                response.status_code,
            )
            return None
        return BytesIO(response.content)
    except Exception:
        current_app.logger.exception("Erro ao baixar imagem remota da OS para PDF.")
        return None


def _prepare_pdf_image_source(image_source, width_mm=165, max_height_mm=110):
    try:
        with Image.open(image_source) as img:
            img = ImageOps.exif_transpose(img)
            source_width, source_height = img.size
            if source_width <= 0 or source_height <= 0:
                return image_source

            draw_width_mm = float(width_mm)
            draw_height_mm = draw_width_mm * (source_height / source_width)
            if max_height_mm and draw_height_mm > float(max_height_mm):
                scale = float(max_height_mm) / draw_height_mm
                draw_width_mm *= scale
                draw_height_mm = float(max_height_mm)

            target_width = max(1, int(round((draw_width_mm / 25.4) * PDF_IMAGE_DPI)))
            target_height = max(1, int(round((draw_height_mm / 25.4) * PDF_IMAGE_DPI)))

            if source_width > target_width or source_height > target_height:
                img.thumbnail((target_width, target_height), Image.Resampling.LANCZOS)

            if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
                rgba = img.convert("RGBA")
                background = Image.new("RGB", img.size, "white")
                background.paste(rgba, mask=rgba.getchannel("A"))
                img = background
            elif img.mode != "RGB":
                img = img.convert("RGB")

            output = BytesIO()
            img.save(
                output,
                format="JPEG",
                quality=PDF_IMAGE_JPEG_QUALITY,
                optimize=True,
                progressive=True,
                dpi=(PDF_IMAGE_DPI, PDF_IMAGE_DPI),
            )
            output.seek(0)
            return output
    except Exception:
        current_app.logger.exception("Erro ao preparar imagem da OS para o PDF.")
        try:
            image_source.seek(0)
        except Exception:
            pass
        return image_source


def _try_make_local_rlimage(rel_path, width_mm=165, max_height_mm=110):
    if not rel_path:
        return None

    image_source = None
    if _is_remote_media_path(rel_path):
        image_source = _download_remote_media_bytes(rel_path)
        if image_source is None:
            return None
    else:
        static_root = os.path.abspath(os.path.join(current_app.root_path, "static"))
        abs_path = os.path.abspath(os.path.join(static_root, str(rel_path).replace("/", os.sep)))
        if os.path.commonpath([static_root, abs_path]) != static_root:
            return None
        if not os.path.exists(abs_path):
            return None
        image_source = abs_path

    try:
        image_source = _prepare_pdf_image_source(image_source, width_mm=width_mm, max_height_mm=max_height_mm)
        image = RLImage(image_source)
        base_width = width_mm * mm
        img_width = float(getattr(image, "imageWidth", 0) or 1)
        img_height = float(getattr(image, "imageHeight", 0) or 1)
        image.drawWidth = base_width
        image.drawHeight = base_width * (img_height / img_width)

        if max_height_mm and image.drawHeight > max_height_mm * mm:
            scale = (max_height_mm * mm) / image.drawHeight
            image.drawHeight = max_height_mm * mm
            image.drawWidth = image.drawWidth * scale
        return image
    except Exception:
        return None


_DATAURL_RE = re.compile(r"^data:image/(?P<fmt>png|jpeg|jpg);base64,(?P<data>.+)$", re.I)


def _dataurl_to_rlimage(dataurl: str, width_mm=80, height_mm=32):
    if not dataurl or not isinstance(dataurl, str):
        return None

    match = _DATAURL_RE.match(dataurl.strip())
    if not match:
        return None

    try:
        raw = base64.b64decode(match.group("data"))
    except Exception:
        return None

    image = RLImage(BytesIO(raw))
    image.drawWidth = width_mm * mm
    image.drawHeight = height_mm * mm
    return image


def _dataurl_to_png_bytes(dataurl: str):
    if not dataurl or not isinstance(dataurl, str):
        return None

    match = _DATAURL_RE.match(dataurl.strip())
    if not match:
        return None

    try:
        return base64.b64decode(match.group("data"))
    except Exception:
        return None


def _try_make_static_map(lat, lng, width_mm=165, height_mm=88, zoom=19, maptype="satellite"):
    lat = _normalize_coord(lat)
    lng = _normalize_coord(lng)
    if not lat or not lng:
        return None

    api_key = current_app.config.get("KEY_API_GOOGLE_MAPS")
    if not api_key:
        return None

    try:
        params = {
            "center": f"{lat},{lng}",
            "zoom": zoom,
            "size": "1200x650",
            "scale": "2",
            "maptype": maptype,
            "markers": f"color:red|label:O|{lat},{lng}",
            "key": api_key,
        }
        url = "https://maps.googleapis.com/maps/api/staticmap?" + urlencode(params)
        response = requests.get(url, timeout=15)
        response.raise_for_status()
        if "image" not in (response.headers.get("Content-Type") or "").lower():
            return None
        image = RLImage(BytesIO(response.content))
        image.drawWidth = width_mm * mm
        image.drawHeight = height_mm * mm
        return image
    except Exception:
        return None


def _load_solicitacao_com_ordem(os_id):
    return (
        Solicitacao.query.options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe),
            joinedload(Solicitacao.ordem_servico),
        )
        .get_or_404(os_id)
    )


def _pdf_styles():
    styles = getSampleStyleSheet()

    title = ParagraphStyle(
        "p_title",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0F3D75"),
        spaceAfter=2,
    )
    subtitle = ParagraphStyle(
        "p_subtitle",
        parent=styles["Normal"],
        fontSize=9.2,
        leading=12.5,
        textColor=colors.HexColor("#667085"),
        spaceAfter=10,
    )
    section = ParagraphStyle(
        "p_section",
        parent=styles["Heading2"],
        fontSize=11.5,
        leading=15,
        textColor=colors.HexColor("#0F3D75"),
        spaceBefore=8,
        spaceAfter=5,
    )
    cell = ParagraphStyle(
        "p_cell",
        parent=styles["BodyText"],
        fontSize=9,
        leading=11.8,
        textColor=colors.HexColor("#111827"),
        wordWrap="CJK",
        splitLongWords=True,
    )
    hint = ParagraphStyle(
        "p_hint",
        parent=styles["Normal"],
        fontSize=8.3,
        leading=11,
        textColor=colors.HexColor("#667085"),
        spaceAfter=5,
    )
    small = ParagraphStyle(
        "p_small",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#475467"),
    )

    return styles, title, subtitle, section, cell, hint, small


def _pdf_header_block(os_id: int, status_txt: str, args):
    styles, title_style, subtitle_style, *_ = _pdf_styles()
    logo = _try_make_logo(args, width_mm=35)
    title = Paragraph(f"OS #{os_id} - Formulario (Admin)", title_style)
    subtitle = Paragraph(
        f"Gerado em {_fmt_dt(datetime.now())} - Status: {_safe(status_txt)}",
        subtitle_style,
    )

    left = [title, subtitle]
    right = [logo] if logo else [Paragraph("", styles["Normal"])]
    table = Table([[left, right]], colWidths=[None, 42 * mm])
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return [table, Spacer(1, 3)]


def _pdf_kv_table_nice(section_title: str, items: list[tuple[str, object]], cell_style, section_style, doc_width, orient="portrait"):
    key_w = 60 * mm if orient == "portrait" else 73 * mm
    val_w = doc_width - key_w

    rows = [[
        Paragraph("<b>Campo</b>", ParagraphStyle("th1", parent=cell_style, textColor=colors.white)),
        Paragraph("<b>Valor</b>", ParagraphStyle("th2", parent=cell_style, textColor=colors.white)),
    ]]
    for key, value in items:
        rows.append([Paragraph(_safe(key), cell_style), Paragraph(_safe(value), cell_style)])

    table = Table(rows, repeatRows=1, colWidths=[key_w, val_w])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.HexColor("#1565C0")),
        ("GRID", (0, 1), (-1, -1), 0.25, colors.HexColor("#DDE3EA")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return [Paragraph(section_title, section_style), table, Spacer(1, 7)]


def _pdf_card(flowables, doc_width, bg="#F8FAFC", border="#D0D5DD", padding=8):
    card = Table([[flowables]], colWidths=[doc_width])
    card.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(bg)),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor(border)),
        ("LEFTPADDING", (0, 0), (-1, -1), padding),
        ("RIGHTPADDING", (0, 0), (-1, -1), padding),
        ("TOPPADDING", (0, 0), (-1, -1), padding),
        ("BOTTOMPADDING", (0, 0), (-1, -1), padding),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return card


def _header_footer_factory_pretty(title):
    def _hf(canvas, doc):
        canvas.saveState()
        _, height = doc.pagesize
        canvas.setFillColor(colors.HexColor("#1565C0"))
        canvas.rect(doc.leftMargin, height - (11 * mm), doc.width, 2.8, fill=1, stroke=0)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawString(doc.leftMargin, 9 * mm, title)
        canvas.drawRightString(doc.leftMargin + doc.width, 9 * mm, f"Pagina {canvas.getPageNumber()}")
        canvas.restoreState()

    return _hf


def build_admin_os_pdf_v2_export(os_id, args):
    orient = args.get("orient", "portrait")
    pagesize = landscape(A4) if orient == "landscape" else A4

    try:
        solicitacao = _load_solicitacao_com_ordem(os_id)
        ordem = solicitacao.ordem_servico

        tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        caminho_pdf = tmp_pdf.name
        tmp_pdf.close()

        doc = SimpleDocTemplate(
            caminho_pdf,
            pagesize=pagesize,
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
        )

        styles, _, _, section_style, cell_style, hint_style, small_style = _pdf_styles()
        story = []
        story += _pdf_header_block(solicitacao.id, solicitacao.status, args)
        story.append(Spacer(1, 2))

        endereco_os = (
            f"{solicitacao.logradouro or ''}, {solicitacao.numero or 'S/N'} - "
            f"{solicitacao.bairro or ''} - {solicitacao.cidade or ''}/{solicitacao.uf or ''}"
        )
        lat = _normalize_coord(getattr(solicitacao, "latitude", None))
        lng = _normalize_coord(getattr(solicitacao, "longitude", None))

        story += _pdf_kv_table_nice("Identificacao", [
            ("Solicitacao ID", solicitacao.id),
            ("Equipe ID", solicitacao.equipe_id or ""),
            ("Equipe", solicitacao.equipe.nome_equipe if solicitacao.equipe else ""),
            ("UVIS", solicitacao.usuario.nome_uvis if solicitacao.usuario else ""),
            ("Endereco", endereco_os),
            ("Data agendamento", _fmt_date(solicitacao.data_agendamento)),
            ("Hora agendamento", _fmt_time(solicitacao.hora_agendamento)),
            ("Foco", solicitacao.foco or ""),
            ("Status", solicitacao.status or ""),
            ("Protocolo", getattr(solicitacao, "protocolo", "") or ""),
        ], cell_style, section_style, doc.width, orient=orient)

        story += _pdf_kv_table_nice("Endereco / Coordenadas", [
            ("CEP", solicitacao.cep or ""),
            ("Logradouro", solicitacao.logradouro or ""),
            ("Numero", solicitacao.numero or ""),
            ("Bairro", solicitacao.bairro or ""),
            ("Cidade", solicitacao.cidade or ""),
            ("UF", solicitacao.uf or ""),
            ("Complemento", solicitacao.complemento or ""),
            ("Latitude", lat or ""),
            ("Longitude", lng or ""),
        ], cell_style, section_style, doc.width, orient=orient)

        if lat and lng:
            maps_link = f"https://www.google.com/maps?q={lat},{lng}"
            map_img = _try_make_static_map(
                lat=lat,
                lng=lng,
                width_mm=175 if orient == "landscape" else 165,
                height_mm=92,
                zoom=19,
                maptype="satellite",
            )
            map_block = [
                Paragraph("Localizacao no mapa", section_style),
                Paragraph(
                    "Visual gerado automaticamente a partir da latitude e longitude da solicitacao.",
                    hint_style,
                ),
                Paragraph(
                    f'Para acessar o Google Maps, clique aqui: <link href="{maps_link}">{maps_link}</link>',
                    small_style,
                ),
                Spacer(1, 5),
            ]
            if map_img:
                map_block.append(map_img)
            else:
                map_block.append(Paragraph("Nao foi possivel gerar a imagem do mapa no momento.", styles["Normal"]))

            story.append(_pdf_card(map_block, doc.width, bg="#F8FAFC", border="#D0D5DD", padding=8))
            story.append(Spacer(1, 8))

        if not ordem:
            story.append(Paragraph("Formulario", section_style))
            story.append(Paragraph("Esta OS nao possui formulario preenchido.", styles["Normal"]))
        else:
            story += _pdf_kv_table_nice("Responsavel / Registro", [
                ("Identificador OS", ordem.identificador_os or ""),
                ("Respondido por", ordem.respondido_por or ""),
                ("Respondido em", _fmt_dt(ordem.respondido_em)),
            ], cell_style, section_style, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Aeronaves - Pulverizacao (Principal)", [
                ("Drone ID", ordem.drone_id or ""),
                ("Prefixo", ordem.prefixo_aeronave_pulverizacao or ""),
                ("Denominacao", ordem.drone_denominacao or ""),
                ("Modelo", ordem.drone_modelo or ""),
                ("N Serie", ordem.drone_numero_serie or ""),
                ("Registro ANATEL", ordem.drone_registro_anatel or ""),
                ("Registro ANAC", ordem.drone_registro_anac or ""),
            ], cell_style, section_style, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Aeronaves - Monitoramento", [
                ("Drone Monitoramento ID", ordem.drone_monitoramento_id or ""),
                ("Prefixo", ordem.prefixo_aeronave_monitoramento or ""),
                ("Denominacao", ordem.drone_monitoramento_denominacao or ""),
                ("Modelo", ordem.drone_monitoramento_modelo or ""),
                ("N Serie", ordem.drone_monitoramento_numero_serie or ""),
                ("Registro ANATEL", ordem.drone_monitoramento_registro_anatel or ""),
                ("Registro ANAC", ordem.drone_monitoramento_registro_anac or ""),
            ], cell_style, section_style, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Aplicacao", [
                ("Situacao da aplicacao", ordem.situacao_aplicacao or ""),
                ("Larva visualizada", ordem.larva_visualizada or ""),
                ("Retornar monitorar larvas", ordem.retornar_proxima_semana_monitorar_larvas or ""),
                ("DA (Distrito)", ordem.distrito_administrativo or ""),
                ("Nome/RF ACE responsavel", ordem.nome_rf_ace_responsavel_os or ""),
                ("Criadouro OS (tipo/volume)", ordem.criadouro_os_tipo_volume or ""),
                ("Data aplicacao", _fmt_date(ordem.data_aplicacao)),
                ("Hora inicio", _fmt_time(ordem.hora_inicio_aplicacao)),
                ("Hora termino", _fmt_time(ordem.hora_termino_aplicacao)),
                ("Tratamento adicional", ordem.tratamento_adicional_realizado or ""),
                ("Quantos / Quais", ordem.quantos_quais or ""),
            ], cell_style, section_style, doc.width, orient=orient)

            planejado_items = _build_planejado_dosagem_items(ordem)
            if planejado_items:
                story += _pdf_kv_table_nice(
                    "Calculo Planejado de Dosagem",
                    planejado_items,
                    cell_style,
                    section_style,
                    doc.width,
                    orient=orient,
                )

            story += _pdf_kv_table_nice("Produto e Parametros", [
                ("Descricao produto", ordem.descricao_produto or ""),
                ("Formulacao", ordem.formulacao_produto or ""),
                ("Dosagem (g/10L)", ordem.dosagem_g_10l or ""),
                ("Tipo aplicacao", ordem.tipo_aplicacao or ""),
                ("Qtd administrada (ml)", ordem.quantidade_produto_administrada_ml or ""),
                ("Pulverizacao area (l/ha)", ordem.pulverizacao_area_l_ha or ""),
                ("Ponta pulverizacao", ordem.ponta_pulverizacao or ""),
            ], cell_style, section_style, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Condicoes Ambientais", [
                ("Imagens registradas", ordem.quantidade_imagens_registradas or ""),
                ("Videos registrados", ordem.quantidade_videos_registradas or ""),
                ("Temperatura (C)", ordem.temperatura_c or ""),
                ("Umidade (%)", ordem.umidade_relativa_pct or ""),
                ("Vento (km/h)", ordem.velocidade_vento_kmh or ""),
            ], cell_style, section_style, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Fechamento", [
                ("Observacoes gerais", ordem.observacoes or ""),
                ("Motivo nao realizacao", ordem.motivo_nao_realizacao or ""),
                ("Piloto", ordem.piloto or ""),
                ("Auxiliar", ordem.auxiliar or ""),
                ("Proprietario/Preposto", ordem.proprietario_ou_preposto or ""),
            ], cell_style, section_style, doc.width, orient=orient)

            if getattr(ordem, "imagem_principal", None):
                image_report_title = ParagraphStyle(
                    "p_image_report_title",
                    parent=styles["Title"],
                    fontSize=14.5,
                    leading=18,
                    alignment=1,
                    textColor=colors.HexColor("#111827"),
                    spaceAfter=8,
                )
                image_report_info = ParagraphStyle(
                    "p_image_report_info",
                    parent=cell_style,
                    alignment=1,
                    fontSize=9.2,
                    leading=11.5,
                )

                uvis_nome = solicitacao.usuario.nome_uvis if solicitacao.usuario else ""
                regiao_nome = (
                    getattr(getattr(solicitacao, "usuario", None), "regiao", None)
                    or getattr(getattr(solicitacao, "equipe", None), "regiao", None)
                    or ""
                )
                data_coleta = (
                    _fmt_date(ordem.data_aplicacao)
                    or _fmt_date(ordem.respondido_em.date() if ordem.respondido_em else None)
                    or _fmt_date(solicitacao.data_agendamento)
                )
                coordenadas = f"{lat}, {lng}" if lat and lng else ""
                imagem_principal = _try_make_local_rlimage(
                    ordem.imagem_principal,
                    width_mm=175 if orient == "portrait" else 230,
                    max_height_mm=120,
                )

                story.append(PageBreak())
                logo_relatorio = _try_make_logo(args, width_mm=28)
                if logo_relatorio:
                    logo_box = Table([[logo_relatorio]], colWidths=[doc.width])
                    logo_box.setStyle(TableStyle([
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]))
                    story.append(logo_box)
                    story.append(Spacer(1, 6))

                story.append(Paragraph("Relatorio de Coleta de Imagens - Operacao Dengue PMSP", image_report_title))

                resumo_topo = Table([[
                    Paragraph(f"<b>UVIS:</b> {uvis_nome or '-'}", image_report_info),
                    Paragraph(f"<b>REGIAO:</b> {regiao_nome or '-'}", image_report_info),
                ]], colWidths=[doc.width / 2, doc.width / 2])
                resumo_topo.setStyle(TableStyle([
                    ("BOX", (0, 0), (-1, -1), 1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]))
                story.append(resumo_topo)
                story.append(Spacer(1, 8))

                if imagem_principal:
                    imagem_box = Table([[imagem_principal]], colWidths=[doc.width])
                    imagem_box.setStyle(TableStyle([
                        ("BOX", (0, 0), (-1, -1), 1, colors.black),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 5),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ]))
                    story.append(imagem_box)
                    story.append(Spacer(1, 6))
                else:
                    story.append(Paragraph("A foto principal foi registrada, mas o arquivo nao foi encontrado para exportacao.", hint_style))
                    story.append(Spacer(1, 6))

                story += _pdf_kv_table_nice("Detalhes do levantamento", [
                    ("Endereco", endereco_os),
                    ("CEP", solicitacao.cep or ""),
                    ("Coordenadas", coordenadas),
                    ("Data de coleta de img", data_coleta),
                    ("Foco", solicitacao.foco or ""),
                    ("Imagens registradas", ordem.quantidade_imagens_registradas or ""),
                ], cell_style, section_style, doc.width, orient=orient)

            ass_piloto = _dataurl_to_rlimage(getattr(ordem, "assinatura_piloto", None), width_mm=82, height_mm=32)
            ass_resp = _dataurl_to_rlimage(
                getattr(ordem, "assinatura_proprietario_ou_preposto", None),
                width_mm=82,
                height_mm=32,
            )

            story.append(Paragraph("Assinaturas", section_style))
            story.append(Paragraph("Exportadas diretamente do formulario.", hint_style))
            story.append(Spacer(1, 4))

            def _sig_card(title_html, who, image_or_none):
                inner = [
                    Paragraph(f"<b>{title_html}</b>", styles["Normal"]),
                    Paragraph(_safe(who) if who else "-", hint_style),
                    Spacer(1, 4),
                    image_or_none if image_or_none else Paragraph("Nao informada.", styles["Normal"]),
                ]
                card = Table([[inner]], colWidths=[doc.width / 2 - 5 * mm])
                card.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#D0D5DD")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 9),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]))
                return card

            cards = Table([[
                _sig_card("Assinatura do Piloto", getattr(ordem, "piloto", ""), ass_piloto),
                _sig_card("Assinatura do Responsavel (Local)", getattr(ordem, "proprietario_ou_preposto", ""), ass_resp),
            ]], colWidths=[doc.width / 2, doc.width / 2])
            cards.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(cards)
            story.append(Spacer(1, 8))

        header_title = f"OS #{solicitacao.id} - Oceano Azul / IJA Drones"
        doc.build(
            story,
            onFirstPage=_header_footer_factory_pretty(header_title),
            onLaterPages=_header_footer_factory_pretty(header_title),
        )
        return caminho_pdf, f"os_{solicitacao.id}_formulario.pdf"
    except Exception:
        db.session.rollback()
        raise


THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEADER = PatternFill("solid", fgColor="0D6EFD")
FILL_SECTION = PatternFill("solid", fgColor="EAF2FF")
FILL_ZEBRA = PatternFill("solid", fgColor="FBFDFF")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_TITLE = Font(bold=True, size=16, color="0D6EFD")
FONT_SUBTITLE = Font(size=10, color="555555")
FONT_SECTION = Font(bold=True, color="0D6EFD")


def _excel_add_title(ws, title: str, subtitle: str = ""):
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    ws["A2"] = subtitle
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18


def _excel_add_section(ws, row: int, title: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = FILL_SECTION
    cell.font = FONT_SECTION
    cell.alignment = Alignment(vertical="center")
    cell.border = BORDER
    ws.cell(row=row, column=2).border = BORDER
    ws.row_dimensions[row].height = 18


def _excel_apply_table_style(ws, header_row: int, end_row: int, col_count: int = 2):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for row in range(header_row + 1, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if (row - header_row) % 2 == 0:
                cell.fill = FILL_ZEBRA


def _excel_write_kv(ws, start_row: int, items: list[tuple[str, object]]):
    ws.cell(row=start_row, column=1, value="Campo")
    ws.cell(row=start_row, column=2, value="Valor")

    row = start_row + 1
    for key, value in items:
        ws.cell(row=row, column=1, value=str(key))
        ws.cell(row=row, column=2, value=_safe(value))
        row += 1

    _excel_apply_table_style(ws, start_row, row - 1, col_count=2)
    return row


def _excel_auto_width(ws, max_col=2, min_w=18, max_w=75):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = 0
        for cell in ws[letter]:
            if cell.value:
                best = max(best, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, best + 2))


def build_admin_os_excel_v2_export(os_id, args):
    try:
        solicitacao = _load_solicitacao_com_ordem(os_id)
        ordem = solicitacao.ordem_servico

        workbook = Workbook()
        ws_os = workbook.active
        ws_os.title = "OS (Resumo)"

        _excel_add_title(
            ws_os,
            f"OS #{solicitacao.id} - Exportacao",
            f"Gerado em {_fmt_dt(datetime.now())} | Status: {_safe(solicitacao.status)}",
        )

        endereco = (
            f"{solicitacao.logradouro or ''}, {solicitacao.numero or 'S/N'} - "
            f"{solicitacao.bairro or ''} - {solicitacao.cidade or ''}/{solicitacao.uf or ''}"
        )

        row = 4
        _excel_add_section(ws_os, row, "Identificacao")
        row += 1
        row = _excel_write_kv(ws_os, row, [
            ("Solicitacao ID", solicitacao.id),
            ("Status", solicitacao.status or ""),
            ("Equipe ID", solicitacao.equipe_id or ""),
            ("Equipe", solicitacao.equipe.nome_equipe if solicitacao.equipe else ""),
            ("UVIS", solicitacao.usuario.nome_uvis if solicitacao.usuario else ""),
            ("Endereco", endereco),
            ("Data agendamento", _fmt_date(solicitacao.data_agendamento)),
            ("Hora agendamento", _fmt_time(solicitacao.hora_agendamento)),
            ("Foco", solicitacao.foco or ""),
            ("Protocolo", getattr(solicitacao, "protocolo", "") or ""),
        ])

        row += 1
        _excel_add_section(ws_os, row, "Endereco / Coordenadas")
        row += 1
        _excel_write_kv(ws_os, row, [
            ("CEP", solicitacao.cep or ""),
            ("Logradouro", solicitacao.logradouro or ""),
            ("Numero", solicitacao.numero or ""),
            ("Bairro", solicitacao.bairro or ""),
            ("Cidade", solicitacao.cidade or ""),
            ("UF", solicitacao.uf or ""),
            ("Complemento", solicitacao.complemento or ""),
            ("Latitude", solicitacao.latitude or ""),
            ("Longitude", solicitacao.longitude or ""),
        ])

        ws_os.freeze_panes = "A5"
        _excel_auto_width(ws_os, max_col=2, min_w=18, max_w=75)

        ws_form = workbook.create_sheet("Formulario")
        _excel_add_title(ws_form, f"Formulario - OS #{solicitacao.id}", "Campos preenchidos pelo piloto")

        row = 4
        if not ordem:
            _excel_add_section(ws_form, row, "Sem formulario")
            row += 1
            _excel_write_kv(ws_form, row, [("Status", "Esta OS nao possui formulario preenchido.")])
        else:
            row = _build_admin_os_excel_form_sheet(ws_form, row, ordem)

        ws_form.freeze_panes = "A5"
        _excel_auto_width(ws_form, max_col=2, min_w=18, max_w=75)

        want_signatures = args.get("assinaturas", "0") == "1"
        if want_signatures and ordem:
            _build_admin_os_signatures_sheet(workbook, solicitacao.id, ordem)

        bio = BytesIO()
        workbook.save(bio)
        bio.seek(0)
        return bio, f"os_{solicitacao.id}_formulario.xlsx"
    except Exception:
        db.session.rollback()
        raise


def _build_admin_os_excel_form_sheet(ws_form, row, ordem):
    _excel_add_section(ws_form, row, "Responsavel / Registro")
    row += 1
    row = _excel_write_kv(ws_form, row, [
        ("Identificador OS", ordem.identificador_os or ""),
        ("Respondido por", ordem.respondido_por or ""),
        ("Respondido em", _fmt_dt(ordem.respondido_em)),
    ])

    row += 1
    _excel_add_section(ws_form, row, "Aeronaves - Pulverizacao (Principal)")
    row += 1
    row = _excel_write_kv(ws_form, row, [
        ("Drone ID", ordem.drone_id or ""),
        ("Prefixo", ordem.prefixo_aeronave_pulverizacao or ""),
        ("Denominacao", ordem.drone_denominacao or ""),
        ("Modelo", ordem.drone_modelo or ""),
        ("N Serie", ordem.drone_numero_serie or ""),
        ("Registro ANATEL", ordem.drone_registro_anatel or ""),
        ("Registro ANAC", ordem.drone_registro_anac or ""),
    ])

    row += 1
    _excel_add_section(ws_form, row, "Aeronaves - Monitoramento")
    row += 1
    row = _excel_write_kv(ws_form, row, [
        ("Drone Monitoramento ID", ordem.drone_monitoramento_id or ""),
        ("Prefixo", ordem.prefixo_aeronave_monitoramento or ""),
        ("Denominacao", ordem.drone_monitoramento_denominacao or ""),
        ("Modelo", ordem.drone_monitoramento_modelo or ""),
        ("N Serie", ordem.drone_monitoramento_numero_serie or ""),
        ("Registro ANATEL", ordem.drone_monitoramento_registro_anatel or ""),
        ("Registro ANAC", ordem.drone_monitoramento_registro_anac or ""),
    ])

    row += 1
    _excel_add_section(ws_form, row, "Aplicacao")
    row += 1
    row = _excel_write_kv(ws_form, row, [
        ("Situacao da aplicacao", ordem.situacao_aplicacao or ""),
        ("Larva visualizada", ordem.larva_visualizada or ""),
        ("Retornar monitorar larvas", ordem.retornar_proxima_semana_monitorar_larvas or ""),
        ("DA (Distrito)", ordem.distrito_administrativo or ""),
        ("Nome/RF ACE responsavel", ordem.nome_rf_ace_responsavel_os or ""),
        ("Criadouro OS (tipo/volume)", ordem.criadouro_os_tipo_volume or ""),
        ("Data aplicacao", _fmt_date(ordem.data_aplicacao)),
        ("Hora inicio", _fmt_time(ordem.hora_inicio_aplicacao)),
        ("Hora termino", _fmt_time(ordem.hora_termino_aplicacao)),
        ("Tratamento adicional", ordem.tratamento_adicional_realizado or ""),
        ("Quantos / Quais", ordem.quantos_quais or ""),
    ])

    row += 1
    planejado_items = _build_planejado_dosagem_items(ordem)
    if planejado_items:
        _excel_add_section(ws_form, row, "Calculo Planejado de Dosagem")
        row += 1
        row = _excel_write_kv(ws_form, row, planejado_items)
        row += 1

    _excel_add_section(ws_form, row, "Produto e Parametros")
    row += 1
    row = _excel_write_kv(ws_form, row, [
        ("Descricao produto", ordem.descricao_produto or ""),
        ("Formulacao", ordem.formulacao_produto or ""),
        ("Dosagem (g/10L)", ordem.dosagem_g_10l or ""),
        ("Tipo aplicacao", ordem.tipo_aplicacao or ""),
        ("Qtd administrada (ml)", ordem.quantidade_produto_administrada_ml or ""),
        ("Pulverizacao area (l/ha)", ordem.pulverizacao_area_l_ha or ""),
        ("Ponta pulverizacao", ordem.ponta_pulverizacao or ""),
    ])

    row += 1
    _excel_add_section(ws_form, row, "Condicoes Ambientais")
    row += 1
    row = _excel_write_kv(ws_form, row, [
        ("Imagens registradas", ordem.quantidade_imagens_registradas or ""),
        ("Videos registrados", ordem.quantidade_videos_registradas or ""),
        ("Temperatura (C)", ordem.temperatura_c or ""),
        ("Umidade (%)", ordem.umidade_relativa_pct or ""),
        ("Vento (km/h)", ordem.velocidade_vento_kmh or ""),
    ])

    row += 1
    _excel_add_section(ws_form, row, "Fechamento")
    row += 1
    row = _excel_write_kv(ws_form, row, [
        ("Observacoes gerais", ordem.observacoes or ""),
        ("Motivo nao realizacao", ordem.motivo_nao_realizacao or ""),
        ("Piloto", ordem.piloto or ""),
        ("Auxiliar", ordem.auxiliar or ""),
        ("Proprietario/Preposto", ordem.proprietario_ou_preposto or ""),
    ])

    row += 1
    _excel_add_section(ws_form, row, "Assinaturas")
    row += 1

    has_piloto = bool(ordem.assinatura_piloto and str(ordem.assinatura_piloto).startswith("data:image"))
    has_resp = bool(
        ordem.assinatura_proprietario_ou_preposto
        and str(ordem.assinatura_proprietario_ou_preposto).startswith("data:image")
    )
    _excel_write_kv(ws_form, row, [
        ("Assinatura piloto", "OK" if has_piloto else "Nao informada"),
        ("Assinatura responsavel", "OK" if has_resp else "Nao informada"),
    ])
    return row


def _build_admin_os_signatures_sheet(workbook, solicitacao_id, ordem):
    ws_sig = workbook.create_sheet("Assinaturas")
    _excel_add_title(ws_sig, f"Assinaturas - OS #{solicitacao_id}", "Imagens exportadas do formulario")

    ws_sig.column_dimensions["A"].width = 26
    ws_sig.column_dimensions["B"].width = 55
    ws_sig["A4"] = "Piloto"
    ws_sig["A4"].font = Font(bold=True)
    ws_sig["A5"] = ordem.piloto or ""
    ws_sig["A8"] = "Responsavel"
    ws_sig["A8"].font = Font(bold=True)
    ws_sig["A9"] = ordem.proprietario_ou_preposto or ""

    png1 = _dataurl_to_png_bytes(getattr(ordem, "assinatura_piloto", None))
    if png1:
        tmp1 = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        tmp1.write(png1)
        tmp1.close()
        img1 = XLImage(tmp1.name)
        img1.width = 420
        img1.height = 140
        ws_sig.add_image(img1, "B4")
    else:
        ws_sig["B4"] = "Assinatura nao informada"

    png2 = _dataurl_to_png_bytes(getattr(ordem, "assinatura_proprietario_ou_preposto", None))
    if png2:
        tmp2 = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        tmp2.write(png2)
        tmp2.close()
        img2 = XLImage(tmp2.name)
        img2.width = 420
        img2.height = 140
        ws_sig.add_image(img2, "B8")
    else:
        ws_sig["B8"] = "Assinatura nao informada"
