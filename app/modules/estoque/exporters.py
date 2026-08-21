from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

STATUS_LABELS = {
    "disponivel_manutencao": "Disponível para manutenção",
    "reservada": "Reservada",
    "baixada": "Baixada",
    "indisponivel": "Indisponível",
}


def _fmt_dt(value):
    if not value:
        return ""
    try:
        return value.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(value)


def _safe(value):
    return "" if value is None else str(value)


def _style_worksheet(ws, *, title):
    max_col = ws.max_column
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color="1F4F82")

    header_fill = PatternFill("solid", fgColor="1F4F82")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D9E2EF"),
        right=Side(style="thin", color="D9E2EF"),
        top=Side(style="thin", color="D9E2EF"),
        bottom=Side(style="thin", color="D9E2EF"),
    )

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def build_estoque_excel(pecas):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Estoque"
    ws.append([""])
    ws.append([
        "ID",
        "Peça",
        "Nº de série",
        "Quantidade",
        "Status",
        "Drone",
        "Modelo drone",
        "Equipe",
        "Observações",
        "Criado em",
        "Atualizado em",
    ])

    for peca in pecas:
        drone = peca.drone
        ws.append([
            peca.id,
            _safe(peca.modelo_peca),
            _safe(peca.numero_serie),
            peca.quantidade or 0,
            STATUS_LABELS.get(peca.status, _safe(peca.status)),
            _safe(getattr(drone, "renomacao", None)),
            _safe(getattr(drone, "modelo", None)),
            _safe(getattr(getattr(drone, "equipe", None), "nome_equipe", None)),
            _safe(peca.observacoes),
            _fmt_dt(peca.criado_em),
            _fmt_dt(peca.atualizado_em),
        ])

    _style_worksheet(ws, title="Estoque geral de peças")
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"estoque_geral_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename
