from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models import Clientes
from app.shared.formatters import (
    format_cep,
    format_documento,
    format_phone_br,
    only_digits,
)
from app.shared.validators import validate_documento, validate_email


def build_endereco_full(cep, logradouro, numero, complemento, bairro, cidade, uf) -> str:
    cep = only_digits(cep)
    logradouro = (logradouro or "").strip()
    numero = (numero or "").strip()
    complemento = (complemento or "").strip()
    bairro = (bairro or "").strip()
    cidade = (cidade or "").strip()
    uf = (uf or "").strip().upper()

    linha_1 = ""
    if logradouro:
        linha_1 += logradouro
    if numero:
        linha_1 += f", {numero}" if linha_1 else numero
    if complemento:
        linha_1 += f" ({complemento})" if linha_1 else complemento

    cidade_uf = f"{cidade}/{uf}" if cidade and uf else cidade or uf
    linha_2 = " - ".join([item for item in [bairro, cidade_uf] if item])

    cep_formatado = format_cep(cep) if cep else ""
    linha_3 = f"CEP {cep_formatado}" if cep_formatado else ""

    return " - ".join([item for item in [linha_1, linha_2, linha_3] if item]).strip()


def build_clientes_query(q: str, documento: str, email: str, telefone: str, sort: str):
    query = Clientes.query

    if documento:
        query = query.filter(Clientes.documento.ilike(f"%{only_digits(documento)}%"))

    if email:
        query = query.filter(Clientes.email.ilike(f"%{email}%"))

    if telefone:
        query = query.filter(Clientes.telefone.ilike(f"%{only_digits(telefone)}%"))

    if q:
        like = f"%{q}%"
        q_digits = only_digits(q)
        query = query.filter(
            db.or_(
                Clientes.nome_cliente.ilike(like),
                Clientes.contato.ilike(like),
                Clientes.email.ilike(like),
                Clientes.endereco.ilike(like),
                Clientes.documento.ilike(f"%{q_digits}%") if q_digits else db.false(),
                Clientes.telefone.ilike(f"%{q_digits}%") if q_digits else db.false(),
            )
        )

    if sort == "nome_desc":
        return query.order_by(Clientes.nome_cliente.desc())
    if sort == "id_desc":
        return query.order_by(Clientes.id.desc())
    if sort == "id_asc":
        return query.order_by(Clientes.id.asc())
    return query.order_by(Clientes.nome_cliente.asc())


def serialize_clientes(clientes):
    return [
        {
            "id": cliente.id,
            "nome_cliente": cliente.nome_cliente,
            "documento_fmt": format_documento(cliente.documento),
            "contato": cliente.contato or "-",
            "telefone_fmt": format_phone_br(cliente.telefone or "") or "-",
            "email": cliente.email or "-",
            "endereco": cliente.endereco or "-",
        }
        for cliente in clientes
    ]


def build_clientes_export(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clientes"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet["A1"] = "Relatorio de Clientes"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sheet["A2"].font = Font(color="6B7280")

    start_row = 4
    headers = ["ID", "Nome", "Documento", "Contato", "Telefone", "E-mail", "Endereco"]

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for row_idx, cliente in enumerate(rows, start=start_row + 1):
        values = [
            cliente.id,
            cliente.nome_cliente,
            format_documento(cliente.documento),
            cliente.contato or "",
            format_phone_br(cliente.telefone or ""),
            cliente.email or "",
            cliente.endereco or "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = center_align if col_idx == 1 else text_align
            if col_idx in (3, 5):
                cell.number_format = "@"

    last_row = start_row + len(rows)
    last_col = len(headers)

    sheet.freeze_panes = sheet["A5"]
    sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(last_col)}{max(last_row, start_row)}"
    sheet.row_dimensions[start_row].height = 22

    max_widths = {1: 8, 2: 28, 3: 22, 4: 18, 5: 18, 6: 26, 7: 45}
    for col_idx in range(1, last_col + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(start_row + 1, last_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))

        width = min(max_len + 2, max_widths.get(col_idx, 40))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    zebra_fill = PatternFill("solid", fgColor="F9FAFB")
    for row_idx in range(start_row + 1, last_row + 1):
        if (row_idx - (start_row + 1)) % 2 == 1:
            for col_idx in range(1, last_col + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = zebra_fill

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"clientes_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    return output, filename
