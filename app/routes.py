# ==========================
# IMPORTS PADRÃO PYTHON
# ==========================
import os
import re
import tempfile
import unicodedata
import math
from datetime import date, datetime
from io import BytesIO
import json
from apscheduler.schedulers.background import BackgroundScheduler
from zoneinfo import ZoneInfo

from flask_login import login_required, current_user
from pyparsing import col
from werkzeug.utils import secure_filename
import uuid
import os
import threading
from flask import render_template, url_for
from flask_login import login_required, current_user

# ==========================
# FLASK
# ==========================
from flask import (Blueprint, after_this_request, current_app, flash, jsonify,
                redirect, render_template, request, send_file,
                send_from_directory, url_for)


from flask_login import current_user , login_required

# ==========================
# EXCEL / PDF
# ==========================
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import landscape
# ==========================
# SQLALCHEMY / BANCO
# ==========================
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload, selectinload

# ==========================
# APP
# ==========================
from app import db
from app.models import (
    Abastecimento,
    Equipe,
    EquipePiloto,
    EquipeUvis,
    LogVeiculo,
    Notificacao,
    OrdemServico,
    Pilotos,
    Solicitacao,
    Usuario,
    Veiculos,
)
from app.modules.clientes import register_routes as register_clientes_routes
from app.modules.equipes import register_routes as register_equipes_routes
from app.modules.pilotos import register_routes as register_pilotos_routes
from app.modules.usuarios import register_routes as register_usuarios_routes
TZ = ZoneInfo("America/Sao_Paulo")
print("--- ROTAS CARREGADAS COM SUCESSO ---")

bp = Blueprint('main', __name__)
register_clientes_routes(bp)
register_equipes_routes(bp)
register_pilotos_routes(bp)
register_usuarios_routes(bp)

# --- 1: GLOBAL CONTEXT  ---
@bp.context_processor
def inject_globals():
    """
    Otimizado e blindado contra erros de transação. 
    Se o banco falhar, retorna 0 notificações mas não derruba o sistema.
    """
    if current_user.is_authenticated:
        try:
            # Consulta de contagem direta no banco
            q = db.session.query(db.func.count(Notificacao.id)).filter(
                Notificacao.lida_em.is_(None),
                Notificacao.apagada_em.is_(None)
            )
            
            if current_user.tipo_usuario not in ["admin", "operario", "visualizar"]:
                q = q.filter(Notificacao.usuario_id == current_user.id)
                
            return dict(notif_count=q.scalar() or 0)
        except Exception as e:
            # Se houver erro de transação (Transaction Aborted), limpamos aqui
            db.session.rollback()
            # Opcional: print(f"Erro no inject_globals (notificações): {e}")
            return dict(notif_count=0) 
            
    return dict(notif_count=0)

def inject_google_key():
    return {"google_maps_key": current_app.config.get("KEY_API_GOOGLE_MAPS") or os.getenv("KEY_API_GOOGLE_MAPS")}


# --- 2: FILTRO DE DATA PARA JINJA2 ---
@bp.app_template_filter('datetimeformat')
def datetimeformat(value, format='%d-%m-%y'):
    """
    Filtro para formatar datas no Jinja2.
    Otimizado para lidar com strings, objetos datetime e valores nulos.
    """
    if value is None:
        return ""
    try:
        if isinstance(value, str):
            # Se for string (ex: '2025-12-31'), converte para objeto date antes de formatar
            return datetime.strptime(value, "%Y-%m-%d").strftime(format)
        return value.strftime(format)
    except Exception:
        return value # Retorna o valor original em caso de erro para não quebrar a página
    
# --- 3: CÁLCULO DE DISTÂNCIA ENTRE COORDENADAS ---
def calcular_distancia(lat1, lon1, lat2, lon2):
    # Raio da Terra em metros
    R = 6371000 
    
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    
    a = math.sin(dphi / 2)**2 + \
        math.cos(phi1) * math.cos(phi2) * \
        math.sin(dlambda / 2)**2
    
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


AREAS_GEOFENCING = [
    {"nome": "Aeroporto de Congonhas (CGH)", "lat": -23.6273, "lng": -46.6565, "raio": 5400},
    {"nome": "Aeroporto Campo de Marte (RTE)", "lat": -23.5092, "lng": -46.6377, "raio": 5400},
    {"nome": "Aeroporto de Guarulhos (GRU)", "lat": -23.4356, "lng": -46.4731, "raio": 9000},
    {"nome": "Aeroporto de Viracopos (VCP)", "lat": -23.0069, "lng": -47.1344, "raio": 9000},
    {"nome": "Zona de Helipontos (Av. Paulista)", "lat": -23.5615, "lng": -46.6559, "raio": 2000},
    {"nome": "Base Aérea de Santos", "lat": -23.9275, "lng": -46.2975, "raio": 5400},
]


def detectar_area_restrita(latitude, longitude):
    if latitude is None or longitude is None:
        return False

    for area in AREAS_GEOFENCING:
        distancia = calcular_distancia(latitude, longitude, area["lat"], area["lng"])
        if distancia < area["raio"]:
            return True

    return False

def get_upload_folder():
    """
    Localiza a pasta de uploads de forma absoluta.
    Garante que a pasta exista sem processamento repetitivo desnecessário.
    """
    # Pasta 'upload-files' no mesmo nível da pasta 'app'
    folder = os.path.join(current_app.root_path, '..', 'upload-files')
    if not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
    return os.path.abspath(folder)

import unicodedata

def normalize_string(value):
    if value:
        return ''.join(
            c for c in unicodedata.normalize('NFD', value)
            if unicodedata.category(c) != 'Mn'
        ).lower()
    return value


def allowed_file(filename: str) -> bool:
    """Verifica se a extensão do arquivo é permitida."""
    ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "doc", "docx", "xls", "xlsx"}
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

from sqlalchemy import extract, cast, Integer, func

def aplicar_filtros_base(query, filtro_data, uvis_id):
    if filtro_data:
        try:
            # filtro_data = "2026-01"
            ano, mes = map(int, filtro_data.split('-'))
            
            # Forçamos a comparação de INTEIRO com INTEIRO
            query = query.filter(
                cast(extract('year', Solicitacao.data_agendamento), Integer) == ano,
                cast(extract('month', Solicitacao.data_agendamento), Integer) == mes
            )
            print(f"DEBUG SQL: Filtrando por Ano={ano} e Mes={mes}")
        except Exception as e:
            print(f"Erro no filtro de data: {e}")

    if uvis_id:
        query = query.filter(Solicitacao.usuario_id == int(uvis_id))
            
    return query

from functools import wraps
from flask import abort
from flask_login import current_user

def roles_required(*roles):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                abort(401)
            if current_user.tipo_usuario not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return deco

import os
from datetime import datetime
from flask import request, redirect, url_for, render_template
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from sqlalchemy import func

# ajuste imports conforme seu projeto:
# from app import db
# from app.models import Solicitacao, Equipe, EquipeUvis, EquipeUvis...

@bp.route('/')
@login_required
def dashboard():
    google_maps_key = os.getenv("KEY_API_GOOGLE_MAPS")  # a mesma do .env

    if current_user.tipo_usuario == 'piloto':
        return redirect(url_for('main.piloto_os'))

    if current_user.tipo_usuario == "equipe_uvis":
        return redirect(url_for("main.dashboard_equipe_uvis"))

    if current_user.tipo_usuario in ['admin', 'operario', 'visualizar']:
        return redirect(url_for('main.admin_dashboard'))

    #  UVIS: só as solicitações dela + carrega equipe para exibir
    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe)
        )
        .filter(Solicitacao.usuario_id == current_user.id)
    )

    #  NÃO MOSTRAR CANCELADAS NO DASHBOARD PRINCIPAL
    query = query.filter(Solicitacao.status != "CANCELADO")

    # =========================
    # FILTROS (status, tipo, foco)
    # =========================
    filtro_status = request.args.get('status')
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    filtro_tipo_visita = request.args.get('tipo_visita')
    if filtro_tipo_visita:
        query = query.filter(Solicitacao.tipo_visita == filtro_tipo_visita)

    filtro_foco = request.args.get('foco')
    if filtro_foco:
        query = query.filter(Solicitacao.foco == filtro_foco)

    # =========================
    #  FILTRO POR DATA (NOVO)
    # =========================
    data_ini = request.args.get("data_ini")  # YYYY-MM-DD
    data_fim = request.args.get("data_fim")  # YYYY-MM-DD

    if data_ini:
        try:
            dt_ini = datetime.strptime(data_ini, "%Y-%m-%d").date()
            query = query.filter(Solicitacao.data_agendamento >= dt_ini)
        except ValueError:
            pass

    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, "%Y-%m-%d").date()
            query = query.filter(Solicitacao.data_agendamento <= dt_fim)
        except ValueError:
            pass

    # =========================
    # PAGINAÇÃO
    # =========================
    page = request.args.get("page", 1, type=int)
    paginacao = (
        query.order_by(Solicitacao.data_criacao.desc())
        .paginate(page=page, per_page=6, error_out=False)
    )

    # =========================
    # EQUIPES (ATIVAS / POR REGIÃO)
    # =========================
    equipes_query = Equipe.query.filter_by(ativa=True)

    if current_user.regiao:
        equipes_query = equipes_query.filter(Equipe.regiao == current_user.regiao)

    equipes = equipes_query.order_by(Equipe.nome_equipe.asc()).all()

    # =========================
    # EQUIPES UVIS (contagem)
    # =========================
    rows = (
        db.session.query(EquipeUvis.nome_equipe, func.count(EquipeUvis.id).label("total"))
        .filter(EquipeUvis.uvis_usuario_id == current_user.id)
        .group_by(EquipeUvis.nome_equipe)
        .order_by(EquipeUvis.nome_equipe.asc())
        .all()
    )

    equipes_uvis = [{"nome_equipe": r[0], "total": int(r[1])} for r in rows]

    return render_template(
        "dashboard.html",
        solicitacoes=paginacao.items,
        paginacao=paginacao,
        google_maps_key=google_maps_key,
        equipes_uvis=equipes_uvis
    )

# --- PAINEL DE GESTÃO (Visualização para todos) ---
from flask_login import login_required, current_user
from datetime import datetime


from sqlalchemy import case # Necessário para a ordenação personalizada
from datetime import datetime
from sqlalchemy.orm import joinedload

from sqlalchemy import case
# ... (se já tiver importado, ignora)
# from flask import request, redirect, url_for, flash, render_template
# from flask_login import login_required, current_user

@bp.route('/admin')
@login_required
def admin_dashboard():
    google_maps_key = os.getenv("KEY_API_GOOGLE_MAPS")

    # 🔐 Controle de acesso
    if current_user.tipo_usuario not in ['admin', 'operario', 'visualizar']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('main.dashboard'))

    # Pode editar apenas admin e operario
    is_editable = current_user.tipo_usuario in ['admin', 'operario']

    # --- Captura filtros ---
    filtro_status = (request.args.get("status") or "").strip()
    filtro_unidade = (request.args.get("unidade") or "").strip()
    filtro_regiao = (request.args.get("regiao") or "").strip()

    #  novo filtro (SIM / NAO)
    filtro_apoio_cet = (request.args.get("apoio_cet") or "").strip().upper()

    # 🔁 Se alguém tentar acessar CANCELADO pelo filtro, redireciona
    if filtro_status == "CANCELADO":
        return redirect(url_for(
            "main.admin_canceladas",
            unidade=filtro_unidade,
            regiao=filtro_regiao,
            apoio_cet=filtro_apoio_cet
        ))

    unidades_select = (
        Usuario.query
        .filter_by(tipo_usuario='uvis')
        .order_by(Usuario.nome_uvis.asc())
        .all()
    )

    # --- Query base ---
    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe)
        )
        .join(Usuario)
    )

    #  NÃO MOSTRAR CANCELADAS NO PAINEL PRINCIPAL
    query = query.filter(Solicitacao.status != "CANCELADO")

    # --- Aplicação dos filtros ---
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_unidade:
        query = query.filter(Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"))

    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))

    #  filtro apoio CET
    if filtro_apoio_cet == "SIM":
        query = query.filter(Solicitacao.apoio_cet.is_(True))
    elif filtro_apoio_cet == "NAO":
        query = query.filter(Solicitacao.apoio_cet.is_(False))

    # --- Ordenação personalizada ---
    ordem_status = case(
        {
            'PENDENTE': 1,
            'EM ANÁLISE': 2,
            'APROVADO COM RECOMENDAÇÕES': 3,
            'APROVADO': 4,
            'NEGADO': 5,
            'CONCLUÍDO': 6,
        },
        value=Solicitacao.status,
        else_=99
    )

    equipes = (
        Equipe.query
        .filter(Equipe.ativa.is_(True))
        .order_by(Equipe.regiao.asc(), Equipe.nome_equipe.asc())
        .all()
    )

    # --- Paginação ---
    page = request.args.get("page", 1, type=int)

    paginacao = (
        query.order_by(ordem_status, Solicitacao.data_criacao.desc())
        .paginate(page=page, per_page=6, error_out=False)
    )

    return render_template(
        'admin.html',
        pedidos=paginacao.items,
        paginacao=paginacao,
        is_editable=is_editable,
        now=datetime.now(),
        equipes=equipes,
        unidades_select=unidades_select,
        google_maps_key=google_maps_key
    )

@bp.route('/admin/exportar_excel')
@login_required
def exportar_excel():

    # 🔐 Permissão: somente admin e operario
    if current_user.tipo_usuario not in ['admin', 'operario' , 'visualizar']:
        flash('Permissão negada para exportar.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    try:
        filtro_status = (request.args.get("status") or "").strip()
        filtro_unidade = (request.args.get("unidade") or "").strip()
        filtro_regiao = (request.args.get("regiao") or "").strip()
        filtro_apoio_cet = (request.args.get("apoio_cet") or "").strip().upper()

        query = (
            db.session.query(Solicitacao)
            .join(Usuario)
            .options(
                joinedload(Solicitacao.usuario),
                joinedload(Solicitacao.equipe)  #  agora puxa equipe
            )
        )

        if filtro_status:
            query = query.filter(Solicitacao.status == filtro_status)

        if filtro_unidade:
            query = query.filter(Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"))

        if filtro_regiao:
            query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))

        if filtro_apoio_cet:
            query = query.filter(Usuario.regiao.ilike(f"%{filtro_apoio_cet}%"))

        pedidos = query.order_by(Solicitacao.data_criacao.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Solicitações"

        headers = [
            "ID", "Unidade", "Região", "Equipe Responsável",  #  adicionada
            "Data Agendada", "Hora",
            "Endereço Completo", "Latitude", "Longitude",
            "Foco", "Tipo Visita", "Altura", "Apoio CET?",
            "Observação", "Status", "Protocolo", "Justificativa"
        ]

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_num, p in enumerate(pedidos, 2):
            uvis_nome = p.usuario.nome_uvis if p.usuario else "Não informado"
            uvis_regiao = p.usuario.regiao if p.usuario else "Não informado"

            equipe_nome = ""
            if getattr(p, "equipe", None):
                equipe_nome = p.equipe.nome_equipe or ""
            elif getattr(p, "equipe_id", None):
                equipe_nome = f"ID #{p.equipe_id}"

            endereco_completo = (
                f"{p.logradouro or ''}, {p.numero or ''} - "
                f"{p.bairro or ''} - "
                f"{(p.cidade or '')}/{(p.uf or '')} - {p.cep or ''}"
            )
            if p.complemento:
                endereco_completo += f" - {p.complemento}"

            data_formatada = ""
            if p.data_agendamento:
                if isinstance(p.data_agendamento, (date, datetime)):
                    data_formatada = p.data_agendamento.strftime("%d/%m/%Y")
                else:
                    data_formatada = str(p.data_agendamento)

            row = [
                p.id,
                uvis_nome,
                uvis_regiao,
                equipe_nome,  #  adicionada
                data_formatada,
                str(p.hora_agendamento or ""),
                endereco_completo,
                p.latitude or "",
                p.longitude or "",
                p.foco,
                p.tipo_visita or "",
                p.altura_voo or "",
                "SIM" if p.apoio_cet else "NÃO",
                p.observacao or "",
                p.status,
                p.protocolo or "",
                p.justificativa or ""
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            download_name="relatorio_solicitacoes.xlsx",
            as_attachment=False,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        db.session.rollback()
        print(f"ERRO EXPORTAR EXCEL: {e}")
        flash("Erro ao gerar o Excel. Verifique se os dados estão corretos.", "danger")
        return redirect(url_for('main.admin_dashboard'))


from flask import request, jsonify, current_app, redirect, url_for, flash
import os, uuid
from werkzeug.utils import secure_filename

@bp.route('/admin/atualizar/<int:id>', methods=['POST'])
@login_required
def atualizar(id):

    # 🔐 Permissão
    if current_user.tipo_usuario not in ['admin', 'operario']:
        if request.accept_mimetypes.accept_html and not request.is_json:
            flash("Permissão negada.", "danger")
            return redirect(request.referrer or url_for("main.admin_dashboard"))
        return jsonify({"error": "Permissão negada"}), 403

    pedido = Solicitacao.query.get_or_404(id)

    # --- Atualização de campos ---
    pedido.protocolo = request.form.get('protocolo')
    pedido.status = request.form.get('status')
    pedido.justificativa = request.form.get('justificativa')
    pedido.latitude = request.form.get('latitude')
    pedido.longitude = request.form.get('longitude')

    #  Atribuição de equipe (opcional)
    equipe_id = request.form.get("equipe_id")

    if equipe_id in (None, "", "null", "undefined"):
        pedido.equipe_id = None
        equipe_nome = None
    else:
        try:
            equipe_id_int = int(equipe_id)
            equipe = Equipe.query.get(equipe_id_int)
            if not equipe:
                flash("Equipe selecionada não existe.", "warning")
                return redirect(request.referrer or url_for("main.admin_dashboard"))

            pedido.equipe_id = equipe_id_int
            equipe_nome = equipe.nome_equipe

        except ValueError:
            flash("Equipe inválida.", "warning")
            return redirect(request.referrer or url_for("main.admin_dashboard"))

    #  Regra de negócio: se aprovou, precisa ter equipe
    status_aprovacao = ["APROVADO", "APROVADO COM RECOMENDAÇÕES"]
    if pedido.status in status_aprovacao and not pedido.equipe_id:
        flash("Para aprovar, selecione uma equipe responsável.", "warning")
        return redirect(request.referrer or url_for("main.admin_dashboard"))

    # Processamento de Anexo
    file = request.files.get("anexo")
    if file and file.filename:
        if allowed_file(file.filename):
            original_filename = secure_filename(file.filename)
            ext = original_filename.rsplit(".", 1)[1].lower()
            unique_name = f"sol_{pedido.id}_{uuid.uuid4().hex}.{ext}"
            upload_folder = get_upload_folder()
            file_path = os.path.join(upload_folder, unique_name)

            try:
                file.save(file_path)
                pedido.anexo_path = f"upload-files/{unique_name}"
                pedido.anexo_nome = original_filename
            except Exception as e:
                current_app.logger.error(f"Erro ao salvar arquivo físico: {e}")
                if request.accept_mimetypes.accept_html and not request.is_json:
                    flash("Falha ao salvar o arquivo no servidor.", "danger")
                    return redirect(request.referrer or url_for("main.admin_dashboard"))
                return jsonify({"error": "Falha ao salvar o arquivo no servidor."}), 500
        else:
            if request.accept_mimetypes.accept_html and not request.is_json:
                flash("Formato de arquivo não permitido.", "warning")
                return redirect(request.referrer or url_for("main.admin_dashboard"))
            return jsonify({"error": "Formato de arquivo não permitido."}), 400

    # commit final
    try:
        db.session.commit()

        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.is_json
        if is_ajax:
            return jsonify({
                "ok": True,
                "message": "Solicitação atualizada com sucesso!",
                "anexo_nome": pedido.anexo_nome,
                "equipe_id": pedido.equipe_id,
                "equipe_nome": equipe_nome,
            }), 200

        flash("Solicitação atualizada com sucesso!", "success")
        return redirect(request.referrer or url_for("main.admin_dashboard"))

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro de Banco (Atualizar ID {id}): {e}")

        if request.accept_mimetypes.accept_html and not request.is_json:
            flash("Erro ao gravar dados no banco de dados.", "danger")
            return redirect(request.referrer or url_for("main.admin_dashboard"))

        return jsonify({"error": "Erro ao gravar dados no banco de dados."}), 500


    

import requests
from urllib.parse import urlencode

def geocode_endereco_google(*, logradouro, numero, bairro, cidade, uf, cep=None):
    """
    Retorna (lat, lng) ou (None, None) se não achar.
    """
    api_key = current_app.config.get("Maps_KEY_BACK") or os.getenv("GOOGLE_MAPS_KEY_BACK")
    
    # Validação de segurança
    if not api_key:
        raise RuntimeError("Maps_KEY_BACK não encontrada nas configurações do App")
    # Monta um endereço bem “forte” pro Google
    partes = [
        (logradouro or "").strip(),
        (numero or "").strip(),
        (bairro or "").strip(),
        (cidade or "").strip(),
        (uf or "").strip(),
    ]
    if cep:
        partes.append((cep or "").strip())
    partes.append("Brasil")

    address = ", ".join([p for p in partes if p])

    params = {
        "address": address,
        "key": api_key,
        "region": "br",   # ajuda a “puxar” para Brasil
    }

    url = "https://maps.googleapis.com/maps/api/geocode/json?" + urlencode(params)
    resp = requests.get(url, timeout=10)
    data = resp.json()

    status = data.get("status")
    if status != "OK":
        return None, None

    results = data.get("results") or []
    if not results:
        return None, None

    loc = results[0]["geometry"]["location"]
    return loc.get("lat"), loc.get("lng")
    
from datetime import date, datetime
from flask import render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user

# -----------------------------------------
# ENDPOINT PARA GEOCODE (AJAX)
# -----------------------------------------
@bp.route("/api/geocode", methods=["POST"], endpoint="api_geocode")
@login_required
def api_geocode():
    try:
        data = request.get_json(silent=True) or {}

        logradouro = (data.get("logradouro") or "").strip()
        numero = (data.get("numero") or "").strip()
        bairro = (data.get("bairro") or "").strip()
        cidade = (data.get("cidade") or "").strip()
        uf = (data.get("uf") or "").strip()
        cep = (data.get("cep") or "").strip()

        # Mínimo necessário
        if not logradouro or not numero or not cidade or not uf:
            return jsonify({"ok": False, "message": "Endereço incompleto"}), 200

        lat, lng = geocode_endereco_google(
            logradouro=logradouro,
            numero=numero,
            bairro=bairro,
            cidade=cidade,
            uf=uf,
            cep=cep,
        )

        if lat is None or lng is None:
            return jsonify({"ok": False, "message": "Não foi possível geocodificar"}), 200

        return jsonify({"ok": True, "lat": lat, "lng": lng}), 200

    except Exception as e:
        print(f"ERRO /api/geocode: {e}")
        return jsonify({"ok": False, "message": "Erro interno"}), 200


# -----------------------------------------
# --- NOVO PEDIDO ---
# -----------------------------------------
@bp.route('/novo_cadastro', methods=['GET', 'POST'], endpoint='novo')
@login_required
def novo():
    hoje = date.today().isoformat()
    key_for_map = current_app.config.get("Maps_KEY_FRONT") or os.getenv("KEY_API_GOOGLE_MAPS")

    # --- NOVO: Busca lista de UVIS para a COVISA/Admin escolher ---
    uvis_lista = []
    if current_user.tipo_usuario in ['admin', 'visualizar']:
        uvis_lista = Usuario.query.filter_by(tipo_usuario='uvis').order_by(Usuario.nome_uvis.asc()).all()

    if request.method == 'POST':
        try:
            # --- Dados de Data/Hora ---
            data_str = request.form.get('data')
            hora_str = request.form.get('hora')
            data_obj = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
            hora_obj = datetime.strptime(hora_str, '%H:%M').time() if hora_str else None

            perimetro_planejado = request.form.get('perimetro_planejado')
            
            # --- LÓGICA DE ATRIBUIÇÃO DE USUÁRIO ---
            # Se for COVISA ou Admin, usa o ID vindo do select. Se for UVIS, usa o próprio ID.
            if current_user.tipo_usuario in ['admin', 'visualizar']:
                uvis_id_final = request.form.get('uvis_responsavel_id')
                if not uvis_id_final:
                    flash("Por favor, selecione a UVIS responsável.", "warning")
                    return render_template('cadastro.html', hoje=hoje, google_maps_key=key_for_map, uvis_lista=uvis_lista)
            else:
                uvis_id_final = current_user.id

            # ... (Restante dos dados de endereço e coordenadas mantidos) ...
            lat_raw = (request.form.get('latitude') or "").strip()
            lng_raw = (request.form.get('longitude') or "").strip()
            latitude = float(lat_raw.replace(",", ".")) if lat_raw else None
            longitude = float(lng_raw.replace(",", ".")) if lng_raw else None
            area_restrita = detectar_area_restrita(latitude, longitude) or request.form.get('risco_aereo') == '1'

            # --- Criação da Solicitação ---
            nova_solicitacao = Solicitacao(
                data_agendamento=data_obj,
                hora_agendamento=hora_obj,
                cep=request.form.get('cep'),
                logradouro=request.form.get('logradouro'),
                bairro=request.form.get('bairro'),
                cidade=request.form.get('cidade'),
                numero=request.form.get('numero'),
                uf=request.form.get('uf'),
                complemento=request.form.get('complemento'),
                foco=request.form.get('foco'),
                tipo_visita=request.form.get('tipo_visita'),
                altura_voo=request.form.get('altura_voo'),
                apoio_cet=request.form.get('apoio_cet') == 'sim',
                observacao=request.form.get('observacao'),
                latitude=latitude,
                longitude=longitude,
                area_restrita=area_restrita,
                perimetro_planejado=perimetro_planejado, 
                usuario_id=uvis_id_final, 
                status='PENDENTE'
            )

            db.session.add(nova_solicitacao)
            db.session.commit()

            flash('Solicitação criada e enviada para a UVIS com sucesso!', 'success')
            return redirect(url_for('main.dashboard'))

        except Exception as e:
            db.session.rollback()
            print(f"ERRO NOVO CADASTRO: {e}")
            flash("Erro ao salvar o pedido.", "danger")

    return render_template('cadastro.html', hoje=hoje, google_maps_key=key_for_map, uvis_lista=uvis_lista)

@bp.route("/forcar_erro")
def forcar_erro():
    1 / 0  # erro proposital
    return "nunca vai chegar aqui"

# Openpyxl (Excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
# ReportLab (PDF)
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                Spacer, Table, TableStyle)

# O objeto 'bp' precisa ser definido (Exemplo: bp = Blueprint('main', __name__))
# E 'Usuario' e 'Solicitacao' precisam ser seus modelos SQLAlchemy

# =======================================================================
# Função Auxiliar de Filtros (Reutilizada em todas as rotas)
# =======================================================================

from datetime import datetime

# =======================================================================
# ROTA 1: Visualização do Relatório (HTML)
# =======================================================================
from flask import redirect, render_template, request, session, url_for

from app import db
from app.models import Solicitacao, Usuario
from sqlalchemy import extract, func
from datetime import datetime

@bp.route('/relatorios/solicitacoes', methods=['GET'], endpoint='relatorios_solicitacoes')
def relatorios():
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))

    try:
        # 1. Inicialize variáveis para evitar erro de 'not defined'
        uvis_disponiveis = []
        
        # 2. Busque as UVIS primeiro (se for admin)
        if current_user.tipo_usuario in ['admin', 'operario', 'visualizar']:
            uvis_disponiveis = (
                db.session.query(Usuario.id, Usuario.nome_uvis)
                .filter(Usuario.tipo_usuario == 'uvis')
                .order_by(Usuario.nome_uvis)
                .all()
            )

        # 3. Capture os parâmetros de Filtro
        mes_atual = request.args.get('mes', datetime.now().month, type=int)
        ano_atual = request.args.get('ano', datetime.now().year, type=int)
        uvis_id = request.args.get('uvis_id', type=int) if current_user.tipo_usuario != 'uvis' else current_user.id

        # 4. Monte a string de data para a função de filtro
        filtro_data = f"{ano_atual}-{mes_atual:02d}"

        # 5. Query base (AQUI USAMOS O EXTRACT AUTOMATICAMENTE ATRAVÉS DA FUNÇÃO)
        base_query = aplicar_filtros_base(
            db.session.query(Solicitacao),
            filtro_data,
            uvis_id
        )

        base_query = aplicar_filtros_base(db.session.query(Solicitacao), filtro_data, uvis_id)
        

        # ADICIONE ISSO AQUI:
        print("SQL EXECUTADO:", str(base_query.statement.compile(dialect=db.engine.dialect)))

        # =====================================================
        # 🔹 TOTAIS POR STATUS (JSON-safe)
        # =====================================================
        status_counts = {
            status: total
            for status, total in (
                base_query
                .with_entities(Solicitacao.status, db.func.count(Solicitacao.id))
                .group_by(Solicitacao.status)
                .all()
            )
        }

        total_solicitacoes = sum(status_counts.values())
        total_aprovadas = status_counts.get("APROVADO", 0)
        total_aprovadas_com_recomendacoes = status_counts.get(
            "APROVADO COM RECOMENDAÇÕES", 0
        )
        total_recusadas = status_counts.get("NEGADO", 0)
        total_analise = status_counts.get("EM ANÁLISE", 0)
        total_pendentes = status_counts.get("PENDENTE", 0)

        # =====================================================
        # 🔹 FUNÇÃO GENÉRICA DE AGRUPAMENTO (JSON-safe)
        # =====================================================
        def agrupar_por(campo):
            resultados = (
                base_query
                .with_entities(campo, db.func.count(Solicitacao.id))
                .group_by(campo)
                .order_by(db.func.count(Solicitacao.id).desc())
                .all()
            )

            return [
                (valor or "Não informado", total)
                for valor, total in resultados
            ]

        dados_status = agrupar_por(Solicitacao.status)
        dados_foco = agrupar_por(Solicitacao.foco)
        dados_tipo_visita = agrupar_por(Solicitacao.tipo_visita)
        dados_altura_voo = agrupar_por(Solicitacao.altura_voo)

        # =====================================================
        # 🔹 AGRUPAMENTOS COM JOIN (Corrigido para usar base_query)
        # =====================================================
        dados_regiao = [
            (regiao or "Não informado", total)
            for regiao, total in (
                base_query.join(Usuario)
                .with_entities(Usuario.regiao, db.func.count(Solicitacao.id))
                .group_by(Usuario.regiao)
                .order_by(db.func.count(Solicitacao.id).desc())
                .all()
            )
        ]

        dados_unidade = [
            (uvis or "Não informado", total)
            for uvis, total in (
                base_query.join(Usuario)
                .filter(Usuario.tipo_usuario == 'uvis')
                .with_entities(Usuario.nome_uvis, db.func.count(Solicitacao.id))
                .group_by(Usuario.nome_uvis)
                .order_by(db.func.count(Solicitacao.id).desc())
                .all()
            )
        ]

        # =====================================================
        # 🔹 HISTÓRICO MENSAL (Independente do filtro atual)
        # =====================================================
        dados_mensais = [
            (f"{int(ano_h):04d}-{int(mes_h):02d}", total)
            for ano_h, mes_h, total in (
                db.session.query(
                    extract('year', Solicitacao.data_agendamento),
                    extract('month', Solicitacao.data_agendamento),
                    db.func.count(Solicitacao.id)
                )
                .group_by(extract('year', Solicitacao.data_agendamento), extract('month', Solicitacao.data_agendamento))
                .order_by(extract('year', Solicitacao.data_agendamento), extract('month', Solicitacao.data_agendamento))
                .all()
            )
        ]

        anos_disponiveis = (
            sorted({m.split('-')[0] for m, _ in dados_mensais}, reverse=True)
            if dados_mensais else [ano_atual]
        )

        print(f"DEBUG FILTRO: Mês selecionado: {mes_atual} | String gerada: {filtro_data}")

        return render_template(
            'relatorios.html',
            total_solicitacoes=total_solicitacoes,
            total_aprovadas=total_aprovadas,
            total_aprovadas_com_recomendacoes=total_aprovadas_com_recomendacoes,
            total_recusadas=total_recusadas,
            total_analise=total_analise,
            total_pendentes=total_pendentes,
            dados_regiao=dados_regiao,
            dados_status=dados_status,
            dados_foco=dados_foco,
            dados_tipo_visita=dados_tipo_visita,
            dados_altura_voo=dados_altura_voo,
            dados_unidade=dados_unidade,
            dados_mensais=dados_mensais,
            mes_selecionado=mes_atual,
            ano_selecionado=ano_atual,
            anos_disponiveis=anos_disponiveis,
            uvis_id_selecionado=uvis_id,
            uvis_disponiveis=uvis_disponiveis,
            filtros={'total': total_solicitacoes}
        )

    except Exception as e:
        db.session.rollback()
        print(f"ERRO NOS RELATÓRIOS: {e}")
        return render_template(
            "erro.html",
            codigo=500,
            titulo="Erro nos Relatórios",
            mensagem="Houve um erro técnico ao processar os dados."
        )
    

@bp.route('/relatorios', methods=['GET'], endpoint='relatorios')
@login_required
def relatorios_menu():
    if current_user.tipo_usuario not in ['admin', 'operario', 'visualizar']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('main.dashboard'))
    return render_template('relatorios_menu.html')


@bp.route('/relatorios-os', methods=['GET'])
@login_required
def relatorios_os():
    try:
        from sqlalchemy import and_, or_

        uvis_disponiveis = []
        if current_user.tipo_usuario in ['admin', 'operario', 'visualizar']:
            uvis_disponiveis = (
                db.session.query(Usuario.id, Usuario.nome_uvis)
                .filter(Usuario.tipo_usuario == 'uvis')
                .order_by(Usuario.nome_uvis)
                .all()
            )

        mes_atual = request.args.get('mes', datetime.now().month, type=int)
        ano_atual = request.args.get('ano', datetime.now().year, type=int)
        uvis_id = request.args.get('uvis_id', type=int) if current_user.tipo_usuario != 'uvis' else current_user.id

        base_query = (
            db.session.query(OrdemServico)
            .join(Solicitacao, Solicitacao.id == OrdemServico.solicitacao_id)
            .join(Usuario, Usuario.id == Solicitacao.usuario_id)
        )

        base_query = base_query.filter(
            or_(
                and_(
                    OrdemServico.respondido_em.isnot(None),
                    extract('year', OrdemServico.respondido_em) == ano_atual,
                    extract('month', OrdemServico.respondido_em) == mes_atual
                ),
                and_(
                    OrdemServico.respondido_em.is_(None),
                    OrdemServico.data_aplicacao.isnot(None),
                    extract('year', OrdemServico.data_aplicacao) == ano_atual,
                    extract('month', OrdemServico.data_aplicacao) == mes_atual
                )
            )
        )

        if uvis_id:
            base_query = base_query.filter(Solicitacao.usuario_id == uvis_id)

        total_os = base_query.count()
        total_concluidas = base_query.filter(Solicitacao.status.in_(["CONCLUÍDO", "CONCLUIDO"])).count()
        total_larva_sim = base_query.filter(func.upper(func.coalesce(OrdemServico.larva_visualizada, "")) == "SIM").count()
        total_tratamento_adicional = base_query.filter(func.upper(func.coalesce(OrdemServico.tratamento_adicional_realizado, "")) == "SIM").count()
        total_nao_realizadas = base_query.filter(func.length(func.trim(func.coalesce(OrdemServico.motivo_nao_realizacao, ""))) > 0).count()

        def agrupar_por(campo):
            return [
                (valor or "Não informado", total)
                for valor, total in (
                    base_query
                    .with_entities(campo, func.count(OrdemServico.id))
                    .group_by(campo)
                    .order_by(func.count(OrdemServico.id).desc())
                    .all()
                )
            ]

        dados_situacao_aplicacao = agrupar_por(OrdemServico.situacao_aplicacao)
        dados_tipo_aplicacao = agrupar_por(OrdemServico.tipo_aplicacao)
        dados_larva = agrupar_por(OrdemServico.larva_visualizada)
        dados_piloto = agrupar_por(OrdemServico.piloto)

        dados_unidade = [
            (uvis or "Não informado", total)
            for uvis, total in (
                base_query
                .with_entities(Usuario.nome_uvis, func.count(OrdemServico.id))
                .group_by(Usuario.nome_uvis)
                .order_by(func.count(OrdemServico.id).desc())
                .all()
            )
        ]

        mensal_query = (
            db.session.query(
                func.coalesce(
                    extract('year', OrdemServico.respondido_em),
                    extract('year', OrdemServico.data_aplicacao)
                ).label("ano_ref"),
                func.coalesce(
                    extract('month', OrdemServico.respondido_em),
                    extract('month', OrdemServico.data_aplicacao)
                ).label("mes_ref"),
                func.count(OrdemServico.id)
            )
            .join(Solicitacao, Solicitacao.id == OrdemServico.solicitacao_id)
            .join(Usuario, Usuario.id == Solicitacao.usuario_id)
        )

        if uvis_id:
            mensal_query = mensal_query.filter(Solicitacao.usuario_id == uvis_id)

        dados_mensais = [
            (f"{int(ano_h):04d}-{int(mes_h):02d}", total)
            for ano_h, mes_h, total in (
                mensal_query
                .filter(
                    or_(
                        OrdemServico.respondido_em.isnot(None),
                        OrdemServico.data_aplicacao.isnot(None)
                    )
                )
                .group_by("ano_ref", "mes_ref")
                .order_by("ano_ref", "mes_ref")
                .all()
            )
            if ano_h and mes_h
        ]

        anos_disponiveis = (
            sorted({m.split('-')[0] for m, _ in dados_mensais}, reverse=True)
            if dados_mensais else [ano_atual]
        )

        return render_template(
            "relatorios_os.html",
            total_os=total_os,
            total_concluidas=total_concluidas,
            total_larva_sim=total_larva_sim,
            total_tratamento_adicional=total_tratamento_adicional,
            total_nao_realizadas=total_nao_realizadas,
            dados_situacao_aplicacao=dados_situacao_aplicacao,
            dados_tipo_aplicacao=dados_tipo_aplicacao,
            dados_larva=dados_larva,
            dados_piloto=dados_piloto,
            dados_unidade=dados_unidade,
            dados_mensais=dados_mensais,
            mes_selecionado=mes_atual,
            ano_selecionado=ano_atual,
            anos_disponiveis=anos_disponiveis,
            uvis_id_selecionado=uvis_id,
            uvis_disponiveis=uvis_disponiveis,
        )
    except Exception as e:
        db.session.rollback()
        print(f"ERRO NOS RELATÓRIOS DE OS: {e}")
        return render_template(
            "erro.html",
            codigo=500,
            titulo="Erro nos Relatórios de OS",
            mensagem="Houve um erro técnico ao processar os dados das ordens de serviço."
        )

import os
import tempfile
from datetime import datetime
from io import BytesIO

from flask import send_file, request
from flask_login import login_required, current_user
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage
)

try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

@bp.route('/admin/exportar_relatorio_pdf')
@login_required
def exportar_relatorio_pdf():
    # -------------------------
    # 1. Parâmetros e filtros (IGUAL ao /relatorios)
    # -------------------------
    mes = int(request.args.get('mes', datetime.now().month))
    ano = int(request.args.get('ano', datetime.now().year))
    orient = request.args.get('orient', default='portrait')  # 'portrait' ou 'landscape'
    filtro_data = f"{ano}-{mes:02d}"

    if current_user.tipo_usuario == 'uvis':
        uvis_id = current_user.id
    else:
        uvis_id = request.args.get('uvis_id', type=int)

    # -------------------------
    # 2. Query base e detalhe
    # -------------------------
    base_query = aplicar_filtros_base(
        db.session.query(Solicitacao),
        filtro_data,
        uvis_id
    )

    query_detalhe = aplicar_filtros_base(
        db.session.query(Solicitacao, Usuario).join(Usuario, Usuario.id == Solicitacao.usuario_id),
        filtro_data,
        uvis_id
    )

    query_results = query_detalhe.order_by(Solicitacao.data_criacao.desc()).all()

    # -------------------------
    # 3. Totais
    # -------------------------
    total_solicitacoes = base_query.count()
    total_aprovadas = base_query.filter(Solicitacao.status == "APROVADO").count()
    total_aprovadas_com_recomendacoes = base_query.filter(
        Solicitacao.status == "APROVADO COM RECOMENDAÇÕES"
    ).count()
    total_recusadas = base_query.filter(Solicitacao.status == "NEGADO").count()
    total_analise = base_query.filter(Solicitacao.status == "EM ANÁLISE").count()
    total_pendentes = base_query.filter(Solicitacao.status == "PENDENTE").count()

    STATUS_COLORS = {
        "APROVADO": "#2ecc71",
        "APROVADO COM RECOMENDAÇÕES": "#ee650a",
        "EM ANÁLISE": "#f1c40f",
        "PENDENTE": "#3498db",
        "NEGADO": "#e74c3c",
    }

    # -------------------------
    # 4. Agrupamentos
    # -------------------------
    dados_regiao = [
        (regiao or "Não informado", total)
        for regiao, total in (
            aplicar_filtros_base(
                db.session.query(Usuario.regiao, db.func.count(Solicitacao.id)).join(Usuario),
                filtro_data,
                uvis_id
            )
            .group_by(Usuario.regiao)
            .all()
        )
    ]

    dados_status = [
        (status or "Não informado", total)
        for status, total in (
            base_query
            .with_entities(Solicitacao.status, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.status)
            .all()
        )
    ]

    dados_foco = [
        (foco or "Não informado", total)
        for foco, total in (
            base_query
            .with_entities(Solicitacao.foco, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.foco)
            .all()
        )
    ]

    dados_tipo_visita = [
        (tipo or "Não informado", total)
        for tipo, total in (
            base_query
            .with_entities(Solicitacao.tipo_visita, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.tipo_visita)
            .all()
        )
    ]

    dados_altura_voo = [
        (altura or "Não informado", total)
        for altura, total in (
            base_query
            .with_entities(Solicitacao.altura_voo, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.altura_voo)
            .all()
        )
    ]

    dados_unidade = [
        (uvis_nome or "Não informado", total)
        for uvis_nome, total in (
            aplicar_filtros_base(
                db.session.query(Usuario.nome_uvis, db.func.count(Solicitacao.id))
                .join(Usuario)
                .filter(Usuario.tipo_usuario == 'uvis'),
                filtro_data,
                uvis_id
            )
            .group_by(Usuario.nome_uvis)
            .all()
        )
    ]

    if db.engine.name == 'postgresql':
        func_mes = db.func.to_char(Solicitacao.data_agendamento, 'YYYY-MM')
    else:
        func_mes = db.func.strftime('%Y-%m', Solicitacao.data_agendamento)

    dados_mensais = [
        tuple(row) for row in (
            db.session.query(func_mes.label('mes'), db.func.count(Solicitacao.id))
            .group_by('mes')
            .order_by('mes')
            .all()
        )
    ]

    # -------------------------
    # 5. Preparar PDF
    # -------------------------
    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    caminho_pdf = tmp_pdf.name
    tmp_pdf.close()

    pagesize = landscape(A4) if orient == 'landscape' else A4

    doc = SimpleDocTemplate(
        caminho_pdf,
        pagesize=pagesize,
        leftMargin=14*mm, rightMargin=14*mm,
        topMargin=16*mm, bottomMargin=16*mm
    )

    styles = getSampleStyleSheet()

    # Tipografia melhor
    title_style = ParagraphStyle(
        'title',
        parent=styles['Title'],
        fontSize=18,
        leading=22,
        alignment=1,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=10
    )

    subtitle_style = ParagraphStyle(
        'subtitle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        alignment=1,
        textColor=colors.HexColor('#555'),
        spaceAfter=12
    )

    section_h = ParagraphStyle(
        'sec',
        parent=styles['Heading2'],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#0d6efd'),
        spaceBefore=10,
        spaceAfter=6
    )

    normal = ParagraphStyle(
        'normal',
        parent=styles['Normal'],
        fontSize=9.5,
        leading=13
    )

    cell_style = ParagraphStyle(
        'cell',
        parent=styles['BodyText'],
        fontSize=8.6,
        leading=11,
        textColor=colors.HexColor('#222'),
        wordWrap='CJK',
        splitLongWords=True
    )

    story = []

    # -------------------------
    # CAPA (Resumo)
    # -------------------------
    story.append(Paragraph(f"Relatório Mensal — {mes:02d}/{ano}", title_style))

    filtro_txt = f"Filtro: {filtro_data}"
    if uvis_id:
        filtro_txt += f" | UVIS ID: {uvis_id}"
    else:
        filtro_txt += " | UVIS: Todas"
    story.append(Paragraph(filtro_txt, subtitle_style))

    # Cards do resumo (bem mais bonito)
    def resumo_cards():
        cards = [
            ("Total", total_solicitacoes, '#0d6efd'),
            ("Aprovadas", total_aprovadas, '#198754'),
            ("Aprov. c/ Recom.", total_aprovadas_com_recomendacoes, '#6c757d'),
            ("Negadas", total_recusadas, '#dc3545'),
            ("Em Análise", total_analise, '#ffc107'),
            ("Pendentes", total_pendentes, '#0dcaf0'),
        ]

        rows = []
        row = []
        for i, (label, value, hexcolor) in enumerate(cards, start=1):
            box = Table(
                [
                    [Paragraph(label, ParagraphStyle('l', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#666')))],
                    [Paragraph(str(value), ParagraphStyle('v', parent=styles['Normal'], fontSize=18, leading=20, textColor=colors.HexColor(hexcolor)))]
                ],
                colWidths=[48*mm] if orient == 'portrait' else [52*mm],
            )
            box.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8f9fa')),
                ('BOX', (0,0), (-1,-1), 0.6, colors.HexColor('#e5e7eb')),
                ('LEFTPADDING', (0,0), (-1,-1), 8),
                ('RIGHTPADDING', (0,0), (-1,-1), 8),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))

            row.append(box)
            if len(row) == 3:
                rows.append(row)
                row = []

        if row:
            # completa a linha
            while len(row) < 3:
                row.append(Spacer(1, 1))
            rows.append(row)

        grid = Table(rows, colWidths=None)
        grid.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        return grid

    story.append(resumo_cards())
    story.append(Spacer(1, 10))

    # -------------------------
    # TABELAS (DADOS ESCRITOS PRIMEIRO)
    # -------------------------
    def add_count_table(titulo, dados, col1="Categoria"):
        story.append(Paragraph(titulo, section_h))

        rows = [
            [Paragraph(col1, ParagraphStyle('th', parent=cell_style, textColor=colors.white, fontSize=9)),
             Paragraph("Total", ParagraphStyle('th2', parent=cell_style, textColor=colors.white, fontSize=9))]
        ]

        for nome, total in (dados or [("Nenhum", 0)]):
            rows.append([Paragraph(str(nome), cell_style), Paragraph(str(total), cell_style)])

        tbl = Table(rows, repeatRows=1, colWidths=[140*mm, 25*mm] if orient == 'portrait' else [190*mm, 30*mm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,0),colors.HexColor('#0d6efd')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,0),9),
            ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#d9dee7')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#fbfdff')]),
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('LEFTPADDING',(0,0),(-1,-1),6),
            ('RIGHTPADDING',(0,0),(-1,-1),6),
            ('TOPPADDING',(0,0),(-1,-1),4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 10))

    # Um “Resumo por agrupamento” em sequência (mais agradável)
    story.append(Paragraph("Resumo por Agrupamentos", section_h))
    story.append(Paragraph("Abaixo estão os agrupamentos do mês selecionado, apresentados em formato de tabela.", normal))
    story.append(Spacer(1, 6))

    add_count_table("Agrupamento — Região", dados_regiao)
    add_count_table("Agrupamento — Status", dados_status)
    add_count_table("Agrupamento — Foco", dados_foco)
    add_count_table("Agrupamento — Tipo de Visita", dados_tipo_visita)
    add_count_table("Agrupamento — Altura do Voo", dados_altura_voo)
    add_count_table("Agrupamento — Unidade (UVIS)", dados_unidade)
    add_count_table("Histórico Mensal (tabela)", dados_mensais, col1="Mês")

    # -------------------------
    #  GRÁFICOS (AGORA DEPOIS DOS DADOS ESCRITOS)
    # -------------------------
    story.append(PageBreak())
    story.append(Paragraph("Gráficos", section_h))
    story.append(Paragraph("Os gráficos abaixo representam visualmente os dados apresentados nas tabelas anteriores.", normal))
    story.append(Spacer(1, 8))

    def safe_img_from_plt(fig, width_mm=170):
        bio = BytesIO()
        fig.tight_layout()
        fig.savefig(bio, format='png', dpi=220, bbox_inches='tight')
        plt.close(fig)
        bio.seek(0)
        return RLImage(bio, width=width_mm*mm)

    if MATPLOTLIB_AVAILABLE:
        try:
            # 1) Donut por status (mais limpo)
            labels = [s for s, _ in dados_status]
            values = [c for _, c in dados_status]
            colors_status = [STATUS_COLORS.get(s, "#bdc3c7") for s in labels]

            fig1, ax1 = plt.subplots(figsize=(6.4, 3.0))
            def autopct(p): return f'{p:.0f}%' if p >= 6 else ''
            wedges, *_ = ax1.pie(
                values or [1],
                labels=None,
                colors=colors_status,
                autopct=autopct,
                startangle=90,
                pctdistance=0.78,
                textprops={'fontsize': 9}
            )
            centre_circle = plt.Circle((0, 0), 0.58, fc='white')
            ax1.add_artist(centre_circle)
            ax1.legend(wedges, labels, loc='center left', bbox_to_anchor=(1.02, 0.5),
                       fontsize=9, frameon=False)
            ax1.set_title('Distribuição por Status', fontsize=11, pad=10)
            ax1.axis('equal')

            story.append(safe_img_from_plt(fig1, width_mm=170))
            story.append(Spacer(1, 10))

            # 2) Top UVIS (barra horizontal)
            u_names = [u for u, _ in dados_unidade[:10]]
            u_vals = [c for _, c in dados_unidade[:10]]

            fig2, ax2 = plt.subplots(figsize=(7.2, 3.0))
            ax2.barh(u_names[::-1] or ['Nenhum'], u_vals[::-1] or [0])
            ax2.set_xlabel('Total', fontsize=9)
            ax2.set_title('Top UVIS', fontsize=11, pad=10)
            ax2.tick_params(axis='both', labelsize=9)
            ax2.grid(axis='x', linestyle=':', linewidth=0.6, alpha=0.6)

            story.append(safe_img_from_plt(fig2, width_mm=180 if orient == 'landscape' else 170))
            story.append(Spacer(1, 10))

            # 3) Histórico mensal (linha)
            months = [m for m, _ in dados_mensais]
            counts = [c for _, c in dados_mensais]

            fig3, ax3 = plt.subplots(figsize=(7.2, 3.0))
            if months:
                ax3.plot(range(len(months)), counts, marker='o', linewidth=1.6)
                ax3.set_xticks(range(len(months)))
                ax3.set_xticklabels(months, rotation=45, ha='right', fontsize=9)
            ax3.set_title('Histórico Mensal', fontsize=11, pad=10)
            ax3.tick_params(axis='y', labelsize=9)
            ax3.grid(axis='y', linestyle=':', linewidth=0.6, alpha=0.6)

            story.append(safe_img_from_plt(fig3, width_mm=185 if orient == 'landscape' else 170))
            story.append(Spacer(1, 8))

        except Exception:
            story.append(Paragraph("Gráficos indisponíveis (erro ao gerar).", normal))
    else:
        story.append(Paragraph("Matplotlib não disponível — gráficos foram omitidos.", normal))

        # -------------------------
    # DETALHES (Registros Detalhados)
    # -------------------------
    story.append(PageBreak())
    story.append(Paragraph("Registros Detalhados", section_h))
    story.append(Paragraph("Listagem completa dos registros retornados pelo filtro selecionado.", normal))
    story.append(Spacer(1, 8))

    registros_header = [
        'Data', 'Hora', 'Unidade', 'Região', 'Protocolo',
        'Status', 'Foco', 'Tipo Visita', 'Altura Voo', 'Observação'
    ]

    hdr_style = ParagraphStyle(
        'hdr',
        parent=cell_style,
        textColor=colors.white,
        fontSize=7.8,
        leading=9.2
    )

    cell_style_small = ParagraphStyle(
        'cell_small',
        parent=cell_style,
        fontSize=7.6,
        leading=9.2,
        wordWrap='CJK',
        splitLongWords=True
    )

    registros_rows = [[Paragraph(h, hdr_style) for h in registros_header]]

    for s, u in query_results:
        data_str = s.data_criacao.strftime("%d/%m/%Y") if getattr(s, 'data_criacao', None) else ''
        hora_str = getattr(s, 'hora_agendamento', '')
        hora_str = hora_str.strftime("%H:%M") if hasattr(hora_str, 'strftime') else str(hora_str or '')

        unidade = getattr(u, 'nome_uvis', '') or "Não informado"
        regiao = getattr(u, 'regiao', '') or "Não informado"

        protocolo = getattr(s, 'protocolo', '') or ''
        status = getattr(s, 'status', '') or ''
        foco = getattr(s, 'foco', '') or ''
        tipo_visita = getattr(s, 'tipo_visita', '') or ''
        altura_voo = getattr(s, 'altura_voo', '') or ''
        obs = getattr(s, 'observacao', '') or ''

        registros_rows.append([
            Paragraph(str(data_str), cell_style_small),
            Paragraph(str(hora_str), cell_style_small),
            Paragraph(str(unidade), cell_style_small),
            Paragraph(str(regiao), cell_style_small),
            Paragraph(str(protocolo), cell_style_small),
            Paragraph(str(status), cell_style_small),
            Paragraph(str(foco), cell_style_small),
            Paragraph(str(tipo_visita), cell_style_small),
            Paragraph(str(altura_voo), cell_style_small),
            Paragraph(str(obs), cell_style_small),
        ])

    #  Larguras base (as suas), mas vamos “encaixar” no doc.width automaticamente
    base_col_widths = [
        18*mm, 14*mm, 28*mm, 22*mm, 22*mm,
        22*mm, 22*mm, 26*mm, 18*mm, 60*mm
    ]

    #  Se a soma estourar a largura útil da página, escala proporcionalmente
    total_w = sum(base_col_widths)
    max_w = doc.width  # largura útil = página - margens

    if total_w > max_w:
        scale = max_w / total_w
        colWidths = [w * scale for w in base_col_widths]
    else:
        colWidths = base_col_widths

    #  Quantidade de linhas por página (ajuste fino)
    chunk_size = 28 if orient == 'landscape' else 24

    # 🔥 renderiza em blocos para não ficar pesado e manter header repetido
    for i in range(0, len(registros_rows), chunk_size):
        chunk = registros_rows[i:i + chunk_size]

        tbl = Table(
            chunk,
            repeatRows=1,
            colWidths=colWidths,
            hAlign='LEFT'  #  evita “puxar” pro centro e cortar laterais
        )

        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),

            ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#d9dee7')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fbfdff')]),

            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (0,0), (-1,0), 'LEFT'),
            ('ALIGN', (0,1), (-1,-1), 'LEFT'),

            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),

            #  reforço de quebra de linha dentro das células
            ('WORDWRAP', (0,0), (-1,-1), 'CJK'),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 6))

        if i + chunk_size < len(registros_rows):
            story.append(PageBreak())


    # -------------------------
    # Header/Footer
    # -------------------------
    def _header_footer(canvas, doc_):
        canvas.saveState()
        w, h = pagesize

        canvas.setFillColor(colors.HexColor('#0d6efd'))
        canvas.rect(doc_.leftMargin, h-(12*mm), doc_.width, 3, fill=1, stroke=0)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor('#777'))
        canvas.drawString(doc_.leftMargin, 9*mm, f"Relatório — {mes:02d}/{ano} — IJASystem")
        canvas.drawRightString(doc_.leftMargin + doc_.width, 9*mm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

    nome_arquivo = f"relatorio_OceanoAzul_{ano}_{mes:02d}"
    if uvis_id:
        nome_arquivo += f"_UVIS_{uvis_id}"

    return send_file(
        caminho_pdf,
        as_attachment=True,
        download_name=f"{nome_arquivo}.pdf",
        mimetype="application/pdf"
    )


# =======================================================================
# ROTA 3: Exportar Excel (Com Filtro UVIS) - Layout “bonito” igual Excel
# =======================================================================
@bp.route('/admin/exportar_relatorio_excel')
@login_required
def exportar_relatorio_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # -------------------------
    # 1. Parâmetros e filtros
    # -------------------------
    mes = request.args.get('mes', datetime.now().month, type=int)
    ano = request.args.get('ano', datetime.now().year, type=int)
    filtro_data = f"{ano}-{mes:02d}"

    # Controle de acesso UVIS
    if current_user.tipo_usuario == 'uvis':
        uvis_id = current_user.id
    else:
        uvis_id = request.args.get('uvis_id', type=int)

    # -------------------------
    # 2. Busca de Dados
    # -------------------------
    query_dados = db.session.query(
        Solicitacao.id,
        Solicitacao.status,
        Solicitacao.foco,
        Solicitacao.tipo_visita,
        Solicitacao.altura_voo,
        Solicitacao.data_agendamento,
        Solicitacao.hora_agendamento,

        # Endereço (campos separados no banco)
        Solicitacao.cep,
        Solicitacao.logradouro,
        Solicitacao.numero,
        Solicitacao.bairro,
        Solicitacao.cidade,
        Solicitacao.uf,

        Solicitacao.latitude,
        Solicitacao.longitude,

        # UVIS
        Usuario.nome_uvis,
        Usuario.regiao
    ).join(Usuario, Usuario.id == Solicitacao.usuario_id)

   # Filtro de data pelo AGENDAMENTO (mês/ano)
    if db.engine.name == 'postgresql':
        query_dados = query_dados.filter(
            Solicitacao.data_agendamento.isnot(None),
            db.func.to_char(Solicitacao.data_agendamento, 'YYYY-MM') == filtro_data
        )
    else:
        query_dados = query_dados.filter(
            Solicitacao.data_agendamento.isnot(None),
            db.func.strftime('%Y-%m', Solicitacao.data_agendamento) == filtro_data
        )
    # Filtro opcional por UVIS
    if uvis_id:
        query_dados = query_dados.filter(Solicitacao.usuario_id == uvis_id)

  # Ordenar pelo agendamento (e hora como critério secundário)
    dados = query_dados.order_by(
        Solicitacao.data_agendamento.desc(),
        Solicitacao.hora_agendamento.desc()
    ).all()

    # Se tiver filtro UVIS, pega o nome pra ajudar no nome do arquivo
    nome_uvis_filtro = None
    if uvis_id:
        nome_uvis_filtro = db.session.query(Usuario.nome_uvis).filter(Usuario.id == uvis_id).scalar()

    # -------------------------
    # 3. Helper: montar endereço em 1 LINHA (compacto igual Excel)
    # -------------------------
    def montar_endereco(row):
        partes_rua = []
        if row.logradouro:
            partes_rua.append(row.logradouro.strip())
        if row.numero is not None and str(row.numero).strip():
            partes_rua.append(str(row.numero).strip())

        rua_numero = ", ".join([p for p in partes_rua if p]).strip()

        cidade_uf = ""
        if row.cidade and row.uf:
            cidade_uf = f"{row.cidade.strip()}/{row.uf.strip()}"
        elif row.cidade:
            cidade_uf = row.cidade.strip()
        elif row.uf:
            cidade_uf = row.uf.strip()

        bairro_cidade = " - ".join([p for p in [(row.bairro or "").strip(), cidade_uf] if p]).strip()
        cep_txt = f"CEP {row.cep.strip()}" if row.cep else ""

        # Formato final: "Rua, 123 | Bairro - Cidade/UF | CEP 00000-000"
        return " | ".join([p for p in [rua_numero, bairro_cidade, cep_txt] if p])

    # -------------------------
    # 4. Criar arquivo Excel
    # -------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    #  UVIS vem no começo agora
    colunas = [
        "UVIS", "Região",
        "ID", "Status", "Foco", "Tipo Visita", "Altura Voo",
        "Data Agendamento", "Hora Agendamento",
        "ENDEREÇO DE AÇÃO",
        "Latitude", "Longitude"
    ]

    # Estilos (bem padrão Excel)
    header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra1 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    zebra2 = PatternFill(start_color="FFF7FBFF", end_color="FFF7FBFF", fill_type="solid")

    #  alinhamento igual ao print (compacto e central vertical)
    center = Alignment(horizontal="center", vertical="center")
    left_center = Alignment(horizontal="left", vertical="center")

    # Cabeçalho
    for col_num, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Altura do cabeçalho (compacto)
    ws.row_dimensions[1].height = 22

    # Preenchimento de linhas
    for row_num, row in enumerate(dados, 2):
        data_agendamento_fmt = row.data_agendamento.strftime("%d/%m/%Y") if row.data_agendamento else ""
        hora_agendamento_fmt = row.hora_agendamento.strftime("%H:%M") if row.hora_agendamento else ""

        endereco_acao = montar_endereco(row)

        values = [
            row.nome_uvis,
            row.regiao,
            row.id,
            row.status,
            row.foco,
            row.tipo_visita,
            row.altura_voo,
            data_agendamento_fmt,
            hora_agendamento_fmt,
            endereco_acao,
            row.latitude,
            row.longitude
        ]

        # Altura das linhas (igual Excel “padrão bonito”)
        ws.row_dimensions[row_num].height = 20

        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_index, value=value)
            cell.border = thin_border
            cell.fill = zebra1 if (row_num % 2 == 0) else zebra2

            # Centraliza campos curtos, texto fica alinhado à esquerda (igual print)
            if col_index in (3, 7, 8, 9, 11, 12):  # ID, Altura, Data, Hora, Lat, Long
                cell.alignment = center
            else:
                cell.alignment = left_center

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Auto-filtro no cabeçalho
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    #  Larguras “na mão” (fica igual ao print / bem organizado)
    larguras = {
        "A": 24,  # UVIS
        "B": 12,  # Região
        "C": 6,   # ID
        "D": 18,  # Status
        "E": 22,  # Foco
        "F": 16,  # Tipo Visita
        "G": 10,  # Altura Voo
        "H": 14,  # Data
        "I": 14,  # Hora
        "J": 90,  # ENDEREÇO DE AÇÃO
        "K": 14,  # Latitude
        "L": 14   # Longitude
    }
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    # -------------------------
    # 5. Gerar arquivo em memória
    # -------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nome do arquivo
    nome_arquivo = f"relatorio_OceanoAzul_{ano}_{mes:02d}"
    if uvis_id:
        safe_nome = (nome_uvis_filtro or f"ID_{uvis_id}").replace(" ", "_")
        nome_arquivo += f"_UVIS_{safe_nome}"

    return send_file(
        output,
        download_name=f"{nome_arquivo}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




from flask import flash, redirect, url_for, render_template, request
from flask_login import login_required, current_user
from app import db
from app.models import Solicitacao, Usuario
from datetime import datetime
from sqlalchemy.orm import joinedload

@bp.route('/solicitacao/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_solicitacao(id):
    # 1️⃣ Busca o pedido com os dados do usuário
    pedido = Solicitacao.query.options(joinedload(Solicitacao.usuario)).get_or_404(id)
    
    is_admin = current_user.tipo_usuario == 'admin'

    # 🔐 Regras de Segurança Reais
    if not is_admin:
        # Verifica se o pedido pertence à UVIS logada
        if pedido.usuario_id != current_user.id:
            flash('Permissão negada. Você só pode editar suas próprias solicitações.', 'danger')
            return redirect(url_for('main.dashboard'))
        
        # Trava de Status: UVIS só edita PENDENTE ou NEGADO
        if pedido.status not in ["PENDENTE", "NEGADO"]:
            flash('Esta solicitação já está em processo de aprovação e não pode ser editada.', 'warning')
            return redirect(url_for('main.dashboard'))

    # Opções para os selects (Pode ser movido para uma constante ou banco)
    status_opcoes = ["PENDENTE", "EM ANÁLISE", "APROVADO", "APROVADO COM RECOMENDAÇÕES", "NEGADO"]
    foco_opcoes = ["Foco 1", "Foco 2", "Foco 3"]
    tipo_visita_opcoes = ["Tipo 1", "Tipo 2", "Tipo 3"]
    uf_opcoes = ["AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG",
                 "PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"]

    if request.method == 'POST':
        try:
            # 2️⃣ Atualização de Datas/Horas (Reaproveitando sua lógica original)
            data_str = request.form.get('data_agendamento')
            hora_str = request.form.get('hora_agendamento')
            pedido.data_agendamento = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
            pedido.hora_agendamento = datetime.strptime(hora_str, '%H:%M').time() if hora_str else None

            # 3️⃣ Atualização de Campos Gerais
            pedido.foco = request.form.get('foco') or pedido.foco
            pedido.tipo_visita = request.form.get('tipo_visita') or pedido.tipo_visita
            pedido.altura_voo = request.form.get('altura_voo') or pedido.altura_voo
            pedido.apoio_cet = request.form.get('apoio_cet', 'não').lower() == 'sim'
            pedido.observacao = request.form.get('observacao') or pedido.observacao
            
            # 4️⃣ Endereço e Localização
            pedido.cep = request.form.get('cep') or pedido.cep
            pedido.logradouro = request.form.get('logradouro') or pedido.logradouro
            pedido.numero = request.form.get('numero') or pedido.numero
            pedido.bairro = request.form.get('bairro') or pedido.bairro
            pedido.cidade = request.form.get('cidade') or pedido.cidade
            pedido.uf = request.form.get('uf') or pedido.uf
            lat_raw = (request.form.get('latitude') or "").strip()
            lng_raw = (request.form.get('longitude') or "").strip()
            pedido.latitude = float(lat_raw.replace(",", ".")) if lat_raw else pedido.latitude
            pedido.longitude = float(lng_raw.replace(",", ".")) if lng_raw else pedido.longitude
            pedido.area_restrita = detectar_area_restrita(pedido.latitude, pedido.longitude) or request.form.get('risco_aereo') == '1'

            # 5️⃣ Lógica de Status e Hierarquia
            if is_admin:
                pedido.status = request.form.get('status') or pedido.status
                pedido.protocolo = request.form.get('protocolo') or pedido.protocolo
                justificativa = (request.form.get('justificativa') or '').strip()
                pedido.justificativa = justificativa or None
            else:
                # Se a UVIS está editando um pedido que estava NEGADO
                if pedido.status == 'NEGADO':
                    # Mantemos o status PENDENTE, mas marcamos a justificativa
                    # Isso ativa o badge "CORREÇÃO RECEBIDA" no Painel do Admin
                    motivo_original = (pedido.justificativa or "").strip()
                    # Evitamos duplicar o prefixo se ela editar várias vezes
                    limpo = re.sub(r'^\s*CORREÇÃO:\s*', '', motivo_original, flags=re.IGNORECASE)
                    pedido.justificativa = "CORREÇÃO: corrigido pela UVIS" if not limpo else f"CORREÇÃO: {limpo}"
                else:
                    # Se era PENDENTE e ela só editou dados (como o CEP), 
                    # mantemos limpo ou como estava.
                    pedido.justificativa = None
                
                pedido.status = "PENDENTE"

            db.session.commit()
            flash('Solicitação atualizada com sucesso!', 'success')
            
            # Redireciona conforme o papel
            return redirect(url_for('main.admin_dashboard' if is_admin else 'main.dashboard'))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar a solicitação: {e}", 'danger')

    return render_template(
        'editar_solicitacao.html', # Renomeie seu arquivo .html também!
        pedido=pedido,
        is_admin=is_admin,
        status_opcoes=status_opcoes,
        foco_opcoes=foco_opcoes,
        tipo_visita_opcoes=tipo_visita_opcoes,
        uf_opcoes=uf_opcoes
    )

from flask_login import current_user, login_required

@bp.route('/admin/deletar/<int:id>', methods=['POST'], endpoint='deletar_registro')
@login_required
def deletar(id):
    # Verifica se é admin
    if current_user.tipo_usuario != 'admin':  # <-- CORRETO: tipo_usuario
        flash('Permissão negada. Apenas administradores podem deletar registros.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    # Busca a solicitação
    pedido = Solicitacao.query.get_or_404(id)
    pedido_id = pedido.id

    # Nome do autor da solicitação
    autor_nome = pedido.usuario.nome_uvis if pedido.usuario else "UVIS"

    try:
        db.session.delete(pedido)
        db.session.commit()
    except Exception:
        db.session.rollback()
        # Não mostra erro ao usuário

    flash(f"Pedido #{pedido_id} da {autor_nome} deletado permanentemente.", "success")
    return redirect(url_for('main.admin_dashboard'))


from flask_login import login_required, current_user

from flask_login import login_required, current_user
import traceback

from flask import request, render_template
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from datetime import datetime
import json
from flask import request, render_template
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from datetime import datetime
import json
import os

@bp.route("/agenda")
@login_required
def agenda():
    try:
        #  Key pro mapa no modal
        google_maps_key = current_app.config.get("Maps_KEY_FRONT") or os.getenv("KEY_API_GOOGLE_MAPS") or ""

        # --- Usuário atual ---
        user_tipo = getattr(current_user, "tipo_usuario", None)
        user_id = current_user.id

        # --- Filtros GET ---
        filtro_status = (request.args.get("status") or "").strip() or None
        filtro_uvis_id = request.args.get("uvis_id", type=int)
        mes = request.args.get("mes", datetime.now().month, type=int)
        ano = request.args.get("ano", datetime.now().year, type=int)

        #  Se não vier "d", abre na data de HOJE (evita cair no dia 01)
        d = (request.args.get("d") or "").strip()
        initial_date = d or datetime.now().strftime("%Y-%m-%d")

        query = (
        Solicitacao.query
        .options(joinedload(Solicitacao.usuario))
        .filter(Solicitacao.status != "CANCELADO")
    )

        # Permissões / filtro UVIS
        if user_tipo not in ["admin", "operario", "visualizar"]:
            query = query.filter(Solicitacao.usuario_id == user_id)
            filtro_uvis_id = None
            pode_filtrar_uvis = False
        else:
            pode_filtrar_uvis = True
            if filtro_uvis_id:
                query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)

        # Filtro status
        if filtro_status:
            query = query.filter(Solicitacao.status == filtro_status)

        # Filtro mês/ano (mantido)
        filtro_mesano = f"{ano}-{mes:02d}"
        if db.engine.name == "postgresql":
            query = query.filter(db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM") == filtro_mesano)
        else:
            query = query.filter(db.func.strftime("%Y-%m", Solicitacao.data_agendamento) == filtro_mesano)

        eventos = query.all()

        # --- Monta eventos para o FullCalendar ---
        agenda_eventos = []

        for e in eventos:
            # data/hora seguros
            if not e.data_agendamento:
                continue

            data = e.data_agendamento.strftime("%Y-%m-%d")
            hora = e.hora_agendamento.strftime("%H:%M") if e.hora_agendamento else "00:00"
            uvis_nome = (e.usuario.nome_uvis if e.usuario else "UVIS") or "UVIS"

            # coords numéricos (tolerante a vírgula)
            lat = None
            lng = None
            try:
                if e.latitude is not None and str(e.latitude).strip() != "":
                    lat = float(str(e.latitude).replace(",", "."))
                if e.longitude is not None and str(e.longitude).strip() != "":
                    lng = float(str(e.longitude).replace(",", "."))
            except Exception:
                lat = None
                lng = None

            # endereço formatado
            logradouro = (e.logradouro or "").strip()
            numero = (e.numero or "S/N")
            bairro = (e.bairro or "").strip()
            cidade = (getattr(e, "cidade", "") or "").strip()
            uf = (getattr(e, "uf", "") or "").strip()
            cep = (e.cep or "").strip()

            partes = []
            if logradouro:
                partes.append(f"{logradouro}, {numero}")
            elif numero:
                partes.append(str(numero))

            if bairro:
                partes.append(bairro)

            if cidade or uf:
                partes.append(f"{cidade}/{uf}".strip("/"))

            endereco_txt = " - ".join([p for p in partes if p and p != "S/N"])
            if cep:
                endereco_txt = (endereco_txt + f" (CEP {cep})").strip()

            # evento do calendário
            agenda_eventos.append({
                "id": str(e.id),
                "title": f"{e.foco} - {uvis_nome}",
                "start": f"{data}T{hora}",
                "color": (
                    "#198754" if e.status == "APROVADO" else
                    "#ffa023" if e.status == "APROVADO COM RECOMENDAÇÕES" else
                    "#dc3545" if e.status == "NEGADO" else
                    "#e9fa05" if e.status == "EM ANÁLISE" else
                    "#0d6efd"
                ),
                "extendedProps": {
                    "foco": e.foco,
                    "uvis": uvis_nome,
                    "hora": hora,
                    "status": e.status,

                    #  padrão único pro JS (modal + rota)
                    "lat": lat,
                    "lng": lng,
                    "endereco": endereco_txt
                }
            })

        # --- Variáveis para filtros ---
        status_opcoes = ["PENDENTE", "EM ANÁLISE", "APROVADO", "APROVADO COM RECOMENDAÇÕES", "NEGADO"]

        uvis_disponiveis = []
        if user_tipo in ["admin", "operario", "visualizar"]:
            uvis_disponiveis = (
                db.session.query(Usuario.id, Usuario.nome_uvis)
                .filter(Usuario.tipo_usuario == "uvis")
                .order_by(Usuario.nome_uvis)
                .all()
            )

        # --- Anos disponíveis ---
        if db.engine.name == "postgresql":
            func_ano = db.func.to_char(Solicitacao.data_agendamento, "YYYY")
        else:
            func_ano = db.func.strftime("%Y", Solicitacao.data_agendamento)

        anos_raw = (
            db.session.query(func_ano)
            .filter(Solicitacao.data_agendamento.isnot(None))
            .distinct()
            .order_by(func_ano.desc())
            .all()
        )
        anos_disponiveis = [int(a[0]) for a in anos_raw if a and a[0]] or [datetime.now().year]

        filtros = {
            "uvis_id": filtro_uvis_id,
            "status": filtro_status,
            "mes": mes,
            "ano": ano
        }

        return render_template(
            "agenda.html",
            eventos_json=json.dumps(agenda_eventos, ensure_ascii=False),
            filtros=filtros,
            status_opcoes=status_opcoes,
            uvis_disponiveis=uvis_disponiveis,
            anos_disponiveis=anos_disponiveis,
            initial_date=initial_date,
            pode_filtrar_uvis=pode_filtrar_uvis,
            google_maps_key=google_maps_key,   #  pro modal com mapa
        )

    except Exception as e:
        import traceback
        print("TRACEBACK COMPLETO:")
        traceback.print_exc()
        return f"ERRO NA AGENDA: {str(e)}"


@bp.route("/agenda/rotas-dia")
@login_required
def agenda_rotas_dia():
    try:
        user_tipo = current_user.tipo_usuario
        user_id = current_user.id

        dia = (request.args.get("dia") or "").strip()
        if not dia:
            return jsonify(ok=False, error="Parâmetro 'dia' é obrigatório (YYYY-MM-DD)."), 400

        try:
            dia_date = datetime.strptime(dia, "%Y-%m-%d").date()
        except ValueError:
            return jsonify(ok=False, error="Formato inválido para 'dia'. Use YYYY-MM-DD."), 400

        # filtros opcionais
        filtro_uvis_id = request.args.get("uvis_id", type=int)

        query = Solicitacao.query.options(joinedload(Solicitacao.usuario))

        # permissão
        if user_tipo not in ["admin", "operario", "visualizar"]:
            query = query.filter(Solicitacao.usuario_id == user_id)
            filtro_uvis_id = None
        else:
            if filtro_uvis_id:
                query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)

        #  SOMENTE status permitidos para rota
        query = query.filter(
            Solicitacao.status.in_(["APROVADO", "APROVADO COM RECOMENDAÇÕES"])
        )

        # dia exato
        query = query.filter(Solicitacao.data_agendamento == dia_date)

        eventos = query.order_by(Solicitacao.hora_agendamento.asc()).all()

        pontos = []
        total_com_coords = 0

        for e in eventos:
            lat = e.latitude
            lng = e.longitude

            # normaliza (string/decimal)
            try:
                if isinstance(lat, str):
                    lat = float(lat.replace(",", "."))
                else:
                    lat = float(lat) if lat is not None else None

                if isinstance(lng, str):
                    lng = float(lng.replace(",", "."))
                else:
                    lng = float(lng) if lng is not None else None
            except:
                lat = None
                lng = None

            if lat is None or lng is None:
                continue

            total_com_coords += 1

            pontos.append({
                "id": e.id,
                "lat": lat,
                "lng": lng,
                "hora": e.hora_agendamento.strftime("%H:%M") if e.hora_agendamento else "00:00",
                "uvis": e.usuario.nome_uvis if e.usuario else "",
                "foco": e.foco or "",
                "status": e.status,
                "endereco": f"{e.logradouro or ''}, {e.numero or 'S/N'} - {e.bairro or ''}".strip()
            })

        return jsonify(
            ok=True,
            dia=dia,
            total_eventos=len(eventos),
            total_com_coordenadas=total_com_coords,
            pontos=pontos
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(ok=False, error=str(e)), 500

@bp.route("/agenda/exportar_excel", endpoint="agenda_exportar_excel")
@login_required
def exportar_agenda_excel():  
    if current_user.tipo_usuario != "admin":
        abort(403)  # Forbidden

    user_tipo = current_user.tipo_usuario
    user_id = current_user.id
    export_all = request.args.get("all") == "1"

    # filtros
    filtro_status = None if export_all else (request.args.get("status") or None)
    filtro_uvis_id = None if export_all else request.args.get("uvis_id", type=int)
    mes = None if export_all else request.args.get("mes", type=int)
    ano = None if export_all else request.args.get("ano", type=int)

    query = Solicitacao.query.options(joinedload(Solicitacao.usuario))

    if filtro_uvis_id:
        query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)
    if mes and ano:
        filtro_mesano = f"{ano}-{mes:02d}"
        if db.engine.name == 'postgresql':
            query = query.filter(db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM") == filtro_mesano)
        else:
            query = query.filter(db.func.strftime("%Y-%m", Solicitacao.data_agendamento) == filtro_mesano)

    query = query.order_by(
        Solicitacao.data_agendamento.desc(),
        Solicitacao.hora_agendamento.desc()
    )
    eventos = query.all()
    # -----------------------------
    # Monta XLSX
    # -----------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda"

    headers = [
        "DATA",
        "HORÁRIO",
        "REGIÃO",
        "UVIS",
        "CET",
        "ENDEREÇO DA AÇÃO",
        "CEP",
        "FOCO DA AÇÃO",
        "COORDENADA GEOGRÁFICA",
        "Altura dos Voos",
        "Protocolo DECA",
        "Status",
    ]
    ws.append(headers)

    for p in eventos:
        endereco_completo = (
            f"{p.logradouro or ''}, {getattr(p, 'numero', '')} - "
            f"{p.bairro or ''} - "
            f"{(p.cidade or '')}/{(p.uf or '')} - "
            f"{p.cep or ''}"
        )
        if getattr(p, "complemento", None):
            endereco_completo += f" - {p.complemento}"

        cet_txt = "SIM" if getattr(p, "apoio_cet", None) else "NÃO"
        data_str = p.data_agendamento.strftime("%d/%m/%Y") if p.data_agendamento else ""
        hora_str = p.hora_agendamento.strftime("%H:%M") if p.hora_agendamento else ""
        uvis_nome = p.usuario.nome_uvis if getattr(p, "usuario", None) else ""
        regiao = p.usuario.regiao if getattr(p, "usuario", None) else ""
        lat = getattr(p, "latitude", "") or ""
        lon = getattr(p, "longitude", "") or ""
        coordenada = f"{lat},{lon}" if (lat or lon) else ""
        protocolo_deca = getattr(p, "protocolo_deca", None) or getattr(p, "protocolo", "") or ""

        ws.append([
            data_str,
            hora_str,
            regiao,
            uvis_nome,
            cet_txt,
            endereco_completo,
            getattr(p, "cep", "") or "",
            getattr(p, "foco", "") or "",
            coordenada,
            getattr(p, "altura_voo", "") or "",
            protocolo_deca,
            getattr(p, "status", "") or "",
        ])

    # -----------------------------
    # Estilo
    # -----------------------------
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = wrap if cell.row > 1 else center

    for col in range(1, ws.max_column + 1):
        max_len = max(len(str(c.value)) if c.value else 0 for c in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 60)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    nome = "agenda_tudo.xlsx" if export_all else "agenda_exportada.xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=nome,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# =================================================
# NOTIFICAÇÕES (Flask-Login: login_required + current_user)
# Requer no topo:
# from flask_login import login_required, current_user
# from flask import redirect, url_for, render_template
# from datetime import datetime, date
# from sqlalchemy.orm import joinedload
# from zoneinfo import ZoneInfo
# =================================================

from zoneinfo import ZoneInfo

TZ_BR = ZoneInfo("America/Sao_Paulo")

def agora_brasilia_naive():
    """
    Retorna datetime no horário de Brasília, mas sem tzinfo (naive),
    para bater com db.DateTime (sem timezone).
    """
    return datetime.now(TZ_BR).replace(tzinfo=None)


# -------------------------------------------------
# CRIAR NOTIFICAÇÃO
# -------------------------------------------------
def criar_notificacao(usuario_id, titulo, mensagem="", link=None, commit=True):
    n = Notificacao(
        usuario_id=usuario_id,
        titulo=titulo,
        mensagem=mensagem or "",
        link=link,
        criada_em=agora_brasilia_naive(),  #  Brasília
    )
    db.session.add(n)
    if commit:
        db.session.commit()
    return n


# -------------------------------------------------
# GARANTIR NOTIFICAÇÕES DO DIA (sem duplicar)
#  REGRA: se já existiu (mesmo apagada), NÃO recria
# -------------------------------------------------
def garantir_notificacoes_do_dia(usuario_id):
    hoje = date.today()

    ags = (
        Solicitacao.query
        .options(joinedload(Solicitacao.usuario))
        .filter_by(usuario_id=usuario_id)
        .filter(Solicitacao.data_agendamento == hoje)
        .all()
    )

    for s in ags:
        hora_fmt = s.hora_agendamento.strftime("%H:%M") if s.hora_agendamento else "00:00"

        # 🔒 chave estável (muda por dia por conta do d=hoje)
        link = url_for("main.agenda", sid=s.id, d=hoje.isoformat())

        #  Se já existe (inclusive apagada), NÃO cria novamente
        ja_existe = (
            Notificacao.query
            .filter_by(usuario_id=usuario_id, link=link)
            .first()
        )
        if ja_existe:
            continue

        criar_notificacao(
            usuario_id=usuario_id,
            titulo="Agendamento para hoje",
            mensagem=f"Você tem um agendamento hoje às {hora_fmt} (Foco: {s.foco}).",
            link=link
        )


# -------------------------------------------------
# LER NOTIFICAÇÃO
# -------------------------------------------------
@bp.route("/notificacoes/<int:notif_id>/ler")
@login_required
def ler_notificacao(notif_id):
    user_tipo = current_user.tipo_usuario

    if user_tipo in ["admin", "operario", "visualizar"]:
        n = Notificacao.query.get_or_404(notif_id)
    else:
        n = (Notificacao.query
             .filter_by(id=notif_id, usuario_id=current_user.id)
             .first_or_404())

    if n.lida_em is None:
        n.lida_em = agora_brasilia_naive()  #  Brasília
        db.session.commit()

    return redirect(n.link or url_for("main.notificacoes"))


# -------------------------------------------------
# LISTAR NOTIFICAÇÕES
# -------------------------------------------------
@bp.route("/notificacoes")
@login_required
def notificacoes():
    user_tipo = current_user.tipo_usuario

    #  só UVIS gera lembrete do dia (pro próprio usuário)
    if user_tipo not in ["admin", "operario", "visualizar"]:
        garantir_notificacoes_do_dia(current_user.id)

    base = Notificacao.query.filter(Notificacao.apagada_em.is_(None))

    #  admin/operário/visualizar vê tudo, uvis só as dela
    if user_tipo in ["admin", "operario", "visualizar"]:
        itens = base.order_by(Notificacao.criada_em.desc()).all()
    else:
        itens = (base
                 .filter_by(usuario_id=current_user.id)
                 .order_by(Notificacao.criada_em.desc())
                 .all())

    return render_template("notificacoes.html", itens=itens)


# -------------------------------------------------
# EXCLUIR UMA NOTIFICAÇÃO (SOFT DELETE)
# -------------------------------------------------
@bp.route("/notificacoes/<int:notif_id>/excluir", methods=["POST"])
@login_required
def excluir_notificacao(notif_id):
    user_tipo = current_user.tipo_usuario

    if user_tipo in ["admin", "operario", "visualizar"]:
        n = Notificacao.query.get_or_404(notif_id)
    else:
        n = (Notificacao.query
             .filter_by(id=notif_id, usuario_id=current_user.id)
             .first_or_404())

    n.apagada_em = agora_brasilia_naive()  #  Brasília
    db.session.commit()

    return redirect(url_for("main.notificacoes"))


# -------------------------------------------------
# LIMPAR TODAS AS NOTIFICAÇÕES (SOFT DELETE EM LOTE)
# -------------------------------------------------
@bp.route("/notificacoes/limpar", methods=["POST"])
@login_required
def limpar_notificacoes():
    user_tipo = current_user.tipo_usuario
    agora = agora_brasilia_naive()  #  Brasília

    q = Notificacao.query.filter(Notificacao.apagada_em.is_(None))

    if user_tipo not in ["admin", "operario", "visualizar"]:
        q = q.filter_by(usuario_id=current_user.id)

    q.update({"apagada_em": agora}, synchronize_session=False)
    db.session.commit()

    return redirect(url_for("main.notificacoes"))


# ==========================
# CHATBOT UVIS (FAQ inteligente)
# ==========================
import unicodedata

from flask import jsonify, request
from flask_login import login_required, current_user


def _norm(text: str) -> str:
    if not text:
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


UVIS_FAQ = [
    {
        "title": "Status da solicitação",
        "keywords": ["status", "pendente", "em analise", "aprovado", "negado", "protocolo"],
        "answer": (
            "📌 **Significado dos status**:\n"
            "- **Pendente**: solicitação registrada e aguardando início do processo.\n"
            "- **Em Análise**: pedido em validação pela equipe responsável.\n"
            "- **Aprovado**: pedido autorizado (pode aparecer o número de protocolo).\n"
            "- **Aprovado com Recomendações**: pedido aprovado com sugestões de melhoria.\n"
            "- **Negado**: pedido não aprovado (o motivo aparece nos detalhes).\n\n"
            "💡 Dica: clique em **Detalhes** para ver justificativa/protocolo."
        ),
    },
    {
        "title": "O que tem na tela 'Minhas Solicitações' (Dashboard)",
        "keywords": ["dashboard", "minhas solicitacoes", "tela inicial", "filtro", "detalhes", "nova solicitacao", "editar", "equipes", "informacoes", "modal", "equipe"],
        "answer": (
            "Na tela **Minhas Solicitações** você encontra:\n"
            "- Botão **Nova Solicitação** (abre o formulário)\n"
            "- **Filtro por status** (Pendente, Em Análise, Aprovado, Aprovado com Recomendações, Negado)\n"
            "- **Tabela** com data/hora, localização e foco\n"
            "- Botão **Detalhes** (abre um modal com informações completas)\n"
            "- Botão **Adicionar/Editar Equipes** (abre um modal para inserir a equipe responsável)\n"
            "- Botão **Editar solicitação** (abre para editar a solicitação apenas quando está pendente ou negada.)\n"
        ),
    },
    {
        "title": "Campos obrigatórios ao criar uma solicitação",
        "keywords": ["novo", "nova solicitacao", "cadastro", "campos", "obrigatorio", "cep", "numero", "tipo de visita", "altura", "foco"],
        "answer": (
            " No cadastro de uma nova solicitação, atenção aos campos:\n"
            "- **Data** e **Hora** (obrigatórios)\n"
            "- **CEP** (8 dígitos) para preencher endereço automático\n"
            "- **Logradouro** (confirmar) e **Número** (preencher manualmente)\n"
            "- **Tipo de visita** (Monitoramento / Aedes / Culex)\n"
            "- **Altura do voo** (10m, 20m, 30m, 40m)\n"
            "- **Foco da ação** (ex.: Imóvel Abandonado, Piscina/Caixa d’água, Terreno Baldio, Ponto Estratégico)\n"
        ),
    },
    {
        "title": "CEP / endereço não encontrado e boas práticas",
        "keywords": ["cep", "endereco", "logradouro", "bairro", "cidade", "uf", "nao encontrado", "boas praticas"],
        "answer": (
            "Se o **CEP não for encontrado**, preencha o endereço manualmente e revise.\n"
            "Boas práticas:\n"
            "- confira se o **CEP** corresponde ao local\n"
            "- verifique logradouro/bairro/cidade/UF\n"
            "- preencha o **número** corretamente\n"
        ),
    },
    {
        "title": "Latitude/Longitude e mapa",
        "keywords": ["latitude", "longitude", "coordenadas", "gps", "mapa", "consulta", "localizacao"],
        "answer": (
            "📍 **Latitude/Longitude** é preenchido automaticamente após inserir o endereço.\n"
            "Tendo coordenadas, o sistema oferecer acesso rápido ao mapa.\n"
            "Você consegue consultar latitude e longitude na opção Geolocalização.\n"        
        ),
    },
    {
        "title": "Notificações e Agenda",
        "keywords": ["notificacao", "notificacoes", "agenda", "calendario", "lembrete"],
        "answer": (
            "🔔 Em **Notificações**, você vê alertas da unidade (lembretes do dia/atualizações).\n"
            "Ao clicar, pode ser direcionado para a **Agenda**, que mostra os agendamentos por mês/lista.\n"
            "Ao clicar em uma solicitação, abre um modal com as informações completas da solicitação com opção de traçar a rota.\n"
        ),
    },
    {
        "title": "Checklist antes de enviar",
        "keywords": ["checklist", "antes de enviar", "enviar pedido", "validar"],
        "answer": (
            "🧾 **Checklist rápido antes de enviar**:\n"
            "☐ Data e hora corretas\n"
            "☐ CEP válido e endereço conferido\n"
            "☐ Número preenchido\n"
            "☐ Tipo de visita e altura do voo selecionados\n"
            "☐ Foco da ação selecionado\n"
            "☐ Endereço válido?\n"
            "☐ Observações (se necessário) com informações objetivas\n"
        ),
    },
    {
        "title": "Suporte",
        "keywords": ["suporte", "erro", "acesso", "login", "senha", "ajuda", "Suporte", "Ajuda"],
        "answer": (            
            "Entre em contato com o time de suporte da Oceano Azul: **suporte@ijadrones.com.br.**"
        ),
    },
]


@bp.route("/api/uvis/chatbot", methods=["POST"])
@login_required
def uvis_chatbot():
    # (opcional) se quiser limitar só para UVIS:
    # if current_user.tipo_usuario != "uvis":
    #     return jsonify({"answer": "Acesso negado."}), 403

    payload = request.get_json(silent=True) or {}
    msg = (payload.get("message") or "").strip()

    if not msg:
        return jsonify({"answer": "Escreva sua dúvida (ex.: “o que significa Em Análise?”)."}), 400

    nmsg = _norm(msg)

    best = None
    best_score = 0

    for item in UVIS_FAQ:
        score = 0
        for kw in item["keywords"]:
            if kw in nmsg:
                score += 1
        if score > best_score:
            best_score = score
            best = item

    if not best or best_score == 0:
        sugestoes = [
            "• “O que significa Pendente/Em Análise/Aprovado/Aprovado com Recomendações/Negado?”",
            "• “Quais campos são obrigatórios na Nova Solicitação?”",
            "• “O que fazer se o CEP não encontrar?”",
            "• “Qual o checklist antes de enviar?”",
            "• “Como funciona Notificações e Agenda?”",
        ]
        return jsonify({
            "answer": (
                "Não encontrei essa dúvida diretamente no manual.\n\n"
                "Tenta uma dessas perguntas:\n" + "\n".join(sugestoes)
            ),
            "matched": None,
            "confidence": 0,
        }), 200

    return jsonify({
        "answer": best["answer"],
        "matched": best["title"],
        "confidence": best_score,
    }), 200


import os
from flask import abort, send_from_directory
from flask_login import login_required, current_user

@bp.route("/solicitacao/<int:id>/anexo", endpoint="baixar_anexo")
@bp.route("/admin/solicitacao/<int:id>/anexo", endpoint="baixar_anexo_admin")
@login_required
def baixar_anexo(id):
    pedido = Solicitacao.query.get_or_404(id)

    # 🔐 permissões
    if current_user.tipo_usuario not in ["admin", "operario", "visualizar", "uvis"]:
        abort(403)
    if current_user.tipo_usuario == "uvis" and pedido.usuario_id != current_user.id:
        abort(403)

    if not pedido.anexo_path:
        abort(404)

    #  mesma pasta do upload
    upload_folder = get_upload_folder()

    #  normaliza o caminho salvo no banco
    rel = (pedido.anexo_path or "").replace("\\", "/")
    if rel.startswith("upload-files/"):
        rel = rel.split("upload-files/", 1)[1]
    rel = os.path.basename(rel)  # segurança

    file_path = os.path.join(upload_folder, rel)
    if not os.path.isfile(file_path):
        abort(404)

    return send_from_directory(
        upload_folder,
        rel,
        as_attachment=False,
        download_name=(pedido.anexo_nome or rel)
    )

@bp.route("/admin/solicitacao/<int:id>/remover_anexo", methods=["POST"])
@login_required
def remover_anexo(id):
    pedido = Solicitacao.query.get_or_404(id)

    # ... lógica de permissão ...

    pedido.anexo_path = None
    pedido.anexo_nome = None
    db.session.commit()

    #  Isso fará o Toast "Removido com sucesso" aparecer no topo igual aos outros deletes
    flash('PDF removido com sucesso!', 'success') 
    return redirect(url_for('main.dashboard'))

@bp.route("/admin/uvis/novo", methods=["GET", "POST"], endpoint="admin_uvis_novo")
@login_required
def admin_uvis_novo():
    # SOMENTE ADMIN
    if current_user.tipo_usuario != "admin":
        abort(403)

    if request.method == "POST":
        nome_uvis = (request.form.get("nome_uvis") or "").strip()
        regiao = (request.form.get("regiao") or "").strip() or None
        codigo_setor = (request.form.get("codigo_setor") or "").strip() or None

        login = (request.form.get("login") or "").strip()
        senha = request.form.get("senha") or ""
        confirmar = request.form.get("confirmar") or ""

        if not nome_uvis or not login or not senha:
            flash("Preencha: Nome da UVIS, Login e Senha.", "warning")
            return render_template("admin_uvis_novo.html")

        if senha != confirmar:
            flash("As senhas não conferem.", "warning")
            return render_template("admin_uvis_novo.html")

        novo_user = Usuario(
            nome_uvis=nome_uvis,
            regiao=regiao,            
            login=login,
            tipo_usuario="uvis",
        )
        novo_user.set_senha(senha)

        try:
            db.session.add(novo_user)
            db.session.commit()
            flash("UVIS cadastrada com sucesso!", "success")
            return redirect(url_for("main.admin_uvis_listar"))
        except IntegrityError:
            db.session.rollback()
            flash("Esse login já está em uso. Escolha outro.", "danger")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar UVIS: {e}", "danger")

    return render_template("admin_uvis_novo.html")


@bp.route("/admin/uvis", methods=["GET"], endpoint="admin_uvis_listar")
@login_required
def admin_uvis_listar():
    if current_user.tipo_usuario is ["admin", "operario", "visualizar"]:
        abort(403)

    q = (request.args.get("q") or "").strip()
    regiao = (request.args.get("regiao") or "").strip()
    codigo_setor = (request.args.get("codigo_setor") or "").strip()

    query = Usuario.query.filter(Usuario.tipo_usuario == "uvis")

    if q:
        query = query.filter(
            db.or_(
                Usuario.nome_uvis.ilike(f"%{q}%"),
                Usuario.login.ilike(f"%{q}%")
            )
        )

    if regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{regiao}%"))

    if codigo_setor:
        query = query.filter(Usuario.codigo_setor.ilike(f"%{codigo_setor}%"))

    total = query.count()
    page = request.args.get("page", 1, type=int)
    paginacao = query.order_by(Usuario.nome_uvis.asc()).paginate(
        page=page, per_page=10, error_out=False
    )

    query = db.session.query(Solicitacao).options(
        joinedload(Solicitacao.usuario),
        joinedload(Solicitacao.piloto)  # 
    ).filter(Solicitacao.usuario_id == current_user.id)


    filters = {
        "q": q,
        "regiao": regiao,        
        "total": total
    }

    return render_template(
        "admin_uvis_listar.html",
        uvis=paginacao.items,
        paginacao=paginacao,
        filters=filters,
        q=q,
        regiao=regiao        
    )

def _admin_only_redirect():
    if current_user.tipo_usuario != "admin":
        flash("Você não tem permissão para acessar esta função.", "danger")
        return redirect(request.referrer or url_for("main.admin_uvis_listar"))
    return None
@bp.route("/admin/uvis/<int:id>/editar", methods=["GET", "POST"], endpoint="admin_uvis_editar")
@login_required
def admin_uvis_editar(id):
    resp = _admin_only_redirect()
    if resp: 
        return resp

    uvis = Usuario.query.get_or_404(id)

    if uvis.tipo_usuario != "uvis":
        flash("Registro inválido para edição.", "danger")
        return redirect(url_for("main.admin_uvis_listar"))

    if request.method == "POST":
        nome_uvis = (request.form.get("nome_uvis") or "").strip()
        regiao = (request.form.get("regiao") or "").strip() or None
        login = (request.form.get("login") or "").strip()

        senha = (request.form.get("senha") or "").strip()
        confirmar = (request.form.get("confirmar") or "").strip()

        if not nome_uvis or not login:
            flash("Preencha: Nome da UVIS e Login.", "warning")
            return render_template("admin_uvis_editar.html", uvis=uvis)

        if senha:
            if senha != confirmar:
                flash("As senhas não conferem.", "warning")
                return render_template("admin_uvis_editar.html", uvis=uvis)
            uvis.set_senha(senha)

        uvis.nome_uvis = nome_uvis
        uvis.regiao = regiao
        uvis.login = login

        try:
            db.session.commit()
            flash("UVIS atualizada com sucesso!", "success")
            return redirect(url_for("main.admin_uvis_listar"))
        except IntegrityError:
            db.session.rollback()
            flash("Esse login já está em uso. Escolha outro.", "danger")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar: {e}", "danger")

    return render_template("admin_uvis_editar.html", uvis=uvis)


@bp.route("/admin/uvis/<int:id>/excluir", methods=["POST"], endpoint="admin_uvis_excluir")
@login_required
def admin_uvis_excluir(id):
    resp = _admin_only_redirect()
    if resp:
        return resp

    uvis = Usuario.query.get_or_404(id)

    if uvis.tipo_usuario != "uvis":
        flash("Registro inválido para exclusão.", "danger")
        return redirect(url_for("main.admin_uvis_listar"))

    existe = Solicitacao.query.filter_by(usuario_id=uvis.id).first()
    if existe:
        flash("Não é possível excluir: esta UVIS possui solicitações vinculadas.", "warning")
        return redirect(url_for("main.admin_uvis_listar"))

    try:
        db.session.delete(uvis)
        db.session.commit()
        flash("UVIS excluída com sucesso!", "success")
    except Exception:
        db.session.rollback()
        flash("Erro ao excluir UVIS.", "danger")

    return redirect(url_for("main.admin_uvis_listar"))


# ==========================
# CHATBOT ADMIN (FAQ inteligente) - Flask-Login
# ==========================
import unicodedata

from flask import jsonify, request
from flask_login import login_required, current_user


def _norm_admin(text: str) -> str:
    if not text:
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_answer(text: str) -> str:
    """Remove markdown simples (**negrito**, `code`, etc) e normaliza."""
    if not text:
        return ""
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)   # remove ** **
    text = text.replace("`", "")                  # remove ` `
    text = re.sub(r"\n{3,}", "\n\n", text)         # evita muitas quebras
    return text.strip()


ADMIN_FAQ = [
    {
        "title": "Perfis e permissões",
        "keywords": ["acesso", "perfil", "permissao", "permissões", "admin", "operario", "operário", "visualizar", "quem pode"],
        "answer": (
            "<b>Perfis do painel:</b><br>"
            "- <b>Administrador</b>: acesso total (<b>editar</b>, <b>excluir</b>, <b>gerenciar UVIS</b>, <b>relatórios</b> e <b>agenda</b>).<br>"
            "- <b>Operário</b>: consegue <b>salvar decisões</b> (<b>status</b>, <b>protocolo</b> e <b>justificativa</b>).<br>"
            "- <b>Visualizar</b>: <b>apenas leitura</b>.<br>"
        ),
    },
    {
        "title": "Filtros no painel",
        "keywords": ["filtro", "filtrar", "status", "unidade", "uvis", "regiao", "região", "buscar", "pesquisar"],
        "answer": (
            "<b>No painel você pode filtrar por:</b><br>"
            "- <b>Status</b><br>"
            "- <b>Unidade (UVIS)</b><br>"
            "- <b>Região</b><br>"
            "Use os <b>filtros</b> para encontrar <b>solicitações específicas</b> rapidamente."
        ),
    },
    {
        "title": "Olá! Como posso ajudar?",
        "keywords": ["olá", "oi", "hello", "hi", "bom dia", "boa tarde", "boa noite", "ajuda", "suporte"],
        "answer": (
            "Olá! Sou o <b>assistente virtual</b> do <b>painel administrativo</b>.<br>"
            "<b>Posso ajudar</b> com dúvidas sobre:<br>"
            "- <b>Perfis e permissões</b><br>"
            "- <b>Filtros no painel</b><br>"
            "- <b>Salvar decisão</b><br>"
            "- <b>Editar completo</b><br>"
            "- <b>Excluir solicitação</b><br>"
            "- <b>Anexos</b><br>"
            "- <b>GPS e mapa</b><br>"
            "- <b>Exportar Excel</b><br>"
            "- <b>Agenda</b><br>"
            "- <b>Relatórios</b><br>"
            "- <b>Gestão de UVIS</b><br>"
            "- <b>Google Maps</b><br>"
            "<b>Como posso ajudar você hoje?</b>"
        ),
    },
    {
        "title": "Salvar decisão",
        "keywords": ["salvar", "decisao", "decisão", "status", "protocolo", "justificativa", "aprovado", "negado", "analise", "recomendacoes", "recomendações"],
        "answer": (
            "Em cada <b>solicitação</b> você pode definir:<br>"
            "- <b>Status</b><br>"
            "- <b>Protocolo</b><br>"
            "- <b>Justificativa</b> (obrigatória ao <b>negar</b> ou <b>orientar</b>)<br>"
            "Se o perfil for <b>Visualizar</b>, fica em <b>somente leitura</b>."
        ),
    },
    {
        "title": "Editar completo",
        "keywords": ["editar", "editar completo", "corrigir", "alterar", "data", "hora", "endereco", "endereço", "agendamento"],
        "answer": (
            "<b>Editar completo</b> serve para <b>corrigir todos os dados</b> do pedido:<br>"
            "<b>Data/Hora</b>, <b>Endereço</b>, <b>Foco</b>, <b>Tipo de visita</b>, <b>Altura</b> e <b>Observações</b>.<br>"
            "Em alguns casos o sistema pode gerar <b>notificação para a unidade</b>."
        ),
    },
    {
        "title": "Excluir solicitação",
        "keywords": ["excluir", "deletar", "apagar", "remover"],
        "answer": (
            "<b>Excluir</b> remove a solicitação <b>definitivamente</b>.<br>"
            "Normalmente é restrito ao perfil <b>Administrador</b> e pede <b>confirmação</b>."
        ),
    },
    {
        "title": "Anexos",
        "keywords": ["anexo", "arquivo", "upload", "baixar", "download", "pdf", "png", "jpg", "doc", "xlsx"],
        "answer": (
            "Você pode <b>anexar arquivos</b> na solicitação e depois <b>baixar</b>.<br>"
            "Se o anexo não aparecer, verifique se foi <b>salvo corretamente</b> e se o <b>formato é permitido</b>."
        ),
    },
    {
        "title": "GPS e mapa",
        "keywords": ["gps", "latitude", "longitude", "coordenadas", "mapa", "google maps", "consulta", "localizacao", "localização", "geolocalizacao", "geolocalização", "Mapas"],
        "answer": (
            "<b>Latitude e Longitude</b> é utilizado para <b>localizar o endereço com precisão</b>.<br>"
            "Quando preenchidas e puxadas corretamente, o botão de <b>mapa</b> abre o local no <b>Google Maps</b>.<br>"
            "O sistema possui mapas de calor para melhorar a visualização das áreas com mais solicitações.<br>"
            "O campo de Geolocalização permite consultar coordenadas a partir do endereço e traçar rotas."
        ),
    },
    {
        "title": "Exportar Excel do painel",
        "keywords": ["exportar", "excel", "xlsx", "planilha", "baixar excel"],
        "answer": (
            "Existe <b>exportação para Excel</b> a partir do painel.<br>"
            "Os <b>filtros aplicados</b> (<b>status</b>, <b>unidade</b>, <b>região</b>) refletem no <b>arquivo exportado</b>."
        ),
    },
    {
        "title": "Agenda",
        "keywords": ["agenda", "calendario", "calendário", "eventos", "mes", "mês", "ano", "exportar agenda"],
        "answer": (
            "A <b>Agenda</b> mostra <b>agendamentos</b> por período.<br>"
            "Você pode <b>filtrar</b> e <b>exportar</b> quando disponível."
            "Você pode <b>traçar as rotas</b> quando disponível mais de 2 solicitações aprovadas naquele dia."
        ),
    },
    {
        "title": "Relatórios",
        "keywords": ["relatorio", "relatórios", "pdf", "grafico", "gráfico", "totais", "mes", "ano"],
        "answer": (
            "<b>Relatórios</b> permitem filtrar por <b>mês</b>, <b>ano</b> e <b>unidade</b>.<br>"
            "Podem ser exportados em <b>PDF</b> e <b>Excel</b>."
        ),
    },
    {
        "title": "Pilotos",
        "keywords": ["piloto", "Piloto", "pilotos", "Pilotos", "copiloto", "auxiliar de piloto", "auxiliar"],
        "answer": (
            "<b>Pilotos</b> são os responsáveis pela <b>execução das solicitações</b>.<br>"
            "O <b>Cadastro de pilotos</b> permite:<br>"
            "- <b>Cadastrar</b><br>"
            "- <b>Editar</b><br>"
            "- <b>Excluir</b><br>"
            "- <b>Listar</b><br>"
            "Cada solicitação pode ter um <b>piloto associado</b>.<br>"
            "As <b>UVIS</b> veem os pilotos da <b>sua região</b>."
        ),
    },
    {
        "title": "Gestão de UVIS",
        "keywords": ["uvis", "cadastrar uvis", "lista uvis", "gerenciar uvis", "unidade", "login", "senha", "codigo setor", "código setor", "regiao", "região"],
        "answer": (
            "<b>Gestão de UVIS</b> inclui:<br>"
            "- <b>Listar UVIS</b><br>"
            "- <b>Cadastrar UVIS</b><br>"
            "- <b>Editar UVIS</b> (inclusive <b>redefinir senha</b>)<br>"
            "<b>Atenção:</b> o <b>login não pode se repetir</b>."
        ),
    },
]



@bp.route("/api/admin/chatbot", methods=["POST"])
@login_required
def admin_chatbot():
    # 🔐 só perfis do painel
    if current_user.tipo_usuario not in ["admin", "operario", "visualizar"]:
        return jsonify({"answer": "Acesso negado para este chatbot."}), 403

    payload = request.get_json(silent=True) or {}
    msg = (payload.get("message") or "").strip()

    if not msg:
        return jsonify({"answer": "Digite sua dúvida (ex.: como exportar Excel?)."}), 400

    nmsg = _norm_admin(msg)

    best = None
    best_score = 0

    for item in ADMIN_FAQ:
        score = 0
        for kw in item["keywords"]:
            if kw in nmsg:
                score += 1
        if score > best_score:
            best_score = score
            best = item

    if not best or best_score == 0:
        sugestoes = [
            "Como filtrar por status/unidade/região?",
            "Como salvar decisão (status/protocolo/justificativa)?",
            "Como editar completo?",
            "Como exportar Excel?",
            "Como funciona Agenda/Relatórios?",
            "Como gerenciar UVIS?",
        ]
        return jsonify({
            "answer": "Não achei essa dúvida direto no guia.\n\nSugestões:\n- " + "\n- ".join(sugestoes),
            "matched": None,
            "confidence": 0,
        }), 200

    return jsonify({
        "answer": _clean_answer(best["answer"]),
        "matched": best["title"],
        "confidence": best_score,
    }), 200



@bp.app_errorhandler(404)
def pagina_nao_encontrada(e):
    return render_template(
        'erro.html', 
        codigo=404, 
        titulo="Página não encontrada", 
        mensagem="Ops! A página que você está procurando não existe ou foi movida."
    ), 404

@bp.app_errorhandler(500)
def erro_interno(e):
    # Opcional: printar o erro no terminal para você ver o que houve
    # print(f"Erro 500 detectado: {e}")
    return render_template(
        'erro.html', 
        codigo=500, 
        titulo="Erro Interno do Servidor", 
        mensagem="Desculpe, algo deu errado do nosso lado. Tente novamente mais tarde."
    ), 500

import re
import requests
from flask import jsonify, current_app
from flask_login import login_required

def only_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")

@bp.route("/api/cep/<cep>", methods=["GET"], endpoint="api_cep")
@login_required
def api_cep(cep):
    cep_digits = only_digits(cep)

    if len(cep_digits) != 8:
        return jsonify(ok=False, error="CEP inválido. Use 8 dígitos."), 400

    def _resp_ok(payload):
        return jsonify(
            ok=True,
            cep=payload.get("cep", ""),
            logradouro=payload.get("logradouro", ""),
            complemento=payload.get("complemento", ""),
            bairro=payload.get("bairro", ""),
            cidade=payload.get("cidade", ""),
            uf=payload.get("uf", ""),
        )

    headers = {"User-Agent": "Mozilla/5.0"}

    # 1) ViaCEP
    try:
        r = requests.get(f"https://viacep.com.br/ws/{cep_digits}/json/", timeout=3, headers=headers, verify=False)
        r.raise_for_status()
        data = r.json()

        if data.get("erro"):
            return jsonify(ok=False, error="CEP não encontrado."), 404

        payload = {
            "cep": data.get("cep", ""),
            "logradouro": data.get("logradouro", ""),
            "complemento": data.get("complemento", ""),
            "bairro": data.get("bairro", ""),
            "cidade": data.get("localidade", ""),
            "uf": data.get("uf", ""),
        }
        return _resp_ok(payload)

    except Exception as e:
        current_app.logger.exception("Falha ViaCEP: %s", e)

        # 2) Fallback: BrasilAPI
        try:
            r2 = requests.get(f"https://brasilapi.com.br/api/cep/v1/{cep_digits}", timeout=3, headers=headers, verify=False)
            r2.raise_for_status()
            data2 = r2.json()

            payload = {
                "cep": data2.get("cep", ""),
                "logradouro": data2.get("street", ""),
                "complemento": "",  # BrasilAPI normalmente não traz
                "bairro": data2.get("neighborhood", ""),
                "cidade": data2.get("city", ""),
                "uf": data2.get("state", ""),
            }
            return _resp_ok(payload)

        except Exception as e2:
            current_app.logger.exception("Falha BrasilAPI: %s", e2)

            # Se estiver em DEBUG, mostra o erro real pra você ver a causa
            if current_app.debug:
                return jsonify(ok=False, error=f"Falha CEP (debug): {repr(e2)}"), 502

            return jsonify(ok=False, error="Falha ao consultar o serviço de CEP."), 502

from flask import request, abort, send_file
from flask_login import login_required, current_user
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@bp.route("/admin/uvis/exportar", methods=["GET"], endpoint="admin_uvis_exportar")
@login_required
def admin_uvis_exportar():
    if current_user.tipo_usuario != "admin":
        abort(403)

    q = (request.args.get("q") or "").strip()
    regiao = (request.args.get("regiao") or "").strip()    

    query = Usuario.query.filter(Usuario.tipo_usuario == "uvis")

    if q:
        query = query.filter(
            db.or_(
                Usuario.nome_uvis.ilike(f"%{q}%"),
                Usuario.login.ilike(f"%{q}%")
            )
        )
    if regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{regiao}%"))
        
    rows = query.order_by(Usuario.nome_uvis.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "UVIS"

    # ---------- ESTILOS ----------
    title_font = Font(bold=True, size=14)
    meta_font = Font(size=10, color="666666")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    zebra_fill = PatternFill("solid", fgColor="F3F6FA")

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    # ---------- TÍTULO / META (FORA DA TABELA) ----------
    ws["A1"] = "UVIS Cadastradas"
    ws["A1"].font = title_font

    ws["A3"] = f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = meta_font

    start_header_row = 5

    # ---------- CABEÇALHO ----------
    headers = ["ID", "Nome", "Região", "Login"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # ---------- DADOS ----------
    start_data_row = start_header_row + 1
    for i, u in enumerate(rows):
        r = start_data_row + i
        values = [u.id, u.nome_uvis, u.regiao, u.login]

        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = center if c == 1 else left

            if i % 2 == 1:
                cell.fill = zebra_fill

    end_data_row = start_data_row + len(rows) - 1

    # ---------- AUTOFILTER (SEGURO) ----------
    if rows:
        ws.auto_filter.ref = f"A{start_header_row}:E{end_data_row}"
        ws.freeze_panes = f"A{start_data_row}"

    # ---------- LARGURAS ----------
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 24

    # ---------- TOTAL ----------
    total_row = end_data_row + 2
    ws.cell(row=total_row, column=1, value="Total de UVIS:").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(rows)).font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"uvis_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route('/sw.js')
def serve_sw():
    return bp.send_static_file('sw.js')

def _uvis_only():
    if getattr(current_user, "tipo_usuario", None) != "uvis":
        abort(403)

def _proximo_slot_equipe_uvis(uvis_usuario_id: int, nome_equipe: str):
    usados = {
        x[0] for x in (
            db.session.query(EquipeUvis.ordem)
            .filter_by(uvis_usuario_id=uvis_usuario_id, nome_equipe=nome_equipe)
            .all()
        )
    }
    for slot in range(1, 6):
        if slot not in usados:
            return slot
    return None  # cheio

from sqlalchemy import func

@bp.route("/uvis/equipes", methods=["GET"], endpoint="listar_equipes_uvis")
@login_required
def listar_equipes_uvis():
    _uvis_only()

    uvis_id = current_user.id

    # 1) total de membros por equipe (pode não existir se equipe ainda não tem membro)
    membros_rows = (
        db.session.query(
            EquipeUvis.nome_equipe.label("nome_equipe"),
            func.count(EquipeUvis.id).label("total")
        )
        .filter(EquipeUvis.uvis_usuario_id == uvis_id)
        .group_by(EquipeUvis.nome_equipe)
        .all()
    )
    membros_map = {r.nome_equipe: int(r.total) for r in membros_rows}

    # 2) contas (login) por equipe
    contas_rows = (
        db.session.query(
            Usuario.equipe_uvis_nome.label("nome_equipe"),
            Usuario.login.label("login")
        )
        .filter(
            Usuario.tipo_usuario == "equipe_uvis",
            Usuario.equipe_uvis_uvis_usuario_id == uvis_id,
            Usuario.equipe_uvis_nome.isnot(None),
        )
        .all()
    )
    login_map = {r.nome_equipe: r.login for r in contas_rows if r.nome_equipe}

    # 3) conjunto final de equipes: as que têm membros OU as que têm conta
    nomes_equipes = sorted(set(membros_map.keys()) | set(login_map.keys()))

    equipes = []
    for nome in nomes_equipes:
        equipes.append({
            "nome_equipe": nome,
            "total": int(membros_map.get(nome, 0)),
            "login": login_map.get(nome),  # pode ser None
        })

    return render_template("uvis_equipes_listar.html", equipes=equipes)

import re
from flask import request, flash, redirect, url_for, abort
from werkzeug.security import generate_password_hash

@bp.route("/uvis/equipes/<string:nome_equipe>/credenciais", methods=["POST"], endpoint="atualizar_credenciais_equipe_uvis")
@login_required
def atualizar_credenciais_equipe_uvis(nome_equipe):
    _uvis_only()

    nome_equipe = (nome_equipe or "").strip()
    if not nome_equipe:
        flash("Equipe inválida.", "danger")
        return redirect(url_for("main.listar_equipes_uvis"))

    # 🔒 pega a conta da equipe (se existir)
    conta = (
        Usuario.query
        .filter(
            Usuario.tipo_usuario == "equipe_uvis",
            Usuario.equipe_uvis_uvis_usuario_id == current_user.id,
            Usuario.equipe_uvis_nome == nome_equipe
        )
        .first()
    )

    if not conta:
        flash("Conta (login) desta equipe não encontrada.", "warning")
        return redirect(url_for("main.listar_equipes_uvis"))

    login_novo = (request.form.get("login_equipe") or "").strip()
    senha = (request.form.get("senha") or "").strip()
    senha2 = (request.form.get("senha2") or "").strip()

    # -------------------
    # valida login (se veio)
    # -------------------
    if login_novo:
        if len(login_novo) < 4:
            flash("Login deve ter pelo menos 4 caracteres.", "warning")
            return redirect(url_for("main.listar_equipes_uvis"))
        if len(login_novo) > 50:
            flash("Login deve ter no máximo 50 caracteres.", "warning")
            return redirect(url_for("main.listar_equipes_uvis"))
        if not re.match(r"^[A-Za-z0-9._\-]+$", login_novo):
            flash("Login inválido: use apenas letras, números, ponto (.), hífen (-) e underscore (_).", "warning")
            return redirect(url_for("main.listar_equipes_uvis"))

        # se mudou, checa duplicado
        if login_novo != conta.login:
            existe = Usuario.query.filter(Usuario.login == login_novo).first()
            if existe:
                flash("Este login já está em uso. Escolha outro.", "danger")
                return redirect(url_for("main.listar_equipes_uvis"))

        conta.login = login_novo

    # -------------------
    # valida senha (só troca se preencher)
    # -------------------
    if senha or senha2:
        if not senha:
            flash("Informe a senha.", "warning")
            return redirect(url_for("main.listar_equipes_uvis"))
        if len(senha) < 6:
            flash("A senha deve ter pelo menos 6 caracteres.", "warning")
            return redirect(url_for("main.listar_equipes_uvis"))
        if senha != senha2:
            flash("As senhas não conferem.", "warning")
            return redirect(url_for("main.listar_equipes_uvis"))

        conta.set_senha(senha)

    db.session.commit()
    flash("Credenciais atualizadas com sucesso!", "success")
    return redirect(url_for("main.listar_equipes_uvis"))

@bp.route("/uvis/equipes/<string:nome_equipe>", methods=["GET"], endpoint="listar_membros_equipe_uvis")
@login_required
def listar_membros_equipe_uvis(nome_equipe):
    _uvis_only()

    nome_equipe = (nome_equipe or "").strip()
    if not nome_equipe:
        abort(404)

    membros = (
        EquipeUvis.query
        .filter_by(uvis_usuario_id=current_user.id, nome_equipe=nome_equipe)
        .order_by(EquipeUvis.ordem.asc())
        .all()
    )

    total = len(membros)
    maximo = 5

    return render_template(
        "uvis_equipe_membros_listar.html",
        nome_equipe=nome_equipe,
        membros=membros,
        total=total,
        maximo=maximo
    )
from flask import request, flash, redirect, url_for

@bp.route("/uvis/equipes/<string:nome_equipe>/adicionar", methods=["GET", "POST"], endpoint="adicionar_membro_equipe_uvis")
@login_required
def adicionar_membro_equipe_uvis(nome_equipe):
    _uvis_only()

    nome_equipe = (nome_equipe or "").strip()
    if not nome_equipe:
        abort(404)

    errors = {}
    form = {"nome_equipe": nome_equipe}

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        funcao = (request.form.get("funcao") or "").strip()
        contato = (request.form.get("contato") or "").strip()

        form.update({"nome": nome, "funcao": funcao, "contato": contato})

        if not nome:
            errors["nome"] = "Informe o nome do membro."

        slot = _proximo_slot_equipe_uvis(current_user.id, nome_equipe)
        if not slot:
            errors["limite"] = "Limite máximo de 5 pessoas nesta equipe atingido."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("uvis_equipe_membro_adicionar.html", form=form, errors=errors, nome_equipe=nome_equipe)

        novo = EquipeUvis(
            uvis_usuario_id=current_user.id,
            nome_equipe=nome_equipe,
            ordem=slot,
            nome=nome,
            funcao=funcao or None,
            contato=contato or None
        )
        db.session.add(novo)
        db.session.commit()

        flash("Membro adicionado com sucesso!", "success")
        return redirect(url_for("main.listar_membros_equipe_uvis", nome_equipe=nome_equipe))

    return render_template("uvis_equipe_membro_adicionar.html", form=form, errors=errors, nome_equipe=nome_equipe)

import re
import unicodedata
from flask import abort
from flask_login import current_user

def _uvis_only():
    if getattr(current_user, "tipo_usuario", None) != "uvis":
        abort(403)

def _slug_upper(text: str) -> str:
    """
    Remove acentos, transforma separadores (/, _, espaços) em '-', remove chars inválidos e deixa MAIÚSCULO.
    """
    text = (text or "").strip()
    if not text:
        return ""

    # normaliza acentos
    text = unicodedata.normalize("NFKD", text)
    text = "".join([c for c in text if not unicodedata.combining(c)])

    text = text.upper()

    # separadores comuns viram hífen
    text = text.replace("/", "-").replace("\\", "-").replace("_", "-")
    text = re.sub(r"\s+", "-", text)

    # mantém só A-Z, 0-9 e hífen
    text = re.sub(r"[^A-Z0-9\-]", "", text)

    # limpa hífens duplicados
    text = re.sub(r"\-+", "-", text).strip("-")

    return text

def _get_first_nonempty(*values) -> str:
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return ""

def _nome_uvis_base() -> str:
    """
    Pega o nome da UVIS e remove prefixos tipo 'UVIS ' / 'UVIS-' para não duplicar no nome final.
    """
    direto = _get_first_nonempty(
        getattr(current_user, "nome_uvis", None),
        getattr(current_user, "nome", None),
        getattr(current_user, "name", None),
        getattr(current_user, "nome_completo", None),
        getattr(current_user, "username", None),
    )

    if not direto:
        u = getattr(current_user, "usuario", None)
        if u is not None:
            direto = _get_first_nonempty(
                getattr(u, "nome_uvis", None),
                getattr(u, "nome", None),
                getattr(u, "name", None),
                getattr(u, "nome_completo", None),
            )

    if not direto:
        email = getattr(current_user, "email", None)
        if email and "@" in str(email):
            direto = str(email).split("@", 1)[0]

    if not direto:
        direto = "SEM-NOME"

    #  remove "UVIS" do começo (ex: "UVIS Lapa/Pinheiros" ou "UVIS-Lapa")
    direto = str(direto).strip()
    direto = re.sub(r"^\s*UVIS\s*[-:\s]*", "", direto, flags=re.IGNORECASE).strip()

    return direto or "SEM-NOME"

def _proximo_nome_equipe_uvis(uvis_usuario_id: int) -> str:
    """
    Formato final: UVIS-<NOME_DA_UVIS>-<N>
    Ex.: UVIS-LAPA-PINHEIROS-1
    """
    nome_uvis = _slug_upper(_nome_uvis_base())
    if not nome_uvis:
        nome_uvis = "SEM-NOME"

    prefixo = f"UVIS-{nome_uvis}-"

    rows = (
        db.session.query(EquipeUvis.nome_equipe)
        .filter(EquipeUvis.uvis_usuario_id == uvis_usuario_id)
        .distinct()
        .all()
    )
    existentes = [r[0] for r in rows if r and r[0]]

    maior = 0
    pattern = re.compile(rf"^{re.escape(prefixo)}(\d+)$", re.IGNORECASE)

    for nome in existentes:
        m = pattern.match(nome.strip())
        if m:
            try:
                n = int(m.group(1))
                if n > maior:
                    maior = n
            except ValueError:
                pass

    return f"{prefixo}{maior + 1}"
#-------------------------------------------------------------
# Rota: criar nova equipe UVIS
#-------------------------------------------------------------
from flask import request, flash, redirect, url_for, render_template
from flask_login import login_required
import secrets
def _login_equipe_sugerido(nome_equipe: str) -> str:
    # sugestão automática, mas editável no template
    return f"EQUIPE-{_slug_upper(nome_equipe)}"[:50]

@bp.route("/uvis/equipes/nova", methods=["GET", "POST"], endpoint="criar_equipe_uvis")
@login_required
def criar_equipe_uvis():
    _uvis_only()

    errors = {}
    form = {}

    # sempre gera o nome automático (imutável)
    nome_equipe = _proximo_nome_equipe_uvis(current_user.id)
    form["nome_equipe"] = nome_equipe

    # sugestão de login (editável)
    form["login_equipe"] = _login_equipe_sugerido(nome_equipe)

    if request.method == "POST":
        # nome da equipe continua automático
        nome_equipe = _proximo_nome_equipe_uvis(current_user.id)
        form["nome_equipe"] = nome_equipe

        login_equipe = (request.form.get("login_equipe") or "").strip()
        senha = (request.form.get("senha") or "").strip()
        senha2 = (request.form.get("senha2") or "").strip()

        form["login_equipe"] = login_equipe  # preserva se der erro

        if not nome_equipe:
            errors["nome_equipe"] = "Não foi possível gerar o nome automático da equipe."

        # ---- valida login
        if not login_equipe:
            errors["login_equipe"] = "Informe o login da equipe."
        elif len(login_equipe) < 4:
            errors["login_equipe"] = "O login deve ter pelo menos 4 caracteres."
        elif len(login_equipe) > 50:
            errors["login_equipe"] = "O login deve ter no máximo 50 caracteres."
        elif not re.match(r"^[A-Za-z0-9._\-]+$", login_equipe):
            errors["login_equipe"] = "Use apenas letras, números, ponto (.), hífen (-) e underscore (_)."
        else:
            existente = Usuario.query.filter_by(login=login_equipe).first()
            if existente:
                errors["login_equipe"] = "Este login já está em uso. Escolha outro."

        # ---- valida senha
        if not senha:
            errors["senha"] = "Informe a senha da equipe."
        elif len(senha) < 6:
            errors["senha"] = "A senha deve ter pelo menos 6 caracteres."
        elif senha != senha2:
            errors["senha2"] = "As senhas não conferem."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("uvis_equipe_criar.html", form=form, errors=errors)

        # cria usuario da equipe (com senha DEFINIDA, não temporária)
        usuario_equipe = Usuario(
            nome_uvis=nome_equipe,
            regiao=current_user.regiao,
            codigo_setor=current_user.codigo_setor,
            login=login_equipe,
            tipo_usuario="equipe_uvis",
            piloto_id=None,
            equipe_uvis_uvis_usuario_id=current_user.id,
            equipe_uvis_nome=nome_equipe,
        )
        usuario_equipe.set_senha(senha)

        db.session.add(usuario_equipe)
        db.session.commit()

        flash("Equipe criada! Login da equipe definido com sucesso.", "success")
        return redirect(url_for("main.adicionar_membro_equipe_uvis", nome_equipe=nome_equipe))

    return render_template("uvis_equipe_criar.html", form=form, errors=errors)
# -------------------------------------------------------------
# Rota: editar membro da equipe UVIS    
#-------------------------------------------------------------

@bp.route("/uvis/equipe-membro/<int:membro_id>/editar", methods=["GET", "POST"], endpoint="editar_membro_equipe_uvis")
@login_required
def editar_membro_equipe_uvis(membro_id):
    _uvis_only()

    membro = EquipeUvis.query.get_or_404(membro_id)

    # 🔒 só a UVIS dona
    if membro.uvis_usuario_id != current_user.id:
        abort(403)

    errors = {}
    form = {}

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        funcao = (request.form.get("funcao") or "").strip()
        contato = (request.form.get("contato") or "").strip()

        form = {"nome": nome, "funcao": funcao, "contato": contato}

        if not nome:
            errors["nome"] = "Informe o nome do membro."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("uvis_equipe_membro_editar.html", membro=membro, form=form, errors=errors)

        membro.nome = nome
        membro.funcao = funcao or None
        membro.contato = contato or None

        db.session.commit()
        flash("Membro atualizado com sucesso!", "success")
        return redirect(url_for("main.listar_membros_equipe_uvis", nome_equipe=membro.nome_equipe))

    form = {
        "nome": membro.nome or "",
        "funcao": membro.funcao or "",
        "contato": membro.contato or "",
    }
    return render_template("uvis_equipe_membro_editar.html", membro=membro, form=form, errors=errors)

#-------------------------------------------------------------
# Rota: deletar membro da equipe UVIS   
#-------------------------------------------------------------
@bp.route("/uvis/equipe-membro/<int:membro_id>/deletar", methods=["POST"], endpoint="deletar_membro_equipe_uvis")
@login_required
def deletar_membro_equipe_uvis(membro_id):
    _uvis_only()

    membro = EquipeUvis.query.get_or_404(membro_id)
    if membro.uvis_usuario_id != current_user.id:
        abort(403)

    nome_equipe = membro.nome_equipe

    db.session.delete(membro)
    db.session.commit()

    flash("Membro removido com sucesso.", "success")

    # se ficou sem ninguém, volta para lista de equipes
    restante = EquipeUvis.query.filter_by(uvis_usuario_id=current_user.id, nome_equipe=nome_equipe).count()
    if restante == 0:
        flash("Equipe ficou sem membros e não será mais exibida.", "info")
        return redirect(url_for("main.listar_equipes_uvis"))

    return redirect(url_for("main.listar_membros_equipe_uvis", nome_equipe=nome_equipe))

@bp.route("/solicitacao/<int:id>/atribuir-equipe-uvis", methods=["POST"], endpoint="atribuir_equipe_uvis_solicitacao")
@login_required
def atribuir_equipe_uvis_solicitacao(id):
    sol = Solicitacao.query.get_or_404(id)

    # 🔒 Segurança: só a UVIS dona da solicitação (ou admin, se quiser permitir)
    if sol.usuario_id != current_user.id and current_user.tipo_usuario != "admin":
        flash("Você não tem permissão para alterar esta solicitação.", "danger")
        return redirect(url_for("main.dashboard"))

    nome_equipe = (request.form.get("nome_equipe") or "").strip()
    if not nome_equipe:
        flash("Selecione uma equipe.", "warning")
        return redirect(url_for("main.dashboard"))

    #  valida se existe ESSA equipe para ESSA UVIS
    existe = (
        db.session.query(EquipeUvis.id)
        .filter(EquipeUvis.uvis_usuario_id == current_user.id)
        .filter(EquipeUvis.nome_equipe == nome_equipe)
        .first()
    )
    if not existe:
        flash("Equipe UVIS não encontrada para seu usuário.", "danger")
        return redirect(url_for("main.dashboard"))

    sol.equipe_uvis_nome = nome_equipe
    db.session.commit()

    flash("Equipe UVIS atribuída com sucesso!", "success")
    return redirect(url_for("main.dashboard"))

from flask import request, jsonify, render_template, current_app
from flask_login import login_required, current_user
from sqlalchemy import extract
import os

# --- API: pontos do mapa ---
@bp.route('/api/heatmap-data')
@login_required
def heatmap_data():
    uvis_id = request.args.get('uvis_id', type=int)   # já converte pra int
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)

    query = Solicitacao.query.filter(
        Solicitacao.latitude.isnot(None),
        Solicitacao.longitude.isnot(None),
        Solicitacao.status.in_(['APROVADO', 'APROVADO COM RECOMENDAÇÕES'])
    )

    # Filtro de data (só aplica se ambos existirem e forem válidos)
    if mes and ano:
        query = query.filter(
            extract('month', Solicitacao.data_agendamento) == mes,
            extract('year', Solicitacao.data_agendamento) == ano
        )

    # Filtro por UVIS
    # - UVIS só vê os próprios
    # - Admin pode filtrar por uvis_id se vier
    if current_user.tipo_usuario == 'uvis':
        query = query.filter(Solicitacao.usuario_id == current_user.id)
    elif current_user.tipo_usuario == 'admin' and uvis_id:
        query = query.filter(Solicitacao.usuario_id == uvis_id)

    solicitacoes = query.all()

    pontos = []
    for s in solicitacoes:
        try:
            lat = float(s.latitude)
            lng = float(s.longitude)
        except (ValueError, TypeError):
            continue

        pontos.append({
            "lat": lat,
            "lng": lng,
            #IMPORTANTE: manter EXATAMENTE o texto do banco (pra bater com o select)
            "foco": (s.foco or "").strip() or "Outros"
        })

    return jsonify(pontos)


# --- Página do mapa ---
@bp.route('/mapa-relatorio')
@login_required
def mapa_relatorio():
    uvis_disponiveis = []
    if current_user.tipo_usuario == 'admin':
        uvis_disponiveis = (
            db.session.query(Usuario.id, Usuario.nome_uvis)
            .filter(Usuario.tipo_usuario == 'uvis')
            .order_by(Usuario.nome_uvis.asc())
            .all()
        )

    google_maps_key = current_app.config.get('Maps_KEY_FRONT') or os.getenv('KEY_API_GOOGLE_MAPS')
    if not google_maps_key:
        current_app.logger.warning("Google Maps API Key não encontrada (Maps_KEY_FRONT / KEY_API_GOOGLE_MAPS).")

    return render_template(
        'mapa_relatorio.html',
        uvis_disponiveis=uvis_disponiveis,
        google_maps_key=google_maps_key
    )

from flask import abort, render_template
from sqlalchemy import func
def _admin_only():
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

# -------------------------------------------------------------
# ADMIN: listar todas as equipes criadas por todas as UVIS
# -------------------------------------------------------------
@bp.route("/admin/uvis/equipes", methods=["GET"], endpoint="admin_listar_equipes_uvis")
@login_required
def admin_listar_equipes_uvis():
    _admin_only()

    # Agrupa por UVIS + nome_equipe
    rows = (
        db.session.query(
            Usuario.id.label("uvis_id"),
            Usuario.nome_uvis.label("uvis_nome"),
            EquipeUvis.nome_equipe.label("nome_equipe"),
            func.count(EquipeUvis.id).label("total")
        )
        .join(Usuario, Usuario.id == EquipeUvis.uvis_usuario_id)
        .filter(Usuario.tipo_usuario == "uvis")
        .group_by(Usuario.id, Usuario.nome_uvis, EquipeUvis.nome_equipe)
        .order_by(Usuario.nome_uvis.asc(), EquipeUvis.nome_equipe.asc())
        .all()
    )

    equipes = [
        {
            "uvis_id": int(r.uvis_id),
            "uvis_nome": r.uvis_nome or "",
            "nome_equipe": r.nome_equipe,
            "total": int(r.total),
        }
        for r in rows
    ]

    # Pode reaproveitar um template específico do admin
    return render_template("admin_uvis_equipes_listar.html", equipes=equipes)

# -------------------------------------------------------------
# ADMIN: listar membros de uma equipe específica de uma UVIS
# -------------------------------------------------------------
@bp.route("/admin/uvis/<int:uvis_id>/equipes/<string:nome_equipe>", methods=["GET"], endpoint="admin_listar_membros_equipe_uvis")
@login_required
def admin_listar_membros_equipe_uvis(uvis_id, nome_equipe):
    _admin_only()

    nome_equipe = (nome_equipe or "").strip()
    if not nome_equipe:
        abort(404)

    membros = (
        EquipeUvis.query
        .filter_by(uvis_usuario_id=uvis_id, nome_equipe=nome_equipe)
        .order_by(EquipeUvis.ordem.asc())
        .all()
    )

    uvis = Usuario.query.get(uvis_id)
    uvis_nome = (uvis.nome_uvis if uvis else "") or ""

    return render_template(
        "admin_uvis_equipe_membros_listar.html",
        uvis_id=uvis_id,
        uvis_nome=uvis_nome,
        nome_equipe=nome_equipe,
        membros=membros,
        total=len(membros),
        maximo=5
    )

@bp.route('/consultar_endereco_geolocalizacao', methods=['GET'])
@login_required
def consultar_endereco_geolocalizacao():
    google_maps_key = os.getenv("KEY_API_GOOGLE_MAPS")    
    return render_template('consultar_endereco_geolocalizacao.html', google_maps_key=google_maps_key)

from flask import abort, current_app
from flask_login import login_required

def _dev_only():
    if not current_app.debug:
        abort(404)

@bp.route("/__test/erro/<int:code>", methods=["GET"], endpoint="test_error_code")
@login_required
def test_error_code(code):
    _dev_only()

    if code == 500:
        raise RuntimeError("Erro 500 forçado para teste")
    abort(code)

import uuid
from flask import render_template, request
from werkzeug.exceptions import HTTPException

def _error_payload(code: int):
    # títulos/mensagens padrão por tipo de erro
    defaults = {
        400: ("Requisição inválida", "A solicitação não pôde ser processada. Verifique os dados e tente novamente."),
        401: ("Não autenticado", "Você precisa fazer login para continuar."),
        403: ("Acesso negado", "Você não tem permissão para acessar este recurso."),
        404: ("Página não encontrada", "O endereço acessado não existe ou foi movido."),
        405: ("Método não permitido", "Essa ação não é permitida para esta rota."),
        408: ("Tempo esgotado", "A solicitação demorou demais. Tente novamente."),
        409: ("Conflito", "Houve um conflito ao processar sua solicitação."),
        410: ("Recurso indisponível", "Esse conteúdo não está mais disponível."),
        415: ("Mídia não suportada", "Formato de arquivo/dados não suportado."),
        422: ("Não foi possível processar", "Verifique os campos informados e tente novamente."),
        429: ("Muitas tentativas", "Você fez muitas solicitações em pouco tempo. Aguarde e tente novamente."),
        500: ("Erro interno", "Ocorreu um erro no servidor. Tente novamente em instantes."),
        502: ("Gateway inválido", "Serviço temporariamente indisponível. Tente novamente."),
        503: ("Serviço indisponível", "Serviço em manutenção ou sobrecarregado. Tente novamente mais tarde."),
    }
    return defaults.get(code, ("Ocorreu um problema", "Não foi possível concluir sua solicitação no momento. Tente novamente."))


def _render_error(code: int, titulo=None, mensagem=None):
    request_id = request.headers.get("X-Request-Id") or str(uuid.uuid4())[:8]
    if not titulo or not mensagem:
        t, m = _error_payload(code)
        titulo = titulo or t
        mensagem = mensagem or m

    return render_template(
        "erro.html",   
        codigo=code,
        titulo=titulo,
        mensagem=mensagem,
        request_id=request_id,
    ), code


#  pega qualquer HTTPException (abort(403), abort(404) etc.)
@bp.app_errorhandler(HTTPException)
def handle_http_exception(e: HTTPException):
    return _render_error(e.code or 500)


#  pega exceptions “reais” (500)
@bp.app_errorhandler(Exception)
def handle_exception(e: Exception):
    # loga no console (ou usa logging)
    try:
        current_app.logger.exception(e)
    except Exception:
        pass
    return _render_error(500)

@bp.post("/solicitacao/<int:id>/cancelar")
@login_required
def cancelar_solicitacao(id):
    s = Solicitacao.query.get_or_404(id)

    # segurança: UVIS só cancela as próprias solicitações
    if current_user.tipo_usuario != "admin" and s.usuario_id != current_user.id:
        abort(403)

    # seta status cancelado
    s.status = "CANCELADO"
    db.session.commit()

    flash("Solicitação cancelada.", "success")
    return redirect(request.referrer or url_for("main.dashboard"))

@bp.route("/canceladas")
@login_required
def solicitacoes_canceladas():
    # se piloto for redirecionado no seu app, mantém sua lógica
    if current_user.tipo_usuario == 'piloto':
        return redirect(url_for('main.piloto_os'))

    # UVIS: só as dela
    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe)
        )
        .filter(Solicitacao.usuario_id == current_user.id)
        .filter(Solicitacao.status == "CANCELADO")
        .order_by(Solicitacao.data_criacao.desc())
    )

    page = request.args.get("page", 1, type=int)
    paginacao = query.paginate(page=page, per_page=6, error_out=False)

    return render_template(
        "dashboard_canceladas.html",
        solicitacoes=paginacao.items,
        paginacao=paginacao
    )

@bp.post("/admin/solicitacao/<int:id>/cancelar")
@login_required
def cancelar_solicitacao_admin(id):
    s = Solicitacao.query.get_or_404(id)

    #  perfis do admin painel podem cancelar tudo
    if current_user.tipo_usuario in ["admin", "operario", "visualizar"]:
        pass
    else:
        #  UVIS só cancela as próprias
        if s.usuario_id != current_user.id:
            abort(403)

    # evita re-cancelar
    if s.status == "CANCELADO":
        flash("Essa solicitação já está cancelada.", "info")
        return redirect(request.referrer or url_for("main.admin_dashboard"))

    s.status = "CANCELADO"
    db.session.commit()

    flash("Solicitação cancelada.", "success")
    return redirect(request.referrer or url_for("main.admin_dashboard"))

@bp.route("/admin/canceladas")
@login_required
def admin_canceladas():
    google_maps_key = os.getenv("KEY_API_GOOGLE_MAPS")

    if current_user.tipo_usuario not in ['admin', 'operario', 'visualizar']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('main.dashboard'))

    # filtros (mantém unidade/região igual ao admin)
    filtro_unidade = (request.args.get("unidade") or "").strip()
    filtro_regiao = (request.args.get("regiao") or "").strip()
    filtro_foco = (request.args.get("foco") or "").strip()

    unidades_select = Usuario.query.filter_by(tipo_usuario='uvis').order_by(Usuario.nome_uvis.asc()).all()

    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe)
        )
        .join(Usuario)
        .filter(Solicitacao.status == "CANCELADO")
    )

    if filtro_unidade:
        query = query.filter(Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"))
    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))
    if filtro_foco: 
        query = query.filter(Solicitacao.foco == filtro_foco)

    page = request.args.get("page", 1, type=int)
    paginacao = query.order_by(Solicitacao.data_criacao.desc()).paginate(page=page, per_page=6, error_out=False)

    return render_template(
        "admin_canceladas.html",
        pedidos=paginacao.items,
        paginacao=paginacao,
        now=datetime.now(),
        unidades_select=unidades_select,
        google_maps_key=google_maps_key,
        foco_selecionado=filtro_foco
    )


@bp.route("/admin/historico-os")
@login_required
def admin_historico_os():
    if current_user.tipo_usuario not in ["admin", "operario", "visualizar"]:
        flash("Acesso restrito.", "danger")
        return redirect(url_for("main.dashboard"))

    filtro_unidade = (request.args.get("unidade") or "").strip()
    filtro_regiao = (request.args.get("regiao") or "").strip()

    unidades_select = (
        Usuario.query
        .filter_by(tipo_usuario="uvis")
        .order_by(Usuario.nome_uvis.asc())
        .all()
    )

    status_concluido = ["CONCLUÍDO", "CONCLUIDO"]

    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe)
        )
        .join(Usuario)
        .filter(Solicitacao.status.in_(status_concluido))
    )

    if filtro_unidade:
        query = query.filter(Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"))
    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))

    page = request.args.get("page", 1, type=int)
    paginacao = (
        query
        .order_by(Solicitacao.data_criacao.desc(), Solicitacao.id.desc())
        .paginate(page=page, per_page=6, error_out=False)
    )

    return render_template(
        "admin_historico_os.html",
        pedidos=paginacao.items,
        paginacao=paginacao,
        unidades_select=unidades_select
    )

import os
import subprocess
import threading
import gzip
import shutil
import dropbox
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from flask import jsonify, render_template, current_app
from apscheduler.schedulers.background import BackgroundScheduler
from flask_login import login_required, current_user

# =========================
# CONFIG
# =========================
TIMEZONE = "America/Sao_Paulo"
TZ = ZoneInfo(TIMEZONE)
BACKUP_DIR = Path(__file__).resolve().parent / "backup"

scheduler = BackgroundScheduler(timezone=TIMEZONE)
_scheduler_started = False

# =========================
# DROPBOX CORE (Versão Única e Correta)
# =========================
def upload_to_dropbox(file_path):
    """Versão simplificada para debugar o erro de credenciais"""
    app_key = os.environ.get('DROPBOX_APP_KEY', '').strip()
    app_secret = os.environ.get('DROPBOX_APP_SECRET', '').strip()
    refresh_token = os.environ.get('DROPBOX_REFRESH_TOKEN', '').strip()

    # Esse print vai nos confirmar se o Python está lendo as chaves do Render
    print(f"DEBUG: Tentando Dropbox com Key: {app_key[:4]}... / Secret: {app_secret[:4]}...")

    if not all([app_key, app_secret, refresh_token]):
        print("❌ ERRO: Faltam variáveis de ambiente do Dropbox no Render.")
        return False

    zipped_file = file_path.with_suffix(file_path.suffix + ".gz")

    try:
        # Compactação
        with open(file_path, 'rb') as f_in:
            with gzip.open(zipped_file, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)

        # Conexão FORÇANDO a renovação do token
        dbx = dropbox.Dropbox(
            app_key=app_key,
            app_secret=app_secret,
            oauth2_refresh_token=refresh_token
        )

        dest_path = f"/backups/{zipped_file.name}"
        with open(zipped_file, "rb") as f:
            meta = dbx.files_upload(f.read(), dest_path, mode=dropbox.files.WriteMode.overwrite)
            print(f" SUCESSO ABSOLUTO! Salvo em: {meta.path_display}")
        
        if zipped_file.exists(): os.remove(zipped_file)
        if file_path.exists(): os.remove(file_path)
        return True

    except Exception as e:
        # Se der erro aqui, saberemos se é credencial ou permissão
        print(f"❌ ERRO NO DROPBOX: {str(e)}")
        return False

# =========================
# BACKUP CORE
# =========================
def _ensure_backup_dir():
    if not BACKUP_DIR.exists():
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        print(f"📁 Pasta de backup criada em: {BACKUP_DIR}")

def _backup_filename():
    stamp = datetime.now(TZ).strftime("%d-%m-%Y_%H-%M")

    project_name = os.getenv("PROJECT_NAME", "backup")
    environment = os.getenv("APP_ENV", "prod")  # prod | staging | dev

    return BACKUP_DIR / f"{project_name}_{environment}_{stamp}.sql"


def _run_postgres_backup():
    """Gera o arquivo de backup REAL (Postgres) e envia para o Dropbox"""
    database_url = os.getenv("DATABASE_URL")
    _ensure_backup_dir()
    output_file = _backup_filename()

    try:
        if not database_url:
            raise RuntimeError("DATABASE_URL não configurada no .env ou no Render.")

        # --- LÓGICA DE BACKUP REAL ---
        # Se estiver no Windows, tentamos localizar o pg_dump.exe
        pg_dump_cmd = "pg_dump"
        if os.name == 'nt':
            print(f"🖥️ Localhost (Windows) detectado. Tentando backup real do banco remoto...")
            # Se o pg_dump não estiver no seu PATH, você pode colocar o caminho completo abaixo:
            # pg_dump_cmd = r"C:\Program Files\PostgreSQL\16\bin\pg_dump.exe"

        cmd = [
            pg_dump_cmd,
            "--no-owner",
            "--no-privileges",
            "--format=plain",
            "--file", str(output_file),
            database_url
        ]

        # Executa o comando real
        # shell=True ajuda o Windows a encontrar o executável se ele estiver no PATH
        result = subprocess.run(cmd, capture_output=True, text=True, shell=(os.name == 'nt'))

        if result.returncode != 0:
            # Se falhar no Windows, pode ser que você não tenha o PostgreSQL instalado localmente
            if os.name == 'nt':
                raise RuntimeError(f"O comando pg_dump falhou. Verifique se o PostgreSQL está instalado no seu Nitro V15. Erro: {result.stderr}")
            else:
                raise RuntimeError(f"pg_dump falhou no servidor: {result.stderr}")

        print(f"⚙️ Backup SQL gerado com sucesso: {output_file.name}")

        # --- ENVIO PARA O DROPBOX ---
        upload_success = upload_to_dropbox(output_file)
        
        if not upload_success:
            print("⚠️ O backup foi gerado, mas o envio ao Dropbox falhou.")
            
        return output_file 

    except Exception as e:
        print(f"❌ Erro crítico no backup real: {e}")
        raise e

# =========================
# SCHEDULER (1x/dia 05:00)
# =========================
def start_daily_backup_scheduler():
    global _scheduler_started
    if _scheduler_started:
        return

    _ensure_backup_dir()

    if scheduler.get_job("daily_backup_0500") is None:
        scheduler.add_job(
            _run_postgres_backup,
            trigger="cron",
            hour=5,
            minute=0,
            id="daily_backup_0500",
            replace_existing=True,
            max_instances=1,
            coalesce=True,
        )

    if not scheduler.running:
        scheduler.start()

    _scheduler_started = True

@bp.record_once
def _on_bp_load(state):
    start_daily_backup_scheduler()

# =========================
# UI BACKUP (admin)
# =========================
_backup_state = {
    "running": False,
    "last_file": None,
    "last_error": None,
    "started_at": None,
    "finished_at": None,
}

def _run_backup_async():
    try:
        _backup_state["running"] = True
        _backup_state["last_error"] = None
        _backup_state["started_at"] = datetime.now(TZ).isoformat()
        _backup_state["finished_at"] = None

        file_path = _run_postgres_backup() 
        _backup_state["last_file"] = f"Enviado para Nuvem: {file_path.name}.gz"

    except Exception as e:
        _backup_state["last_error"] = str(e)
    finally:
        _backup_state["running"] = False
        _backup_state["finished_at"] = datetime.now(TZ).isoformat()

def _list_backups():
    """Lista os arquivos diretamente da pasta /backups no Dropbox"""
    app_key = os.environ.get('DROPBOX_APP_KEY').strip()
    app_secret = os.environ.get('DROPBOX_APP_SECRET').strip()
    refresh_token = os.environ.get('DROPBOX_REFRESH_TOKEN').strip()

    backups = []
    try:
        dbx = dropbox.Dropbox(
            app_key=app_key,
            app_secret=app_secret,
            oauth2_refresh_token=refresh_token
        )

        # Lista os arquivos da pasta /backups
        # Se a pasta estiver vazia ou não existir, ele cai no except
        result = dbx.files_list_folder('/backups')

        for entry in sorted(result.entries, key=lambda x: x.name, reverse=True):
            if isinstance(entry, dropbox.files.FileMetadata):
                backups.append({
                    "name": entry.name,
                    "path": entry.path_display,
                    "size_bytes": entry.size,
                    "modified_at": entry.client_modified.replace(tzinfo=ZoneInfo("UTC")).astimezone(TZ),
                    "is_cloud": True # Flag para sabermos que está na nuvem
                })
        
    except Exception as e:
        print(f"⚠️ Erro ao listar Dropbox (Pasta pode estar vazia): {e}")
        # Se falhar a nuvem, tenta olhar a pasta local por garantia
        _ensure_backup_dir()
        files = sorted(BACKUP_DIR.glob("backup_*"), key=lambda p: p.stat().st_mtime, reverse=True)
        for p in files:
            st = p.stat()
            backups.append({
                "name": p.name,
                "path": str(p),
                "size_bytes": st.st_size,
                "modified_at": datetime.fromtimestamp(st.st_mtime, tz=TZ),
                "is_cloud": False
            })
            
    return backups
# (As rotas /backup, /backup/status e /backups permanecem iguais)

@bp.route("/backup", methods=["GET"])
@login_required
def backup_page():
    if getattr(current_user, "tipo_usuario", None) != "admin":
        return render_template(
            "backup_aguarde.html",
            codigo=403,
            titulo="Acesso negado",
            mensagem="Apenas administradores podem gerar backup do banco.",
            is_error=True,
        ), 403

    # dispara o backup async se não estiver rodando
    if not _backup_state["running"]:
        t = threading.Thread(target=_run_backup_async, daemon=True)
        t.start()

    return render_template(
        "backup_aguarde.html",
        codigo="Backup",
        titulo="Gerando backup do banco de dados",
        mensagem="Aguarde alguns segundos. Ao finalizar, você poderá ver a lista de backups gerados.",
        is_error=False,
    )


@bp.route("/backup/status", methods=["GET"])
@login_required
def backup_status():
    if getattr(current_user, "tipo_usuario", None) != "admin":
        return jsonify({"ok": False, "error": "forbidden"}), 403

    return jsonify({
        "running": _backup_state["running"],
        "last_file": _backup_state["last_file"],
        "last_error": _backup_state["last_error"],
        "started_at": _backup_state["started_at"],
        "finished_at": _backup_state["finished_at"],
    })


@bp.route("/backups", methods=["GET"])
@login_required
def backups_list_page():
    if getattr(current_user, "tipo_usuario", None) != "admin":
        return render_template(
            "backup_lista.html",
            codigo=403,
            titulo="Acesso negado",
            mensagem="Apenas administradores podem visualizar os backups.",
            backups=[],
            is_error=True,
        ), 403

    backups = _list_backups()
    return render_template(
        "backup_lista.html",
        codigo="Backups",
        titulo="Backups do Banco",
        mensagem="Lista de backups gerados automaticamente (05:00) e manuais.",
        backups=backups,
        is_error=False,
    )

from app.models import Drones , Baterias, Equipamentos

@bp.route('/equipamentos', methods=['GET'])
@login_required
def listar_equipamentos():
    # Usamos o campo 'tipo_equipamento' que definimos no polimorfismo para contar
    total_drones = Equipamentos.query.filter_by(tipo_equipamento='drones').count()
    total_baterias = Equipamentos.query.filter_by(tipo_equipamento='baterias').count()
    em_manutencao = Equipamentos.query.filter_by(status='Em Manutenção').count()
    
    # Busca todos para a tabela
    todos = Equipamentos.query.order_by(Equipamentos.criado_em.desc()).all()

    return render_template('equipamentos_listar.html', 
                            equipamentos=todos,
                            total_drones=total_drones,
                            total_baterias=total_baterias,
                            em_manutencao=em_manutencao)

@bp.route('/equipamentos/drones', methods=['GET'])
@login_required
def listar_drones():
    # Filtra apenas por Drones
    drones = Drones.query.all()
    
    # Define se o usuário tem permissão de administrador
    is_admin = current_user.tipo_usuario == 'admin'
    
    return render_template('drones_listar.html', 
                            drones=drones, 
                            is_admin=is_admin)

@bp.route('/equipamentos/baterias', methods=['GET'])
@login_required
def listar_baterias():

    # Busca todas as baterias (para ver estoque geral)
    baterias = Baterias.query.all()
    is_admin = current_user.tipo_usuario == 'admin'

    return render_template('baterias_listar.html', baterias=baterias, is_admin=is_admin)


# -----------------------------
# Rota: cadastrar drone
# -----------------------------
@bp.route('/drones/cadastrar', methods=['GET', 'POST'], endpoint='cadastrar_drone')
@login_required
def cadastrar_drone():
    # Segurança: só admin
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    errors = {}
    form = {}

    # lista de equipes (opcional no form)
    equipes = Equipe.query.filter_by(ativa=True).order_by(Equipe.nome_equipe.asc()).all()

    if request.method == "POST":
        modelo = (request.form.get("modelo") or "").strip()
        renomacao = (request.form.get("renomacao") or "").strip()
        categoria = (request.form.get("categoria") or "").strip()
        status = (request.form.get("status") or "Ativo").strip()

        ano_fabricacao_raw = (request.form.get("ano_fabricacao") or "").strip()
        numero_serie = (request.form.get("numero_serie") or "").strip()

        registro_anatel = (request.form.get("registro_anatel") or "").strip()
        registro_anac = (request.form.get("registro_anac") or "").strip()

        pmd_kg_raw = (request.form.get("pmd_kg") or "").strip()
        equipe_id_raw = (request.form.get("equipe_id") or "").strip()
        ultima_manutencao_raw = (request.form.get("ultima_manutencao") or "").strip()

        # Mantém valores pra re-render do form
        form = {
            "modelo": modelo,
            "renomacao": renomacao,
            "categoria": categoria,
            "status": status,
            "ano_fabricacao": ano_fabricacao_raw,
            "numero_serie": numero_serie,
            "registro_anatel": registro_anatel,
            "registro_anac": registro_anac,
            "pmd_kg": pmd_kg_raw,
            "equipe_id": equipe_id_raw,
            "ultima_manutencao": ultima_manutencao_raw,
        }

        # Obrigatórios
        if not modelo:
            errors["modelo"] = "Informe o modelo do drone."
        if not renomacao:
            errors["renomacao"] = "Informe a renomação do drone."
        if not registro_anatel:
            errors["registro_anatel"] = "Informe o registro ANATEL."
        if not registro_anac:
            errors["registro_anac"] = "Informe o registro ANAC."
        if not pmd_kg_raw:
            errors["pmd_kg"] = "Informe o PMD (kg)."

        # PMD
        pmd_kg = None
        if pmd_kg_raw:
            try:
                pmd_kg = float(pmd_kg_raw.replace(",", "."))
                if pmd_kg <= 0:
                    errors["pmd_kg"] = "PMD deve ser maior que 0."
            except ValueError:
                errors["pmd_kg"] = "PMD inválido. Use um número (ex: 25.5)."

        # Ano fabricação (opcional)
        ano_fabricacao = None
        if ano_fabricacao_raw:
            try:
                ano_fabricacao = int(ano_fabricacao_raw)
                if ano_fabricacao < 1900 or ano_fabricacao > 2100:
                    errors["ano_fabricacao"] = "Ano de fabricação inválido."
            except ValueError:
                errors["ano_fabricacao"] = "Ano de fabricação inválido."

        # Equipe (opcional)
        equipe_id = None
        if equipe_id_raw:
            try:
                equipe_id = int(equipe_id_raw)
                equipe_ok = Equipe.query.filter_by(id=equipe_id, ativa=True).first()
                if not equipe_ok:
                    errors["equipe_id"] = "Equipe inválida."
            except ValueError:
                errors["equipe_id"] = "Equipe inválida."

        # Última manutenção (opcional)
        ultima_manutencao = None
        if ultima_manutencao_raw:
            try:
                ultima_manutencao = datetime.strptime(ultima_manutencao_raw, "%Y-%m-%d").date()
            except ValueError:
                errors["ultima_manutencao"] = "Data inválida."

        # Unique checks (antes do commit)
        if registro_anac and not errors.get("registro_anac"):
            existe_anac = Drones.query.filter_by(registro_anac=registro_anac).first()
            if existe_anac:
                errors["registro_anac"] = "Já existe um drone com esse Registro ANAC."

        if numero_serie:
            # numero_serie é unique em Equipamentos, mas aqui checamos via Drones já ajuda (se quiser 100%,
            # melhor checar no modelo base Equipamentos também)
            existe_ns = Drones.query.filter_by(numero_serie=numero_serie).first()
            if existe_ns:
                errors["numero_serie"] = "Já existe um equipamento com esse Número de Série."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("cadastrar_drone.html", form=form, errors=errors, equipes=equipes)

        novo = Drones(
            tipo_equipamento="drones",  # importante pro polimorfismo
            status=status,
            modelo=modelo,
            renomacao=renomacao,
            categoria=categoria or None,
            ano_fabricacao=ano_fabricacao,
            numero_serie=numero_serie or None,
            ultima_manutencao=ultima_manutencao,

            equipe_id=equipe_id,

            registro_anatel=registro_anatel,
            registro_anac=registro_anac,
            pmd_kg=pmd_kg,
        )

        try:
            db.session.add(novo)
            db.session.commit()
            flash("Drone cadastrado com sucesso!", "success")
            # troque para sua rota real de listagem quando existir
            return redirect(url_for("main.cadastrar_drone"))
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar drone: {str(e)}", "danger")
            return render_template("cadastrar_drone.html", form=form, errors=errors, equipes=equipes)

    return render_template("cadastrar_drone.html", form=form, errors=errors, equipes=equipes)

# -----------------------------
# Rota: editar drone
# -----------------------------
@bp.route('/drones/<int:drone_id>/editar', methods=['GET', 'POST'], endpoint='editar_drone')
@login_required
def editar_drone(drone_id):

    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    drone = Drones.query.get_or_404(drone_id)

    errors = {}
    form = {}

    equipes = Equipe.query.filter_by(ativa=True).order_by(Equipe.nome_equipe.asc()).all()

    if request.method == "POST":

        modelo = (request.form.get("modelo") or "").strip()
        renomacao = (request.form.get("renomacao") or "").strip()
        categoria = (request.form.get("categoria") or "").strip()
        status = (request.form.get("status") or "Ativo").strip()

        ano_raw = (request.form.get("ano_fabricacao") or "").strip()
        numero_serie = (request.form.get("numero_serie") or "").strip()
        registro_anatel = (request.form.get("registro_anatel") or "").strip()
        registro_anac = (request.form.get("registro_anac") or "").strip()
        pmd_raw = (request.form.get("pmd_kg") or "").strip()
        equipe_id_raw = (request.form.get("equipe_id") or "").strip()
        manut_raw = (request.form.get("ultima_manutencao") or "").strip()

        form = request.form.to_dict()

        # Validações básicas
        if not modelo:
            errors["modelo"] = "Informe o modelo."
        if not renomacao:
            errors["renomacao"] = "Informe a renomação."
        if not registro_anatel:
            errors["registro_anatel"] = "Informe o registro ANATEL."
        if not registro_anac:
            errors["registro_anac"] = "Informe o registro ANAC."

        # PMD
        try:
            pmd_kg = float(pmd_raw.replace(",", "."))
            if pmd_kg <= 0:
                raise ValueError()
        except:
            errors["pmd_kg"] = "PMD inválido."

        # Unique ANAC (exceto ele mesmo)
        existe_anac = Drones.query.filter(
            Drones.registro_anac == registro_anac,
            Drones.id != drone.id
        ).first()
        if existe_anac:
            errors["registro_anac"] = "Já existe outro drone com esse ANAC."

        # Unique Número de Série
        if numero_serie:
            existe_ns = Drones.query.filter(
                Drones.numero_serie == numero_serie,
                Drones.id != drone.id
            ).first()
            if existe_ns:
                errors["numero_serie"] = "Número de série já utilizado."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("editar_drone.html", drone=drone, form=form, errors=errors, equipes=equipes)

        # Atualiza
        drone.modelo = modelo
        drone.renomacao = renomacao
        drone.categoria = categoria or None
        drone.status = status
        drone.numero_serie = numero_serie or None
        drone.registro_anatel = registro_anatel
        drone.registro_anac = registro_anac
        drone.pmd_kg = pmd_kg

        drone.equipe_id = int(equipe_id_raw) if equipe_id_raw else None

        if ano_raw:
            drone.ano_fabricacao = int(ano_raw)
        else:
            drone.ano_fabricacao = None

        if manut_raw:
            drone.ultima_manutencao = datetime.strptime(manut_raw, "%Y-%m-%d").date()
        else:
            drone.ultima_manutencao = None

        db.session.commit()
        flash("Drone atualizado com sucesso!", "success")
        return redirect(url_for("main.listar_drones"))

    # GET
    form = {
        "modelo": drone.modelo,
        "renomacao": drone.renomacao,
        "categoria": drone.categoria,
        "status": drone.status,
        "ano_fabricacao": drone.ano_fabricacao,
        "numero_serie": drone.numero_serie,
        "registro_anatel": drone.registro_anatel,
        "registro_anac": drone.registro_anac,
        "pmd_kg": drone.pmd_kg,
        "equipe_id": drone.equipe_id,
        "ultima_manutencao": drone.ultima_manutencao
    }

    return render_template("editar_drone.html", drone=drone, form=form, errors=errors, equipes=equipes)

# -----------------------------
# Rota: deletar drone (robusta)
# -----------------------------
@bp.route('/drones/<int:drone_id>/deletar', methods=['POST'], endpoint='deletar_drone')
@login_required
def deletar_drone(drone_id):

    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    drone = Drones.query.get_or_404(drone_id)

    try:
        # 1) desvincula baterias antes
        for bateria in list(drone.baterias):
            bateria.drone_id = None

        # força o UPDATE das baterias acontecer antes do DELETE do drone
        db.session.flush()

        # 2) agora remove o drone
        db.session.delete(drone)
        db.session.commit()

        flash("Drone removido com sucesso.", "success")

    except Exception as e:
        db.session.rollback()
        # importante: logar no console também (Render)
        print("ERRO AO DELETAR DRONE:", repr(e))
        flash("Erro ao remover drone. Verifique vínculos (baterias/OS) e tente novamente.", "danger")

    return redirect(url_for("main.listar_drones"))


# -----------------------------
# Rota: cadastrar bateria (vincula a um drone)
# -----------------------------
@bp.route('/baterias/cadastrar', methods=['GET', 'POST'], endpoint='cadastrar_bateria')
@login_required
def cadastrar_bateria():
    # Segurança: só admin
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    errors = {}
    form = {}

    # lista de drones para dropdown
    drones = Drones.query.order_by(Drones.renomacao.asc()).all()

    # opcional: permitir pré-seleção por querystring ?drone_id=123
    drone_id_pre = request.args.get("drone_id", type=int)

    if request.method == "POST":
        modelo = (request.form.get("modelo") or "").strip()
        renomacao = (request.form.get("renomacao") or "").strip()
        status = (request.form.get("status") or "Ativo").strip()

        categoria = (request.form.get("categoria") or "").strip()  # opcional
        ano_raw = (request.form.get("ano_fabricacao") or "").strip()
        numero_serie = (request.form.get("numero_serie") or "").strip()

        ciclo_raw = (request.form.get("ciclo") or "").strip()
        drone_id_raw = (request.form.get("drone_id") or "").strip()
        manut_raw = (request.form.get("ultima_manutencao") or "").strip()

        # mantém dados
        form = {
            "modelo": modelo,
            "renomacao": renomacao,
            "status": status,
            "categoria": categoria,
            "ano_fabricacao": ano_raw,
            "numero_serie": numero_serie,
            "ciclo": ciclo_raw,
            "drone_id": drone_id_raw,
            "ultima_manutencao": manut_raw,
        }

        # obrigatórios mínimos
        if not modelo:
            errors["modelo"] = "Informe o modelo da bateria."
        if not renomacao:
            errors["renomacao"] = "Informe a renomação (ex: BAT-01)."

        # ciclo (opcional, mas se vier tem que ser inteiro >= 0)
        ciclo = 0
        if ciclo_raw:
            try:
                ciclo = int(ciclo_raw)
                if ciclo < 0:
                    errors["ciclo"] = "Ciclo não pode ser negativo."
            except ValueError:
                errors["ciclo"] = "Ciclo inválido (use número inteiro)."

        # ano fabricação (opcional)
        ano_fabricacao = None
        if ano_raw:
            try:
                ano_fabricacao = int(ano_raw)
                if ano_fabricacao < 1900 or ano_fabricacao > 2100:
                    errors["ano_fabricacao"] = "Ano de fabricação inválido."
            except ValueError:
                errors["ano_fabricacao"] = "Ano de fabricação inválido."

        # ultima manutenção (opcional)
        ultima_manutencao = None
        if manut_raw:
            try:
                ultima_manutencao = datetime.strptime(manut_raw, "%Y-%m-%d").date()
            except ValueError:
                errors["ultima_manutencao"] = "Data inválida."

        # drone_id (opcional)
        drone_id = None
        if drone_id_raw:
            try:
                drone_id = int(drone_id_raw)
                d_ok = Drones.query.get(drone_id)
                if not d_ok:
                    errors["drone_id"] = "Drone inválido."
            except ValueError:
                errors["drone_id"] = "Drone inválido."

        # numero_serie é unique em Equipamentos (se preencher, validar duplicidade)
        if numero_serie:
            existe_ns = Equipamentos.query.filter_by(numero_serie=numero_serie).first()
            if existe_ns:
                errors["numero_serie"] = "Já existe um equipamento com esse Número de Série."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("cadastrar_bateria.html", form=form, errors=errors, drones=drones)

        nova = Baterias(
            tipo_equipamento="baterias",
            status=status,
            modelo=modelo,
            renomacao=renomacao,
            categoria=categoria or None,
            ano_fabricacao=ano_fabricacao,
            numero_serie=numero_serie or None,
            ultima_manutencao=ultima_manutencao,
            ciclo=ciclo,
            drone_id=drone_id
        )

        try:
            db.session.add(nova)
            db.session.commit()
            flash("Bateria cadastrada com sucesso!", "success")
            # ajuste para sua rota de listagem, se existir
            return redirect(url_for("main.listar_baterias"))
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar bateria: {str(e)}", "danger")
            return render_template("cadastrar_bateria.html", form=form, errors=errors, drones=drones)

    # GET: se veio drone_id na querystring, já deixa selecionado
    if drone_id_pre and Drones.query.get(drone_id_pre):
        form["drone_id"] = str(drone_id_pre)

    return render_template("cadastrar_bateria.html", form=form, errors=errors, drones=drones)


# -----------------------------
# Rota: editar bateria
# -----------------------------
@bp.route('/baterias/<int:bateria_id>/editar', methods=['GET', 'POST'], endpoint='editar_bateria')
@login_required
def editar_bateria(bateria_id):
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    bateria = Baterias.query.get_or_404(bateria_id)
    drones = Drones.query.order_by(Drones.renomacao.asc()).all()

    errors = {}
    form = {}

    if request.method == "POST":
        modelo = (request.form.get("modelo") or "").strip()
        renomacao = (request.form.get("renomacao") or "").strip()
        status = (request.form.get("status") or "Ativo").strip()

        categoria = (request.form.get("categoria") or "").strip()
        ano_raw = (request.form.get("ano_fabricacao") or "").strip()
        numero_serie = (request.form.get("numero_serie") or "").strip()

        ciclo_raw = (request.form.get("ciclo") or "").strip()
        drone_id_raw = (request.form.get("drone_id") or "").strip()
        manut_raw = (request.form.get("ultima_manutencao") or "").strip()

        form = request.form.to_dict()

        # obrigatórios
        if not modelo:
            errors["modelo"] = "Informe o modelo da bateria."
        if not renomacao:
            errors["renomacao"] = "Informe a renomação (ex: BAT-01)."

        # ciclo
        ciclo = bateria.ciclo or 0
        if ciclo_raw == "":
            ciclo = 0
        else:
            try:
                ciclo = int(ciclo_raw)
                if ciclo < 0:
                    errors["ciclo"] = "Ciclo não pode ser negativo."
            except ValueError:
                errors["ciclo"] = "Ciclo inválido (use número inteiro)."

        # ano fabricação (opcional)
        ano_fabricacao = None
        if ano_raw:
            try:
                ano_fabricacao = int(ano_raw)
                if ano_fabricacao < 1900 or ano_fabricacao > 2100:
                    errors["ano_fabricacao"] = "Ano de fabricação inválido."
            except ValueError:
                errors["ano_fabricacao"] = "Ano de fabricação inválido."

        # última manutenção (opcional)
        ultima_manutencao = None
        if manut_raw:
            try:
                ultima_manutencao = datetime.strptime(manut_raw, "%Y-%m-%d").date()
            except ValueError:
                errors["ultima_manutencao"] = "Data inválida."

        # drone (opcional)
        drone_id = None
        if drone_id_raw:
            try:
                drone_id = int(drone_id_raw)
                d_ok = Drones.query.get(drone_id)
                if not d_ok:
                    errors["drone_id"] = "Drone inválido."
            except ValueError:
                errors["drone_id"] = "Drone inválido."

        # numero_serie unique em Equipamentos (exceto ele mesmo)
        if numero_serie:
            existe_ns = Equipamentos.query.filter(
                Equipamentos.numero_serie == numero_serie,
                Equipamentos.id != bateria.id
            ).first()
            if existe_ns:
                errors["numero_serie"] = "Já existe outro equipamento com esse Número de Série."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("editar_bateria.html", bateria=bateria, form=form, errors=errors, drones=drones)

        # atualiza
        bateria.modelo = modelo
        bateria.renomacao = renomacao
        bateria.status = status
        bateria.categoria = categoria or None
        bateria.numero_serie = numero_serie or None
        bateria.ciclo = ciclo
        bateria.drone_id = drone_id
        bateria.ano_fabricacao = ano_fabricacao
        bateria.ultima_manutencao = ultima_manutencao

        try:
            db.session.commit()
            flash("Bateria atualizada com sucesso!", "success")
            return redirect(url_for("main.listar_baterias"))
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar bateria: {str(e)}", "danger")
            return render_template("editar_bateria.html", bateria=bateria, form=form, errors=errors, drones=drones)

    # GET (preenche form)
    form = {
        "modelo": bateria.modelo,
        "renomacao": bateria.renomacao,
        "status": bateria.status,
        "categoria": bateria.categoria,
        "ano_fabricacao": bateria.ano_fabricacao,
        "numero_serie": bateria.numero_serie,
        "ciclo": bateria.ciclo,
        "drone_id": bateria.drone_id,
        "ultima_manutencao": bateria.ultima_manutencao.strftime("%Y-%m-%d") if bateria.ultima_manutencao else ""
    }

    return render_template("editar_bateria.html", bateria=bateria, form=form, errors=errors, drones=drones)


# -----------------------------
# Rota: deletar bateria
# -----------------------------
@bp.route('/baterias/<int:bateria_id>/deletar', methods=['POST'], endpoint='deletar_bateria')
@login_required
def deletar_bateria(bateria_id):
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    bateria = Baterias.query.get_or_404(bateria_id)

    try:
        # se estiver vinculada, desvincula antes
        bateria.drone_id = None
        db.session.flush()

        db.session.delete(bateria)
        db.session.commit()

        flash("Bateria removida com sucesso.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao remover bateria: {str(e)}", "danger")

    return redirect(url_for("main.listar_baterias"))

# -----------------------------
# Rota: Equipamentos em Manutenção
# -----------------------------
@bp.route("/equipamentos/em-manutencao", methods=["GET"], endpoint="equipamentos_manutencao")
@login_required
def equipamentos_manutencao():

    # opcional: restringir só admin
    # if getattr(current_user, "tipo_usuario", None) != "admin":
    #     abort(403)

    equipamentos = (
        Equipamentos.query
        .filter(Equipamentos.status == "Em Manutenção")
        .order_by(Equipamentos.criado_em.desc())
        .all()
    )

    return render_template(
        "equipamentos_manutencao.html",
        equipamentos=equipamentos,
        total=len(equipamentos)
    )

@bp.route('/equipamentos/baterias/update_ciclos/<int:id>', methods=['POST'])
@login_required
def update_ciclos(id):
    bateria = Baterias.query.get_or_404(id)
    data = request.get_json()
    
    quantidade = data.get('quantidade', 1)
    operacao = data.get('operacao', 'add')

    if operacao == 'add':
        bateria.ciclo += quantidade
    else:
        bateria.ciclo = max(0, bateria.ciclo - quantidade) # Impede ciclos negativos

    db.session.commit()
    
    # Retorna o novo valor e a cor para o JavaScript atualizar a tela
    return {
        'novo_ciclo': bateria.ciclo,
        'cor': 'bg-danger' if bateria.ciclo > 200 else 'bg-success'
    }


# -----------------------------
# Rota: enviar drone para manutenção
# -----------------------------
@bp.route('/drones/<int:drone_id>/manutencao', methods=['POST'], endpoint='enviar_manutencao_drone')
@login_required
def enviar_manutencao_drone(drone_id):
    # Segurança: só admin
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    drone = Drones.query.get_or_404(drone_id)

    try:
        # já está em manutenção? só avisa e volta
        if (drone.status or "").strip() == "Em Manutenção":
            flash("Este drone já está em manutenção.", "warning")
            return redirect(url_for("main.listar_drones"))

        drone.status = "Em Manutenção"

        # opcional: registra a data (boa prática)
        drone.ultima_manutencao = date.today()

        db.session.commit()
        flash(f"Drone {drone.renomacao} enviado para manutenção.", "success")

    except Exception as e:
        db.session.rollback()
        print("ERRO AO ENVIAR DRONE PARA MANUTENÇÃO:", repr(e))
        flash("Erro ao enviar o drone para manutenção.", "danger")

    return redirect(url_for("main.listar_drones"))


def get_responsaveis_choices():
    """
    Retorna lista de opções para o select de responsável.
    Inclui:
      - Todos os Pilotos cadastrados
      - E marca se ele aparece como 'piloto' e/ou 'auxiliar' em alguma equipe
    """
    # Mapa piloto_id -> set(papeis)
    papeis_por_piloto = {}
    for row in db.session.query(EquipePiloto.piloto_id, EquipePiloto.papel).all():
        papeis_por_piloto.setdefault(row.piloto_id, set()).add((row.papel or "").lower())

    pilotos = Pilotos.query.order_by(Pilotos.nome_piloto.asc()).all()

    opts = []
    for p in pilotos:
        papeis = papeis_por_piloto.get(p.id, set())

        # etiqueta: se tiver os dois papéis, mostra ambos
        if "piloto" in papeis and "auxiliar" in papeis:
            label = f"{p.nome_piloto} (Piloto/Aux)"
        elif "auxiliar" in papeis:
            label = f"{p.nome_piloto} (Auxiliar)"
        else:
            # default: piloto cadastrado (mesmo que não esteja em equipe ainda)
            label = f"{p.nome_piloto} (Piloto)"

        # value salvo no banco (string). Salva só o nome limpo.
        value = p.nome_piloto

        opts.append({"value": value, "label": label})

    return opts


from flask import render_template, request, abort, make_response
from flask_login import login_required, current_user
from app import db
from app.models import Veiculos
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter


from flask import render_template, request, abort, make_response
from flask_login import login_required, current_user
from app import db
from app.models import Veiculos
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


@bp.route("/veiculos", methods=["GET"], endpoint="listar_veiculos")
@login_required
def listar_veiculos():
    tipo = getattr(current_user, "tipo_usuario", None)

    # Permissões
    if tipo not in ("admin", "visualizar", "operario", "uvis", "piloto"):
        abort(403)

    # Filtros
    q = (request.args.get("q") or "").strip()
    operacao = (request.args.get("operacao") or "").strip().upper()
    frota = (request.args.get("frota") or "").strip().upper()
    status = (request.args.get("status") or "").strip()

    export = (request.args.get("export") or "").strip()  # se vier "1" exporta

    query = Veiculos.query

    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                Veiculos.modelo.ilike(like),
                Veiculos.placa.ilike(like),
                Veiculos.responsavel.ilike(like),
            )
        )

    if operacao:
        query = query.filter(Veiculos.operacao == operacao)

    if frota:
        query = query.filter(Veiculos.frota == frota)

    if status:
        query = query.filter(Veiculos.status == status)

    veiculos = query.order_by(Veiculos.criado_em.desc()).all()

    # -----------------------------
    # EXPORTAR EXCEL (design próximo ao print)
    # -----------------------------
    if export in ("1", "true", "yes", "xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Veículos"

        # Estilos
        fill_title = PatternFill("solid", fgColor="FFD966")   # amarelo forte (título)
        fill_header = PatternFill("solid", fgColor="FFF2CC")  # amarelo claro (cabeçalho)
        fill_green  = PatternFill("solid", fgColor="C6EFCE")  # verde claro
        fill_yellow = PatternFill("solid", fgColor="FFEB9C")  # amarelo alerta
        fill_red    = PatternFill("solid", fgColor="FFC7CE")  # vermelho claro
        fill_none   = PatternFill()  # sem fill

        font_bold = Font(bold=True)
        font_title = Font(bold=True, size=12)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        thin = Side(style="thin", color="000000")
        border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

        num_format = '#,##0.00'

        headers = ["MODELO", "ANO", "FROTA", "OPERAÇÃO", "PLACA", "RESPONSAVEL", "KM ATUAL", "PROX REVISAO", "OBS"]

        def write_section(title, rows, start_row):
            # faixa do título (mescla B..H)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=8)
            cell = ws.cell(row=start_row, column=2, value=title)
            cell.font = font_title
            cell.alignment = align_center
            cell.fill = fill_title

            # pinta faixa B..H
            for c in range(2, 9):
                ws.cell(row=start_row, column=c).fill = fill_title
                ws.cell(row=start_row, column=c).border = border_thin

            # cabeçalho (A..I)
            header_row = start_row + 1
            for col_idx, h in enumerate(headers, start=1):
                ch = ws.cell(row=header_row, column=col_idx, value=h)
                ch.font = font_bold
                ch.alignment = align_center
                ch.fill = fill_header
                ch.border = border_thin

            # dados
            r = header_row + 1
            for v in rows:
                falt = v.km_restante_revisao

                obs = ""
                if v.revisao_marcada_em:
                    obs = "MARCADO " + v.revisao_marcada_em.strftime("%d/%m %H:%M")
                elif v.revisao_obs:
                    obs = v.revisao_obs

                data = [
                    v.modelo or "",
                    v.ano_fabricacao or "",
                    v.frota or "",
                    v.operacao or "",
                    v.placa or "",
                    v.responsavel or "",
                    float(v.km_atual or 0),
                    float(v.km_prox_revisao) if v.km_prox_revisao is not None else "",
                    obs,
                ]

                for col_idx, value in enumerate(data, start=1):
                    c = ws.cell(row=r, column=col_idx, value=value)
                    c.border = border_thin
                    c.alignment = align_left if col_idx in (1, 5, 6, 9) else align_center

                    # números
                    if col_idx in (7, 8) and isinstance(value, (int, float)):
                        c.number_format = num_format

                    # KM ATUAL verde
                    if col_idx == 7 and isinstance(value, (int, float)):
                        c.fill = fill_green

                    # PROX REVISAO por status
                    if col_idx == 8:
                        if value == "" or falt is None:
                            c.fill = fill_none
                        else:
                            if falt < 0:
                                c.fill = fill_red
                            elif falt <= 2000:
                                c.fill = fill_yellow
                            else:
                                c.fill = fill_green

                r += 1

            return r + 2  # respiro

        # separar por operação
        by_op = {}
        for v in veiculos:
            op = (v.operacao or "OUTROS").upper()
            by_op.setdefault(op, []).append(v)

        # ordem PMSP, AGRO, resto
        ops_order = []
        for k in ("PMSP", "AGRO"):
            if k in by_op:
                ops_order.append(k)
        for k in sorted(by_op.keys()):
            if k not in ops_order:
                ops_order.append(k)

        current_row = 2
        for op in ops_order:
            current_row = write_section(f"VEICULOS {op}", by_op[op], current_row)

        # larguras
        col_widths = {
            1: 16,  # MODELO
            2: 8,   # ANO
            3: 12,  # FROTA
            4: 12,  # OPERAÇÃO
            5: 14,  # PLACA
            6: 18,  # RESPONSAVEL
            7: 14,  # KM ATUAL
            8: 14,  # PROX REVISAO
            9: 26,  # OBS
        }
        for col_idx, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # gerar arquivo
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"veiculos_{ts}.xlsx"

        response = make_response(file_stream.getvalue())
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # -----------------------------
    # RENDER NORMAL
    # -----------------------------
    is_admin = (tipo == "admin")

    filters = {
        "q": q,
        "operacao": operacao,
        "frota": frota,
        "status": status,
        "total": len(veiculos),
    }

    ultimos_logs = {}
    if veiculos:
        veiculo_ids = [v.id for v in veiculos]
        logs = (
            LogVeiculo.query
            .options(selectinload(LogVeiculo.abastecimentos_detalhados))
            .filter(LogVeiculo.veiculo_id.in_(veiculo_ids))
            .order_by(LogVeiculo.veiculo_id.asc(), LogVeiculo.data_registro.desc())
            .all()
        )
        for log in logs:
            if log.veiculo_id not in ultimos_logs:
                ultimos_logs[log.veiculo_id] = log

    return render_template(
        "veiculos_listar.html",
        veiculos=veiculos,
        is_admin=is_admin,
        filters=filters,
        ultimos_logs=ultimos_logs
    )

from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.rule import CellIsRule


def _ultima_movimentacao_log_subquery():
    return (
        db.session.query(
            Abastecimento.log_veiculo_id.label("log_id"),
            db.func.max(Abastecimento.data_hora).label("ultima_movimentacao_em"),
        )
        .group_by(Abastecimento.log_veiculo_id)
        .subquery()
    )


@bp.route("/veiculos/logs", methods=["GET"], endpoint="veiculos_logs")
@login_required
def veiculos_logs():
    tipo = getattr(current_user, "tipo_usuario", None)
    if tipo not in ("admin", "visualizar", "operario"):
        abort(403)

    q = (request.args.get("q") or "").strip()
    data_inicio = (request.args.get("data_inicio") or "").strip()
    data_fim = (request.args.get("data_fim") or "").strip()
    page = request.args.get("page", 1, type=int)
    ultima_movimentacao_subq = _ultima_movimentacao_log_subquery()
    ultima_movimentacao_expr = db.func.coalesce(
        ultima_movimentacao_subq.c.ultima_movimentacao_em,
        LogVeiculo.data_registro,
    )

    query = (
        LogVeiculo.query
        .options(
            joinedload(LogVeiculo.veiculo),
            joinedload(LogVeiculo.piloto),
            selectinload(LogVeiculo.abastecimentos_detalhados),
        )
        .outerjoin(ultima_movimentacao_subq, ultima_movimentacao_subq.c.log_id == LogVeiculo.id)
        .join(Veiculos, LogVeiculo.veiculo_id == Veiculos.id)
        .join(Pilotos, LogVeiculo.piloto_id == Pilotos.id)
    )

    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                Veiculos.modelo.ilike(like),
                Veiculos.placa.ilike(like),
                Veiculos.responsavel.ilike(like),
                Pilotos.nome_piloto.ilike(like),
            )
        )

    if data_inicio:
        try:
            dt_ini = datetime.strptime(data_inicio, "%Y-%m-%d")
            query = query.filter(ultima_movimentacao_expr >= dt_ini)
        except ValueError:
            pass

    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(ultima_movimentacao_expr <= dt_fim)
        except ValueError:
            pass

    query = query.order_by(ultima_movimentacao_expr.desc(), LogVeiculo.id.desc())
    paginacao = query.paginate(page=page, per_page=20, error_out=False)
    logs = paginacao.items

    filters = {"q": q, "data_inicio": data_inicio, "data_fim": data_fim}

    return render_template(
        "veiculos_logs.html",
        logs=logs,
        paginacao=paginacao,
        total_logs=query.count(),
        total_abastecido=sum((l.total_valor_abastecido or 0) for l in logs),
        filters=filters,
    )

from flask import request, abort, send_file
from flask_login import login_required, current_user
from datetime import datetime
from io import BytesIO

from sqlalchemy.orm import joinedload

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


@bp.route("/veiculos/logs/exportar", methods=["GET"], endpoint="exportar_logs_veiculos_xlsx")
@login_required
def exportar_logs_veiculos_xlsx():
    tipo = getattr(current_user, "tipo_usuario", None)
    if tipo not in ("admin", "visualizar", "operario"):
        abort(403)

    q = (request.args.get("q") or "").strip()
    data_inicio = (request.args.get("data_inicio") or "").strip()
    data_fim = (request.args.get("data_fim") or "").strip()
    ultima_movimentacao_subq = _ultima_movimentacao_log_subquery()
    ultima_movimentacao_expr = db.func.coalesce(
        ultima_movimentacao_subq.c.ultima_movimentacao_em,
        LogVeiculo.data_registro,
    )

    query = (
        LogVeiculo.query
        .options(
            joinedload(LogVeiculo.veiculo),
            joinedload(LogVeiculo.piloto),
            selectinload(LogVeiculo.abastecimentos_detalhados),
        )
        .outerjoin(ultima_movimentacao_subq, ultima_movimentacao_subq.c.log_id == LogVeiculo.id)
        .join(Veiculos, LogVeiculo.veiculo_id == Veiculos.id)
        .join(Pilotos, LogVeiculo.piloto_id == Pilotos.id)
    )

    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                Veiculos.modelo.ilike(like),
                Veiculos.placa.ilike(like),
                Veiculos.responsavel.ilike(like),
                Pilotos.nome_piloto.ilike(like),
            )
        )

    if data_inicio:
        try:
            dt_ini = datetime.strptime(data_inicio, "%Y-%m-%d")
            query = query.filter(ultima_movimentacao_expr >= dt_ini)
        except ValueError:
            pass

    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(ultima_movimentacao_expr <= dt_fim)
        except ValueError:
            pass

    logs = query.order_by(ultima_movimentacao_expr.desc(), LogVeiculo.id.desc()).all()

    # -----------------------------
    # Helpers de estilo/estrutura
    # -----------------------------
    header_fill = PatternFill("solid", fgColor="1F4E79")  # azul escuro
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row=1, height=28):
        ws.row_dimensions[row].height = height
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

    def auto_width(ws, max_col, min_w=10, max_w=48):
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                s = str(cell.value)
                max_len = max(max_len, len(s))
            ws.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))

    def center_cols(ws, cols, start_row, end_row):
        for c in cols:
            for r in range(start_row, end_row + 1):
                ws.cell(row=r, column=c).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

    # -----------------------------
    # Workbook / Abas
    # -----------------------------
    wb = Workbook()

    # 1) Aba DETALHAMENTO
    ws = wb.active
    ws.title = "Detalhamento"

    # >>> Campos "direitinho" (sem KM no Abastecimento) <<<
    headers = [
        "Ultima Movimentacao",
        "Veículo",
        "Placa",
        "Responsável",
        "Piloto",
        "Check Diário",
        "KM Inicial",
        "KM Final",
        "KM Rodado",              # calculado
        "Abasteceu",
        "Qtd. Abastecimentos",
        "Tipos de Abastecimento",
        "Litros",
        "Valor Abastecimento (R$)",
        "Valor por Litro (R$)",   # calculado
        "Custo por KM (R$)",      # calculado
        "Assinatura",
        "Observação",
    ]
    ws.append(headers)
    style_header(ws, row=1)
    ws.freeze_panes = "A2"

    # Escreve dados
    for log in logs:
        ws.append([
            log.ultima_movimentacao_em.strftime("%d/%m/%Y %H:%M") if log.ultima_movimentacao_em else "",
            (log.veiculo.modelo if log.veiculo else "") or "",
            (log.veiculo.placa if log.veiculo else "") or "",
            (log.veiculo.responsavel if log.veiculo else "") or "",
            (log.piloto.nome_piloto if log.piloto else "") or "",
            "SIM" if log.check_diario else "NÃO",
            float(log.km_inicial or 0),
            "" if log.km_final is None else float(log.km_final),
            None,  # KM Rodado (formula)
            "SIM" if log.teve_abastecimento else "NÃO",
            int(log.qtd_abastecimentos or 0),
            log.tipos_abastecimento_resumo or "",
            float(log.total_litros_abastecidos or 0),
            float(log.total_valor_abastecido or 0),
            None,  # Valor por Litro (formula)
            None,  # Custo por KM (formula)
            "SIM" if log.assinatura_piloto else "NÃO",
            (log.observacao or ""),
        ])

    last_row = ws.max_row
    last_col = ws.max_column

    # Filtro no range real (sem Table, compatível com openpyxl antigo)
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

    # Colunas (1-based) conforme a lista de headers acima
    COL_DATA = 1
    COL_CHECK = 6
    COL_KM_INI = 7
    COL_KM_FIM = 8
    COL_KM_ROD = 9
    COL_ABAST = 10
    COL_QTD_AB = 11
    COL_LITROS = 13
    COL_VALOR = 14
    COL_VAL_LITRO = 15
    COL_CUSTO_KM = 16
    COL_ASS = 17

    # Formatação numérica / fórmulas por linha + bordas/alinhamento
    for r in range(2, last_row + 1):
        # KM Rodado
        ws.cell(r, COL_KM_ROD).value = (
            f'=IF({get_column_letter(COL_KM_FIM)}{r}="","",'
            f'{get_column_letter(COL_KM_FIM)}{r}-{get_column_letter(COL_KM_INI)}{r})'
        )

        # Valor por Litro
        ws.cell(r, COL_VAL_LITRO).value = (
            f'=IF(OR({get_column_letter(COL_LITROS)}{r}="",{get_column_letter(COL_LITROS)}{r}=0),"",'
            f'{get_column_letter(COL_VALOR)}{r}/{get_column_letter(COL_LITROS)}{r})'
        )

        # Custo por KM
        ws.cell(r, COL_CUSTO_KM).value = (
            f'=IF(OR({get_column_letter(COL_KM_ROD)}{r}="",{get_column_letter(COL_KM_ROD)}{r}=0),"",'
            f'{get_column_letter(COL_VALOR)}{r}/{get_column_letter(COL_KM_ROD)}{r})'
        )

        # formatos numéricos
        ws.cell(r, COL_KM_INI).number_format = "#,##0.00"
        ws.cell(r, COL_KM_FIM).number_format = "#,##0.00"
        ws.cell(r, COL_KM_ROD).number_format = "#,##0.00"
        ws.cell(r, COL_QTD_AB).number_format = "0"
        ws.cell(r, COL_LITROS).number_format = "#,##0.00"
        ws.cell(r, COL_VALOR).number_format = '"R$" #,##0.00'
        ws.cell(r, COL_VAL_LITRO).number_format = '"R$" #,##0.00'
        ws.cell(r, COL_CUSTO_KM).number_format = '"R$" #,##0.00'

        # bordas + wrap
        for c in range(1, last_col + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Centraliza colunas "SIM/NÃO" e data
    center_cols(ws, cols=[COL_DATA, COL_CHECK, COL_ABAST, COL_ASS], start_row=2, end_row=last_row)

    # Zebra + destaque de colunas de custo (bem fácil de ler)
    stripe_fill = PatternFill("solid", fgColor="F2F2F2")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    highlight_fill = PatternFill("solid", fgColor="FFF2CC")  # amarelo claro
    highlight_cols = {COL_LITROS, COL_VALOR, COL_VAL_LITRO, COL_CUSTO_KM}

    for r in range(2, last_row + 1):
        row_fill = stripe_fill if (r % 2 == 0) else white_fill
        for c in range(1, last_col + 1):
            cell = ws.cell(r, c)
            cell.fill = row_fill
            if c in highlight_cols and cell.value not in (None, ""):
                cell.fill = highlight_fill

    # Condicional: "SIM" verde e "NÃO" vermelho (Check Diário, Abasteceu, Assinatura)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(color="006100", bold=True)
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006", bold=True)

    for col in (COL_CHECK, COL_ABAST, COL_ASS):
        col_letter = get_column_letter(col)
        rng = f"{col_letter}2:{col_letter}{last_row}"

        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'{col_letter}2="SIM"'], fill=green_fill, font=green_font)
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'{col_letter}2="NÃO"'], fill=red_fill, font=red_font)
        )

    # Ajusta largura
    auto_width(ws, last_col, min_w=10, max_w=48)

    # 2) Aba RESUMO (médias e indicadores) - layout bonito e explicativo
    ws2 = wb.create_sheet("Resumo (Médias)")

    # Cores auxiliares
    title_fill = PatternFill("solid", fgColor="0B2F4F")
    card_fill  = PatternFill("solid", fgColor="E7EFF8")
    info_fill  = PatternFill("solid", fgColor="F8F8F8")
    kpi_fill   = PatternFill("solid", fgColor="D9E1F2")
    warn_fill  = PatternFill("solid", fgColor="FFF2CC")
    ok_fill    = PatternFill("solid", fgColor="C6EFCE")
    bad_fill   = PatternFill("solid", fgColor="FFC7CE")

    title_font = Font(color="FFFFFF", bold=True, size=14)
    section_font = Font(color="1F4E79", bold=True, size=12)
    label_font = Font(color="1F4E79", bold=True)
    small_font = Font(color="404040", size=10)
    big_font = Font(color="1F4E79", bold=True, size=16)

    # ranges fixos no Detalhamento
    has_data = last_row >= 2
    rng_km_rod = f"Detalhamento!{get_column_letter(COL_KM_ROD)}2:{get_column_letter(COL_KM_ROD)}{last_row}"
    rng_valor  = f"Detalhamento!{get_column_letter(COL_VALOR)}2:{get_column_letter(COL_VALOR)}{last_row}"
    rng_litros = f"Detalhamento!{get_column_letter(COL_LITROS)}2:{get_column_letter(COL_LITROS)}{last_row}"
    rng_vl     = f"Detalhamento!{get_column_letter(COL_VAL_LITRO)}2:{get_column_letter(COL_VAL_LITRO)}{last_row}"
    rng_ckm    = f"Detalhamento!{get_column_letter(COL_CUSTO_KM)}2:{get_column_letter(COL_CUSTO_KM)}{last_row}"
    rng_ab     = f"Detalhamento!{get_column_letter(COL_ABAST)}2:{get_column_letter(COL_ABAST)}{last_row}"
    rng_qtd_ab = f"Detalhamento!{get_column_letter(COL_QTD_AB)}2:{get_column_letter(COL_QTD_AB)}{last_row}"

    # Cabeçalho
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "RESUMO — MÉDIAS E INDICADORES (FROTA)"
    ws2["A1"].fill = title_fill
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:H2")
    ws2["A2"] = "Painel de custos, consumo e produtividade com base nos registros da aba “Detalhamento”."
    ws2["A2"].font = small_font
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 18

    # Cards KPI
    ws2.merge_cells("A4:C4")
    ws2.merge_cells("A5:C6")
    ws2["A4"] = "Total de Registros"
    ws2["A5"] = (f"={max(0, last_row-1)}" if has_data else "")
    ws2["A4"].font = label_font
    ws2["A5"].font = big_font
    ws2["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A4"].fill = kpi_fill
    ws2["A5"].fill = card_fill

    ws2.merge_cells("D4:F4")
    ws2.merge_cells("D5:F6")
    ws2["D4"] = "Total Abastecido (R$)"
    ws2["D5"] = (f"=SUM({rng_valor})" if has_data else "")
    ws2["D4"].font = label_font
    ws2["D5"].font = big_font
    ws2["D4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["D5"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["D4"].fill = kpi_fill
    ws2["D5"].fill = card_fill
    ws2["D5"].number_format = '"R$" #,##0.00'

    ws2.merge_cells("G4:H4")
    ws2.merge_cells("G5:H6")
    ws2["G4"] = "Total de Litros"
    ws2["G5"] = (f"=SUM({rng_litros})" if has_data else "")
    ws2["G4"].font = label_font
    ws2["G5"].font = big_font
    ws2["G4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["G5"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["G4"].fill = kpi_fill
    ws2["G5"].fill = card_fill
    ws2["G5"].number_format = "#,##0.00"

    for r in range(4, 7):
        for c in range(1, 9):
            ws2.cell(r, c).border = border
            ws2.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center")

    # Seção Médias
    ws2["A8"] = "MÉDIAS PRINCIPAIS"
    ws2["A8"].font = section_font
    ws2.merge_cells("A8:H8")

    ws2["A9"] = "Métrica"
    ws2["D9"] = "Resultado"
    ws2["F9"] = "Como interpretar"
    ws2.merge_cells("A9:C9")
    ws2.merge_cells("D9:E9")
    ws2.merge_cells("F9:H9")

    for cell in ws2["A9:H9"][0]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws2.row_dimensions[9].height = 22

    metrics = [
        ("Média de KM Rodado", f'=AVERAGEIF({rng_km_rod},">0")',
         "Média de km por registro (desconsidera valores zerados)."),
        ("Média Valor Abastecimento (R$)", f'=AVERAGEIF({rng_valor},">0")',
         "Valor médio por abastecimento (considera apenas valores > 0)."),
        ("Média Valor por Litro (R$)", f'=AVERAGEIF({rng_vl},">0")',
         "Preço médio pago por litro (filtra valores > 0)."),
        ("Média Custo por KM (R$)", f'=AVERAGEIF({rng_ckm},">0")',
         "Custo médio por km: (valor abastecido / km rodado)."),
        ("Qtd. de Abastecimentos", f"=SUM({rng_qtd_ab})",
         "Quantidade total de abastecimentos registrados nos turnos filtrados."),
    ]

    base_r = 10
    for i, (name, formula, tip) in enumerate(metrics):
        r = base_r + i
        ws2.merge_cells(f"A{r}:C{r}")
        ws2.merge_cells(f"D{r}:E{r}")
        ws2.merge_cells(f"F{r}:H{r}")

        ws2[f"A{r}"] = name
        ws2[f"D{r}"] = (formula if has_data else "")
        ws2[f"F{r}"] = tip

        ws2[f"A{r}"].font = Font(bold=True, color="1F4E79")
        ws2[f"F{r}"].font = small_font

        for c in range(1, 9):
            cell = ws2.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = info_fill if (i % 2 == 0) else PatternFill("solid", fgColor="FFFFFF")

    if has_data:
        ws2[f"D{base_r}"].number_format = "#,##0.00"
        ws2[f"D{base_r+1}"].number_format = '"R$" #,##0.00'
        ws2[f"D{base_r+2}"].number_format = '"R$" #,##0.00'
        ws2[f"D{base_r+3}"].number_format = '"R$" #,##0.00'
        ws2[f"D{base_r+4}"].number_format = "0"

    # Seção Alertas
    alert_r = base_r + len(metrics) + 2
    ws2[f"A{alert_r}"] = "ALERTAS (LEITURA RÁPIDA)"
    ws2[f"A{alert_r}"].font = section_font
    ws2.merge_cells(f"A{alert_r}:H{alert_r}")

    limiar_ckm = 1.50  # ajuste livre
    ws2.merge_cells(f"A{alert_r+1}:E{alert_r+1}")
    ws2.merge_cells(f"F{alert_r+1}:H{alert_r+1}")
    ws2[f"A{alert_r+1}"] = f"Custo por KM acima de R$ {limiar_ckm:.2f}"
    ws2[f"F{alert_r+1}"] = (
        f'=IF(AVERAGEIF({rng_ckm},">0")>{limiar_ckm},"ATENÇÃO: ALTO","OK")'
        if has_data else ""
    )
    ws2[f"A{alert_r+1}"].font = Font(bold=True, color="1F4E79")
    ws2[f"F{alert_r+1}"].font = Font(bold=True)

    for c in range(1, 9):
        cell = ws2.cell(alert_r + 1, c)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.fill = warn_fill

    rng_alert = f"F{alert_r+1}:H{alert_r+1}"
    ws2.conditional_formatting.add(
        rng_alert,
        FormulaRule(formula=[f'F{alert_r+1}="OK"'], fill=ok_fill, font=Font(color="006100", bold=True))
    )
    ws2.conditional_formatting.add(
        rng_alert,
        FormulaRule(formula=[f'F{alert_r+1}<>"OK"'], fill=bad_fill, font=Font(color="9C0006", bold=True))
    )

    # Ajustes finais Resumo
    ws2.freeze_panes = "A10"
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 22
    ws2.column_dimensions["G"].width = 14
    ws2.column_dimensions["H"].width = 14

    # -----------------------------
    # Salva e retorna
    # -----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"logs_veiculos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@bp.route("/veiculos/cadastrar", methods=["GET", "POST"], endpoint="cadastrar_veiculo")
@login_required
def cadastrar_veiculo():
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    errors = {}
    form = {}

    responsaveis = get_responsaveis_choices()

    if request.method == "POST":
        modelo = (request.form.get("modelo") or "").strip()
        ano_raw = (request.form.get("ano_fabricacao") or "").strip()
        frota = (request.form.get("frota") or "").strip().upper()
        operacao = (request.form.get("operacao") or "").strip().upper()
        placa = (request.form.get("placa") or "").strip().upper()

        # vem do select
        responsavel = (request.form.get("responsavel") or "").strip()

        km_atual_raw = (request.form.get("km_atual") or "").strip()
        km_prox_raw = (request.form.get("km_prox_revisao") or "").strip()
        status = (request.form.get("status") or "Ativo").strip()
        revisao_marcada_raw = (request.form.get("revisao_marcada_em") or "").strip()
        revisao_obs = (request.form.get("revisao_obs") or "").strip()

        form = {
            "modelo": modelo,
            "ano_fabricacao": ano_raw,
            "frota": frota,
            "operacao": operacao,
            "placa": placa,
            "responsavel": responsavel,
            "km_atual": km_atual_raw,
            "km_prox_revisao": km_prox_raw,
            "status": status,
            "revisao_marcada_em": revisao_marcada_raw,
            "revisao_obs": revisao_obs,
        }

        # validações básicas
        if not modelo:
            errors["modelo"] = "Informe o modelo."
        if not ano_raw:
            errors["ano_fabricacao"] = "Informe o ano."
        if frota not in ("PROPRIA", "ALUGADA"):
            errors["frota"] = "Selecione PROPRIA ou ALUGADA."
        if not operacao:
            errors["operacao"] = "Informe a operação (ex: PMSP / AGRO)."
        if not placa:
            errors["placa"] = "Informe a placa."

        # valida responsável: se preencher, tem que estar na lista
        if responsavel:
            valid_values = {r["value"] for r in responsaveis}
            if responsavel not in valid_values:
                errors["responsavel"] = "Selecione um responsável válido."

        # ano
        ano_fabricacao = None
        if ano_raw:
            try:
                ano_fabricacao = int(ano_raw)
                if ano_fabricacao < 1900 or ano_fabricacao > 2100:
                    errors["ano_fabricacao"] = "Ano inválido."
            except ValueError:
                errors["ano_fabricacao"] = "Ano inválido."

        # km_atual
        km_atual = 0
        if km_atual_raw:
            try:
                km_atual = float(km_atual_raw.replace(",", "."))
                if km_atual < 0:
                    errors["km_atual"] = "KM atual não pode ser negativo."
            except ValueError:
                errors["km_atual"] = "KM atual inválido."

        # km_prox_revisao
        km_prox_revisao = None
        if km_prox_raw:
            try:
                km_prox_revisao = float(km_prox_raw.replace(",", "."))
                if km_prox_revisao < 0:
                    errors["km_prox_revisao"] = "Próx revisão inválida."
            except ValueError:
                errors["km_prox_revisao"] = "Próx revisão inválida."

        # revisao marcada
        revisao_marcada_em = None
        if revisao_marcada_raw:
            try:
                revisao_marcada_em = datetime.strptime(revisao_marcada_raw, "%Y-%m-%dT%H:%M")
            except ValueError:
                errors["revisao_marcada_em"] = "Data/hora inválida."

        # placa única
        if placa and not errors.get("placa"):
            existe = Veiculos.query.filter_by(placa=placa).first()
            if existe:
                errors["placa"] = "Já existe um veículo com essa placa."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template(
                "cadastrar_veiculo.html",
                form=form,
                errors=errors,
                responsaveis=responsaveis
            )

        novo = Veiculos(
            tipo_equipamento="veiculos",
            status=status,
            modelo=modelo,
            ano_fabricacao=ano_fabricacao,
            renomacao=placa,
            categoria=None,
            numero_serie=None,
            ultima_manutencao=None,
            frota=frota,
            operacao=operacao,
            placa=placa,
            responsavel=responsavel or None,
            km_atual=km_atual,
            km_prox_revisao=km_prox_revisao,
            revisao_marcada_em=revisao_marcada_em,
            revisao_obs=revisao_obs or None,
        )

        try:
            db.session.add(novo)
            db.session.commit()
            flash("Veículo cadastrado com sucesso!", "success")
            return redirect(url_for("main.listar_veiculos"))
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar veículo: {str(e)}", "danger")
            return render_template(
                "cadastrar_veiculo.html",
                form=form,
                errors=errors,
                responsaveis=responsaveis
            )

    # GET
    return render_template(
        "cadastrar_veiculo.html",
        form=form,
        errors=errors,
        responsaveis=responsaveis
    )
@bp.route("/veiculos/<int:veiculo_id>/editar", methods=["GET", "POST"], endpoint="editar_veiculo")
@login_required
def editar_veiculo(veiculo_id):
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    v = Veiculos.query.get_or_404(veiculo_id)
    errors = {}

    responsaveis = get_responsaveis_choices()
    valid_values = {r["value"] for r in responsaveis}

    if request.method == "POST":
        modelo = (request.form.get("modelo") or "").strip()
        ano_raw = (request.form.get("ano_fabricacao") or "").strip()
        frota = (request.form.get("frota") or "").strip().upper()
        operacao = (request.form.get("operacao") or "").strip().upper()
        placa = (request.form.get("placa") or "").strip().upper()

        responsavel = (request.form.get("responsavel") or "").strip()

        km_atual_raw = (request.form.get("km_atual") or "").strip()
        km_prox_raw = (request.form.get("km_prox_revisao") or "").strip()
        status = (request.form.get("status") or "Ativo").strip()
        revisao_marcada_raw = (request.form.get("revisao_marcada_em") or "").strip()
        revisao_obs = (request.form.get("revisao_obs") or "").strip()

        if not modelo:
            errors["modelo"] = "Informe o modelo."
        if not ano_raw:
            errors["ano_fabricacao"] = "Informe o ano."
        if frota not in ("PROPRIA", "ALUGADA"):
            errors["frota"] = "Selecione PROPRIA ou ALUGADA."
        if not operacao:
            errors["operacao"] = "Informe a operação."
        if not placa:
            errors["placa"] = "Informe a placa."

        if responsavel and responsavel not in valid_values:
            errors["responsavel"] = "Selecione um responsável válido."

        # ano
        ano_fabricacao = None
        if ano_raw:
            try:
                ano_fabricacao = int(ano_raw)
                if ano_fabricacao < 1900 or ano_fabricacao > 2100:
                    errors["ano_fabricacao"] = "Ano inválido."
            except ValueError:
                errors["ano_fabricacao"] = "Ano inválido."

        # km_atual
        km_atual = v.km_atual or 0
        if km_atual_raw:
            try:
                km_atual = float(km_atual_raw.replace(",", "."))
                if km_atual < 0:
                    errors["km_atual"] = "KM atual inválido."
            except ValueError:
                errors["km_atual"] = "KM atual inválido."

        # km_prox_revisao
        km_prox_revisao = None
        if km_prox_raw:
            try:
                km_prox_revisao = float(km_prox_raw.replace(",", "."))
            except ValueError:
                errors["km_prox_revisao"] = "Próx revisão inválida."

        # revisao marcada
        revisao_marcada_em = None
        if revisao_marcada_raw:
            try:
                revisao_marcada_em = datetime.strptime(revisao_marcada_raw, "%Y-%m-%dT%H:%M")
            except ValueError:
                errors["revisao_marcada_em"] = "Data/hora inválida."

        # placa única (exceto ele mesmo)
        if placa and not errors.get("placa"):
            existe = Veiculos.query.filter(Veiculos.placa == placa, Veiculos.id != v.id).first()
            if existe:
                errors["placa"] = "Já existe um veículo com essa placa."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            form = {
                "modelo": modelo,
                "ano_fabricacao": ano_raw,
                "frota": frota,
                "operacao": operacao,
                "placa": placa,
                "responsavel": responsavel,
                "km_atual": km_atual_raw,
                "km_prox_revisao": km_prox_raw,
                "status": status,
                "revisao_marcada_em": revisao_marcada_raw,
                "revisao_obs": revisao_obs,
            }
            return render_template(
                "cadastrar_veiculo.html",
                form=form,
                errors=errors,
                veiculo=v,
                responsaveis=responsaveis
            )

        v.modelo = modelo
        v.ano_fabricacao = ano_fabricacao
        v.frota = frota
        v.operacao = operacao
        v.placa = placa
        v.responsavel = responsavel or None
        v.km_atual = km_atual
        v.km_prox_revisao = km_prox_revisao
        v.status = status
        v.revisao_marcada_em = revisao_marcada_em
        v.revisao_obs = revisao_obs or None
        v.renomacao = placa

        try:
            db.session.commit()
            flash("Veículo atualizado!", "success")
            return redirect(url_for("main.listar_veiculos"))
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar: {str(e)}", "danger")

    # GET
    form = {
        "modelo": v.modelo or "",
        "ano_fabricacao": str(v.ano_fabricacao or ""),
        "frota": v.frota or "",
        "operacao": v.operacao or "",
        "placa": v.placa or "",
        "responsavel": v.responsavel or "",
        "km_atual": str(v.km_atual or ""),
        "km_prox_revisao": str(v.km_prox_revisao or "") if v.km_prox_revisao is not None else "",
        "status": v.status or "Ativo",
        "revisao_marcada_em": v.revisao_marcada_em.strftime("%Y-%m-%dT%H:%M") if v.revisao_marcada_em else "",
        "revisao_obs": v.revisao_obs or "",
    }
    return render_template(
        "cadastrar_veiculo.html",
        form=form,
        errors=errors,
        veiculo=v,
        responsaveis=responsaveis
    )

# -----------------------------
# DELETAR VEÍCULO
# -----------------------------
@bp.route("/veiculos/<int:veiculo_id>/deletar", methods=["POST"], endpoint="deletar_veiculo")
@login_required
def deletar_veiculo(veiculo_id):
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    v = Veiculos.query.get_or_404(veiculo_id)
    try:
        db.session.delete(v)
        db.session.commit()
        flash("Veículo removido!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao remover: {str(e)}", "danger")

    return redirect(url_for("main.listar_veiculos"))

def _piloto_nome_logado():
    return (getattr(current_user, "nome_uvis", None) or "").strip()


def _parse_decimal_form(raw_value):
    raw_value = (raw_value or "").strip()
    if not raw_value:
        return None

    if "," in raw_value and "." in raw_value:
        raw_value = raw_value.replace(".", "").replace(",", ".")
    else:
        raw_value = raw_value.replace(",", ".")

    return float(raw_value)


def _veiculo_do_piloto_logado(veiculo_id, nome_piloto):
    veiculo = Veiculos.query.get_or_404(veiculo_id)
    if (veiculo.responsavel or "").strip().lower() != nome_piloto.lower():
        abort(403)
    return veiculo


def _buscar_turno_aberto_piloto(veiculo_id, incluir_abastecimentos=False):
    query = LogVeiculo.query.filter(
        LogVeiculo.veiculo_id == veiculo_id,
        LogVeiculo.piloto_id == current_user.piloto_id,
        LogVeiculo.km_final.is_(None),
    )
    if incluir_abastecimentos:
        query = query.options(selectinload(LogVeiculo.abastecimentos_detalhados))
    return query.order_by(LogVeiculo.data_registro.desc()).first()


def _salvar_upload_veiculo(arquivo, subpasta, prefixo, placa):
    if not arquivo or not arquivo.filename:
        return None

    pasta_base = os.path.join(current_app.root_path, "static", "uploads", "veiculos")
    pasta_destino = os.path.join(pasta_base, subpasta)
    os.makedirs(pasta_destino, exist_ok=True)

    ext = os.path.splitext(secure_filename(arquivo.filename))[1] or ".jpg"
    stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    nome = secure_filename(f"{prefixo}_{placa}_{stamp}{ext}")
    arquivo.save(os.path.join(pasta_destino, nome))

    return f"uploads/veiculos/{subpasta}/{nome}"


@bp.route("/piloto/veiculos", methods=["GET"], endpoint="piloto_veiculos")
@login_required
@roles_required("piloto")
def piloto_veiculos():
    nome_piloto = _piloto_nome_logado()

    if not nome_piloto or not getattr(current_user, "piloto_id", None):
        flash("Seu usuário piloto está sem vínculo completo. Contate o administrador.", "warning")
        return render_template("piloto_veiculos.html", veiculos=[], turnos_abertos={})

    veiculos = (
        Veiculos.query
        .filter(db.func.lower(Veiculos.responsavel) == nome_piloto.lower())
        .order_by(Veiculos.operacao.asc(), Veiculos.modelo.asc())
        .all()
    )

    turnos_abertos = {}
    veiculo_ids = [v.id for v in veiculos]

    if veiculo_ids:
        logs_abertos = (
            LogVeiculo.query
            .options(selectinload(LogVeiculo.abastecimentos_detalhados))
            .filter(
                LogVeiculo.piloto_id == current_user.piloto_id,
                LogVeiculo.veiculo_id.in_(veiculo_ids),
                LogVeiculo.km_final.is_(None),
            )
            .order_by(LogVeiculo.veiculo_id.asc(), LogVeiculo.data_registro.desc())
            .all()
        )
        for log in logs_abertos:
            if log.veiculo_id not in turnos_abertos:
                turnos_abertos[log.veiculo_id] = log

    return render_template("piloto_veiculos.html", veiculos=veiculos, turnos_abertos=turnos_abertos)


@bp.route("/piloto/veiculos/<int:veiculo_id>/km", methods=["POST"], endpoint="piloto_atualizar_km_veiculo")
@login_required
@roles_required("piloto")
def piloto_atualizar_km_veiculo(veiculo_id):
    nome_piloto = _piloto_nome_logado()
    if not nome_piloto or not getattr(current_user, "piloto_id", None):
        abort(403)

    v = _veiculo_do_piloto_logado(veiculo_id, nome_piloto)

    try:
        km_inicial = _parse_decimal_form(request.form.get("km_inicial"))
    except ValueError:
        flash("Kilometragem inicial invalida.", "warning")
        return redirect(url_for("main.piloto_veiculos"))

    assinatura_b64 = request.form.get("assinatura_b64")
    foto_painel = request.files.get("foto_painel")

    if km_inicial is None or not assinatura_b64 or not foto_painel or not foto_painel.filename:
        flash("Kilometragem inicial, foto do painel e assinatura sao obrigatorias.", "warning")
        return redirect(url_for("main.piloto_veiculos"))

    try:
        km_atual_veiculo = v.km_atual or 0

        if km_inicial < km_atual_veiculo:
            flash(f"KM inicial ({km_inicial:.0f}) menor que o KM atual do veiculo.", "danger")
            return redirect(url_for("main.piloto_veiculos"))

        turno_aberto = _buscar_turno_aberto_piloto(v.id)
        if turno_aberto:
            flash("Ja existe um turno aberto para este veiculo. Finalize-o antes de iniciar outro.", "warning")
            return redirect(url_for("main.piloto_veiculos"))

        novo_log = LogVeiculo(
            veiculo_id=v.id,
            piloto_id=current_user.piloto_id,
            km_inicial=km_inicial,
            km_final=None,
            check_diario=True,
            assinatura_piloto=assinatura_b64,
            data_registro=datetime.now(),
        )
        novo_log.foto_painel_path = _salvar_upload_veiculo(foto_painel, "paineis", "painel", v.placa)

        db.session.add(novo_log)
        v.km_atual = km_inicial

        db.session.commit()
        flash(f"Turno de {v.modelo} iniciado com sucesso!", "success")

    except IntegrityError as e:
        db.session.rollback()
        current_app.logger.exception("Erro de integridade ao iniciar turno de veiculo.")
        if "km_final" in str(e.orig).lower():
            flash("Erro de banco: o campo km_final ainda nao aceita vazio. Rode a migracao para permitir NULL.", "danger")
        else:
            flash("Erro de integridade ao salvar. Verifique os dados e a estrutura do banco.", "danger")
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Erro tecnico ao iniciar turno de veiculo.")
        flash("Erro tecnico ao salvar.", "danger")

    return redirect(url_for("main.piloto_veiculos"))


@bp.route("/piloto/veiculos/<int:veiculo_id>/abastecimento", methods=["POST"], endpoint="piloto_registrar_abastecimento_turno")
@login_required
@roles_required("piloto")
def piloto_registrar_abastecimento_turno(veiculo_id):
    nome_piloto = _piloto_nome_logado()
    if not nome_piloto or not getattr(current_user, "piloto_id", None):
        abort(403)

    v = _veiculo_do_piloto_logado(veiculo_id, nome_piloto)
    log = _buscar_turno_aberto_piloto(v.id, incluir_abastecimentos=True)

    if not log:
        flash("Nenhum turno aberto encontrado para registrar abastecimento.", "warning")
        return redirect(url_for("main.piloto_veiculos"))

    try:
        km_registro = _parse_decimal_form(request.form.get("km_abastecimento"))
        litros = _parse_decimal_form(request.form.get("litros"))
        valor_total = _parse_decimal_form(request.form.get("valor_abastecimento"))
    except ValueError:
        flash("Os dados do abastecimento estao invalidos.", "warning")
        return redirect(url_for("main.piloto_veiculos"))

    tipo_abastecimento = (request.form.get("tipo_abastecimento") or "").strip()
    foto_nf = request.files.get("foto_nf")

    if (
        km_registro is None
        or litros is None
        or valor_total is None
        or not tipo_abastecimento
        or not foto_nf
        or not foto_nf.filename
    ):
        flash("KM, tipo, litros, valor total e foto da nota sao obrigatorios no abastecimento.", "warning")
        return redirect(url_for("main.piloto_veiculos"))

    if len(tipo_abastecimento) > 100:
        flash("O tipo de abastecimento deve ter no maximo 100 caracteres.", "warning")
        return redirect(url_for("main.piloto_veiculos"))

    ultimo_km_turno = log.ultimo_km_registrado or 0
    if km_registro < ultimo_km_turno:
        flash(f"O KM do abastecimento nao pode ser menor que o ultimo KM do turno ({ultimo_km_turno:.0f}).", "danger")
        return redirect(url_for("main.piloto_veiculos"))

    try:
        novo_abastecimento = Abastecimento(
            log_veiculo_id=log.id,
            data_hora=datetime.now(),
            km_registro=km_registro,
            tipo_abastecimento=tipo_abastecimento,
            litros=litros,
            valor_total=valor_total,
            foto_nf_path=_salvar_upload_veiculo(foto_nf, "notas", "nf", v.placa),
        )

        db.session.add(novo_abastecimento)
        v.km_atual = max(v.km_atual or 0, km_registro)
        db.session.commit()
        flash("Abastecimento registrado com sucesso!", "success")

    except Exception:
        db.session.rollback()
        current_app.logger.exception("Erro tecnico ao registrar abastecimento do turno.")
        flash("Erro tecnico ao registrar abastecimento.", "danger")

    return redirect(url_for("main.piloto_veiculos"))


@bp.route("/piloto/veiculos/<int:veiculo_id>/encerrar", methods=["POST"])
@login_required
@roles_required("piloto")
def piloto_encerrar_turno(veiculo_id):
    nome_piloto = _piloto_nome_logado()
    if not nome_piloto or not getattr(current_user, "piloto_id", None):
        abort(403)

    v = _veiculo_do_piloto_logado(veiculo_id, nome_piloto)

    try:
        km_final = _parse_decimal_form(request.form.get("km_final"))
    except ValueError:
        flash("Kilometragem final invalida.", "warning")
        return redirect(url_for("main.piloto_veiculos"))

    qtd_fazendas_enderecos = request.form.get("qtd_fazendas_enderecos", type=int)
    observacao = (request.form.get("observacao") or "").strip() or None

    if km_final is None:
        flash("Informe a kilometragem final para encerrar o turno.", "warning")
        return redirect(url_for("main.piloto_veiculos"))

    log = _buscar_turno_aberto_piloto(veiculo_id, incluir_abastecimentos=True)

    if not log:
        flash("Nenhum turno aberto encontrado.", "warning")
        return redirect(url_for("main.piloto_veiculos"))

    ultimo_km_turno = log.ultimo_km_registrado or 0
    if km_final < ultimo_km_turno:
        flash(f"KM final nao pode ser menor que o ultimo KM registrado no turno ({ultimo_km_turno:.0f}).", "danger")
        return redirect(url_for("main.piloto_veiculos"))

    log.qtd_fazendas_enderecos = qtd_fazendas_enderecos
    log.km_final = km_final
    log.observacao = observacao
    v.km_atual = km_final

    db.session.commit()
    flash("Turno encerrado com sucesso!", "success")
    return redirect(url_for("main.piloto_veiculos"))

@bp.route('/piloto/os')
@login_required
@roles_required('piloto')
def piloto_os():
    if not current_user.piloto_id:
        flash("Piloto sem vínculo cadastrado.", "danger")
        return redirect(url_for('main.dashboard'))

    google_maps_key = os.getenv("KEY_API_GOOGLE_MAPS") or current_app.config.get("GOOGLE_MAPS_API_KEY", "")
    status_ok = ["APROVADO", "APROVADO COM RECOMENDAÇÕES", "APROVADA", "APROVADA COM RECOMENDAÇÕES"]

    vinculo = (
        EquipePiloto.query
        .join(Equipe, Equipe.id == EquipePiloto.equipe_id)
        .options(joinedload(EquipePiloto.equipe))
        .filter(
            EquipePiloto.piloto_id == current_user.piloto_id,
            Equipe.ativa.is_(True)
        )
        .order_by(
            db.case((EquipePiloto.papel == "piloto", 0), else_=1),
            EquipePiloto.criado_em.desc()
        )
        .first()
    )

    if not vinculo or not vinculo.equipe_id:
        flash("Você ainda não está vinculado a nenhuma equipe ativa.", "warning")
        return render_template(
            "piloto_os.html",
            pedidos=[],
            paginacao=None,
            status_ok=status_ok,
            pilot_team_nome=None,
            pilot_team_regiao=None,
            pilot_team_papel=None,
            google_maps_key=google_maps_key,
            drones_equipe=[],
            baterias_equipe=[],
            veiculos_equipe=[],
        )

    pilot_team_nome = vinculo.equipe.nome_equipe if vinculo.equipe else None
    pilot_team_regiao = vinculo.equipe.regiao if vinculo.equipe else None
    pilot_team_papel = (vinculo.papel or "").lower()

    drones_equipe = (
        Drones.query
        .options(joinedload(Drones.equipe))
        .filter(Drones.equipe_id == vinculo.equipe_id)
        .order_by(Drones.renomacao.asc())
        .all()
    )

    baterias_equipe = (
        Baterias.query
        .join(Drones, Baterias.drone_id == Drones.id)
        .filter(Drones.equipe_id == vinculo.equipe_id)
        .order_by(Baterias.renomacao.asc())
        .all()
    )

    #  VEÍCULOS vinculados à equipe ativa do piloto
    veiculos_equipe = (
        Veiculos.query
        .filter(Veiculos.equipe_id == vinculo.equipe_id)
        .order_by(Veiculos.operacao.asc(), Veiculos.modelo.asc())
        .all()
    )

    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe)
        )
        .filter(
            Solicitacao.equipe_id == vinculo.equipe_id,
            Solicitacao.status.in_(status_ok)
        )
    )

    filtro_data = request.args.get("data")
    uvis_id = request.args.get("uvis_id")
    query = aplicar_filtros_base(query, filtro_data, uvis_id)

    page = request.args.get("page", 1, type=int)
    paginacao = query.order_by(
        Solicitacao.data_agendamento.asc(),
        Solicitacao.hora_agendamento.asc()
    ).paginate(page=page, per_page=6, error_out=False)

    return render_template(
        "piloto_os.html",
        pedidos=paginacao.items,
        paginacao=paginacao,
        status_ok=status_ok,
        pilot_team_nome=pilot_team_nome,
        pilot_team_regiao=pilot_team_regiao,
        pilot_team_papel=pilot_team_papel,
        google_maps_key=google_maps_key,
        drones_equipe=drones_equipe,
        baterias_equipe=baterias_equipe,
        veiculos_equipe=veiculos_equipe,  
    )


@bp.route("/piloto/os/historico")
@login_required
@roles_required("piloto")
def piloto_os_historico():
    if not current_user.piloto_id:
        flash("Piloto sem vínculo cadastrado.", "danger")
        return redirect(url_for("main.dashboard"))

    status_concluido = ["CONCLUÍDO", "CONCLUIDO"]

    equipes_vinculadas = (
        db.session.query(EquipePiloto.equipe_id)
        .filter(
            EquipePiloto.piloto_id == current_user.piloto_id,
            EquipePiloto.equipe_id.isnot(None)
        )
        .distinct()
    )

    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe)
        )
        .filter(
            Solicitacao.equipe_id.in_(equipes_vinculadas),
            Solicitacao.status.in_(status_concluido)
        )
    )

    page = request.args.get("page", 1, type=int)
    paginacao = (
        query
        .order_by(Solicitacao.data_criacao.desc(), Solicitacao.id.desc())
        .paginate(page=page, per_page=6, error_out=False)
    )

    return render_template(
        "piloto_os_historico.html",
        pedidos=paginacao.items,
        paginacao=paginacao
    )

from datetime import datetime
from flask import render_template, request, redirect, url_for, flash, current_app
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload

from app import db
# ajuste imports conforme sua estrutura
from app.models import Solicitacao, OrdemServico, EquipePiloto, Equipe


# ============================================================
# CONCLUIR OS (mantive seu código)
# ============================================================
@bp.route('/piloto/os/<int:os_id>/concluir', methods=['POST'])
@login_required
@roles_required('piloto')
def piloto_concluir_os(os_id):

    s = Solicitacao.query.get_or_404(os_id)

    status_ok = ["APROVADO", "APROVADO COM RECOMENDAÇÕES", "APROVADA", "APROVADA COM RECOMENDAÇÕES"]
    if s.status not in status_ok:
        flash("A OS não está aprovada.", "warning")
        return redirect(url_for('main.piloto_os'))

    if not s.equipe_id:
        flash("Esta OS não possui equipe atribuída.", "danger")
        return redirect(url_for('main.piloto_os'))

    #  valida se o piloto logado faz parte da equipe da OS
    vinculo = (
        EquipePiloto.query
        .join(Equipe, Equipe.id == EquipePiloto.equipe_id)
        .options(joinedload(EquipePiloto.equipe))
        .filter(
            EquipePiloto.equipe_id == s.equipe_id,
            EquipePiloto.piloto_id == current_user.piloto_id,
            Equipe.ativa.is_(True)
        )
        .first()
    )

    if not vinculo:
        flash("Você não faz parte da equipe atribuída a esta OS.", "danger")
        return redirect(url_for('main.piloto_os'))

    equipe_nome = vinculo.equipe.nome_equipe if vinculo.equipe else None
    papel = (vinculo.papel or "").lower() if vinculo.papel else None

    s.status = "CONCLUÍDO"
    db.session.commit()

    if equipe_nome and papel:
        flash(f"OS #{s.id} concluída! Equipe: {equipe_nome} | Papel: {papel}.", "success")
    elif equipe_nome:
        flash(f"OS #{s.id} concluída! Equipe: {equipe_nome}.", "success")
    else:
        flash(f"OS #{s.id} concluída com sucesso!", "success")

    return redirect(url_for('main.piloto_os'))


from datetime import datetime, timedelta
from flask import render_template, request, redirect, url_for, flash, current_app, jsonify
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload

from app import db
from app.models import (
    Solicitacao, OrdemServico,
    Equipe, EquipePiloto,
    Drones, Baterias, Veiculos,
    ChecklistSemanalVeiculo, ChecklistSemanalDrone
)

# se seu decorator roles_required estiver em outro módulo:
# from app.decorators import roles_required

# ============================================================
# HELPERS DE PARSE (PT-BR friendly)
# ============================================================
def _clean(v):
    """
    Mantido para compatibilidade:
    - retorna None quando vazio
    """
    if v is None:
        return None
    v = str(v).strip()
    return v if v != "" else None

def _clean_str(v):
    """
     Para campos de TEXTO (String/Text):
    - retorna "" quando vazio (nunca None)
    """
    if v is None:
        return ""
    return str(v).strip()  # se ficar "", ok

def _to_int(v):
    """
    Para INT:
    - vazio -> None
    - inválido -> None
    """
    v = _clean(v)
    if v is None:
        return None
    try:
        return int(float(v))
    except Exception:
        return None

def _to_float(v):
    """
    Para FLOAT:
    - vazio -> None
    - inválido -> None
    """
    v = _clean(v)
    if v is None:
        return None

    # aceita "1.234,56" ou "1234,56" ou "1234.56"
    if "," in v and "." in v:
        v = v.replace(".", "").replace(",", ".")
    else:
        v = v.replace(",", ".")

    try:
        return float(v)
    except Exception:
        return None

def _to_date(v):
    """
    Para DATE:
    - vazio -> None
    """
    v = _clean(v)
    if v is None:
        return None
    try:
        return datetime.strptime(v, "%Y-%m-%d").date()
    except Exception:
        return None

def _to_time(v):
    """
    Para TIME:
    - vazio -> None
    """
    v = _clean(v)
    if v is None:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(v, fmt).time()
        except Exception:
            pass
    return None

def _to_datetime_local(v):
    """
    input type="datetime-local" vem como "YYYY-MM-DDTHH:MM"
    - vazio -> None
    """
    v = _clean(v)
    if v is None:
        return None
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(v, fmt)
        except Exception:
            pass
    return None



# ============================================================
# ROTA ANTIGA (AGORA SÓ REDIRECIONA)
# /piloto/os/formulario?os_id=123
# ============================================================
@bp.route("/piloto/os/formulario", methods=["GET"])
@login_required
@roles_required("piloto")
def piloto_os_formulario_redirect():
    os_id = request.args.get("os_id", type=int) or request.args.get("solicitacao_id", type=int)

    if not os_id:
        flash("Selecione uma OS para preencher o formulário.", "info")
        return redirect(url_for("main.piloto_os"))

    return redirect(url_for("main.piloto_os_formulario_view", os_id=os_id))



from datetime import datetime, date, timedelta

def criar_solicitacao_retorno_monitoramento(solicitacao_original, ordem_atual):
    """
    Cria uma nova solicitação 7 dias após a data da aplicação
    para retorno de monitoramento de larvas.
    """

    data_base = (
        ordem_atual.data_aplicacao
        or solicitacao_original.data_agendamento
        or date.today()
    )
    nova_data = data_base + timedelta(days=7)

    observacao_original = (solicitacao_original.observacao or "").strip()
    complemento = f"Retorno automático para monitoramento de larvas gerado a partir da solicitação #{solicitacao_original.id}."
    nova_observacao = f"{observacao_original}\n{complemento}".strip() if observacao_original else complemento

    nova_solicitacao = Solicitacao(
        data_agendamento=nova_data,
        hora_agendamento=solicitacao_original.hora_agendamento,
        foco=solicitacao_original.foco,
        tipo_visita=solicitacao_original.tipo_visita,
        altura_voo=solicitacao_original.altura_voo,
        criadouro=solicitacao_original.criadouro,
        apoio_cet=solicitacao_original.apoio_cet,
        observacao=nova_observacao,
        area_restrita=solicitacao_original.area_restrita,
        cep=solicitacao_original.cep,
        logradouro=solicitacao_original.logradouro,
        bairro=solicitacao_original.bairro,
        cidade=solicitacao_original.cidade,
        uf=solicitacao_original.uf,
        numero=solicitacao_original.numero,
        complemento=solicitacao_original.complemento,
        latitude=solicitacao_original.latitude,
        longitude=solicitacao_original.longitude,
        perimetro_planejado=solicitacao_original.perimetro_planejado,
        perimetro_executado=None,
        anexo_path=solicitacao_original.anexo_path,
        anexo_nome=solicitacao_original.anexo_nome,
        protocolo=None,
        justificativa=None,
        equipe_uvis_nome=solicitacao_original.equipe_uvis_nome,
        status="PENDENTE",  # troque para APROVADO se quiser liberar direto
        usuario_id=solicitacao_original.usuario_id,
        piloto_id=solicitacao_original.piloto_id,
        equipe_id=solicitacao_original.equipe_id,
        origem_retorno_id=solicitacao_original.id,
        gerada_automaticamente=True,
    )

    db.session.add(nova_solicitacao)
    db.session.flush()  # garante ID da nova solicitacao

    # opcional: já cria uma OrdemServico "espelho" com os mesmos dados-base
    nova_ordem = OrdemServico(
        solicitacao_id=nova_solicitacao.id,
        equipe_id=solicitacao_original.equipe_id,

        identificador_os="",
        respondido_por="",
        respondido_em=None,

        situacao_aplicacao="",
        larva_visualizada="",
        retornar_proxima_semana_monitorar_larvas="NAO",

        distrito_administrativo=ordem_atual.distrito_administrativo,
        nome_rf_ace_responsavel_os=ordem_atual.nome_rf_ace_responsavel_os,
        criadouro_os_tipo_volume=ordem_atual.criadouro_os_tipo_volume,

        data_aplicacao=None,
        hora_inicio_aplicacao=None,
        hora_termino_aplicacao=None,

        tratamento_adicional_realizado="",
        quantos_quais="",

        descricao_produto=ordem_atual.descricao_produto,
        formulacao_produto=ordem_atual.formulacao_produto,
        dosagem_g_10l=ordem_atual.dosagem_g_10l,
        tipo_aplicacao=ordem_atual.tipo_aplicacao,
        quantidade_produto_administrada_ml=None,
        pulverizacao_area_l_ha=ordem_atual.pulverizacao_area_l_ha,

        prefixo_aeronave_pulverizacao=ordem_atual.prefixo_aeronave_pulverizacao,
        prefixo_aeronave_monitoramento=ordem_atual.prefixo_aeronave_monitoramento,

        quantidade_videos_registradas=None,
        quantidade_imagens_registradas=None,
        ponta_pulverizacao=ordem_atual.ponta_pulverizacao,

        temperatura_c=None,
        umidade_relativa_pct=None,
        velocidade_vento_kmh=None,

        motivo_nao_realizacao="",
        observacoes="",

        piloto=ordem_atual.piloto,
        assinatura_piloto="",
        auxiliar=ordem_atual.auxiliar,
        proprietario_ou_preposto="",
        assinatura_proprietario_ou_preposto="",

        drone_id=ordem_atual.drone_id,
        drone_monitoramento_id=ordem_atual.drone_monitoramento_id,

        drone_denominacao=ordem_atual.drone_denominacao,
        drone_modelo=ordem_atual.drone_modelo,
        drone_numero_serie=ordem_atual.drone_numero_serie,
        drone_registro_anatel=ordem_atual.drone_registro_anatel,
        drone_registro_anac=ordem_atual.drone_registro_anac,

        drone_monitoramento_denominacao=ordem_atual.drone_monitoramento_denominacao,
        drone_monitoramento_modelo=ordem_atual.drone_monitoramento_modelo,
        drone_monitoramento_numero_serie=ordem_atual.drone_monitoramento_numero_serie,
        drone_monitoramento_registro_anatel=ordem_atual.drone_monitoramento_registro_anatel,
        drone_monitoramento_registro_anac=ordem_atual.drone_monitoramento_registro_anac,
    )

    db.session.add(nova_ordem)
    return nova_solicitacao


# ============================================================
# ROTA REAL DO FORMULÁRIO (GET + POST)
# /piloto/os/<os_id>/formulario
# ============================================================
@bp.route("/piloto/os/<int:os_id>/formulario", methods=["GET", "POST"])
@login_required
@roles_required("piloto")
def piloto_os_formulario_view(os_id):

    if not current_user.piloto_id:
        flash("Piloto sem vínculo cadastrado.", "danger")
        return redirect(url_for("main.dashboard"))

    status_ok = ["APROVADO", "APROVADO COM RECOMENDAÇÕES", "APROVADA", "APROVADA COM RECOMENDAÇÕES"]
    status_permitidos = set(status_ok + ["CONCLUÍDO"])

    # Carrega OS + usuario (UVIS) + equipe
    s = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe),
        )
        .get_or_404(os_id)
    )

    if s.status not in status_permitidos:
        flash("Esta OS não está liberada para preenchimento do formulário.", "warning")
        return redirect(url_for("main.piloto_os"))

    if not s.equipe_id:
        flash("Esta OS não possui equipe atribuída.", "danger")
        return redirect(url_for("main.piloto_os"))

    # Segurança: piloto precisa estar na equipe ativa da OS
    vinculo = (
        EquipePiloto.query
        .join(Equipe, Equipe.id == EquipePiloto.equipe_id)
        .options(joinedload(EquipePiloto.equipe))
        .filter(
            EquipePiloto.equipe_id == s.equipe_id,
            EquipePiloto.piloto_id == current_user.piloto_id,
            Equipe.ativa.is_(True)
        )
        .first()
    )

    if not vinculo:
        flash("Você não tem permissão para acessar esta OS.", "danger")
        return redirect(url_for("main.piloto_os"))

    equipe = vinculo.equipe
    ordem = s.ordem_servico  # 1:1

    from sqlalchemy.orm import aliased
    d_alias = aliased(Drones, flat=True)

    drones_equipe = (
        db.session.query(d_alias)
        .filter(
            d_alias.equipe_id == s.equipe_id,
            d_alias.status == "Ativo"
        )
        .order_by(d_alias.renomacao.asc())
        .all()
    )

    # Defaults puxados das tabelas
    uvis_nome = s.usuario.nome_uvis if s.usuario else ""
    endereco_os = f"{s.logradouro or ''}, {s.numero or 'S/N'} - {s.bairro or ''} - {s.cidade or ''}/{s.uf or ''}"

    piloto_padrao = (equipe.piloto_titular.nome_piloto if equipe and equipe.piloto_titular else "") if equipe else ""
    auxiliar_padrao = (equipe.piloto_auxiliar.nome_piloto if equipe and equipe.piloto_auxiliar else "") if equipe else ""

    respondido_por_padrao = ""
    if getattr(current_user, "piloto", None) and current_user.piloto:
        respondido_por_padrao = current_user.piloto.nome_piloto or ""
    else:
        respondido_por_padrao = getattr(current_user, "nome_uvis", "") or ""

    # datetime-local
    if ordem and ordem.respondido_em:
        respondido_em_value = ordem.respondido_em.strftime("%Y-%m-%dT%H:%M")
    else:
        respondido_em_value = datetime.now().strftime("%Y-%m-%dT%H:%M")

    os_concluida = (s.status or "").strip().upper() in {"CONCLUÍDO", "CONCLUIDO"}
    modo_visualizacao = os_concluida

    # ----------------------------
    # POST (salvar)
    # ----------------------------
    if request.method == "POST":
        if modo_visualizacao:
            flash("Esta OS já foi concluída e não pode mais ser editada pelo piloto.", "warning")
            return redirect(url_for("main.piloto_os_formulario_view", os_id=os_id))

        # cria registro se não existir
        if ordem is None:
            ordem = OrdemServico(
                solicitacao_id=s.id,
                equipe_id=s.equipe_id
            )
            db.session.add(ordem)

        # ----------------------------
        #  Drones selecionados (PULVERIZAÇÃO E MONITORAMENTO)
        # ----------------------------
        # ⚠️ IMPORTANTE:
        # Ajuste o name do select principal no HTML pra "drone_pulv_id"
        # Aqui eu deixei compatível com seu código atual (drone_id) + o do monitoramento.
        drone_pulv_id = request.form.get("drone_id", type=int)                 # principal (pulv)
        drone_monit_id = request.form.get("drone_monitoramento_id", type=int) # monitoramento

        # --- PROCESSA DRONE DE PULVERIZAÇÃO (PRINCIPAL) ---
        if drone_pulv_id:
            drone_p = Drones.query.get(drone_pulv_id)
            if drone_p and drone_p.equipe_id == s.equipe_id:
                ordem.drone_id = drone_p.id
                # Snapshot para histórico
                ordem.drone_denominacao = drone_p.renomacao
                ordem.drone_modelo = drone_p.modelo
                ordem.drone_numero_serie = drone_p.numero_serie
                ordem.drone_registro_anatel = drone_p.registro_anatel
                ordem.drone_registro_anac = drone_p.registro_anac
                # Automação de Prefixo
                ordem.prefixo_aeronave_pulverizacao = drone_p.renomacao
        else:
            ordem.drone_id = None
            ordem.drone_denominacao = ""          #  vazio ao invés de None
            ordem.drone_modelo = ""               #  vazio ao invés de None
            ordem.drone_numero_serie = ""         #  vazio ao invés de None
            ordem.drone_registro_anatel = ""      #  vazio ao invés de None
            ordem.drone_registro_anac = ""        #  vazio ao invés de None
            ordem.prefixo_aeronave_pulverizacao = _clean_str(request.form.get("prefixo_aeronave_pulverizacao"))

        # --- PROCESSA DRONE DE MONITORAMENTO ---
        if drone_monit_id:
            drone_m = Drones.query.get(drone_monit_id)
            if drone_m and drone_m.equipe_id == s.equipe_id:
                ordem.drone_monitoramento_id = drone_m.id
                ordem.drone_monitoramento_denominacao = drone_m.renomacao
                ordem.drone_monitoramento_modelo = drone_m.modelo
                ordem.drone_monitoramento_numero_serie = drone_m.numero_serie
                ordem.drone_monitoramento_registro_anatel = drone_m.registro_anatel
                ordem.drone_monitoramento_registro_anac = drone_m.registro_anac
                ordem.prefixo_aeronave_monitoramento = drone_m.renomacao
        else:
            ordem.drone_monitoramento_id = None
            ordem.drone_monitoramento_denominacao = ""     #  vazio
            ordem.drone_monitoramento_modelo = ""          #  vazio
            ordem.drone_monitoramento_numero_serie = ""    #  vazio
            ordem.drone_monitoramento_registro_anatel = "" #  vazio
            ordem.drone_monitoramento_registro_anac = ""   #  vazio
            ordem.prefixo_aeronave_monitoramento = _clean_str(request.form.get("prefixo_aeronave_monitoramento"))

        # --- Continuação normal dos campos ---
        #  Textos -> _clean_str (vazio vira "")
        #  Números/datas/horas -> _to_* (vazio vira None)

        ordem.identificador_os = _clean_str(request.form.get("identificador_os"))
        ordem.respondido_por = _clean_str(request.form.get("respondido_por")) or respondido_por_padrao
        ordem.respondido_em = _to_datetime_local(request.form.get("respondido_em")) or datetime.now()

        ordem.situacao_aplicacao = _clean_str(request.form.get("situacao_aplicacao"))
        ordem.larva_visualizada = _clean_str(request.form.get("larva_visualizada"))
        ordem.retornar_proxima_semana_monitorar_larvas = _clean_str(request.form.get("retornar_proxima_semana_monitorar_larvas"))

        # template usa name="da" mas model é distrito_administrativo
        ordem.distrito_administrativo = _clean_str(request.form.get("da")) or _clean_str(request.form.get("distrito_administrativo"))

        ordem.nome_rf_ace_responsavel_os = _clean_str(request.form.get("nome_rf_ace_responsavel_os"))
        ordem.criadouro_os_tipo_volume = _clean_str(request.form.get("criadouro_os_tipo_volume"))

        ordem.data_aplicacao = _to_date(request.form.get("data_aplicacao"))
        ordem.hora_inicio_aplicacao = _to_time(request.form.get("hora_inicio_aplicacao"))
        ordem.hora_termino_aplicacao = _to_time(request.form.get("hora_termino_aplicacao"))

        ordem.tratamento_adicional_realizado = _clean_str(request.form.get("tratamento_adicional_realizado"))
        ordem.quantos_quais = _clean_str(request.form.get("quantos_quais"))

        ordem.descricao_produto = _clean_str(request.form.get("descricao_produto"))
        ordem.formulacao_produto = _clean_str(request.form.get("formulacao_produto"))
        ordem.dosagem_g_10l = _clean_str(request.form.get("dosagem_g_10l"))

        ordem.tipo_aplicacao = _clean_str(request.form.get("tipo_aplicacao"))
        ordem.quantidade_produto_administrada_ml = _to_float(request.form.get("quantidade_produto_administrada_ml"))
        ordem.pulverizacao_area_l_ha = _to_float(request.form.get("pulverizacao_area_l_ha"))
        ordem.pulverizacao_foco_tempo_estimado_segundos = _to_float(request.form.get("pulverizacao_foco_tempo_estimado_segundos"))
        ordem.pulverizacao_foco_l_min = _to_float(request.form.get("pulverizacao_foco_l_min"))

        ordem.quantidade_imagens_registradas = _to_int(request.form.get("quantidade_imagens_registradas"))
        ordem.quantidade_videos_registradas = _to_int(request.form.get("quantidade_videos_registradas"))

        ordem.ponta_pulverizacao = _clean_str(request.form.get("ponta_pulverizacao"))
        ordem.temperatura_c = _to_float(request.form.get("temperatura_c"))
        ordem.umidade_relativa_pct = _to_float(request.form.get("umidade_relativa_pct"))
        ordem.velocidade_vento_kmh = _to_float(request.form.get("velocidade_vento_kmh"))

        ordem.motivo_nao_realizacao = _clean_str(request.form.get("motivo_nao_realizacao"))

        obs = _clean_str(request.form.get("observacoes"))
        ordem.observacoes = obs

        ordem.piloto = _clean_str(request.form.get("piloto")) or piloto_padrao
        ordem.assinatura_piloto = _clean_str(request.form.get("assinatura_piloto"))
        ordem.auxiliar = _clean_str(request.form.get("auxiliar")) or auxiliar_padrao

        ordem.proprietario_ou_preposto = _clean_str(request.form.get("proprietario_ou_preposto"))
        ordem.assinatura_proprietario_ou_preposto = _clean_str(request.form.get("assinatura_proprietario_ou_preposto"))

        try:
            gerar_retorno = (
                (ordem.retornar_proxima_semana_monitorar_larvas or "").strip().upper() == "SIM"
            )

            current_app.logger.warning(
                "OS %s - retornar_proxima_semana_monitorar_larvas=%s",
                os_id,
                ordem.retornar_proxima_semana_monitorar_larvas
            )
            current_app.logger.warning("OS %s - gerar_retorno=%s", os_id, gerar_retorno)

            if gerar_retorno:
                retorno_existente = Solicitacao.query.filter_by(origem_retorno_id=s.id).first()
                current_app.logger.warning(
                    "OS %s - retorno_existente=%s",
                    os_id,
                    bool(retorno_existente)
                )

                if not retorno_existente:
                    nova = criar_solicitacao_retorno_monitoramento(s, ordem)
                    current_app.logger.warning(
                        "OS %s - nova solicitacao de retorno criada id=%s",
                        os_id,
                        nova.id
                    )

            db.session.commit()

            if gerar_retorno:
                flash("Formulário salvo com sucesso! Uma nova OS de retorno foi criada para 7 dias depois.", "success")
            else:
                flash("Formulário salvo com sucesso!", "success")

            return redirect(url_for("main.piloto_os"))

        except Exception:
            db.session.rollback()
            current_app.logger.exception("Erro ao salvar formulário da OS %s", os_id)
            flash("Erro ao salvar o formulário. Verifique os campos e tente novamente.", "danger")

    # ----------------------------
    # GET (render)
    # ----------------------------
    return render_template(
        "piloto_os_formulario.html",
        solicitacao=s,
        equipe=equipe,
        ordem=ordem,
        modo_visualizacao=modo_visualizacao,
        uvis_nome=uvis_nome,
        endereco_os=endereco_os,
        piloto_padrao=piloto_padrao,
        auxiliar_padrao=auxiliar_padrao,
        respondido_por_padrao=respondido_por_padrao,
        respondido_em_value=respondido_em_value,
        drones_equipe=drones_equipe,
        url_voltar=url_for("main.piloto_os"),
        form_action=url_for("main.piloto_os_formulario_view", os_id=os_id),
    )


@bp.route("/admin/os/<int:os_id>/formulario", methods=["GET"])
@login_required
def admin_os_formulario_view(os_id):
    if current_user.tipo_usuario not in ["admin", "operario", "visualizar"]:
        flash("Acesso restrito.", "danger")
        return redirect(url_for("main.dashboard"))

    s = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe),
            joinedload(Solicitacao.ordem_servico),
        )
        .get_or_404(os_id)
    )

    equipe = s.equipe
    ordem = s.ordem_servico

    drones_equipe = []
    if s.equipe_id:
        from sqlalchemy.orm import aliased
        d_alias = aliased(Drones, flat=True)
        drones_equipe = (
            db.session.query(d_alias)
            .filter(d_alias.equipe_id == s.equipe_id)
            .order_by(d_alias.renomacao.asc())
            .all()
        )

    uvis_nome = s.usuario.nome_uvis if s.usuario else ""
    endereco_os = f"{s.logradouro or ''}, {s.numero or 'S/N'} - {s.bairro or ''} - {s.cidade or ''}/{s.uf or ''}"
    piloto_padrao = (equipe.piloto_titular.nome_piloto if equipe and equipe.piloto_titular else "") if equipe else ""
    auxiliar_padrao = (equipe.piloto_auxiliar.nome_piloto if equipe and equipe.piloto_auxiliar else "") if equipe else ""
    respondido_por_padrao = ""
    respondido_em_value = ordem.respondido_em.strftime("%Y-%m-%dT%H:%M") if ordem and ordem.respondido_em else ""

    return render_template(
        "piloto_os_formulario.html",
        solicitacao=s,
        equipe=equipe,
        ordem=ordem,
        modo_visualizacao=True,
        uvis_nome=uvis_nome,
        endereco_os=endereco_os,
        piloto_padrao=piloto_padrao,
        auxiliar_padrao=auxiliar_padrao,
        respondido_por_padrao=respondido_por_padrao,
        respondido_em_value=respondido_em_value,
        drones_equipe=drones_equipe,
        url_voltar=url_for("main.admin_dashboard"),
        form_action="#",
    )


# ============================================================
# API: detalhes do drone (para preencher campos do form)
# ============================================================
@bp.route("/piloto/api/drone/<int:drone_id>", methods=["GET"])
@login_required
@roles_required("piloto")
def piloto_api_drone(drone_id):
    if not current_user.piloto_id:
        return jsonify({"error": "Piloto sem vínculo."}), 403

    drone = Drones.query.get_or_404(drone_id)

    # segurança: drone precisa estar vinculado a uma equipe
    if not drone.equipe_id:
        return jsonify({"error": "Drone sem equipe."}), 403

    # segurança: piloto precisa estar na equipe ativa desse drone
    vinculo = (
        EquipePiloto.query
        .join(Equipe, Equipe.id == EquipePiloto.equipe_id)
        .filter(
            EquipePiloto.equipe_id == drone.equipe_id,
            EquipePiloto.piloto_id == current_user.piloto_id,
            Equipe.ativa.is_(True)
        )
        .first()
    )
    if not vinculo:
        return jsonify({"error": "Sem permissão."}), 403

    return jsonify({
        "id": drone.id,
        "renomacao": drone.renomacao,
        "modelo": drone.modelo,
        "numero_serie": drone.numero_serie,
        "registro_anatel": drone.registro_anatel,
        "registro_anac": drone.registro_anac,
        "status": drone.status,
        "categoria": drone.categoria,
        "pmd_kg": drone.pmd_kg,
        "ano_fabricacao": drone.ano_fabricacao,
    })




import os
from datetime import datetime
from flask import request, render_template, redirect, url_for, flash
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload

@bp.route("/equipe-uvis", methods=["GET"], endpoint="dashboard_equipe_uvis")
@login_required
def dashboard_equipe_uvis():
    google_maps_key = os.getenv("KEY_API_GOOGLE_MAPS")

    # 🔒 só conta de equipe
    if getattr(current_user, "tipo_usuario", None) != "equipe_uvis":
        return redirect(url_for("main.dashboard"))

    uvis_id = getattr(current_user, "equipe_uvis_uvis_usuario_id", None)
    nome_equipe = (getattr(current_user, "equipe_uvis_nome", "") or "").strip()

    if not uvis_id or not nome_equipe:
        flash("Conta de equipe sem vínculo com UVIS/equipe. Contate o administrador.", "danger")
        return redirect(url_for("auth.login"))

    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.equipe)
        )
        .filter(Solicitacao.usuario_id == uvis_id)
        .filter(Solicitacao.equipe_uvis_nome == nome_equipe)
        .filter(Solicitacao.status != "CANCELADO")
    )

    # =========================
    # FILTROS (status, tipo, foco)
    # =========================
    filtro_status = request.args.get('status')
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    filtro_tipo_visita = request.args.get('tipo_visita')
    if filtro_tipo_visita:
        query = query.filter(Solicitacao.tipo_visita == filtro_tipo_visita)

    filtro_foco = request.args.get('foco')
    if filtro_foco:
        query = query.filter(Solicitacao.foco == filtro_foco)

    # =========================
    # FILTRO POR DATA
    # =========================
    data_ini = request.args.get("data_ini")  # YYYY-MM-DD
    data_fim = request.args.get("data_fim")  # YYYY-MM-DD

    if data_ini:
        try:
            dt_ini = datetime.strptime(data_ini, "%Y-%m-%d").date()
            query = query.filter(Solicitacao.data_agendamento >= dt_ini)
        except ValueError:
            pass

    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, "%Y-%m-%d").date()
            query = query.filter(Solicitacao.data_agendamento <= dt_fim)
        except ValueError:
            pass

    # =========================
    # PAGINAÇÃO
    # =========================
    page = request.args.get("page", 1, type=int)
    paginacao = (
        query.order_by(Solicitacao.data_criacao.desc())
        .paginate(page=page, per_page=6, error_out=False)
    )

    return render_template(
        "dashboard_equipe_uvis.html",
        solicitacoes=paginacao.items,
        paginacao=paginacao,
        google_maps_key=google_maps_key,
        nome_equipe=nome_equipe,
    )

import os
import base64
import re
import tempfile
from datetime import datetime
from io import BytesIO
from urllib.parse import urlencode

import requests

from flask import send_file, request, current_app
from flask_login import login_required
from sqlalchemy.orm import joinedload

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)
from reportlab.platypus import Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors

# Seus imports reais:
# from app import db
# from app.models import Solicitacao, OrdemServico
# from app.decorators import roles_required


def _fmt_dt(v):
    if not v:
        return ""
    try:
        return v.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(v)


def _fmt_date(v):
    if not v:
        return ""
    try:
        return v.strftime("%d/%m/%Y")
    except Exception:
        return str(v)


def _fmt_time(v):
    if not v:
        return ""
    try:
        return v.strftime("%H:%M")
    except Exception:
        return str(v)


def _safe(v):
    if v is None:
        return ""
    return str(v)


def _normalize_coord(v):
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    return s.replace(",", ".")


# -------------------------
#  Logo (static/img/...)
# -------------------------
def _get_logo_path():
    logo_mode = (request.args.get("logo") or "light").strip().lower()
    filename = "img/logo_oceano_azul_dark.png" if logo_mode == "dark" else "img/logo_oceano_azul_light.png"
    return os.path.join(current_app.root_path, "static", filename)


def _try_make_logo(width_mm=34):
    try:
        p = _get_logo_path()
        if not os.path.exists(p):
            return None
        img = RLImage(p)
        img.drawWidth = width_mm * mm
        img.drawHeight = (width_mm * 0.55) * mm
        return img
    except Exception:
        return None


# -------------------------
#  Assinatura dataURL -> imagem
# -------------------------
_DATAURL_RE = re.compile(r"^data:image/(?P<fmt>png|jpeg|jpg);base64,(?P<data>.+)$", re.I)


def _dataurl_to_rlimage(dataurl: str, width_mm=80, height_mm=32):
    if not dataurl or not isinstance(dataurl, str):
        return None

    m = _DATAURL_RE.match(dataurl.strip())
    if not m:
        return None

    try:
        raw = base64.b64decode(m.group("data"))
    except Exception:
        return None

    bio = BytesIO(raw)
    img = RLImage(bio)
    img.drawWidth = width_mm * mm
    img.drawHeight = height_mm * mm
    return img


# -------------------------
#  Google Maps Static API -> imagem
# -------------------------
def _try_make_static_map(lat, lng, width_mm=165, height_mm=88, zoom=19, maptype="satellite"):
    lat = _normalize_coord(lat)
    lng = _normalize_coord(lng)

    if not lat or not lng:
        return None

    api_key = current_app.config.get("KEY_API_GOOGLE_MAPS")
    if not api_key:
        return None

    try:
        params = {
            "center": f"{lat},{lng}",
            "zoom": zoom,
            "size": "1200x650",
            "scale": "2",
            "maptype": maptype,
            "markers": f"color:red|label:O|{lat},{lng}",
            "key": api_key,
        }

        url = "https://maps.googleapis.com/maps/api/staticmap?" + urlencode(params)

        r = requests.get(url, timeout=15)
        r.raise_for_status()

        content_type = (r.headers.get("Content-Type") or "").lower()
        if "image" not in content_type:
            return None

        bio = BytesIO(r.content)
        img = RLImage(bio)
        img.drawWidth = width_mm * mm
        img.drawHeight = height_mm * mm
        return img

    except Exception:
        return None


# -------------------------
#  Componentes visuais do PDF
# -------------------------
def _pdf_styles():
    styles = getSampleStyleSheet()

    title = ParagraphStyle(
        "p_title",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0F3D75"),
        spaceAfter=2
    )

    subtitle = ParagraphStyle(
        "p_subtitle",
        parent=styles["Normal"],
        fontSize=9.2,
        leading=12.5,
        textColor=colors.HexColor("#667085"),
        spaceAfter=10
    )

    section = ParagraphStyle(
        "p_section",
        parent=styles["Heading2"],
        fontSize=11.5,
        leading=15,
        textColor=colors.HexColor("#0F3D75"),
        spaceBefore=8,
        spaceAfter=5
    )

    cell = ParagraphStyle(
        "p_cell",
        parent=styles["BodyText"],
        fontSize=9,
        leading=11.8,
        textColor=colors.HexColor("#111827"),
        wordWrap="CJK",
        splitLongWords=True
    )

    hint = ParagraphStyle(
        "p_hint",
        parent=styles["Normal"],
        fontSize=8.3,
        leading=11,
        textColor=colors.HexColor("#667085"),
        spaceAfter=5
    )

    small = ParagraphStyle(
        "p_small",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#475467"),
    )

    return styles, title, subtitle, section, cell, hint, small


def _pdf_header_block(os_id: int, status_txt: str):
    styles, title_s, subtitle_s, *_ = _pdf_styles()

    logo = _try_make_logo(width_mm=35)
    title = Paragraph(f"OS #{os_id} — Formulário (Admin)", title_s)
    subtitle = Paragraph(
        f"Gerado em {_fmt_dt(datetime.now())} • Status: {_safe(status_txt)}",
        subtitle_s
    )

    left = [title, subtitle]
    right = [logo] if logo else [Paragraph("", styles["Normal"])]

    tbl = Table([[left, right]], colWidths=[None, 42 * mm])
    tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    return [tbl, Spacer(1, 3)]


def _pdf_kv_table_nice(section_title: str, items: list[tuple[str, object]], cell_style, section_style, doc_width, orient="portrait"):
    key_w = 60 * mm if orient == "portrait" else 73 * mm
    val_w = doc_width - key_w

    rows = [[
        Paragraph("<b>Campo</b>", ParagraphStyle("th1", parent=cell_style, textColor=colors.white)),
        Paragraph("<b>Valor</b>", ParagraphStyle("th2", parent=cell_style, textColor=colors.white)),
    ]]

    for k, v in items:
        rows.append([
            Paragraph(_safe(k), cell_style),
            Paragraph(_safe(v), cell_style),
        ])

    tbl = Table(rows, repeatRows=1, colWidths=[key_w, val_w])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.HexColor("#1565C0")),
        ("GRID", (0, 1), (-1, -1), 0.25, colors.HexColor("#DDE3EA")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    return [
        Paragraph(section_title, section_style),
        tbl,
        Spacer(1, 7)
    ]


def _pdf_card(flowables, doc_width, bg="#F8FAFC", border="#D0D5DD", padding=8):
    card = Table([[flowables]], colWidths=[doc_width])
    card.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(bg)),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor(border)),
        ("LEFTPADDING", (0, 0), (-1, -1), padding),
        ("RIGHTPADDING", (0, 0), (-1, -1), padding),
        ("TOPPADDING", (0, 0), (-1, -1), padding),
        ("BOTTOMPADDING", (0, 0), (-1, -1), padding),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return card


def _header_footer_factory_pretty(title):
    def _hf(canvas, doc):
        canvas.saveState()
        w, h = doc.pagesize

        canvas.setFillColor(colors.HexColor("#1565C0"))
        canvas.rect(doc.leftMargin, h - (11 * mm), doc.width, 2.8, fill=1, stroke=0)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawString(doc.leftMargin, 9 * mm, title)
        canvas.drawRightString(doc.leftMargin + doc.width, 9 * mm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()

    return _hf


# ============================================================
#  ROTA PDF V2 (bonita com logo + coordenadas + mapa)
# ============================================================
@bp.route("/admin/os/<int:os_id>/export/pdf/v2", methods=["GET"])
@login_required
@roles_required("admin")
def admin_export_os_pdf_v2(os_id):
    orient = request.args.get("orient", "portrait")
    pagesize = landscape(A4) if orient == "landscape" else A4

    try:
        s = (
            Solicitacao.query
            .options(
                joinedload(Solicitacao.usuario),
                joinedload(Solicitacao.equipe),
                joinedload(Solicitacao.ordem_servico),
            )
            .get_or_404(os_id)
        )
        ordem = s.ordem_servico

        tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        caminho_pdf = tmp_pdf.name
        tmp_pdf.close()

        doc = SimpleDocTemplate(
            caminho_pdf,
            pagesize=pagesize,
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm
        )

        styles, title_s, subtitle_s, section_s, cell_s, hint_s, small_s = _pdf_styles()

        story = []
        story += _pdf_header_block(s.id, s.status)
        story.append(Spacer(1, 2))

        endereco_os = f"{s.logradouro or ''}, {s.numero or 'S/N'} - {s.bairro or ''} - {s.cidade or ''}/{s.uf or ''}"
        lat = _normalize_coord(getattr(s, "latitude", None))
        lng = _normalize_coord(getattr(s, "longitude", None))

        story += _pdf_kv_table_nice("Identificação", [
            ("Solicitação ID", s.id),
            ("Equipe ID", s.equipe_id or ""),
            ("Equipe", (s.equipe.nome_equipe if s.equipe else "")),
            ("UVIS", (s.usuario.nome_uvis if s.usuario else "")),
            ("Endereço", endereco_os),
            ("Data agendamento", _fmt_date(s.data_agendamento)),
            ("Hora agendamento", _fmt_time(s.hora_agendamento)),
            ("Foco", s.foco or ""),
            ("Status", s.status or ""),
            ("Protocolo", getattr(s, "protocolo", "") or ""),
        ], cell_s, section_s, doc.width, orient=orient)

        story += _pdf_kv_table_nice("Endereço / Coordenadas", [
            ("CEP", s.cep or ""),
            ("Logradouro", s.logradouro or ""),
            ("Número", s.numero or ""),
            ("Bairro", s.bairro or ""),
            ("Cidade", s.cidade or ""),
            ("UF", s.uf or ""),
            ("Complemento", s.complemento or ""),
            ("Latitude", lat or ""),
            ("Longitude", lng or ""),
        ], cell_s, section_s, doc.width, orient=orient)

        if lat and lng:
            maps_link = f"https://www.google.com/maps?q={lat},{lng}"
            map_img = _try_make_static_map(
                lat=lat,
                lng=lng,
                width_mm=175 if orient == "landscape" else 165,
                height_mm=92,
                zoom=19,
                maptype="satellite"
            )

            map_block = [
                Paragraph("Localização no mapa", section_s),
                Paragraph(
                    "Visual gerado automaticamente a partir da latitude e longitude da solicitação.",
                    hint_s
                ),
                Paragraph(
                    f'Para acessar o Google Maps, clique aqui: <link href="{maps_link}">{maps_link}</link>',
                    small_s
                ),
                Spacer(1, 5),
            ]

            if map_img:
                map_block.append(map_img)
            else:
                map_block.append(Paragraph(
                    "Não foi possível gerar a imagem do mapa no momento.",
                    styles["Normal"]
                ))

            story.append(_pdf_card(map_block, doc.width, bg="#F8FAFC", border="#D0D5DD", padding=8))
            story.append(Spacer(1, 8))

        if not ordem:
            story.append(Paragraph("Formulário", section_s))
            story.append(Paragraph("Esta OS não possui formulário preenchido.", styles["Normal"]))
        else:
            story += _pdf_kv_table_nice("Responsável / Registro", [
                ("Identificador OS", ordem.identificador_os or ""),
                ("Respondido por", ordem.respondido_por or ""),
                ("Respondido em", _fmt_dt(ordem.respondido_em)),
            ], cell_s, section_s, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Aeronaves — Pulverização (Principal)", [
                ("Drone ID", ordem.drone_id or ""),
                ("Prefixo", ordem.prefixo_aeronave_pulverizacao or ""),
                ("Denominação", ordem.drone_denominacao or ""),
                ("Modelo", ordem.drone_modelo or ""),
                ("Nº Série", ordem.drone_numero_serie or ""),
                ("Registro ANATEL", ordem.drone_registro_anatel or ""),
                ("Registro ANAC", ordem.drone_registro_anac or ""),
            ], cell_s, section_s, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Aeronaves — Monitoramento", [
                ("Drone Monitoramento ID", ordem.drone_monitoramento_id or ""),
                ("Prefixo", ordem.prefixo_aeronave_monitoramento or ""),
                ("Denominação", ordem.drone_monitoramento_denominacao or ""),
                ("Modelo", ordem.drone_monitoramento_modelo or ""),
                ("Nº Série", ordem.drone_monitoramento_numero_serie or ""),
                ("Registro ANATEL", ordem.drone_monitoramento_registro_anatel or ""),
                ("Registro ANAC", ordem.drone_monitoramento_registro_anac or ""),
            ], cell_s, section_s, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Aplicação", [
                ("Situação da aplicação", ordem.situacao_aplicacao or ""),
                ("Larva visualizada", ordem.larva_visualizada or ""),
                ("Retornar monitorar larvas", ordem.retornar_proxima_semana_monitorar_larvas or ""),
                ("DA (Distrito)", ordem.distrito_administrativo or ""),
                ("Nome/RF ACE responsável", ordem.nome_rf_ace_responsavel_os or ""),
                ("Criadouro OS (tipo/volume)", ordem.criadouro_os_tipo_volume or ""),
                ("Data aplicação", _fmt_date(ordem.data_aplicacao)),
                ("Hora início", _fmt_time(ordem.hora_inicio_aplicacao)),
                ("Hora término", _fmt_time(ordem.hora_termino_aplicacao)),
                ("Tratamento adicional", ordem.tratamento_adicional_realizado or ""),
                ("Quantos / Quais", ordem.quantos_quais or ""),
            ], cell_s, section_s, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Produto e Parâmetros", [
                ("Descrição produto", ordem.descricao_produto or ""),
                ("Formulação", ordem.formulacao_produto or ""),
                ("Dosagem (g/10L)", ordem.dosagem_g_10l or ""),
                ("Tipo aplicação", ordem.tipo_aplicacao or ""),
                ("Qtd administrada (ml)", ordem.quantidade_produto_administrada_ml or ""),
                ("Pulverização área (l/ha)", ordem.pulverizacao_area_l_ha or ""),
                ("Ponta pulverização", ordem.ponta_pulverizacao or ""),
            ], cell_s, section_s, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Condições Ambientais", [
                ("Imagens registradas", ordem.quantidade_imagens_registradas or ""),
                ("Vídeos registrados", ordem.quantidade_videos_registradas or ""),
                ("Temperatura (°C)", ordem.temperatura_c or ""),
                ("Umidade (%)", ordem.umidade_relativa_pct or ""),
                ("Vento (km/h)", ordem.velocidade_vento_kmh or ""),
            ], cell_s, section_s, doc.width, orient=orient)

            story += _pdf_kv_table_nice("Fechamento", [
                ("Observações gerais", ordem.observacoes or ""),
                ("Motivo não realização", ordem.motivo_nao_realizacao or ""),
                ("Piloto", ordem.piloto or ""),
                ("Auxiliar", ordem.auxiliar or ""),
                ("Proprietário/Preposto", ordem.proprietario_ou_preposto or ""),
            ], cell_s, section_s, doc.width, orient=orient)

            ass_piloto = _dataurl_to_rlimage(
                getattr(ordem, "assinatura_piloto", None),
                width_mm=82,
                height_mm=32
            )
            ass_resp = _dataurl_to_rlimage(
                getattr(ordem, "assinatura_proprietario_ou_preposto", None),
                width_mm=82,
                height_mm=32
            )

            story.append(Paragraph("Assinaturas", section_s))
            story.append(Paragraph("Exportadas diretamente do formulário.", hint_s))
            story.append(Spacer(1, 4))

            def _sig_card(title_html, who, img_or_none):
                inner = [
                    Paragraph(f"<b>{title_html}</b>", styles["Normal"]),
                    Paragraph(_safe(who) if who else "—", hint_s),
                    Spacer(1, 4),
                    img_or_none if img_or_none else Paragraph("Não informada.", styles["Normal"]),
                ]
                card = Table([[inner]], colWidths=[doc.width / 2 - 5 * mm])
                card.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#D0D5DD")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 9),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]))
                return card

            card_left = _sig_card(
                "Assinatura do Piloto",
                getattr(ordem, "piloto", ""),
                ass_piloto
            )
            card_right = _sig_card(
                "Assinatura do Responsável (Local)",
                getattr(ordem, "proprietario_ou_preposto", ""),
                ass_resp
            )

            cards = Table([[card_left, card_right]], colWidths=[doc.width / 2, doc.width / 2])
            cards.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(cards)
            story.append(Spacer(1, 8))

        header_title = f"OS #{s.id} — Oceano Azul / IJA Drones"

        doc.build(
            story,
            onFirstPage=_header_footer_factory_pretty(header_title),
            onLaterPages=_header_footer_factory_pretty(header_title)
        )

        nome_arquivo = f"os_{s.id}_formulario.pdf"
        return send_file(
            caminho_pdf,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/pdf"
        )

    except Exception:
        db.session.rollback()
        raise
# ============================================================
#  EXPORTAÇÃO OS (ADMIN) — EXCEL V2 (BONITO) + ASSINATURAS EM IMAGEM
# - OpenPyXL (sem libs extras)
# - Layout com títulos, seções, cabeçalho azul, zebra, bordas
# - Assinaturas:
#     - Aba "Formulário": OK / Não informada
#     - Aba "Assinaturas" (imagem): habilita com ?assinaturas=1
# ============================================================

import tempfile
import base64
import re
from datetime import datetime
from io import BytesIO

from flask import send_file, request
from flask_login import login_required
from sqlalchemy.orm import joinedload

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

# Seus imports reais:
# from app import db
# from app.models import Solicitacao, OrdemServico
# from app.decorators import roles_required
# from app.routes import bp  (ou onde estiver seu blueprint)


# ------------------------------------------------------------
# Helpers básicos
# ------------------------------------------------------------
def _fmt_dt(v):
    if not v:
        return ""
    try:
        return v.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(v)

def _fmt_date(v):
    if not v:
        return ""
    try:
        return v.strftime("%d/%m/%Y")
    except Exception:
        return str(v)

def _fmt_time(v):
    if not v:
        return ""
    try:
        return v.strftime("%H:%M")
    except Exception:
        return str(v)

def _safe(v):
    if v is None:
        return ""
    return str(v)


# ------------------------------------------------------------
#  Assinatura (dataURL base64 -> bytes)
# ------------------------------------------------------------
_DATAURL_RE = re.compile(r"^data:image/(?P<fmt>png|jpeg|jpg);base64,(?P<data>.+)$", re.I)

def _dataurl_to_png_bytes(dataurl: str):
    """
    Recebe 'data:image/png;base64,...' ou 'data:image/jpeg;base64,...'
    e devolve os bytes decodificados (PNG/JPEG).
    Retorna None se inválido.
    """
    if not dataurl or not isinstance(dataurl, str):
        return None

    m = _DATAURL_RE.match(dataurl.strip())
    if not m:
        return None

    try:
        return base64.b64decode(m.group("data"))
    except Exception:
        return None


# ------------------------------------------------------------
#  Excel helpers (layout bonito)
# ------------------------------------------------------------
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_HEADER = PatternFill("solid", fgColor="0D6EFD")
FILL_SECTION = PatternFill("solid", fgColor="EAF2FF")
FILL_ZEBRA = PatternFill("solid", fgColor="FBFDFF")

FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_TITLE = Font(bold=True, size=16, color="0D6EFD")
FONT_SUBTITLE = Font(size=10, color="555555")
FONT_SECTION = Font(bold=True, color="0D6EFD")

def _excel_add_title(ws, title: str, subtitle: str = ""):
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    ws["A2"] = subtitle
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

def _excel_add_section(ws, row: int, title: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = FILL_SECTION
    c.font = FONT_SECTION
    c.alignment = Alignment(vertical="center")
    c.border = BORDER
    ws.cell(row=row, column=2).border = BORDER
    ws.row_dimensions[row].height = 18

def _excel_apply_table_style(ws, header_row: int, end_row: int, col_count: int = 2):
    # header row
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # body zebra + borders
    for r in range(header_row + 1, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if (r - header_row) % 2 == 0:
                cell.fill = FILL_ZEBRA

def _excel_write_kv(ws, start_row: int, items: list[tuple[str, object]]):
    ws.cell(row=start_row, column=1, value="Campo")
    ws.cell(row=start_row, column=2, value="Valor")

    r = start_row + 1
    for k, v in items:
        ws.cell(row=r, column=1, value=str(k))
        ws.cell(row=r, column=2, value=_safe(v))
        r += 1

    _excel_apply_table_style(ws, start_row, r - 1, col_count=2)
    return r

def _excel_auto_width(ws, max_col=2, min_w=18, max_w=75):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = 0
        for cell in ws[letter]:
            if cell.value:
                best = max(best, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, best + 2))


# ============================================================
#  ROTA — EXCEL V2 COMPLETA (como você pediu)
# ============================================================
@bp.route("/admin/os/<int:os_id>/export/excel/v2")
@login_required
@roles_required("admin")
def admin_export_os_excel_v2(os_id):
    """
    Excel mais organizado e bonito (seções + estilo).
    (Opcional) para incluir imagens das assinaturas na aba "Assinaturas":
      /admin/os/<id>/export/excel/v2?assinaturas=1
    """
    try:
        s = (
            Solicitacao.query
            .options(
                joinedload(Solicitacao.usuario),
                joinedload(Solicitacao.equipe),
                joinedload(Solicitacao.ordem_servico),
            )
            .get_or_404(os_id)
        )
        ordem = s.ordem_servico

        wb = Workbook()

        # =========================
        # ABA 1 — OS (Resumo)
        # =========================
        ws_os = wb.active
        ws_os.title = "OS (Resumo)"

        _excel_add_title(
            ws_os,
            f"OS #{s.id} — Exportação",
            f"Gerado em {_fmt_dt(datetime.now())} | Status: {_safe(s.status)}"
        )

        endereco = f"{s.logradouro or ''}, {s.numero or 'S/N'} - {s.bairro or ''} - {s.cidade or ''}/{s.uf or ''}"

        r = 4
        _excel_add_section(ws_os, r, "Identificação")
        r += 1
        r = _excel_write_kv(ws_os, r, [
            ("Solicitação ID", s.id),
            ("Status", s.status or ""),
            ("Equipe ID", s.equipe_id or ""),
            ("Equipe", (s.equipe.nome_equipe if s.equipe else "")),
            ("UVIS", (s.usuario.nome_uvis if s.usuario else "")),
            ("Endereço", endereco),
            ("Data agendamento", _fmt_date(s.data_agendamento)),
            ("Hora agendamento", _fmt_time(s.hora_agendamento)),
            ("Foco", s.foco or ""),
            ("Protocolo", getattr(s, "protocolo", "") or ""),
        ])

        r += 1
        _excel_add_section(ws_os, r, "Endereço / Coordenadas")
        r += 1
        r = _excel_write_kv(ws_os, r, [
            ("CEP", s.cep or ""),
            ("Logradouro", s.logradouro or ""),
            ("Número", s.numero or ""),
            ("Bairro", s.bairro or ""),
            ("Cidade", s.cidade or ""),
            ("UF", s.uf or ""),
            ("Complemento", s.complemento or ""),
            ("Latitude", s.latitude or ""),
            ("Longitude", s.longitude or ""),
        ])

        ws_os.freeze_panes = "A5"
        _excel_auto_width(ws_os, max_col=2, min_w=18, max_w=75)

        # =========================
        # ABA 2 — Formulário
        # =========================
        ws_f = wb.create_sheet("Formulário")
        _excel_add_title(ws_f, f"Formulário — OS #{s.id}", "Campos preenchidos pelo piloto")

        r = 4
        if not ordem:
            _excel_add_section(ws_f, r, "Sem formulário")
            r += 1
            r = _excel_write_kv(ws_f, r, [("Status", "Esta OS não possui formulário preenchido.")])
        else:
            _excel_add_section(ws_f, r, "Responsável / Registro")
            r += 1
            r = _excel_write_kv(ws_f, r, [
                ("Identificador OS", ordem.identificador_os or ""),
                ("Respondido por", ordem.respondido_por or ""),
                ("Respondido em", _fmt_dt(ordem.respondido_em)),
            ])

            r += 1
            _excel_add_section(ws_f, r, "Aeronaves — Pulverização (Principal)")
            r += 1
            r = _excel_write_kv(ws_f, r, [
                ("Drone ID", ordem.drone_id or ""),
                ("Prefixo", ordem.prefixo_aeronave_pulverizacao or ""),
                ("Denominação", ordem.drone_denominacao or ""),
                ("Modelo", ordem.drone_modelo or ""),
                ("Nº Série", ordem.drone_numero_serie or ""),
                ("Registro ANATEL", ordem.drone_registro_anatel or ""),
                ("Registro ANAC", ordem.drone_registro_anac or ""),
            ])

            r += 1
            _excel_add_section(ws_f, r, "Aeronaves — Monitoramento")
            r += 1
            r = _excel_write_kv(ws_f, r, [
                ("Drone Monitoramento ID", ordem.drone_monitoramento_id or ""),
                ("Prefixo", ordem.prefixo_aeronave_monitoramento or ""),
                ("Denominação", ordem.drone_monitoramento_denominacao or ""),
                ("Modelo", ordem.drone_monitoramento_modelo or ""),
                ("Nº Série", ordem.drone_monitoramento_numero_serie or ""),
                ("Registro ANATEL", ordem.drone_monitoramento_registro_anatel or ""),
                ("Registro ANAC", ordem.drone_monitoramento_registro_anac or ""),
            ])

            r += 1
            _excel_add_section(ws_f, r, "Aplicação")
            r += 1
            r = _excel_write_kv(ws_f, r, [
                ("Situação da aplicação", ordem.situacao_aplicacao or ""),
                ("Larva visualizada", ordem.larva_visualizada or ""),
                ("Retornar monitorar larvas", ordem.retornar_proxima_semana_monitorar_larvas or ""),
                ("DA (Distrito)", ordem.distrito_administrativo or ""),
                ("Nome/RF ACE responsável", ordem.nome_rf_ace_responsavel_os or ""),
                ("Criadouro OS (tipo/volume)", ordem.criadouro_os_tipo_volume or ""),
                ("Data aplicação", _fmt_date(ordem.data_aplicacao)),
                ("Hora início", _fmt_time(ordem.hora_inicio_aplicacao)),
                ("Hora término", _fmt_time(ordem.hora_termino_aplicacao)),
                ("Tratamento adicional", ordem.tratamento_adicional_realizado or ""),
                ("Quantos / Quais", ordem.quantos_quais or ""),
            ])

            r += 1
            _excel_add_section(ws_f, r, "Produto e Parâmetros")
            r += 1
            r = _excel_write_kv(ws_f, r, [
                ("Descrição produto", ordem.descricao_produto or ""),
                ("Formulação", ordem.formulacao_produto or ""),
                ("Dosagem (g/10L)", ordem.dosagem_g_10l or ""),
                ("Tipo aplicação", ordem.tipo_aplicacao or ""),
                ("Qtd administrada (ml)", ordem.quantidade_produto_administrada_ml or ""),
                ("Pulverização área (l/ha)", ordem.pulverizacao_area_l_ha or ""),               
                ("Ponta pulverização", ordem.ponta_pulverizacao or ""),
            ])

            r += 1
            _excel_add_section(ws_f, r, "Condições Ambientais")
            r += 1
            r = _excel_write_kv(ws_f, r, [
                ("Imagens registradas", ordem.quantidade_imagens_registradas or ""),
                ("Vídeos registrados", ordem.quantidade_videos_registradas or ""),
                ("Temperatura (°C)", ordem.temperatura_c or ""),
                ("Umidade (%)", ordem.umidade_relativa_pct or ""),
                ("Vento (km/h)", ordem.velocidade_vento_kmh or ""),
            ])

            r += 1
            _excel_add_section(ws_f, r, "Fechamento")
            r += 1
            r = _excel_write_kv(ws_f, r, [
                ("Observações gerais", ordem.observacoes or ""),
                ("Motivo não realização", ordem.motivo_nao_realizacao or ""),
                ("Piloto", ordem.piloto or ""),
                ("Auxiliar", ordem.auxiliar or ""),
                ("Proprietário/Preposto", ordem.proprietario_ou_preposto or ""),
            ])

            #  Assinaturas (texto: OK / Não informada)
            r += 1
            _excel_add_section(ws_f, r, "Assinaturas")
            r += 1

            has_piloto = bool(ordem.assinatura_piloto and str(ordem.assinatura_piloto).startswith("data:image"))
            has_resp = bool(ordem.assinatura_proprietario_ou_preposto and str(ordem.assinatura_proprietario_ou_preposto).startswith("data:image"))

            r = _excel_write_kv(ws_f, r, [
                ("Assinatura piloto", "OK" if has_piloto else "Não informada"),
                ("Assinatura responsável", "OK" if has_resp else "Não informada"),
            ])

        ws_f.freeze_panes = "A5"
        _excel_auto_width(ws_f, max_col=2, min_w=18, max_w=75)

        # =========================
        # ABA 3 — Assinaturas (imagem)   (nova aba)
        # =========================
        want_sigs = request.args.get("assinaturas", "0") == "1"
        if want_sigs and ordem:
            ws_sig = wb.create_sheet("Assinaturas")
            _excel_add_title(ws_sig, f"Assinaturas — OS #{s.id}", "Imagens exportadas do formulário")

            ws_sig.column_dimensions["A"].width = 26
            ws_sig.column_dimensions["B"].width = 55

            ws_sig["A4"] = "Piloto"
            ws_sig["A4"].font = Font(bold=True)
            ws_sig["A5"] = ordem.piloto or ""

            ws_sig["A8"] = "Responsável"
            ws_sig["A8"].font = Font(bold=True)
            ws_sig["A9"] = ordem.proprietario_ou_preposto or ""

            # Piloto
            png1 = _dataurl_to_png_bytes(getattr(ordem, "assinatura_piloto", None))
            if png1:
                tmp1 = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                tmp1.write(png1)
                tmp1.close()

                img1 = XLImage(tmp1.name)
                img1.width = 420
                img1.height = 140
                ws_sig.add_image(img1, "B4")
            else:
                ws_sig["B4"] = "Assinatura não informada"

            # Responsável
            png2 = _dataurl_to_png_bytes(getattr(ordem, "assinatura_proprietario_ou_preposto", None))
            if png2:
                tmp2 = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                tmp2.write(png2)
                tmp2.close()

                img2 = XLImage(tmp2.name)
                img2.width = 420
                img2.height = 140
                ws_sig.add_image(img2, "B8")
            else:
                ws_sig["B8"] = "Assinatura não informada"

        # -------------------------
        # enviar em memória
        # -------------------------
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        nome_arquivo = f"os_{s.id}_formulario.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception:
        db.session.rollback()
        raise

    # ============================================================
# ✅ EXPORTAÇÃO RELATÓRIO OS — EXCEL + PDF (RESPEITA FILTROS)
# Rotas novas:
#   /relatorios-os/export/excel
#   /relatorios-os/export/pdf
# ============================================================

import tempfile
from datetime import datetime
from io import BytesIO

from flask import send_file, request
from flask_login import login_required, current_user

from sqlalchemy import and_, or_
from sqlalchemy.sql import func
from sqlalchemy.sql.expression import extract

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Seus imports reais:
# from app import db
# from app.models import Usuario, Solicitacao, OrdemServico
# from app.decorators import roles_required  (se quiser restringir)
# from app.routes import bp


# ------------------------------------------------------------
# helpers básicos (se já tiver, pode reaproveitar)
# ------------------------------------------------------------
def _fmt_dt(v):
    if not v:
        return ""
    try:
        return v.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(v)

def _safe(v):
    if v is None:
        return ""
    return str(v)


# ------------------------------------------------------------
# ✅ Excel helpers (bonito) — mesmo padrão do seu v2
# ------------------------------------------------------------
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_HEADER = PatternFill("solid", fgColor="0D6EFD")
FILL_SECTION = PatternFill("solid", fgColor="EAF2FF")
FILL_ZEBRA = PatternFill("solid", fgColor="FBFDFF")

FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_TITLE = Font(bold=True, size=16, color="0D6EFD")
FONT_SUBTITLE = Font(size=10, color="555555")
FONT_SECTION = Font(bold=True, color="0D6EFD")

def _excel_add_title(ws, title: str, subtitle: str = ""):
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    ws["A2"] = subtitle
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

def _excel_add_section(ws, row: int, title: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = FILL_SECTION
    c.font = FONT_SECTION
    c.alignment = Alignment(vertical="center")
    c.border = BORDER
    ws.cell(row=row, column=2).border = BORDER
    ws.row_dimensions[row].height = 18

def _excel_apply_table_style(ws, header_row: int, end_row: int, col_count: int = 2):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for r in range(header_row + 1, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if (r - header_row) % 2 == 0:
                cell.fill = FILL_ZEBRA

def _excel_write_kv(ws, start_row: int, items: list[tuple[str, object]]):
    ws.cell(row=start_row, column=1, value="Campo")
    ws.cell(row=start_row, column=2, value="Valor")

    r = start_row + 1
    for k, v in items:
        ws.cell(row=r, column=1, value=str(k))
        ws.cell(row=r, column=2, value=_safe(v))
        r += 1

    _excel_apply_table_style(ws, start_row, r - 1, col_count=2)
    return r

def _excel_write_table(ws, start_row: int, headers: list[str], rows: list[tuple], col_widths=None):
    # headers
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)

    # body
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=_safe(val))
        r += 1

    # style
    _excel_apply_table_style(ws, start_row, r - 1, col_count=len(headers))

    # widths
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return r

def _excel_auto_width(ws, max_col=2, min_w=12, max_w=60):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = 0
        for cell in ws[letter]:
            if cell.value:
                best = max(best, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, best + 2))


# ------------------------------------------------------------
# ✅ PDF helpers simples e bonitos
# ------------------------------------------------------------
def _pdf_header_footer_factory(title: str):
    def _hf(canvas, doc):
        canvas.saveState()
        w, h = doc.pagesize

        canvas.setFillColor(colors.HexColor("#0d6efd"))
        canvas.rect(doc.leftMargin, h - 12 * mm, doc.width, 3, fill=1, stroke=0)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#666"))
        canvas.drawString(doc.leftMargin, 9 * mm, title)
        canvas.drawRightString(doc.leftMargin + doc.width, 9 * mm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()
    return _hf

def _pdf_table(title: str, rows: list[list[str]], styles, col_widths=None):
    section = ParagraphStyle(
        "sec",
        parent=styles["Heading2"],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#0d6efd"),
        spaceBefore=10,
        spaceAfter=6
    )
    story = [Paragraph(title, section)]

    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9dee7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fbfdff")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 8))
    return story


# ------------------------------------------------------------
# ✅ Builder do relatório (MESMA LÓGICA do relatorios_os)
# ------------------------------------------------------------
def _build_relatorio_os_data():
    """
    Lê filtros (mes/ano/uvis_id) e retorna dicionário com tudo:
    totals + agrupamentos (situação/tipo/larva/piloto/unidade/mensal) + labels do filtro.
    """
    mes_atual = request.args.get("mes", datetime.now().month, type=int)
    ano_atual = request.args.get("ano", datetime.now().year, type=int)

    # regra igual a sua:
    uvis_id = request.args.get("uvis_id", type=int) if current_user.tipo_usuario != "uvis" else current_user.id

    base_query = (
        db.session.query(OrdemServico)
        .join(Solicitacao, Solicitacao.id == OrdemServico.solicitacao_id)
        .join(Usuario, Usuario.id == Solicitacao.usuario_id)
    )

    base_query = base_query.filter(
        or_(
            and_(
                OrdemServico.respondido_em.isnot(None),
                extract("year", OrdemServico.respondido_em) == ano_atual,
                extract("month", OrdemServico.respondido_em) == mes_atual
            ),
            and_(
                OrdemServico.respondido_em.is_(None),
                OrdemServico.data_aplicacao.isnot(None),
                extract("year", OrdemServico.data_aplicacao) == ano_atual,
                extract("month", OrdemServico.data_aplicacao) == mes_atual
            )
        )
    )

    if uvis_id:
        base_query = base_query.filter(Solicitacao.usuario_id == uvis_id)

    total_os = base_query.count()
    total_concluidas = base_query.filter(Solicitacao.status.in_(["CONCLUÍDO", "CONCLUIDO"])).count()
    total_larva_sim = base_query.filter(func.upper(func.coalesce(OrdemServico.larva_visualizada, "")) == "SIM").count()
    total_tratamento_adicional = base_query.filter(func.upper(func.coalesce(OrdemServico.tratamento_adicional_realizado, "")) == "SIM").count()
    total_nao_realizadas = base_query.filter(func.length(func.trim(func.coalesce(OrdemServico.motivo_nao_realizacao, ""))) > 0).count()

    def agrupar_por(campo):
        return [
            (valor or "Não informado", total)
            for valor, total in (
                base_query
                .with_entities(campo, func.count(OrdemServico.id))
                .group_by(campo)
                .order_by(func.count(OrdemServico.id).desc())
                .all()
            )
        ]

    dados_situacao_aplicacao = agrupar_por(OrdemServico.situacao_aplicacao)
    dados_tipo_aplicacao = agrupar_por(OrdemServico.tipo_aplicacao)
    dados_larva = agrupar_por(OrdemServico.larva_visualizada)
    dados_piloto = agrupar_por(OrdemServico.piloto)

    dados_unidade = [
        (uvis or "Não informado", total)
        for uvis, total in (
            base_query
            .with_entities(Usuario.nome_uvis, func.count(OrdemServico.id))
            .group_by(Usuario.nome_uvis)
            .order_by(func.count(OrdemServico.id).desc())
            .all()
        )
    ]

    mensal_query = (
        db.session.query(
            func.coalesce(
                extract("year", OrdemServico.respondido_em),
                extract("year", OrdemServico.data_aplicacao)
            ).label("ano_ref"),
            func.coalesce(
                extract("month", OrdemServico.respondido_em),
                extract("month", OrdemServico.data_aplicacao)
            ).label("mes_ref"),
            func.count(OrdemServico.id)
        )
        .join(Solicitacao, Solicitacao.id == OrdemServico.solicitacao_id)
        .join(Usuario, Usuario.id == Solicitacao.usuario_id)
    )

    if uvis_id:
        mensal_query = mensal_query.filter(Solicitacao.usuario_id == uvis_id)

    dados_mensais = [
        (f"{int(ano_h):04d}-{int(mes_h):02d}", total)
        for ano_h, mes_h, total in (
            mensal_query
            .filter(or_(OrdemServico.respondido_em.isnot(None), OrdemServico.data_aplicacao.isnot(None)))
            .group_by("ano_ref", "mes_ref")
            .order_by("ano_ref", "mes_ref")
            .all()
        )
        if ano_h and mes_h
    ]

    # nome da UVIS selecionada (só pra título)
    nome_uvis = None
    if uvis_id:
        nome_uvis = db.session.query(Usuario.nome_uvis).filter(Usuario.id == uvis_id).scalar()

    return {
        "mes": mes_atual,
        "ano": ano_atual,
        "uvis_id": uvis_id,
        "uvis_nome": nome_uvis or "Todas as Unidades",

        "total_os": total_os,
        "total_concluidas": total_concluidas,
        "total_larva_sim": total_larva_sim,
        "total_tratamento_adicional": total_tratamento_adicional,
        "total_nao_realizadas": total_nao_realizadas,

        "dados_situacao_aplicacao": dados_situacao_aplicacao,
        "dados_tipo_aplicacao": dados_tipo_aplicacao,
        "dados_larva": dados_larva,
        "dados_piloto": dados_piloto,
        "dados_unidade": dados_unidade,
        "dados_mensais": dados_mensais,
    }


# ============================================================
# ✅ ROTA NOVA — EXCEL
# ============================================================
@bp.route("/relatorios-os/export/excel", methods=["GET"])
@login_required
def relatorios_os_export_excel():
    data = _build_relatorio_os_data()

    wb = Workbook()

    # -------------------------
    # Aba 1 — Resumo
    # -------------------------
    ws = wb.active
    ws.title = "Resumo"

    _excel_add_title(
        ws,
        "Relatório Geral de OS",
        f"Filtro: {data['mes']:02d}/{data['ano']} | Unidade: {data['uvis_nome']} | Gerado em {_fmt_dt(datetime.now())}"
    )

    r = 4
    _excel_add_section(ws, r, "Indicadores")
    r += 1
    r = _excel_write_kv(ws, r, [
        ("Total OS", data["total_os"]),
        ("Concluídas", data["total_concluidas"]),
        ("Larva (SIM)", data["total_larva_sim"]),
        ("Tratamento adicional", data["total_tratamento_adicional"]),
        ("Não realizadas", data["total_nao_realizadas"]),
    ])

    ws.freeze_panes = "A5"
    _excel_auto_width(ws, max_col=2, min_w=18, max_w=70)

    # -------------------------
    # Aba 2 — Agrupamentos
    # -------------------------
    ws2 = wb.create_sheet("Detalhamento")
    _excel_add_title(ws2, "Detalhamento do Relatório", "Agrupamentos por campos")

    r = 4
    _excel_add_section(ws2, r, "Situação da Aplicação")
    r += 1
    r = _excel_write_table(ws2, r, ["Situação", "Total"], data["dados_situacao_aplicacao"], col_widths=[45, 12])

    r += 1
    _excel_add_section(ws2, r, "Tipo de Aplicação")
    r += 1
    r = _excel_write_table(ws2, r, ["Tipo", "Total"], data["dados_tipo_aplicacao"], col_widths=[45, 12])

    r += 1
    _excel_add_section(ws2, r, "Larva Visualizada")
    r += 1
    r = _excel_write_table(ws2, r, ["Resposta", "Total"], data["dados_larva"], col_widths=[45, 12])

    r += 1
    _excel_add_section(ws2, r, "Pilotos (Top 10)")
    r += 1
    r = _excel_write_table(ws2, r, ["Piloto", "Total"], data["dados_piloto"][:10], col_widths=[45, 12])

    r += 1
    _excel_add_section(ws2, r, "OS por Unidade")
    r += 1
    r = _excel_write_table(ws2, r, ["Unidade (UVIS)", "Total"], data["dados_unidade"], col_widths=[45, 12])

    r += 1
    _excel_add_section(ws2, r, "Histórico Mensal")
    r += 1
    r = _excel_write_table(ws2, r, ["Mês", "Total"], data["dados_mensais"], col_widths=[18, 12])

    ws2.freeze_panes = "A5"

    # -------------------------
    # enviar
    # -------------------------
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    nome = f"relatorio_os_{data['ano']}_{data['mes']:02d}"
    if data["uvis_id"]:
        nome += f"_uvis_{data['uvis_id']}"
    nome += ".xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=nome,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ============================================================
# ✅ ROTA NOVA — PDF
# ============================================================
@bp.route("/relatorios-os/export/pdf", methods=["GET"])
@login_required
def relatorios_os_export_pdf():
    data = _build_relatorio_os_data()

    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    path = tmp_pdf.name
    tmp_pdf.close()

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "title",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        alignment=1,
        textColor=colors.HexColor("#0d6efd"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "sub",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        alignment=1,
        textColor=colors.HexColor("#555"),
        spaceAfter=12,
    )

    story = []
    story.append(Paragraph("Relatório Geral de OS", title_style))
    story.append(Paragraph(
        f"Filtro: {data['mes']:02d}/{data['ano']} | Unidade: {data['uvis_nome']} | Gerado em {_fmt_dt(datetime.now())}",
        subtitle_style
    ))

    # Indicadores
    story += _pdf_table(
        "Indicadores",
        rows=[
            ["Indicador", "Total"],
            ["Total OS", str(data["total_os"])],
            ["Concluídas", str(data["total_concluidas"])],
            ["Larva (SIM)", str(data["total_larva_sim"])],
            ["Tratamento adicional", str(data["total_tratamento_adicional"])],
            ["Não realizadas", str(data["total_nao_realizadas"])],
        ],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )

    # Agrupamentos (Top para não ficar gigante)
    story += _pdf_table(
        "Situação da Aplicação",
        rows=[["Situação", "Total"]] + [[a, str(b)] for a, b in data["dados_situacao_aplicacao"]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )

    story += _pdf_table(
        "Tipo de Aplicação",
        rows=[["Tipo", "Total"]] + [[a, str(b)] for a, b in data["dados_tipo_aplicacao"]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )

    story += _pdf_table(
        "Larva Visualizada",
        rows=[["Resposta", "Total"]] + [[a, str(b)] for a, b in data["dados_larva"]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )

    story += _pdf_table(
        "Pilotos (Top 10)",
        rows=[["Piloto", "Total"]] + [[a, str(b)] for a, b in data["dados_piloto"][:10]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )

    story += _pdf_table(
        "OS por Unidade",
        rows=[["Unidade (UVIS)", "Total"]] + [[a, str(b)] for a, b in data["dados_unidade"]],
        styles=styles,
        col_widths=[120 * mm, 50 * mm],
    )

    story += _pdf_table(
        "Histórico Mensal",
        rows=[["Mês", "Total"]] + [[a, str(b)] for a, b in data["dados_mensais"]],
        styles=styles,
        col_widths=[60 * mm, 110 * mm],
    )

    header_title = f"Relatório OS — {data['mes']:02d}/{data['ano']}"
    doc.build(story, onFirstPage=_pdf_header_footer_factory(header_title), onLaterPages=_pdf_header_footer_factory(header_title))

    nome = f"relatorio_os_{data['ano']}_{data['mes']:02d}"
    if data["uvis_id"]:
        nome += f"_uvis_{data['uvis_id']}"
    nome += ".pdf"

    return send_file(
        path,
        as_attachment=True,
        download_name=nome,
        mimetype="application/pdf"
    )



CHECKLIST_VEICULO_BOOL_FIELDS = [
    "farois_funcionando",
    "setas_funcionando",
    "lanternas_funcionando",
    "piscaalerta_funcionando",
    "luz_painel",
    "limpador_parabrisa",
    "agua_radiador",
    "fluido_freio",
    "oleo_motor",
    "vidros",
    "retrovisores",
    "pneus",
    "estepe",
    "macaco",
    "triangulo",
    "chave_roda",
    "extintor",
    "cinto_seguranca",
    "alarme",
    "ar_condicionado",
    "radio",
    "giroflex",
    "isqueiro",
    "carregador",
    "lataria_frontal",
    "lataria_lateral",
    "lataria_traseira",
    "lataria_porta_frontal",
    "lataria_porta_traseira",
    "lataria_porta_lateral",
    "parachoque_frontal",
    "parachoque_traseiro",
]

CHECKLIST_VEICULO_TEXT_FIELDS = [
    "condicao_luzes_direcao",
    "condicao_luz_painel",
    "condicao_itens_manutencao",
    "condicao_vidros_retrovisores",
    "condicao_pneus_estepe",
    "condicao_itens_seguranca",
    "condicao_itens_carro_interno",
    "condicao_giroflex_isqueiro_carregador",
    "condicao_lataria",
    "condicao_lataria_portas",
    "condicao_itens_carro_externo",
]

CHECKLIST_DRONE_BOOL_FIELDS = [
    "helices_status",
    "tanque",
    "trem_pouso",
    "cameras",
    "carregador_controle",
    "baterias",
    "cabos_carregador",
    "correia_pescoco",
]

CHECKLIST_DRONE_TEXT_FIELDS = [
    "condicao_helices",
    "condicao_estrutura",
    "condicao_carregador_bateria",
    "condicao_cabos_correia",
    "observacoes_equipamento",
]

CHECKLIST_VEICULO_BOOL_LABELS = [
    ("farois_funcionando", "Farois"),
    ("setas_funcionando", "Setas"),
    ("lanternas_funcionando", "Lanternas"),
    ("piscaalerta_funcionando", "Pisca-alerta"),
    ("luz_painel", "Luz do painel"),
    ("limpador_parabrisa", "Limpador de parabrisa"),
    ("agua_radiador", "Agua do radiador"),
    ("fluido_freio", "Fluido de freio"),
    ("oleo_motor", "Oleo do motor"),
    ("vidros", "Vidros"),
    ("retrovisores", "Retrovisores"),
    ("pneus", "Pneus"),
    ("estepe", "Estepe"),
    ("macaco", "Macaco"),
    ("triangulo", "Triangulo"),
    ("chave_roda", "Chave de roda"),
    ("extintor", "Extintor"),
    ("cinto_seguranca", "Cinto de seguranca"),
    ("alarme", "Alarme"),
    ("ar_condicionado", "Ar-condicionado"),
    ("radio", "Radio"),
    ("giroflex", "Giroflex"),
    ("isqueiro", "Isqueiro"),
    ("carregador", "Carregador"),
    ("lataria_frontal", "Lataria frontal"),
    ("lataria_lateral", "Lataria lateral"),
    ("lataria_traseira", "Lataria traseira"),
    ("lataria_porta_frontal", "Porta frontal"),
    ("lataria_porta_traseira", "Porta traseira"),
    ("lataria_porta_lateral", "Porta lateral"),
    ("parachoque_frontal", "Parachoque frontal"),
    ("parachoque_traseiro", "Parachoque traseiro"),
]

CHECKLIST_VEICULO_TEXT_LABELS = [
    ("condicao_luzes_direcao", "Condicao luzes / direcao"),
    ("condicao_luz_painel", "Condicao luz do painel"),
    ("condicao_itens_manutencao", "Condicao manutencao preventiva"),
    ("condicao_vidros_retrovisores", "Condicao vidros / retrovisores"),
    ("condicao_pneus_estepe", "Condicao pneus / estepe"),
    ("condicao_itens_seguranca", "Condicao itens de seguranca"),
    ("condicao_itens_carro_interno", "Condicao itens internos"),
    ("condicao_giroflex_isqueiro_carregador", "Condicao giroflex / isqueiro / carregador"),
    ("condicao_lataria", "Condicao lataria"),
    ("condicao_lataria_portas", "Condicao lataria portas"),
    ("condicao_itens_carro_externo", "Condicao itens externos"),
]

CHECKLIST_DRONE_BOOL_LABELS = [
    ("helices_status", "Helices"),
    ("tanque", "Tanque"),
    ("trem_pouso", "Trem de pouso"),
    ("cameras", "Cameras"),
    ("carregador_controle", "Carregador do controle"),
    ("baterias", "Baterias"),
    ("cabos_carregador", "Cabos do carregador"),
    ("correia_pescoco", "Correia de pescoco"),
]

CHECKLIST_DRONE_TEXT_LABELS = [
    ("condicao_helices", "Condicao helices"),
    ("condicao_estrutura", "Condicao estrutura"),
    ("condicao_carregador_bateria", "Condicao carregador / bateria"),
    ("condicao_cabos_correia", "Condicao cabos / correia"),
    ("observacoes_equipamento", "Observacoes do equipamento"),
]


def _format_km_admin(value):
    try:
        return f"{float(value or 0):.0f} km"
    except Exception:
        return "-"


def _checklist_status_items(checklist, labels):
    itens = []
    falhas = 0

    for field, label in labels:
        ok = bool(getattr(checklist, field))
        if not ok:
            falhas += 1
        itens.append({"label": label, "ok": ok})

    return itens, falhas


def _checklist_notes_items(checklist, labels):
    observacoes = []

    for field, label in labels:
        valor = _clean_str(getattr(checklist, field))
        if valor:
            observacoes.append({"label": label, "value": valor})

    return observacoes


def _campos_defeituosos_checklist(checklist, labels):
    defeitos = []
    for field, label in labels:
        if not bool(getattr(checklist, field)):
            defeitos.append(label)
    return defeitos


def _identificacao_checklist_veiculo(checklist):
    veiculo = checklist.veiculo
    if veiculo and veiculo.placa:
        return veiculo.placa
    if veiculo and veiculo.modelo:
        return veiculo.modelo
    return f"ID {checklist.veiculo_id}"


def _identificacao_checklist_drone(checklist):
    drone = checklist.drone
    if drone and drone.renomacao:
        return drone.renomacao
    if drone and drone.modelo:
        return drone.modelo
    return f"ID {checklist.drone_id}"


def _coletar_pendencias_checklists_semanais(piloto_id, inicio_semana_dt, proxima_semana_dt):
    pendencias = []

    checklists_veiculo = (
        ChecklistSemanalVeiculo.query
        .options(joinedload(ChecklistSemanalVeiculo.veiculo))
        .filter(
            ChecklistSemanalVeiculo.piloto_id == piloto_id,
            ChecklistSemanalVeiculo.data_registro >= inicio_semana_dt,
            ChecklistSemanalVeiculo.data_registro < proxima_semana_dt,
        )
        .order_by(ChecklistSemanalVeiculo.data_registro.desc())
        .all()
    )
    for checklist in checklists_veiculo:
        defeitos = _campos_defeituosos_checklist(checklist, CHECKLIST_VEICULO_BOOL_LABELS)
        if defeitos:
            pendencias.append(
                f"Veiculo {_identificacao_checklist_veiculo(checklist)}: {', '.join(defeitos)}"
            )

    checklists_drone = (
        ChecklistSemanalDrone.query
        .options(joinedload(ChecklistSemanalDrone.drone))
        .filter(
            ChecklistSemanalDrone.piloto_id == piloto_id,
            ChecklistSemanalDrone.data_registro >= inicio_semana_dt,
            ChecklistSemanalDrone.data_registro < proxima_semana_dt,
        )
        .order_by(ChecklistSemanalDrone.data_registro.desc())
        .all()
    )
    for checklist in checklists_drone:
        defeitos = _campos_defeituosos_checklist(checklist, CHECKLIST_DRONE_BOOL_LABELS)
        if defeitos:
            pendencias.append(
                f"Drone {_identificacao_checklist_drone(checklist)}: {', '.join(defeitos)}"
            )

    return pendencias


def _sincronizar_notificacoes_pendencia_checklist(admin_ids, link, titulo, mensagem=None):
    if not admin_ids or not link:
        return

    existentes = {}
    for notif in (
        Notificacao.query
        .filter(
            Notificacao.usuario_id.in_(admin_ids),
            Notificacao.link == link,
        )
        .order_by(Notificacao.id.desc())
        .all()
    ):
        if notif.usuario_id not in existentes:
            existentes[notif.usuario_id] = notif

    if mensagem:
        agora = agora_brasilia_naive()
        for admin_id in admin_ids:
            notif = existentes.get(admin_id)
            if notif:
                notif.titulo = titulo
                notif.mensagem = mensagem
                notif.criada_em = agora
                notif.lida_em = None
                notif.apagada_em = None
            else:
                criar_notificacao(
                    usuario_id=admin_id,
                    titulo=titulo,
                    mensagem=mensagem,
                    link=link,
                    commit=False,
                )
        return

    agora = agora_brasilia_naive()
    for notif in existentes.values():
        if notif.apagada_em is None:
            notif.apagada_em = agora


def _normalize_checklist_veiculo_admin(checklist):
    itens, falhas = _checklist_status_items(checklist, CHECKLIST_VEICULO_BOOL_LABELS)
    observacoes = _checklist_notes_items(checklist, CHECKLIST_VEICULO_TEXT_LABELS)
    veiculo = checklist.veiculo
    piloto = checklist.piloto

    meta = [
        {"label": "Placa", "value": (veiculo.placa if veiculo else "") or "-"},
        {"label": "Operacao", "value": (veiculo.operacao if veiculo else "") or "-"},
        {"label": "Responsavel", "value": (veiculo.responsavel if veiculo else "") or "-"},
        {"label": "KM lido", "value": _format_km_admin(checklist.km_leitura)},
    ]

    total_itens = len(CHECKLIST_VEICULO_BOOL_LABELS)
    return {
        "id": checklist.id,
        "tipo": "veiculo",
        "tipo_label": "Veiculo",
        "data_registro": checklist.data_registro,
        "piloto_id": checklist.piloto_id,
        "titulo": (veiculo.modelo if veiculo else "") or "Veiculo sem identificacao",
        "subtitulo": (veiculo.placa if veiculo else "") or "-",
        "complemento": (veiculo.responsavel if veiculo else "") or "",
        "piloto_nome": (piloto.nome_piloto if piloto else "") or "-",
        "status_label": "Conforme" if falhas == 0 else f"{falhas} pendencia(s)",
        "status_class": "success" if falhas == 0 else "warning",
        "falhas": falhas,
        "itens_ok": total_itens - falhas,
        "itens_total": total_itens,
        "observacoes_total": len(observacoes),
        "meta": meta,
        "detalhes_itens": itens,
        "observacoes": observacoes,
        "assinatura": checklist.assinatura_piloto or "",
    }


def _normalize_checklist_drone_admin(checklist):
    itens, falhas = _checklist_status_items(checklist, CHECKLIST_DRONE_BOOL_LABELS)
    observacoes = _checklist_notes_items(checklist, CHECKLIST_DRONE_TEXT_LABELS)
    drone = checklist.drone
    piloto = checklist.piloto

    meta = [
        {"label": "Renomacao", "value": (drone.renomacao if drone else "") or "-"},
        {"label": "Modelo", "value": (drone.modelo if drone else "") or "-"},
        {"label": "Serie", "value": (drone.numero_serie if drone else "") or "-"},
        {"label": "Baterias", "value": str(int(checklist.num_baterias or 0))},
        {"label": "Baterias WB", "value": str(int(checklist.num_baterias_wb or 0))},
        {"label": "Responsavel informado", "value": checklist.nome_responsavel or "-"},
    ]

    total_itens = len(CHECKLIST_DRONE_BOOL_LABELS)
    return {
        "id": checklist.id,
        "tipo": "drone",
        "tipo_label": "Drone",
        "data_registro": checklist.data_registro,
        "piloto_id": checklist.piloto_id,
        "titulo": (drone.renomacao if drone else "") or "Drone sem identificacao",
        "subtitulo": (drone.modelo if drone else "") or "-",
        "complemento": (drone.numero_serie if drone else "") or "",
        "piloto_nome": (piloto.nome_piloto if piloto else "") or "-",
        "status_label": "Conforme" if falhas == 0 else f"{falhas} pendencia(s)",
        "status_class": "success" if falhas == 0 else "warning",
        "falhas": falhas,
        "itens_ok": total_itens - falhas,
        "itens_total": total_itens,
        "observacoes_total": len(observacoes),
        "meta": meta,
        "detalhes_itens": itens,
        "observacoes": observacoes,
        "assinatura": checklist.assinatura_piloto or "",
    }


def _group_admin_checklists_by_week(registros):
    grupos = {}

    for item in registros:
        data_registro = item.get("data_registro")
        if not data_registro:
            continue

        data_base = data_registro.date()
        semana_inicio = data_base - timedelta(days=data_base.weekday())
        semana_fim = semana_inicio + timedelta(days=6)
        grupo_key = (item.get("piloto_id"), semana_inicio.isoformat())

        grupo = grupos.setdefault(
            grupo_key,
            {
                "piloto_id": item.get("piloto_id"),
                "piloto_nome": item.get("piloto_nome") or "-",
                "semana_inicio": semana_inicio,
                "semana_fim": semana_fim,
                "ultima_movimentacao": data_registro,
                "veiculos": [],
                "drones": [],
                "falhas": 0,
                "itens_ok": 0,
                "itens_total": 0,
                "observacoes_total": 0,
            },
        )

        if data_registro > (grupo["ultima_movimentacao"] or datetime.min):
            grupo["ultima_movimentacao"] = data_registro

        if item.get("tipo") == "veiculo":
            grupo["veiculos"].append(item)
        elif item.get("tipo") == "drone":
            grupo["drones"].append(item)

        grupo["falhas"] += int(item.get("falhas") or 0)
        grupo["itens_ok"] += int(item.get("itens_ok") or 0)
        grupo["itens_total"] += int(item.get("itens_total") or 0)
        grupo["observacoes_total"] += int(item.get("observacoes_total") or 0)

    grupos_lista = []
    for grupo in grupos.values():
        grupo["veiculos"].sort(key=lambda item: item.get("data_registro") or datetime.min, reverse=True)
        grupo["drones"].sort(key=lambda item: item.get("data_registro") or datetime.min, reverse=True)
        grupo["status_label"] = "Conforme" if grupo["falhas"] == 0 else f"{grupo['falhas']} pendencia(s)"
        grupo["resumo_label"] = f"{len(grupo['veiculos'])} veiculo(s) • {len(grupo['drones'])} drone(s)"
        grupos_lista.append(grupo)

    grupos_lista.sort(key=lambda item: item["ultima_movimentacao"] or datetime.min, reverse=True)
    return grupos_lista


def _piloto_vinculo_ativo():
    if not getattr(current_user, "piloto_id", None):
        return None

    return (
        EquipePiloto.query
        .join(Equipe, Equipe.id == EquipePiloto.equipe_id)
        .options(joinedload(EquipePiloto.equipe))
        .filter(
            EquipePiloto.piloto_id == current_user.piloto_id,
            Equipe.ativa.is_(True)
        )
        .order_by(
            db.case((EquipePiloto.papel == "piloto", 0), else_=1),
            EquipePiloto.criado_em.desc()
        )
        .first()
    )


def _bool_from_form(value, default=True):
    value = _clean(value)
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "on", "sim", "ok", "bom"}


def _serialize_checklist_veiculo(checklist):
    data = {}
    for field in CHECKLIST_VEICULO_BOOL_FIELDS + CHECKLIST_VEICULO_TEXT_FIELDS:
        data[field] = getattr(checklist, field)
    data["km_leitura"] = checklist.km_leitura
    data["assinatura_piloto"] = checklist.assinatura_piloto or ""
    return data


def _serialize_checklist_drone(checklist):
    data = {}
    for field in CHECKLIST_DRONE_BOOL_FIELDS + CHECKLIST_DRONE_TEXT_FIELDS:
        data[field] = getattr(checklist, field)
    data["num_baterias"] = checklist.num_baterias
    data["num_baterias_wb"] = checklist.num_baterias_wb
    data["assinatura_piloto"] = checklist.assinatura_piloto or ""
    data["nome_responsavel"] = checklist.nome_responsavel or ""
    return data


@bp.route("/piloto/checklists/semanais", methods=["GET", "POST"], endpoint="piloto_checklist_semanal")
@login_required
@roles_required("piloto")
def piloto_checklist_semanal():
    # Bloco principal para capturar erros de transação no GET
    try:
        vinculo = _piloto_vinculo_ativo()
        if not vinculo or not vinculo.equipe_id:
            flash("Você ainda não está vinculado a nenhuma equipe ativa.", "warning")
            return redirect(url_for("main.piloto_os"))

        equipe = vinculo.equipe
        piloto_nome = (
            current_user.piloto.nome_piloto
            if getattr(current_user, "piloto", None) and current_user.piloto
            else (getattr(current_user, "nome_uvis", "") or "")
        )

        # Busca de Veículos
        veiculos_equipe = (
            Veiculos.query
            .filter(
                db.or_(
                    Veiculos.equipe_id == equipe.id,
                    db.func.lower(Veiculos.responsavel) == piloto_nome.lower()
                )
            )
            .distinct()
            .order_by(Veiculos.operacao.asc(), Veiculos.modelo.asc(), Veiculos.placa.asc())
            .all()
        )

        # Busca de Drones
        drones_equipe = (
            Drones.query
            .filter(
                Drones.equipe_id == equipe.id,
                Drones.status == "Ativo"
            )
            .order_by(Drones.renomacao.asc())
            .all()
        )

        veiculo_ids = [item.id for item in veiculos_equipe]
        drone_ids = [item.id for item in drones_equipe]

        # Metadados para o Frontend
        veiculo_meta = {
            str(item.id): {
                "id": item.id,
                "label": f"{item.modelo} - {item.placa}",
                "km_atual": float(item.km_atual or 0),
                "operacao": item.operacao or "",
                "responsavel": item.responsavel or "",
            }
            for item in veiculos_equipe
        }

        baterias_por_drone = {}
        if drone_ids:
            baterias_por_drone = {
                int(drone_id): total
                for drone_id, total in (
                    db.session.query(Baterias.drone_id, db.func.count(Baterias.id))
                    .filter(Baterias.drone_id.in_(drone_ids))
                    .group_by(Baterias.drone_id)
                    .all()
                )
            }

        drone_meta = {
            str(item.id): {
                "id": item.id,
                "label": f"{item.renomacao} - {item.modelo}",
                "renomacao": item.renomacao or "",
                "modelo": item.modelo or "",
                "numero_serie": item.numero_serie or "",
                "registro_anatel": item.registro_anatel or "",
                "registro_anac": item.registro_anac or "",
                "num_baterias": int(baterias_por_drone.get(item.id, 0) or 0),
            }
            for item in drones_equipe
        }

        # Preenchimento automático (Prefill) - Onde o erro costuma ocorrer
        veiculo_prefill = {}
        if veiculo_ids:
            ultimos_veiculos = (
                ChecklistSemanalVeiculo.query
                .filter(ChecklistSemanalVeiculo.veiculo_id.in_(veiculo_ids))
                .order_by(
                    ChecklistSemanalVeiculo.veiculo_id.asc(),
                    ChecklistSemanalVeiculo.data_registro.desc()
                )
                .all()
            )
            for item in ultimos_veiculos:
                key = str(item.veiculo_id)
                if key not in veiculo_prefill:
                    veiculo_prefill[key] = _serialize_checklist_veiculo(item)

        drone_prefill = {}
        if drone_ids:
            ultimos_drones = (
                ChecklistSemanalDrone.query
                .filter(ChecklistSemanalDrone.drone_id.in_(drone_ids))
                .order_by(
                    ChecklistSemanalDrone.drone_id.asc(),
                    ChecklistSemanalDrone.data_registro.desc()
                )
                .all()
            )
            for item in ultimos_drones:
                key = str(item.drone_id)
                if key not in drone_prefill:
                    drone_prefill[key] = _serialize_checklist_drone(item)

    except Exception as e:
        db.session.rollback() # Limpa a transação para o inject_globals não quebrar
        current_app.logger.error(f"Erro no carregamento dos checklists (GET): {str(e)}")
        flash("Erro interno ao carregar dados do checklist. Tente novamente.", "danger")
        return redirect(url_for("main.piloto_os"))

    hoje = date.today()
    inicio_semana = hoje - timedelta(days=hoje.weekday())
    fim_semana = inicio_semana + timedelta(days=6)
    inicio_semana_dt = datetime.combine(inicio_semana, datetime.min.time())
    proxima_semana_dt = inicio_semana_dt + timedelta(days=7)

    veiculo_padrao_id = request.args.get("veiculo_id", type=int)
    if veiculo_padrao_id not in veiculo_ids:
        veiculo_padrao_id = veiculos_equipe[0].id if len(veiculos_equipe) == 1 else None

    drone_padrao_id = request.args.get("drone_id", type=int)
    if drone_padrao_id not in drone_ids:
        drone_padrao_id = drones_equipe[0].id if len(drones_equipe) == 1 else None

    # --- LÓGICA DO POST ---
    if request.method == "POST":
        veiculo_id = request.form.get("veiculo_id", type=int)
        drone_id = request.form.get("drone_id", type=int)
        assinatura_piloto = _clean_str(request.form.get("assinatura_piloto"))
        nome_responsavel = _clean_str(request.form.get("nome_responsavel")) or piloto_nome

        # Validações Básicas
        if veiculos_equipe and veiculo_id and veiculo_id not in veiculo_ids:
            flash("Selecione um veículo válido da sua equipe.", "warning")
            return redirect(url_for("main.piloto_checklist_semanal"))

        if drones_equipe and drone_id and drone_id not in drone_ids:
            flash("Selecione um drone válido da sua equipe.", "warning")
            return redirect(url_for("main.piloto_checklist_semanal"))

        if not veiculo_id and not drone_id:
            flash("Selecione ao menos um veículo ou um drone para registrar o checklist.", "warning")
            return redirect(url_for("main.piloto_checklist_semanal"))

        if not assinatura_piloto: # Se o nome da var no seu form for assinatura_piloto, use ela
             if not assinatura_piloto:
                flash("A assinatura do piloto é obrigatória.", "warning")
                return redirect(url_for("main.piloto_checklist_semanal"))

        try:
            # Salvando Veículo
            if veiculo_id:
                veiculo = next((item for item in veiculos_equipe if item.id == veiculo_id), None)
                checklist_veiculo = (
                    ChecklistSemanalVeiculo.query
                    .filter(
                        ChecklistSemanalVeiculo.veiculo_id == veiculo_id,
                        ChecklistSemanalVeiculo.piloto_id == current_user.piloto_id,
                        ChecklistSemanalVeiculo.data_registro >= inicio_semana_dt,
                        ChecklistSemanalVeiculo.data_registro < proxima_semana_dt,
                    )
                    .first()
                )
                if not checklist_veiculo:
                    checklist_veiculo = ChecklistSemanalVeiculo(veiculo_id=veiculo_id, piloto_id=current_user.piloto_id)
                    db.session.add(checklist_veiculo)

                checklist_veiculo.data_registro = datetime.now()
                checklist_veiculo.km_leitura = float(veiculo.km_atual or 0)
                
                for field in CHECKLIST_VEICULO_BOOL_FIELDS:
                    setattr(checklist_veiculo, field, _bool_from_form(request.form.get(field), default=True))
                for field in CHECKLIST_VEICULO_TEXT_FIELDS:
                    setattr(checklist_veiculo, field, _clean_str(request.form.get(field)))
                
                checklist_veiculo.assinatura_piloto = assinatura_piloto

            # Salvando Drone
            if drone_id:
                checklist_drone = (
                    ChecklistSemanalDrone.query
                    .filter(
                        ChecklistSemanalDrone.drone_id == drone_id,
                        ChecklistSemanalDrone.piloto_id == current_user.piloto_id,
                        ChecklistSemanalDrone.data_registro >= inicio_semana_dt,
                        ChecklistSemanalDrone.data_registro < proxima_semana_dt,
                    )
                    .first()
                )
                if not checklist_drone:
                    checklist_drone = ChecklistSemanalDrone(drone_id=drone_id, piloto_id=current_user.piloto_id)
                    db.session.add(checklist_drone)

                checklist_drone.data_registro = datetime.now()
                
                for field in CHECKLIST_DRONE_BOOL_FIELDS:
                    setattr(checklist_drone, field, _bool_from_form(request.form.get(field), default=True))
                for field in CHECKLIST_DRONE_TEXT_FIELDS:
                    setattr(checklist_drone, field, _clean_str(request.form.get(field)))

                # Lógica de Baterias
                default_baterias = baterias_por_drone.get(drone_id, 0)
                checklist_drone.num_baterias = _to_int(request.form.get("num_baterias")) or int(default_baterias or 0)
                checklist_drone.num_baterias_wb = _to_int(request.form.get("num_baterias_wb")) or 0

                checklist_drone.assinatura_piloto = assinatura_piloto
                checklist_drone.nome_responsavel = nome_responsavel
                checklist_drone.assinatura_responsavel = assinatura_piloto

            db.session.flush()

            pendencias_semanais = _coletar_pendencias_checklists_semanais(
                current_user.piloto_id,
                inicio_semana_dt,
                proxima_semana_dt,
            )

            detalhe_link = url_for(
                "main.admin_checklist_semanal_detalhe",
                piloto_id=current_user.piloto_id,
                semana_inicio=inicio_semana.isoformat(),
            )
            titulo_notificacao = f"Pendências no checklist semanal de {piloto_nome}"

            admin_ids = [
                row[0]
                for row in (
                    db.session.query(Usuario.id)
                    .filter(Usuario.tipo_usuario == "admin")
                    .all()
                )
            ]

            _sincronizar_notificacoes_pendencia_checklist(
                admin_ids=admin_ids,
                link=detalhe_link,
                titulo=titulo_notificacao,
                mensagem=" | ".join(pendencias_semanais) if pendencias_semanais else None,
            )

            db.session.commit()

            if pendencias_semanais:
                flash(
                    "Checklist salvo com pendências: " + " | ".join(pendencias_semanais),
                    "warning",
                )
            flash("Checklist semanal salvo com sucesso.", "success")
            return redirect(url_for("main.piloto_checklist_semanal", veiculo_id=veiculo_id, drone_id=drone_id))

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Erro ao salvar checklist (POST): {str(e)}")
            flash(f"Erro ao salvar: {str(e)}", "danger")

    return render_template(
        "piloto_checklist_semanal.html",
        equipe=equipe,
        piloto_nome=piloto_nome,
        papel_equipe=(vinculo.papel or "").lower(),
        veiculos_equipe=veiculos_equipe,
        drones_equipe=drones_equipe,
        veiculo_padrao_id=veiculo_padrao_id,
        drone_padrao_id=drone_padrao_id,
        veiculo_meta=veiculo_meta,
        drone_meta=drone_meta,
        veiculo_prefill=veiculo_prefill,
        drone_prefill=drone_prefill,
        semana_inicio=inicio_semana.strftime("%d/%m/%Y"),
        semana_fim=fim_semana.strftime("%d/%m/%Y"),
        agora=datetime.now().strftime("%d/%m/%Y %H:%M"),
    )


@bp.route("/admin/checklists/semanais", methods=["GET"], endpoint="admin_checklists_semanais")
@login_required
@roles_required("admin")
def admin_checklists_semanais():
    q = (request.args.get("q") or "").strip()
    data_inicio = (request.args.get("data_inicio") or "").strip()
    data_fim = (request.args.get("data_fim") or "").strip()

    registros = []

    query_veiculos = (
        ChecklistSemanalVeiculo.query
        .options(
            joinedload(ChecklistSemanalVeiculo.veiculo),
            joinedload(ChecklistSemanalVeiculo.piloto),
        )
        .join(Veiculos, ChecklistSemanalVeiculo.veiculo_id == Veiculos.id)
        .join(Pilotos, ChecklistSemanalVeiculo.piloto_id == Pilotos.id)
    )

    if q:
        like = f"%{q}%"
        query_veiculos = query_veiculos.filter(
            db.or_(
                Veiculos.modelo.ilike(like),
                Veiculos.placa.ilike(like),
                Veiculos.responsavel.ilike(like),
                Pilotos.nome_piloto.ilike(like),
            )
        )

    if data_inicio:
        try:
            dt_ini = datetime.strptime(data_inicio, "%Y-%m-%d")
            query_veiculos = query_veiculos.filter(ChecklistSemanalVeiculo.data_registro >= dt_ini)
        except ValueError:
            pass

    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query_veiculos = query_veiculos.filter(ChecklistSemanalVeiculo.data_registro <= dt_fim)
        except ValueError:
            pass

    registros.extend(
        _normalize_checklist_veiculo_admin(item)
        for item in query_veiculos.order_by(ChecklistSemanalVeiculo.data_registro.desc()).all()
    )

    query_drones = (
        ChecklistSemanalDrone.query
        .options(
            joinedload(ChecklistSemanalDrone.drone),
            joinedload(ChecklistSemanalDrone.piloto),
        )
        .join(Drones, ChecklistSemanalDrone.drone_id == Drones.id)
        .join(Pilotos, ChecklistSemanalDrone.piloto_id == Pilotos.id)
    )

    if q:
        like = f"%{q}%"
        query_drones = query_drones.filter(
            db.or_(
                Drones.renomacao.ilike(like),
                Drones.modelo.ilike(like),
                Drones.numero_serie.ilike(like),
                Pilotos.nome_piloto.ilike(like),
            )
        )

    if data_inicio:
        try:
            dt_ini = datetime.strptime(data_inicio, "%Y-%m-%d")
            query_drones = query_drones.filter(ChecklistSemanalDrone.data_registro >= dt_ini)
        except ValueError:
            pass

    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query_drones = query_drones.filter(ChecklistSemanalDrone.data_registro <= dt_fim)
        except ValueError:
            pass

    registros.extend(
        _normalize_checklist_drone_admin(item)
        for item in query_drones.order_by(ChecklistSemanalDrone.data_registro.desc()).all()
    )

    grupos = _group_admin_checklists_by_week(registros)

    hoje = datetime.now().date()
    inicio_semana = hoje - timedelta(days=hoje.weekday())

    totais = {
        "geral": len(grupos),
        "veiculo": sum(len(item["veiculos"]) for item in grupos),
        "drone": sum(len(item["drones"]) for item in grupos),
        "pendencias": sum(1 for item in grupos if item["falhas"] > 0),
        "semana_atual": sum(1 for item in grupos if item["semana_inicio"] == inicio_semana),
    }

    return render_template(
        "admin_checklists_semanais.html",
        grupos=grupos,
        totais=totais,
        filters={
            "q": q,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
        },
    )


@bp.route(
    "/admin/checklists/semanais/<int:piloto_id>/<string:semana_inicio>",
    methods=["GET"],
    endpoint="admin_checklist_semanal_detalhe",
)
@login_required
@roles_required("admin")
def admin_checklist_semanal_detalhe(piloto_id, semana_inicio):
    try:
        semana_inicio_date = datetime.strptime(semana_inicio, "%Y-%m-%d").date()
    except ValueError:
        abort(404)

    semana_inicio_dt = datetime.combine(semana_inicio_date, datetime.min.time())
    semana_fim_dt = semana_inicio_dt + timedelta(days=7)

    veiculos = [
        _normalize_checklist_veiculo_admin(item)
        for item in (
            ChecklistSemanalVeiculo.query
            .options(
                joinedload(ChecklistSemanalVeiculo.veiculo),
                joinedload(ChecklistSemanalVeiculo.piloto),
            )
            .filter(
                ChecklistSemanalVeiculo.piloto_id == piloto_id,
                ChecklistSemanalVeiculo.data_registro >= semana_inicio_dt,
                ChecklistSemanalVeiculo.data_registro < semana_fim_dt,
            )
            .order_by(ChecklistSemanalVeiculo.data_registro.desc())
            .all()
        )
    ]

    drones = [
        _normalize_checklist_drone_admin(item)
        for item in (
            ChecklistSemanalDrone.query
            .options(
                joinedload(ChecklistSemanalDrone.drone),
                joinedload(ChecklistSemanalDrone.piloto),
            )
            .filter(
                ChecklistSemanalDrone.piloto_id == piloto_id,
                ChecklistSemanalDrone.data_registro >= semana_inicio_dt,
                ChecklistSemanalDrone.data_registro < semana_fim_dt,
            )
            .order_by(ChecklistSemanalDrone.data_registro.desc())
            .all()
        )
    ]

    if not veiculos and not drones:
        abort(404)

    piloto_nome = "-"
    if veiculos:
        piloto_nome = veiculos[0]["piloto_nome"]
    elif drones:
        piloto_nome = drones[0]["piloto_nome"]

    ultima_movimentacao = max(
        [item["data_registro"] for item in (veiculos + drones) if item.get("data_registro")],
        default=None,
    )

    totais = {
        "veiculo": len(veiculos),
        "drone": len(drones),
        "pendencias": sum(item["falhas"] for item in (veiculos + drones)),
        "itens_ok": sum(item["itens_ok"] for item in (veiculos + drones)),
        "itens_total": sum(item["itens_total"] for item in (veiculos + drones)),
    }

    return render_template(
        "admin_checklist_semanal_detalhe.html",
        piloto_id=piloto_id,
        piloto_nome=piloto_nome,
        semana_inicio=semana_inicio_date,
        semana_fim=semana_inicio_date + timedelta(days=6),
        ultima_movimentacao=ultima_movimentacao,
        veiculos=veiculos,
        drones=drones,
        totais=totais,
    )